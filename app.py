import base64, re, io
from PIL import Image
import pytesseract
import os
import psycopg2
import re
import time
import uuid
import json
import imaplib
import email as email_lib
from email.header import decode_header
import requests
from pywebpush import webpush, WebPushException
import threading
import boto3
from botocore.config import Config
from datetime import datetime, timedelta, date, timezone

# Colombia esta siempre en UTC-5 (no tiene horario de verano) -- se usa
# un desfase FIJO en vez de ZoneInfo("America/Bogota"), porque esa
# depende de que el sistema tenga instalada la base de datos de zonas
# horarias (tzdata) -- la imagen de Railway NO la tiene por defecto, y
# eso tumbaba toda la aplicacion al arrancar (ZoneInfoNotFoundError).
# Un desfase fijo no depende de nada externo y para Colombia siempre es
# correcto (nunca cambia por horario de verano).
TZ_COLOMBIA = timezone(timedelta(hours=-5))
from flask import Flask, request, jsonify, send_file
from playwright.sync_api import sync_playwright
from pypdf import PdfWriter
from flask_cors import CORS

app = Flask(__name__)  # v54.1
CORS(app, resources={r"/*": {"origins": "*"}})

TIMEOUT = 10000
MSG_NO_MATRICULADO = "El vehiculo no se encuentra matriculado en la Secretaria de Movilidad"
AÑO_ACTUAL = str(datetime.now().year)

TWOCAPTCHA_API_KEY = os.environ.get("TWOCAPTCHA_API_KEY", "")

# --- Cloudflare R2 (almacenamiento de documentos generados, ej. FUN) ---
R2_ACCESS_KEY_ID = os.environ.get("R2_ACCESS_KEY_ID", "")
R2_SECRET_ACCESS_KEY = os.environ.get("R2_SECRET_ACCESS_KEY", "")
R2_ENDPOINT = os.environ.get("R2_ENDPOINT", "")
R2_BUCKET_NAME = os.environ.get("R2_BUCKET_NAME", "")

# --- Notificaciones Push ---
VAPID_PUBLIC_KEY  = os.environ.get("VAPID_PUBLIC_KEY", "")

# --- Proxy residencial (IPRoyal) -- se usa SOLO los sabados durante el
# monitoreo intensivo de citas de Medellin, para evitar que el sitio
# bloquee la IP fija del servidor por exceso de peticiones seguidas.
IPROYAL_HOST = os.environ.get("IPROYAL_HOST", "geo.iproyal.com")
IPROYAL_PORT = os.environ.get("IPROYAL_PORT", "12321")
IPROYAL_USER = os.environ.get("IPROYAL_USER", "")
IPROYAL_PASS = os.environ.get("IPROYAL_PASS", "")
# DataImpulse -- proxy residencial usado para el monitoreo/reserva de
# citas de Envigado, para evitar que el sitio detecte el trafico
# repetido del servidor como sospechoso y escale la dificultad del
# captcha. Host/puerto por defecto segun la documentacion de
# DataImpulse -- ajustar via variables de entorno si difieren.
DATAIMPULSE_HOST = os.environ.get("DATAIMPULSE_HOST", "gw.dataimpulse.com")
DATAIMPULSE_PORT = os.environ.get("DATAIMPULSE_PORT", "823")
DATAIMPULSE_USER = os.environ.get("DATAIMPULSE_USER", "")
DATAIMPULSE_PASS = os.environ.get("DATAIMPULSE_PASS", "")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_CLAIMS = {"sub": "mailto:soporte@tramy.app"}


def guardar_suscripcion_push(endpoint, p256dh, auth):
    """Guarda (o actualiza) una suscripcion push en la base de datos."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO push_subscriptions (endpoint, p256dh, auth)
            VALUES (%s, %s, %s)
            ON CONFLICT (endpoint) DO UPDATE SET p256dh = EXCLUDED.p256dh, auth = EXCLUDED.auth
        """, (endpoint, p256dh, auth))
        conn.commit()
        cur.close(); conn.close()
        return True
    except Exception as e:
        print(f"Error guardando suscripcion push: {e}")
        return False


def enviar_notificacion_push(titulo, cuerpo, url="/"):
    """Manda una notificacion push a TODAS las suscripciones guardadas
    (todos los dispositivos que hayan activado las notificaciones). Si
    una suscripcion ya no es valida (ej. la persona desinstalo la app o
    el navegador la bloqueo), se borra sola de la base de datos."""
    print(f"[PUSH] Intentando enviar notificacion: '{titulo}' -- '{cuerpo}'", flush=True)
    if not VAPID_PRIVATE_KEY:
        print("[PUSH] No configurado (falta VAPID_PRIVATE_KEY) -- no se envia notificacion.", flush=True)
        return

    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, endpoint, p256dh, auth FROM push_subscriptions")
        suscripciones = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        print(f"[PUSH] Error leyendo suscripciones push: {e}", flush=True)
        return

    print(f"[PUSH] {len(suscripciones)} suscripcion(es) encontrada(s) en la base de datos.", flush=True)
    if not suscripciones:
        print("[PUSH] No hay ninguna suscripcion guardada -- nadie ha activado las notificaciones, o se borraron por invalidas.", flush=True)

    payload = json.dumps({"title": titulo, "body": cuerpo, "url": url, "icon": "/icon-192.png"})

    for sub_id, endpoint, p256dh, auth in suscripciones:
        try:
            webpush(
                subscription_info={
                    "endpoint": endpoint,
                    "keys": {"p256dh": p256dh, "auth": auth}
                },
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims=dict(VAPID_CLAIMS),
            )
            print(f"[PUSH] ✓ Enviada correctamente a la suscripcion id={sub_id}.", flush=True)
        except WebPushException as e:
            print(f"[PUSH] Suscripcion id={sub_id} invalida, se borra: {e}", flush=True)
            try:
                conn2 = get_db_conn()
                cur2 = conn2.cursor()
                cur2.execute("DELETE FROM push_subscriptions WHERE id = %s", (sub_id,))
                conn2.commit()
                cur2.close(); conn2.close()
            except Exception:
                pass
        except Exception as e:
            print(f"[PUSH] Error enviando a la suscripcion id={sub_id}: {e}", flush=True)

R2_PUBLIC_BASE_URL = os.environ.get("R2_PUBLIC_BASE_URL", "")  # ej: https://docs.tramy.app

def _r2_client():
    return boto3.client(
        "s3",
        endpoint_url=R2_ENDPOINT,
        aws_access_key_id=R2_ACCESS_KEY_ID,
        aws_secret_access_key=R2_SECRET_ACCESS_KEY,
        config=Config(signature_version="s3v4"),
        region_name="auto",
    )

def subir_a_r2(ruta_local, nombre_archivo_remoto, content_type="application/pdf", nombre_descarga=None):
    """Sube un archivo a R2 y devuelve la URL publica de descarga.
    El bucket debe tener acceso publico de lectura habilitado (o un dominio
    personalizado conectado) para que la URL funcione directo en el navegador.
    Si se da 'nombre_descarga', el navegador sugiere ese nombre al guardar
    el archivo (no el nombre interno/tecnico usado en la URL)."""
    cliente = _r2_client()
    extra_args = {"ContentType": content_type}
    if nombre_descarga:
        extra_args["ContentDisposition"] = f'inline; filename="{nombre_descarga}"'
    cliente.upload_file(
        ruta_local, R2_BUCKET_NAME, nombre_archivo_remoto,
        ExtraArgs=extra_args
    )
    return f"{R2_PUBLIC_BASE_URL}/{nombre_archivo_remoto}"


def borrar_de_r2(url):
    """Borra un archivo de R2 a partir de su URL publica (la que devuelve
    subir_a_r2). Se usa para limpiar PDFs de declaraciones que ya
    vencieron y no sirven para otro dia."""
    if not url or not url.startswith(R2_PUBLIC_BASE_URL):
        return
    nombre_archivo_remoto = url[len(R2_PUBLIC_BASE_URL):].lstrip("/")
    try:
        cliente = _r2_client()
        cliente.delete_object(Bucket=R2_BUCKET_NAME, Key=nombre_archivo_remoto)
    except Exception as e:
        print(f"Error borrando de R2 ({url}): {e}")


EMTRASUR_SITE_KEY  = "6Leshn4sAAAAAIas9tkeW3vKPg0a4uYqw-7fG7Pn"
EMTRASUR_URL       = "https://sistematizacion.emtrasur.com.co/"
ANTIOQUIA_SITE_KEY = "0x4AAAAAACJy_BR2tRNN1cnv"
ANTIOQUIA_URL      = "https://www.vehiculosantioquia.com.co/impuestosweb/#/public"
ANTIOQUIA_API      = "https://www.vehiculosantioquia.com.co/raiz-backimpuestosweb/backimpuestosweb"

# ============================================================
#  TABLA DE TIPOS DE DOCUMENTO ANTIOQUIA
# ============================================================
ANTIOQUIA_TIPOS_DOCUMENTO = {
    "1":  {"abreviatura": "CC",    "nombre": "Cédula de Ciudadanía"},
    "8":  {"abreviatura": "CD",    "nombre": "Carnet Diplomático"},
    "5":  {"abreviatura": "CE",    "nombre": "Cédula de Extranjería"},
    "2":  {"abreviatura": "NIT",   "nombre": "NIT"},
    "4":  {"abreviatura": "PASAP", "nombre": "Pasaporte"},
    "29": {"abreviatura": "PPT",   "nombre": "Permiso por protección temporal"},
    "7":  {"abreviatura": "RC",    "nombre": "Registro Civil"},
    "6":  {"abreviatura": "TI",    "nombre": "Tarjeta de Identidad"},
}

# Mapa de abreviatura entrante → id numérico
ANTIOQUIA_TIPO_DOC_MAP = {
    "CC": "1", "NIT": "2", "PASAP": "4", "CE": "5",
    "TI": "6", "RC": "7", "CD": "8", "PPT": "29"
}

ANTIOQUIA_LIMITE_VIGENCIAS = 20


def get_db_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def job_actualizar(job_id, mensaje, estado='procesando', datos_parciales=None):
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        if datos_parciales is not None:
            cur.execute("""
                INSERT INTO consulta_jobs (job_id, estado, mensaje, resultado, actualizado_en)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (job_id) DO UPDATE SET estado=%s, mensaje=%s, resultado=%s, actualizado_en=NOW()
            """, (job_id, estado, mensaje, json.dumps({"parcial": datos_parciales}),
                  estado, mensaje, json.dumps({"parcial": datos_parciales})))
        else:
            cur.execute("""
                INSERT INTO consulta_jobs (job_id, estado, mensaje, actualizado_en)
                VALUES (%s, %s, %s, NOW())
                ON CONFLICT (job_id) DO UPDATE SET estado=%s, mensaje=%s, actualizado_en=NOW()
            """, (job_id, estado, mensaje, estado, mensaje))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error job: {e}")

def job_terminar(job_id, resultado):
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE consulta_jobs SET estado='listo', mensaje='Consulta finalizada.', resultado=%s, actualizado_en=NOW()
            WHERE job_id=%s
        """, (json.dumps(resultado), job_id))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error job terminar: {e}")

def job_error(job_id, mensaje_error):
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE consulta_jobs SET estado='error', mensaje=%s, actualizado_en=NOW()
            WHERE job_id=%s
        """, (mensaje_error, job_id))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error job error: {e}")


# ============================================================
#  CACHE IMPUESTOS ANTIOQUIA
# ============================================================

def _dataimpulse_proxy_config(etiqueta=""):
    """Construye la configuracion de proxy residencial de DataImpulse
    para usar con Playwright (browser.new_context(proxy=...)). Devuelve
    None si faltan credenciales (y avisa por log), para que el llamador
    pueda decidir seguir sin proxy en vez de fallar."""
    if DATAIMPULSE_USER and DATAIMPULSE_PASS:
        print(f"{etiqueta} Usando proxy residencial de DataImpulse.", flush=True)
        return {
            "server": f"http://{DATAIMPULSE_HOST}:{DATAIMPULSE_PORT}",
            "username": DATAIMPULSE_USER,
            "password": DATAIMPULSE_PASS,
        }
    print(f"{etiqueta} *** ALERTA: se pidio usar el proxy de DataImpulse, pero faltan las credenciales (DATAIMPULSE_USER/DATAIMPULSE_PASS) -- esta sesion va SIN proxy, usando la IP normal del servidor.", flush=True)
    return None


def _avaluo_declaracion_mas_reciente(data3):
    """Devuelve el avaluoComercial de la declaracion MAS RECIENTE
    (la de mayor vigencia) dentro de listaDetallePagos -- este es el
    avaluo real que aparece en el certificado (ej. 10.979.000 para la
    vigencia 2026), a diferencia de estadoCuenta.avaluoComercial, que es
    un campo GENERAL de la Gobernacion que no siempre coincide con el
    avaluo de la ultima declaracion presentada (se confirmo con un caso
    real: el campo general traia $12.639.000 mientras la declaracion
    2026 real era $10.979.000)."""
    lista = data3.get("listaDetallePagos", []) or []
    mejor = None
    mejor_vigencia = -1
    for d in lista:
        try:
            vig = int(d.get("vigencia", 0) or 0)
        except (TypeError, ValueError):
            continue
        if vig > mejor_vigencia:
            mejor_vigencia = vig
            mejor = d
    if mejor and mejor.get("avaluoComercial"):
        return mejor["avaluoComercial"]
    # Respaldo: si no hay declaraciones (caso raro), se usa el campo
    # general -- mejor un valor aproximado que nada.
    return (data3.get("estadoCuenta", {}) or {}).get("avaluoComercial", 0) or 0


def cache_antioquia_buscar(placa):
    """Busca PAZ_Y_SALVO en caché para el año actual."""
    try:
        anio_actual = datetime.now().year
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT estado, total_pagar, avaluo_comercial, retefuente, vigencia
            FROM cache_impuestos_antioquia
            WHERE placa = %s AND vigencia = %s AND estado = 'PAZ_Y_SALVO'
              AND (expira_en IS NULL OR expira_en >= NOW())
            ORDER BY creado_en DESC LIMIT 1
        """, (placa.upper(), str(anio_actual)))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return {
                "estado":     row[0],
                "total":      row[1] or 0,
                "avaluo":     row[2] or 0,
                "retefuente": row[3] or 0,
                "vigencia":   row[4],
            }
        return None
    except Exception as e:
        print(f"Error cache buscar: {e}")
        return None


def cache_antioquia_buscar_vigencia(placa, anio):
    """Busca el valor cacheado de una vigencia específica con deuda."""
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT total_pagar, avaluo_comercial, retefuente
            FROM cache_impuestos_antioquia
            WHERE placa = %s AND vigencia = %s AND estado = 'CON_DEUDA'
              AND (expira_en IS NULL OR expira_en >= NOW())
            ORDER BY creado_en DESC LIMIT 1
        """, (placa.upper(), int(anio)))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return {"total_pagar": row[0] or 0, "avaluo": row[1] or 0, "retefuente": row[2] or 0}
        return None
    except Exception as e:
        print(f"Error cache buscar vigencia: {e}")
        return None


def cache_antioquia_eliminar_vigencia(placa, anio):
    """Elimina del caché una vigencia que ya fue pagada."""
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            DELETE FROM cache_impuestos_antioquia
            WHERE placa = %s AND vigencia = %s AND estado = 'CON_DEUDA'
        """, (placa.upper(), int(anio)))
        conn.commit()
        cur.close(); conn.close()
        print(f"  → Caché eliminado para {placa} vigencia {anio} (fue pagada)")
    except Exception as e:
        print(f"Error cache eliminar vigencia: {e}")


ENVIGADO_CITAS_SEDES = {
    "Sede Principal": 5,  # idSubsede confirmado por log real (antes se tenia 1, incorrecto)
    "City Plaza": 3,
}
ENVIGADO_CITAS_ID_SERVICIO = "90"
ENVIGADO_PUNTOS_API = "https://movilidad.envigado.gov.co/backavit/avit/citas/getPuntosAtencionServiciosLowcode"
ENVIGADO_HORAS_API = "https://movilidad.envigado.gov.co/backavit/avit/citas/getHorasDisponibles"
ENVIGADO_VALIDAR_API = "https://movilidad.envigado.gov.co/backavit/avit/seguridad/preguntas/validarMostrarPreguntasSeguridad"
# Documento y placa "de prueba" -- no corresponden a un ciudadano real que
# pueda verse afectado, siguiendo la misma logica que ya usa el usuario
# manualmente (un numero que el sistema no tiene registrado).
ENVIGADO_CITAS_DOCUMENTO = "711263131313"
ENVIGADO_CITAS_TIPO_DOC = "2"
ENVIGADO_CITAS_PLACA = "RST37B"


ENVIGADO_CITAS_NOMBRES = "TRAMY"
ENVIGADO_CITAS_APELLIDOS = "MONITOR"
ENVIGADO_CITAS_EMAIL = "monitor.tramy@gmail.com"
ENVIGADO_CITAS_CELULAR = "3000000000"
# Nombre EXACTO del tramite tal como aparece en el desplegable del sitio
# (confirmado con una captura real del monitor de turnos, que mostraba
# este mismo nombre como "nombreServicio").
ENVIGADO_CITAS_TRAMITE_TEXTO = "Comprador/Vendedor - Traspaso"


def _envigado_hay_aviso_sin_agenda(page):
    """Revisa si en este momento aparece el aviso 'El trámite seleccionado
    no tiene agenda disponible para el punto de atención' en cualquier
    parte visible de la pagina. Envigado muestra este aviso en momentos
    DISTINTOS del proceso segun el caso (a veces apenas se entra, a veces
    al elegir la sede, a veces solo al final al hacer clic en 'Agregar
    servicio') -- probablemente para dificultar la automatizacion. Por
    eso se revisa este texto varias veces durante el flujo, en vez de
    confiar en un solo punto fijo."""
    try:
        texto = page.inner_text("body")
        return "no tiene agenda disponible" in texto.lower()
    except Exception:
        return False


MEDELLIN_REGISTRO_URL = "https://www.medellin.gov.co/irj/servlet/prt/portal/prtroot/pcd!3aportal_content!2fMunicipioMedellin!2fPCM!2fadmin!2froles!2fmedellin!2futilMedellin!2fauthexterna!2fSelfRegistrationExterno"

# Mapeo de los valores reales de cada <select>, confirmados leyendo el
# HTML real del formulario (no son solo el texto visible -- cada opcion
# tiene un codigo antes del guion, ej "1-Cedula de Ciudadania").
MEDELLIN_TIPO_SOCIEDAD = {"Persona Natural": "N-Persona Natural", "Persona Juridica": "J-Persona Juridica"}
MEDELLIN_TIPO_IDENTIFICACION = {
    "Cedula de ciudadania": "1-Cedula de Ciudadania",
    "Tarjeta de identidad": "2-Tarjeta de identidad",
    "Cedula de extranjeria": "3-Cedula de Extranjeria",
    "NIT": "4-NIT",
}
MEDELLIN_GENERO = {"Masculino": "m", "Femenino": "f", "Otro": "o"}


def medellin_crear_usuario(datos, usar_proxy=True):
    """Crea un usuario nuevo en el portal 'Movilidad en Linea' de la
    Alcaldia de Medellin (formulario de auto-registro). Los selectores
    estan confirmados con el HTML real del formulario (no son una
    suposicion) -- es un formulario HTML clasico con jQuery, no un SPA
    moderno. 'datos' es un dict con: tipo_sociedad, tipo_identificacion,
    numero_identificacion, nombre, apellidos, genero, email, direccion,
    telefono (todos como texto, tal como se ven en el formulario).
    Pais/Departamento/Ciudad se dejan en su valor por defecto (Colombia/
    Antioquia/Medellin), que ya vienen preseleccionados.
    'usar_proxy' (True por defecto): el sitio de Medellin ha bloqueado la
    IP fija del servidor varias veces por exceso de peticiones -- se usa
    el proxy residencial de DataImpulse por defecto para evitarlo."""
    resultado = {"exito": False, "mensaje": ""}
    etiqueta = f"[MEDELLIN-{uuid.uuid4().hex[:6]}]"  # para poder filtrar SOLO estos logs entre los de otros procesos concurrentes

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True, args=[
            "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
            "--single-process", "--no-zygote", "--disable-setuid-sandbox"
        ])
        proxy_config = _dataimpulse_proxy_config(etiqueta) if usar_proxy else None
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            viewport={"width": 1366, "height": 900},
            proxy=proxy_config,
        )
        page = context.new_page()

        # DIAGNOSTICO TEMPORAL -- se capturan todas las peticiones de red
        # (especialmente las de validacion, tipo AJAX/XHR) para entender
        # que esta pasando exactamente con la validacion del numero de
        # identificacion, que nunca pasa.
        peticiones_capturadas = []
        def _capturar_peticion(response):
            try:
                if response.request.resource_type in ("xhr", "fetch"):
                    cuerpo_pedido = None
                    cuerpo_respuesta = None
                    try:
                        cuerpo_pedido = response.request.post_data
                    except Exception:
                        pass
                    try:
                        cuerpo_respuesta = response.text()[:1500]
                    except Exception:
                        pass
                    peticiones_capturadas.append({
                        "url": response.url,
                        "status": response.status,
                        "metodo": response.request.method,
                        "cuerpo_pedido": cuerpo_pedido,
                        "cuerpo_respuesta": cuerpo_respuesta,
                    })
            except Exception:
                pass
        page.on("response", _capturar_peticion)

        MAX_INTENTOS_COMPLETOS = 3  # si el formulario se queda pegado, se recarga la pagina y se vuelve a llenar todo desde cero, hasta esta cantidad de veces
        for intento_completo in range(MAX_INTENTOS_COMPLETOS):
          try:
            if intento_completo > 0:
                print(f"{etiqueta} *** Reintento completo {intento_completo+1}/{MAX_INTENTOS_COMPLETOS}: recargando la pagina y llenando el formulario desde cero...", flush=True)
                peticiones_capturadas.clear()

            page.goto(MEDELLIN_REGISTRO_URL, wait_until="load", timeout=60000)
            page.wait_for_timeout(2000)

            # 1. Tipo de Sociedad -- esto dispara (via jQuery) que se
            # muestren/oculten otros campos (Tipo de Entidad para
            # Juridica, Apellidos/Genero para Natural).
            valor_sociedad = MEDELLIN_TIPO_SOCIEDAD.get(datos["tipo_sociedad"], "N-Persona Natural")
            page.select_option("#cTipoSociedad", value=valor_sociedad)
            page.wait_for_timeout(500)

            # 2. Tipo de Entidad -- solo aplica si es Persona Juridica
            if datos["tipo_sociedad"] == "Persona Juridica" and datos.get("tipo_entidad"):
                page.select_option("#cTipoEntidad", value=datos["tipo_entidad"])
                page.wait_for_timeout(300)

            # 3. Tipo y Numero de Identificacion
            valor_tipo_id = MEDELLIN_TIPO_IDENTIFICACION.get(datos["tipo_identificacion"], "1-Cedula de Ciudadania")
            page.select_option("#cTipoIdentificacion", value=valor_tipo_id)
            # Se escribe caracter por caracter (no fill() de un solo golpe)
            # porque este campo no dispara ninguna peticion al servidor
            # para validarse -- sospecha: usa un evento tipo "keyup" que
            # solo se dispara escribiendo tecla por tecla, como lo haria
            # una persona real.
            page.click("#cNumeroIdentificacion")
            page.keyboard.type(datos["numero_identificacion"], delay=80)
            # El sitio valida el numero de identificacion en tiempo real
            # (por eso la casilla "tdOk" al lado) -- se espera a que esa
            # validacion termine antes de seguir (puede tardar, es una
            # consulta al servidor).
            page.wait_for_timeout(3000)

            # 4. Nombre / Razon Social
            page.fill("#cNombre", datos["nombre"])

            # 5. Apellidos y Genero -- solo aplican para Persona Natural
            if datos["tipo_sociedad"] == "Persona Natural":
                page.fill("#cApellidos", datos["apellidos"])
                valor_genero = MEDELLIN_GENERO.get(datos["genero"], "m")
                page.select_option("#cGenero", value=valor_genero)

            # 6. Correo y Direccion
            page.fill("#cCorreoElectronico", datos["email"])
            page.fill("#cDireccionResidencia", datos["direccion"])

            # 7. Aceptar politicas de uso (obligatorio) y autorizar
            # notificaciones (opcional, pero conviene para que lleguen
            # avisos del tramite). El <input type="checkbox"> real esta
            # oculto por CSS (checkbox personalizado con su propio
            # estilo) -- se hace clic en la etiqueta visible en vez del
            # checkbox directamente, que es lo que ve y usa una persona.
            # OJO: se hace ESTO primero, y Telefono/Ciudad AL FINAL --
            # se detecto que hacer clic en estos checkboxes reseteaba el
            # campo Ciudad si se seleccionaba antes.
            # El checkbox "Acepto" tiene una etiqueta larga con varios
            # links de por medio (politicas de uso, proteccion de
            # datos) -- el clic normal a veces no cae en el lugar
            # correcto de esa etiqueta y no llega a marcar el checkbox.
            # Como el checkbox real esta oculto por CSS (sin tamaño en
            # pantalla), tampoco se puede "forzar" un clic ahi -- se
            # marca directo por JavaScript, disparando los eventos que
            # el sitio pueda estar escuchando.
            page.evaluate("""() => {
                function marcar(id) {
                    var el = document.getElementById(id);
                    if (!el) return;
                    el.checked = true;
                    el.dispatchEvent(new Event('click', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                }
                marcar('cAcepto');
                marcar('cNotifica');
            }""")
            page.wait_for_timeout(500)

            # DIAGNOSTICO -- confirmar que los checkboxes realmente
            # quedaron marcados (el diagnostico de validacion de campos
            # no los revisa, porque su fila no tiene la misma estructura
            # que los demas campos).
            try:
                estado_checks = page.evaluate("""() => {
                    return {
                        acepto: document.getElementById('cAcepto').checked,
                        notifica: document.getElementById('cNotifica').checked,
                    };
                }""")
                print(f"{etiqueta} === Estado real de los checkboxes: {estado_checks}", flush=True)
            except Exception as e_check:
                print(f"{etiqueta} No se pudo revisar los checkboxes: {e_check}", flush=True)

            # 8. Telefono -- tambien parece validarse contra el servidor
            # (igual que el numero de identificacion), asi que se le da
            # tiempo de sobra despues de escribirlo.
            # Mismo motivo que Numero de Identificacion: se escribe
            # caracter por caracter para disparar el evento de tecleo que
            # el sitio necesita para marcarlo como valido.
            page.click("#cTelefono")
            page.keyboard.type(datos["telefono"], delay=80)
            if datos.get("celular"):
                page.fill("#cCelular", datos["celular"])
            page.wait_for_timeout(3000)

            # 9. Pais, Departamento y Ciudad -- se seleccionan explicito
            # (Colombia / Antioquia / Medellin), en vez de confiar en que
            # ya vienen preseleccionados asi por defecto en el HTML.
            # OJO: se detecto que Ciudad se resetea solo unos ~500ms
            # despues de seleccionarla, sin importar el orden con los
            # checkboxes -- sospecha: el propio sitio, al cambiar
            # Departamento, dispara con un poco de retraso un reinicio
            # automatico de Ciudad (patron comun de "selects en cascada").
            # Se prueba dejando que ese reinicio (si existe) ocurra ANTES,
            # dejando Ciudad como lo ULTIMO que se toca del formulario.
            page.select_option("#cPais", value="CO")
            page.select_option("#cDepartamento", value="05-ANTIOQUIA")
            page.wait_for_timeout(2500)  # dejar que cualquier reinicio automatico de Ciudad ya ocurra aqui

            try:
                v0 = page.evaluate("() => document.querySelector('#cCiudad').value")
                print(f"{etiqueta} === CIUDAD antes de seleccionarla (tras esperar por Departamento): '{v0}'", flush=True)
            except Exception as e0:
                print(f"{etiqueta} No se pudo leer ciudad (momento 0): {e0}", flush=True)

            page.select_option("#cCiudad", value="05001-MEDELLÍN")

            # DIAGNOSTICO TEMPORAL -- revisar el valor de Ciudad en varios
            # momentos seguidos, para ver exactamente cuando se resetea.
            try:
                v1 = page.evaluate("() => document.querySelector('#cCiudad').value")
                print(f"{etiqueta} === CIUDAD justo despues de select_option: '{v1}'", flush=True)
            except Exception as e1:
                print(f"{etiqueta} No se pudo leer ciudad (momento 1): {e1}", flush=True)

            page.wait_for_timeout(500)

            try:
                v2 = page.evaluate("() => document.querySelector('#cCiudad').value")
                print(f"{etiqueta} === CIUDAD despues de esperar 500ms: '{v2}'", flush=True)
            except Exception as e2:
                print(f"{etiqueta} No se pudo leer ciudad (momento 2): {e2}", flush=True)

            # Se le da mas tiempo de sobra a la validacion del numero de
            # identificacion tambien, por si seguia pendiente.
            page.wait_for_timeout(2000)

            try:
                v3 = page.evaluate("() => document.querySelector('#cCiudad').value")
                print(f"{etiqueta} === CIUDAD despues de esperar 2000ms mas: '{v3}'", flush=True)
            except Exception as e3:
                print(f"{etiqueta} No se pudo leer ciudad (momento 3): {e3}", flush=True)

            print(f"{etiqueta} === DIAGNOSTICO: peticiones de red (xhr/fetch) capturadas hasta ahora ===", flush=True)
            for p in peticiones_capturadas:
                print(etiqueta, p, flush=True)
            print(f"{etiqueta} === FIN peticiones capturadas ===", flush=True)

            # DIAGNOSTICO -- antes de dar clic en "Siguiente", se revisa
            # si algun campo quedo marcado como invalido (la casilla
            # tdOk al lado de cada campo se llena con una imagen "ok"
            # cuando el campo pasa la validacion del sitio).
            try:
                estado_validacion = page.evaluate("""() => {
                    var filas = document.querySelectorAll('.tableContFormRegistro tr');
                    var resultado = [];
                    filas.forEach(function(fila){
                        var tdOk = fila.querySelector('.tdOk');
                        var label = fila.querySelector('.tdIzq');
                        var campoInput = fila.querySelector('input, select');
                        if (label) {
                            resultado.push({
                                campo: label.innerText.trim(),
                                valorActual: campoInput ? campoInput.value : null,
                                tieneOk: tdOk ? !!tdOk.querySelector('img') : null,
                                filaHtmlCompleto: fila.innerHTML.substring(0, 400)
                            });
                        }
                    });
                    return resultado;
                }""")
                print(f"{etiqueta} === DIAGNOSTICO: estado de validacion de cada campo antes de Siguiente ===", flush=True)
                for c in estado_validacion:
                    print(etiqueta, c, flush=True)
                print(f"{etiqueta} === FIN DIAGNOSTICO validacion ===", flush=True)
            except Exception as e_diag:
                print(f"{etiqueta} No se pudo revisar el estado de validacion: {e_diag}", flush=True)

            # 9. Clic en "Siguiente" -- CON REINTENTOS. Confirmado (con
            # una persona real llenando el formulario a mano sin ningun
            # problema) que el sitio SI funciona bien -- el problema es
            # que la automatizacion a veces le da clic demasiado rapido,
            # antes de que alguna validacion interna del sitio termine
            # de asentarse (algo que un humano no hace, porque
            # naturalmente se toma mas tiempo entre acciones). Por eso:
            # se espera un poco extra antes del primer clic, y si
            # despues de darle clic la pagina no cambio de contenido
            # (senal de que se quedo pegada en el mismo paso), se
            # reintenta con esperas cada vez mas largas.
            page.wait_for_timeout(2000)  # dar tiempo de sobra a que termine de asentarse la validacion de todos los campos
            contenido_antes_siguiente = page.inner_text("body")
            esperas_reintento = [3000, 5000, 8000]  # reintentos rapidos aqui -- si esto no alcanza, se recarga TODO desde cero (ver mas abajo)
            for intento_siguiente, espera_ms in enumerate(esperas_reintento):
                if page.locator("#inpBtnNext").count() == 0:
                    print(f"{etiqueta} El boton 'Siguiente' ya no existe -- la pagina avanzo.", flush=True)
                    break
                try:
                    page.click("#inpBtnNext", force=True, timeout=5000)
                except Exception as e_click_sig:
                    print(f"{etiqueta} No se pudo hacer clic en Siguiente (probablemente deshabilitado temporalmente): {e_click_sig}", flush=True)
                page.wait_for_timeout(espera_ms)
                contenido_despues_siguiente = page.inner_text("body")
                if contenido_despues_siguiente != contenido_antes_siguiente:
                    print(f"{etiqueta} Clic en Siguiente avanzo la pagina (intento {intento_siguiente+1}, espero {espera_ms}ms).", flush=True)
                    break
                print(f"{etiqueta} Clic en Siguiente NO parece haber avanzado (intento {intento_siguiente+1}/{len(esperas_reintento)}, espero {espera_ms}ms), reintentando...", flush=True)
            else:
                if intento_completo < MAX_INTENTOS_COMPLETOS - 1:
                    print(f"{etiqueta} *** Clic en Siguiente no avanzo despues de {len(esperas_reintento)} intentos -- se recarga la pagina y se reintenta el registro COMPLETO desde cero.", flush=True)
                    continue
                print(f"{etiqueta} *** Clic en Siguiente no avanzo despues de {len(esperas_reintento)} intentos, y ya se agotaron los {MAX_INTENTOS_COMPLETOS} reintentos completos -- se continua de todas formas.", flush=True)

            # DIAGNOSTICO -- que paso despues de dar clic en Siguiente
            # (puede ser un segundo paso del formulario, un mensaje de
            # error, o una confirmacion de que el registro se completo).
            print(f"{etiqueta} === DIAGNOSTICO: despues de clic en Siguiente ===", flush=True)
            print(f"{etiqueta} URL actual:", page.url, flush=True)
            try:
                print(f"{etiqueta} Texto visible de la pagina (primeros 2000 caracteres):", flush=True)
                print(etiqueta, page.inner_text("body")[:2000], flush=True)
            except Exception as e_txt:
                print(f"{etiqueta} No se pudo leer el texto de la pagina:", e_txt, flush=True)
            print(f"{etiqueta} === FIN DIAGNOSTICO despues de Siguiente ===", flush=True)

            # 10. SEGUNDO PASO -- preguntas de Verdadero/Falso. IMPORTANTE:
            # estas preguntas son FIJAS (no se despliegan preguntas nuevas
            # segun las respuestas que se den), pero la CANTIDAD que
            # aparece varia entre 0 y 3 segun cada registro -- a veces no
            # sale ninguna, a veces solo 1 o 2. Por eso el codigo no
            # asume ids fijos (radio0/radio1/etc.), sino que lee el texto
            # de cada pregunta que SI aparezca y responde segun ese
            # texto.
            try:
                hay_preguntas = page.evaluate("() => document.querySelectorAll('.divPregunta').length > 0")
                if not hay_preguntas:
                    print(f"{etiqueta} No aparecio ninguna pregunta de Verdadero/Falso en este registro.", flush=True)
                    print(f"{etiqueta} === DIAGNOSTICO: estado de la pagina sin preguntas ===", flush=True)
                    print(f"{etiqueta} URL actual:", page.url, flush=True)
                    try:
                        print(etiqueta, page.inner_text("body")[:2000], flush=True)
                    except Exception as e_txt3:
                        print(f"{etiqueta} No se pudo leer el texto de la pagina:", e_txt3, flush=True)
                    # Si hay un boton visible que parezca de continuar/
                    # confirmar/registrar (distinto a "Siguiente", que ya
                    # se uso), se reporta para saber si falta darle clic.
                    try:
                        botones_visibles = page.evaluate("""() => {
                            var botones = document.querySelectorAll('input[type=button], input[type=submit], button');
                            var resultado = [];
                            botones.forEach(function(b) {
                                var estilo = window.getComputedStyle(b);
                                if (estilo.display !== 'none' && estilo.visibility !== 'hidden') {
                                    resultado.push({id: b.id, valor: b.value || b.innerText, disabled: b.disabled});
                                }
                            });
                            return resultado;
                        }""")
                        print(f"{etiqueta} === Botones visibles en la pagina: {botones_visibles} ===", flush=True)
                    except Exception as e_btns:
                        print(f"{etiqueta} No se pudo revisar los botones visibles: {e_btns}", flush=True)
                    print(f"{etiqueta} === FIN DIAGNOSTICO sin preguntas ===", flush=True)
                else:
                    preguntas = page.evaluate("""() => {
                        var divs = document.querySelectorAll('.divPregunta');
                        var resultado = [];
                        divs.forEach(function(div) {
                            var ul = div.nextElementSibling;
                            var nombreRadio = null;
                            if (ul) {
                                var primerInput = ul.querySelector('input[type=radio]');
                                if (primerInput) nombreRadio = primerInput.name;
                            }
                            resultado.push({texto: div.innerText.trim(), nombreRadio: nombreRadio});
                        });
                        return resultado;
                    }""")
                    print(f"{etiqueta} === Preguntas detectadas en este registro: {preguntas} ===", flush=True)

                    perfil = (datos.get("perfil_usuario") or "").strip().lower()  # "propietario" o "comprador"
                    placa_propietario = (datos.get("placa_propietario") or "").strip().upper()

                    for p in preguntas:
                        texto = p["texto"].upper()
                        nombre_radio = p["nombreRadio"]
                        if not nombre_radio:
                            print(f"{etiqueta} Pregunta sin radios detectados (revisar a mano): {p['texto']}", flush=True)
                            continue

                        if "ACUERDOS DE PAGO" in texto:
                            respuesta = "Falso"
                        elif "VEHICULOS REGISTRADOS" in texto or "VEHÍCULOS REGISTRADOS" in texto:
                            respuesta = "Verdadero" if perfil == "propietario" else "Falso"
                        elif "PROPIEDAD" in texto and ("CUAL" in texto or "CÚAL" in texto or "CUÁL" in texto):
                            # "¿Cual de los siguientes vehiculos es de su
                            # propiedad?" -- las opciones son las placas
                            # REALES que el sitio ya tiene asociadas a esa
                            # cedula (varian por persona) mas un "N/A".
                            # Comprador -> N/A. Propietario -> la placa
                            # que coincida con placa_propietario.
                            respuesta = "N/A" if perfil != "propietario" else placa_propietario
                        else:
                            print(f"{etiqueta} *** Pregunta NO reconocida (no se responde, revisar a mano): {p['texto']}", flush=True)
                            continue

                        page.evaluate(f"""() => {{
                            var opciones = document.querySelectorAll('input[name="{nombre_radio}"]');
                            opciones.forEach(function(op) {{
                                if (op.value === "{respuesta}") {{
                                    op.checked = true;
                                    op.dispatchEvent(new Event('click', {{bubbles: true}}));
                                    op.dispatchEvent(new Event('change', {{bubbles: true}}));
                                }}
                            }});
                        }}""")
                        print(f"{etiqueta} Pregunta '{p['texto'][:60]}...' respondida: {respuesta}", flush=True)

                    page.wait_for_timeout(500)

                    # Confirmar que las respuestas realmente quedaron
                    # marcadas (el mismo problema que tuvimos antes con
                    # los checkboxes de Acepto/Notifica -- el clic en la
                    # etiqueta a veces no marca el radio de verdad).
                    try:
                        estado_radios_final = page.evaluate("""() => {
                            var resultado = {};
                            document.querySelectorAll('input[type=radio]:checked').forEach(function(r) {
                                resultado[r.name] = r.value;
                            });
                            return resultado;
                        }""")
                        print(f"{etiqueta} === Estado real de las respuestas marcadas: {estado_radios_final} ===", flush=True)
                    except Exception as e_check_radios:
                        print(f"{etiqueta} No se pudo confirmar el estado de las respuestas: {e_check_radios}", flush=True)

                    # DIAGNOSTICO -- volcar el HTML completo de la zona de
                    # preguntas, por si aparecio la pregunta de la placa
                    # (o cualquier otra que no se haya visto antes), para
                    # poder revisar su estructura exacta.
                    try:
                        html_preguntas = page.evaluate("""() => {
                            var cont = document.querySelector('.divContPreguntas');
                            return cont ? cont.outerHTML : null;
                        }""")
                        print(f"{etiqueta} === HTML completo de la zona de preguntas ===", flush=True)
                        print(etiqueta, html_preguntas, flush=True)
                    except Exception as e_html:
                        print(f"{etiqueta} No se pudo volcar el HTML de preguntas: {e_html}", flush=True)

                    # Clic en "Validar respuestas" (si existe el boton --
                    # puede que la pantalla sea distinta si hay una
                    # pregunta de placa sin responder). CON REINTENTOS,
                    # igual que "Siguiente" -- el sitio a veces no
                    # avanza al primer clic.
                    if page.locator("#inpBtnQ").count() > 0:
                        try:
                            disabled_antes = page.evaluate("() => document.querySelector('#inpBtnQ').disabled")
                            print(f"{etiqueta} Boton 'Validar respuestas' deshabilitado ANTES del clic: {disabled_antes}", flush=True)
                        except Exception:
                            pass

                        page.wait_for_timeout(2000)  # dar tiempo de sobra a que se asienten las respuestas marcadas
                        contenido_antes_validar = page.inner_text("body")
                        esperas_reintento_validar = [3000, 5000, 8000]
                        for intento_validar, espera_ms_v in enumerate(esperas_reintento_validar):
                            # Si el boton ya no existe en la pagina, es
                            # señal de que SI avanzo (la pantalla cambio
                            # de verdad), no hace falta seguir intentando.
                            if page.locator("#inpBtnQ").count() == 0:
                                print(f"{etiqueta} El boton 'Validar respuestas' ya no existe -- la pagina avanzo.", flush=True)
                                break
                            # force=True porque el boton a veces queda
                            # deshabilitado brevemente mientras el sitio
                            # revisa las respuestas -- sin esto, Playwright
                            # se queda hasta 30 segundos esperando a que se
                            # habilite y termina fallando con timeout.
                            try:
                                page.click("#inpBtnQ", force=True, timeout=5000)
                            except Exception as e_click_validar:
                                print(f"{etiqueta} No se pudo hacer clic (boton probablemente deshabilitado temporalmente): {e_click_validar}", flush=True)
                            page.wait_for_timeout(espera_ms_v)
                            contenido_despues_validar = page.inner_text("body")
                            if contenido_despues_validar != contenido_antes_validar:
                                print(f"{etiqueta} Clic en Validar respuestas avanzo la pagina (intento {intento_validar+1}, espero {espera_ms_v}ms).", flush=True)
                                break
                            print(f"{etiqueta} Clic en Validar respuestas NO parece haber avanzado (intento {intento_validar+1}/{len(esperas_reintento_validar)}, espero {espera_ms_v}ms), reintentando...", flush=True)
                        else:
                            print(f"{etiqueta} *** Clic en Validar respuestas no avanzo despues de {len(esperas_reintento_validar)} intentos -- se continua de todas formas.", flush=True)

                        try:
                            disabled_despues = page.evaluate("() => document.querySelector('#inpBtnQ') ? document.querySelector('#inpBtnQ').disabled : 'boton ya no existe'")
                            print(f"{etiqueta} Boton 'Validar respuestas' deshabilitado DESPUES del clic: {disabled_despues}", flush=True)
                        except Exception:
                            pass

                        print(f"{etiqueta} === DIAGNOSTICO: despues de Validar respuestas ===", flush=True)
                        print(f"{etiqueta} URL actual:", page.url, flush=True)
                        try:
                            print(etiqueta, page.inner_text("body")[:2000], flush=True)
                        except Exception as e_txt2:
                            print(f"{etiqueta} No se pudo leer el texto de la pagina:", e_txt2, flush=True)
                        print(f"{etiqueta} === FIN DIAGNOSTICO despues de Validar respuestas ===", flush=True)
            except Exception as e_paso2:
                print(f"{etiqueta} Error revisando/respondiendo las preguntas: {e_paso2}", flush=True)

            resultado["exito"] = True
            resultado["mensaje"] = "Formulario enviado -- revisa los logs para confirmar el resultado real (puede haber un tercer paso)."
            break  # exito -- no hace falta reintentar el registro completo de nuevo

          except Exception as e:
            print(f"{etiqueta} Error en el flujo de Playwright para registro Medellin: {e}", flush=True)
            resultado["mensaje"] = str(e)
            try:
                print(f"{etiqueta} === DIAGNOSTICO: texto visible al momento del error ===", flush=True)
                print(etiqueta, page.inner_text("body")[:2000], flush=True)
            except Exception:
                pass
            if intento_completo < MAX_INTENTOS_COMPLETOS - 1:
                print(f"{etiqueta} *** Se reintenta el registro completo desde cero tras este error.", flush=True)
                continue
        # El navegador se cierra UNA sola vez, despues de todos los
        # intentos completos (no dentro del bucle, para no cerrarlo a
        # mitad de un reintento -- "continue" adentro del bucle deja el
        # mismo navegador/pagina abiertos para el siguiente intento).
        context.close(); browser.close()

    return resultado


def medellin_leer_credenciales_temporales(email_cuenta, password_app_email, cedula, minutos_maximo_espera=5, etiqueta=""):
    """Se conecta por IMAP a la bandeja de correo (Gmail, usando una
    'Contraseña de aplicacion', no la contraseña normal de la cuenta) y
    busca el correo de activacion que manda la Alcaldia de Medellin al
    completar el registro. Ese correo trae USUARIO y CONTRASEÑA
    temporales en texto plano dentro del cuerpo -- se extraen con
    expresiones regulares.

    Reintenta cada 15 segundos hasta 'minutos_maximo_espera', porque el
    correo puede tardar un poco en llegar despues del registro.

    Devuelve (usuario, password_temporal) o (None, None) si no se
    encontro el correo a tiempo."""
    intentos_maximos = max(1, (minutos_maximo_espera * 60) // 15)

    for intento in range(intentos_maximos):
        try:
            imap = imaplib.IMAP4_SSL("imap.gmail.com")
            imap.login(email_cuenta, password_app_email)
            imap.select("inbox")

            # Buscar correos recientes de la Alcaldia de Medellin (por
            # remitente), sin filtrar por leido/no leido -- puede haber
            # varios registros seguidos en la misma cuenta de correo.
            status, datos_busqueda = imap.search(None, 'FROM "medellin.gov.co"')
            if status != "OK":
                imap.logout()
                time.sleep(15)
                continue

            ids_correos = datos_busqueda[0].split()
            # Revisar del mas reciente al mas antiguo
            for id_correo in reversed(ids_correos[-10:]):  # solo los ultimos 10, para no revisar toda la bandeja
                status, datos_correo = imap.fetch(id_correo, "(RFC822)")
                if status != "OK":
                    continue
                mensaje = email_lib.message_from_bytes(datos_correo[0][1])

                cuerpo = ""
                if mensaje.is_multipart():
                    for parte in mensaje.walk():
                        tipo = parte.get_content_type()
                        if tipo in ("text/plain", "text/html"):
                            try:
                                cuerpo += parte.get_payload(decode=True).decode(parte.get_content_charset() or "utf-8", errors="ignore")
                            except Exception:
                                pass
                else:
                    try:
                        cuerpo = mensaje.get_payload(decode=True).decode(mensaje.get_content_charset() or "utf-8", errors="ignore")
                    except Exception:
                        cuerpo = str(mensaje.get_payload())

                # Confirmar que este correo es el de ESTA cedula
                # especifica (por si hay varios registros mezclados en
                # la misma bandeja).
                if cedula not in cuerpo:
                    continue

                m_usuario = re.search(r'Usuario:\s*([^\s<]+)', cuerpo)
                m_password = re.search(r'Contrase[ñn]a:\s*([^\s<]+)', cuerpo)
                if m_usuario and m_password:
                    usuario_temp = m_usuario.group(1).strip()
                    password_temp = m_password.group(1).strip()
                    imap.logout()
                    print(f"{etiqueta} Credenciales temporales encontradas en el correo para cedula {cedula}.", flush=True)
                    return usuario_temp, password_temp

            imap.logout()
        except Exception as e:
            print(f"{etiqueta} Error leyendo correo (intento {intento+1}/{intentos_maximos}): {e}", flush=True)

        time.sleep(15)

    print(f"{etiqueta} No se encontro el correo de activacion despues de {minutos_maximo_espera} minutos.", flush=True)
    return None, None


def medellin_activar_cuenta(usuario_temporal, password_temporal, nueva_password, usar_proxy=False):
    """Inicia sesion con las credenciales TEMPORALES que manda el correo
    de activacion, y completa el cambio de contraseña obligatorio que
    pide el sitio la primera vez. Reutiliza el mismo patron de login
    real por Playwright que ya usamos para revisar citas.
    'usar_proxy': si es True, la conexion pasa por el proxy residencial
    de IPRoyal en vez de la IP fija del servidor -- util porque el sitio
    de Medellin ha bloqueado esa IP fija varias veces por exceso de
    peticiones."""
    etiqueta = f"[MEDELLIN-ACTIVAR-{uuid.uuid4().hex[:6]}]"
    resultado = {"exito": False, "mensaje": ""}

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        proxy_config = None
        if usar_proxy:
            if IPROYAL_USER and IPROYAL_PASS:
                proxy_config = {
                    "server": f"http://{IPROYAL_HOST}:{IPROYAL_PORT}",
                    "username": IPROYAL_USER,
                    "password": IPROYAL_PASS,
                }
                print(f"{etiqueta} Usando proxy residencial de IPRoyal para esta consulta.", flush=True)
            else:
                print(f"{etiqueta} *** ALERTA: se pidio usar el proxy, pero faltan las credenciales de IPRoyal en las variables de entorno (IPROYAL_USER/IPROYAL_PASS) -- esta consulta va SIN proxy, usando la IP normal del servidor.", flush=True)
        context = browser.new_context(proxy=proxy_config)
        page = context.new_page()

        try:
            print(f"{etiqueta} Iniciando sesion con credenciales temporales (usuario: {usuario_temporal})...", flush=True)
            page.goto(MEDELLIN_LOGIN_URL, timeout=30000)
            page.fill('input[name="user"]', usuario_temporal)
            page.fill('input[name="passw"]', password_temporal)
            with page.expect_navigation(timeout=30000):
                page.press('input[name="passw"]', "Enter")
            page.wait_for_timeout(3000)

            print(f"{etiqueta} === DIAGNOSTICO: pagina despues del login con credenciales temporales ===", flush=True)
            print(f"{etiqueta} URL actual:", page.url, flush=True)
            try:
                print(etiqueta, page.inner_text("body")[:3000], flush=True)
            except Exception as e_txt:
                print(f"{etiqueta} No se pudo leer el texto de la pagina:", e_txt, flush=True)

            # DIAGNOSTICO -- volcar el HTML completo, ya que todavia no
            # conocemos la estructura exacta de la pantalla de cambio de
            # contraseña (nombres de los campos, boton, etc.).
            try:
                html_completo = page.content()
                print(f"{etiqueta} === HTML completo de la pagina (para programar el cambio de contraseña) ===", flush=True)
                print(etiqueta, html_completo[:8000], flush=True)
            except Exception as e_html:
                print(f"{etiqueta} No se pudo volcar el HTML: {e_html}", flush=True)

            resultado["exito"] = True
            resultado["mensaje"] = "Login con credenciales temporales realizado -- revisa los logs para ver la pantalla de cambio de contraseña y terminar de programarla."
            resultado["nueva_password_a_usar"] = nueva_password

        except Exception as e:
            print(f"{etiqueta} Error activando cuenta: {e}", flush=True)
            resultado["mensaje"] = str(e)
            try:
                print(etiqueta, page.inner_text("body")[:2000], flush=True)
            except Exception:
                pass
        finally:
            context.close(); browser.close()

    return resultado



MEDELLIN_LOGIN_URL = "https://www.medellin.gov.co/irj/servlet/prt/portal/prtroot/pcd!3aportal_content!2fMunicipioMedellin!2fPCM!2fadmin!2froles!2fmedellin!2futilMedellin!2fauth"
MEDELLIN_INICIO_SESION_URL = "https://www.medellin.gov.co/portal-movilidad/index.html#/inicio-sesion"
MEDELLIN_AVIT_API = "https://www.medellin.gov.co/backavit/avit"
MEDELLIN_SERVICIO_TRASPASO = "1"  # confirmado con un HAR real: idServicio=1 -> nombreServicio="Traspaso"


def medellin_hay_citas_disponibles(usuario, password, placa, id_servicio=MEDELLIN_SERVICIO_TRASPASO, sede_deseada=None, usar_proxy=False):
    """Revisa si hay citas disponibles para un servicio (por defecto
    'Traspaso') en el portal 'Movilidad en Linea' de Medellin. A
    diferencia de Envigado, este portal EXIGE iniciar sesion antes de
    poder consultar nada. La secuencia completa de peticiones (incluidos
    los nombres exactos de cada endpoint y sus payloads) se confirmo con
    un HAR real capturado por el usuario, asi que se replican las
    llamadas directamente (dentro de una sesion real de Playwright, para
    que las cookies de la sesion se manejen igual que en un navegador
    real) en vez de interactuar con la interfaz paso a paso.
    'usuario' es el numero de documento con el que se inicia sesion (el
    campo 'username' del login coincide con el numero de documento).
    'sede_deseada' (opcional): si se manda, solo se avisa cuando ESA
    sede especifica tiene citas -- las demas sedes se ignoran por
    completo (ni siquiera se revisan). Coincide por texto parcial, sin
    distinguir mayusculas/minusculas (ej. 'sao paulo' coincide con
    'Punto de atención Sao Paulo'). Si se deja vacio, se revisan todas
    las sedes como antes.
    'usar_proxy' (opcional): si es True, la conexion pasa por el proxy
    residencial de IPRoyal (IP distinta en cada peticion) en vez de la
    IP fija del servidor -- pensado para las ventanas de monitoreo muy
    frecuente (ej. cada 30 segundos), donde el sitio de Medellin puede
    bloquear la IP fija por exceso de peticiones seguidas.
    Devuelve una tupla (hay_citas: bool, detalle: dict|None)."""
    etiqueta = f"[MEDELLIN-CITAS-{uuid.uuid4().hex[:6]}]"

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True, args=[
            "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
            "--single-process", "--no-zygote", "--disable-setuid-sandbox"
        ])
        proxy_config = None
        if usar_proxy:
            if IPROYAL_USER and IPROYAL_PASS:
                proxy_config = {
                    "server": f"http://{IPROYAL_HOST}:{IPROYAL_PORT}",
                    "username": IPROYAL_USER,
                    "password": IPROYAL_PASS,
                }
                print(f"{etiqueta} Usando proxy residencial de IPRoyal para esta consulta.", flush=True)
            else:
                print(f"{etiqueta} *** ALERTA: se pidio usar el proxy, pero faltan las credenciales de IPRoyal en las variables de entorno (IPROYAL_USER/IPROYAL_PASS) -- esta consulta va SIN proxy, usando la IP normal del servidor.", flush=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
            viewport={"width": 1366, "height": 900},
            proxy=proxy_config,
        )
        page = context.new_page()

        # Los proxies residenciales rotan a una IP distinta (de una
        # persona real) en cada conexion nueva -- algunas de esas IPs
        # fallan o estan desconectadas, sin que eso signifique que el
        # proxy este mal configurado. Si la PRIMERA carga de pagina falla
        # por el tunel del proxy, se cierra el contexto y se abre uno
        # NUEVO (que debería tocar una IP distinta al azar), hasta 3
        # veces, antes de rendirse.
        ERRORES_TUNEL_PROXY = ("ERR_TUNNEL_CONNECTION_FAILED", "ERR_PROXY_CONNECTION_FAILED", "ERR_PROXY_AUTH_UNSUPPORTED")
        for intento_proxy in range(3):
            try:
                page.goto(MEDELLIN_INICIO_SESION_URL, wait_until="load", timeout=45000)
                break  # la pagina cargo bien, no hace falta reintentar
            except Exception as e_goto_inicial:
                if proxy_config and any(err in str(e_goto_inicial) for err in ERRORES_TUNEL_PROXY) and intento_proxy < 2:
                    print(f"{etiqueta} Fallo el tunel del proxy (intento {intento_proxy+1}/3) -- probando con una IP nueva...", flush=True)
                    context.close()
                    context = browser.new_context(
                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
                        viewport={"width": 1366, "height": 900},
                        proxy=proxy_config,
                    )
                    page = context.new_page()
                    continue
                raise  # no es un error de tunel, o ya se agotaron los intentos -- se deja que falle normal

        try:
            page.wait_for_timeout(2000)

            # 2. Login -- OJO: el login real del sitio es un ENVIO DE
            # FORMULARIO NORMAL (navegacion completa, Sec-Fetch-Mode:
            # navigate), NO una peticion tipo API/AJAX. Se probo primero
            # con page.request.post() y el servidor respondia distinto
            # (usuario "ANONIMO"), asi que se llena el formulario real y
            # se envia con Enter, para que el navegador dispare la MISMA
            # navegacion completa que dispara una persona real.
            try:
                page.fill('input[name="user"]', usuario)
                page.fill('input[name="passw"]', password)
            except Exception as e_login_campos:
                print(f"{etiqueta} No se encontraron los campos de login con los selectores esperados: {e_login_campos}", flush=True)
                try:
                    campos_login = page.evaluate("""() => {
                        var els = document.querySelectorAll('input');
                        var resultado = [];
                        els.forEach(function(el){
                            resultado.push({name: el.name || null, id: el.id || null, type: el.type || null});
                        });
                        return resultado;
                    }""")
                    print(f"{etiqueta} Campos input reales encontrados en la pagina: {campos_login}", flush=True)
                except Exception:
                    pass
                raise
            with page.expect_navigation(wait_until="load", timeout=20000):
                page.press('input[name="passw"]', "Enter")
            page.wait_for_timeout(2000)

            # 3. Confirmar que la sesion quedo autenticada (revisa el
            # username devuelto, debe coincidir con el que se uso para
            # entrar). OJO: el sitio antepone "UE" al numero de
            # documento (ej "UE1036614666"), asi que no se compara
            # exacto -- solo se revisa que NO sea "ANONIMO" (el valor
            # que devuelve cuando el login de verdad fallo).
            resp_jwt = page.request.get(f"{MEDELLIN_AVIT_API}/login/JWT/")
            datos_sesion = resp_jwt.json()
            print(f"{etiqueta} Usuario autenticado segun el sitio: {datos_sesion.get('username')}", flush=True)
            if not datos_sesion.get("username") or datos_sesion.get("username") == "ANONIMO":
                return False, {"error": "El login no parece haber funcionado (sesión sigue anónima)."}

            headers_json = {"Content-Type": "application/json"}

            # 4. Preguntas de "seguridad" -- igual que en Envigado, no son
            # preguntas reales, solo validan documento+placa.
            resp_preg = page.request.post(
                f"{MEDELLIN_AVIT_API}/seguridad/preguntas/validarMostrarPreguntasSeguridad",
                data=json.dumps([{
                    "documento": None, "respuesta": None, "placa": placa,
                    "homePublic": False, "tipoDocumento": None, "nroDocumento": None,
                    "bloqueoPermanente": False, "paramValidar": 1844,
                }]),
                headers=headers_json,
            )
            print(f"{etiqueta} Preguntas de seguridad: {resp_preg.status} {resp_preg.text()[:200]}", flush=True)

            # 5. Puntos de atencion que ofrecen este servicio.
            resp_puntos = page.request.post(
                f"{MEDELLIN_AVIT_API}/citas/getPuntosAtencionServiciosLowcode",
                data=json.dumps({"idServicios": str(id_servicio), "cantidadTramites": 1}),
                headers=headers_json,
            )
            puntos = resp_puntos.json()
            print(f"{etiqueta} Puntos de atencion encontrados: {puntos}", flush=True)
            nombres_sedes_reales = [p.get("nombreSubsede", "") for p in puntos]
            print(f"{etiqueta} === Nombres EXACTOS de sede que devuelve el sitio ahora mismo: {nombres_sedes_reales} ===", flush=True)
            if sede_deseada:
                print(f"{etiqueta} Sede que se esta buscando (filtro): '{sede_deseada}'", flush=True)

            if not puntos:
                return False, {"mensaje": "No hay ningun punto de atencion ofreciendo este servicio en este momento."}

            # 6. Para cada punto de atencion, revisar fechas disponibles.
            # Si se pidio una sede especifica, se saltan las demas por
            # completo (ni siquiera se consultan sus fechas).
            for punto in puntos:
                nombre_sede = punto.get("nombreSubsede", "")
                if sede_deseada and sede_deseada.strip().lower() not in nombre_sede.lower():
                    continue

                id_subsede = punto.get("idSubsede")
                resp_fechas = page.request.post(
                    f"{MEDELLIN_AVIT_API}/citas/getFechasDisponibles",
                    data=json.dumps({
                        "idServicios": str(id_servicio), "idSubsede": id_subsede,
                        "cantidadTramites": 1, "cantidadDiasMostrar": 211,
                        "tiempoAtencion": 20, "cantidadLimiteLibracion": "1",
                        "isIncluirHorariosCancelados": "S",
                    }),
                    headers=headers_json,
                )
                fechas = resp_fechas.json()
                print(f"{etiqueta} Fechas disponibles en '{nombre_sede}': {fechas}", flush=True)
                if fechas:
                    return True, {"sede": nombre_sede, "id_subsede": id_subsede, "fechas": fechas}

            if sede_deseada:
                return False, {"mensaje": f"Se reviso la sede '{sede_deseada}', no tiene fechas disponibles ahora mismo."}
            return False, {"mensaje": "Se revisaron todos los puntos de atencion, ninguno tiene fechas disponibles ahora mismo."}

        except Exception as e:
            print(f"{etiqueta} Error: {e}", flush=True)
            return False, {"error": str(e)}
        finally:
            context.close(); browser.close()


def _envigado_proximo_dia_habil():
    """Calcula el proximo dia habil (lunes a viernes) a partir de hoy, en
    formato 'DD/MM/YYYY' -- igual al formato que usa la API de Envigado
    para 'diaAtencion'. Si hoy es viernes o fin de semana, salta al
    proximo lunes."""
    hoy = datetime.now().date()
    siguiente = hoy + timedelta(days=1)
    while siguiente.weekday() >= 5:  # 5=sabado, 6=domingo
        siguiente += timedelta(days=1)
    return siguiente.strftime("%d/%m/%Y")


def envigado_hay_puntos_disponibles(usar_proxy=True):
    """Revisa, para CADA sede, si el PROXIMO DIA HABIL especificamente
    tiene fechas de atencion disponibles -- no solo si el servicio esta
    listado (eso casi siempre es cierto y causaba falsos positivos, ya
    que 'getPuntosAtencionServiciosLowcode' solo dice que sedes OFRECEN
    el tramite, sin decir si tienen cupo). Usa un NAVEGADOR REAL
    (Playwright) en vez de peticiones HTTP directas -- se probo con
    peticiones directas primero (replicando payload y encabezados
    exactos capturados de un HAR real), pero el servidor seguia
    devolviendo error 500 solo en este endpoint especifico (el mismo
    enfoque SI funciona para otros endpoints de Envigado, como el
    monitor de turnos) -- lo mas probable es que este endpoint en
    particular tenga alguna proteccion anti-bot que detecta que la
    conexion no viene de un navegador real. Se llena el formulario igual
    que lo haria una persona, con datos de prueba que no corresponden a
    ningun ciudadano real.
    'usar_proxy' (True por defecto): el trafico repetido del monitoreo
    desde la misma IP del servidor puede hacer que el sitio escale la
    dificultad del captcha -- se usa el proxy residencial de DataImpulse
    por defecto para evitarlo.
    Devuelve una lista de {"sede": nombre, "fecha": "DD/MM/YYYY"} -- una
    entrada por cada sede que SI tiene el proximo dia habil disponible
    (vacia [] si ninguna sede lo tiene), o None si algo fallo."""
    fecha_objetivo = _envigado_proximo_dia_habil()
    resultado_final = []
    respuestas_capturadas = {}
    todas_las_urls_vistas = []  # diagnostico -- para ver TODAS las peticiones de red relacionadas a citas, sin importar el nombre exacto
    etiqueta = f"[ENVIGADO-CHEQUEO-{uuid.uuid4().hex[:6]}]"

    def _capturar_respuesta(response):
        if "backavit" in response.url or "citas" in response.url:
            todas_las_urls_vistas.append(response.url)
        for nombre_endpoint in ["getPuntosAtencionServiciosLowcode", "getFechasDisponibles"]:
            if nombre_endpoint in response.url:
                try:
                    respuestas_capturadas[nombre_endpoint] = response.json()
                except Exception:
                    pass

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True, args=[
            "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
            "--single-process", "--no-zygote", "--disable-setuid-sandbox"
        ])
        proxy_config = _dataimpulse_proxy_config(etiqueta) if usar_proxy else None
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Linux; Android 15; Pixel 9) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Mobile Safari/537.36",
            viewport={"width": 390, "height": 844},
            proxy=proxy_config,
        )
        page = context.new_page()
        page.on("response", _capturar_respuesta)

        try:
            page.goto("https://movilidad.envigado.gov.co/portal-servicios/#/agendar-cita-publica",
                      wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(3000)  # dar tiempo a que la app Angular termine de armarse

            if _envigado_hay_aviso_sin_agenda(page):
                print("Aviso 'sin agenda disponible' detectado apenas se cargo la pagina -- confirmado, no hay citas.", flush=True)
                return []

            # DIAGNOSTICO TEMPORAL -- en vez de seguir adivinando campo por
            # campo, se listan TODOS los <input>/<select>/<textarea> reales
            # de la pagina de una sola vez (name, id, type, placeholder),
            # para ajustar todos los selectores juntos. Quitar despues.
            try:
                campos_reales = page.evaluate("""() => {
                    var els = document.querySelectorAll('input, select, textarea');
                    var resultado = [];
                    els.forEach(function(el){
                        resultado.push({
                            tag: el.tagName,
                            name: el.name || null,
                            id: el.id || null,
                            type: el.type || null,
                            placeholder: el.placeholder || null,
                            visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                        });
                    });
                    return resultado;
                }""")
                print("=== DIAGNOSTICO: campos reales del formulario ===", flush=True)
                for c in campos_reales:
                    print(c, flush=True)
                print("=== FIN DIAGNOSTICO campos ===", flush=True)
            except Exception as e_diag:
                print(f"No se pudo listar los campos: {e_diag}", flush=True)

            # Tipo de documento -- CONFIRMADO por diagnostico real: es un
            # <select> nativo de AngularJS clasico, con id="tipoDocumento"
            # y name="tipoDocumento" (no un componente moderno). Se usa
            # select_option (no click), que es lo correcto para un
            # <select> nativo.
            try:
                page.select_option("#tipoDocumento", label="CC", timeout=5000)
            except Exception as e_tipodoc:
                print(f"No se pudo seleccionar tipo de documento (se sigue de todas formas): {e_tipodoc}", flush=True)

            # El resto de campos: como es AngularJS clasico (ng-model), lo
            # mas probable es que usen el atributo "name" igual que el de
            # tipoDocumento -- se usa ese patron en vez de adivinar
            # etiquetas visibles (que resultaron no estar bien asociadas
            # a sus campos en este formulario).
            page.locator('input[name="numeroDocumento"]').fill(ENVIGADO_CITAS_DOCUMENTO, timeout=10000)
            page.locator('input[name="nombres"]').fill(ENVIGADO_CITAS_NOMBRES, timeout=8000)
            page.locator('input[name="apellidos"]').fill(ENVIGADO_CITAS_APELLIDOS, timeout=8000)

            # Nombres reales confirmados por diagnostico (no eran "email"
            # ni "celular" como se habia adivinado antes).
            page.locator('input[name="emailSolicitante"]').fill(ENVIGADO_CITAS_EMAIL, timeout=8000)
            page.locator('input[name="confirmarEmail"]').fill(ENVIGADO_CITAS_EMAIL, timeout=8000)
            page.locator('input[name="phone"]').fill(ENVIGADO_CITAS_CELULAR, timeout=8000)

            # Seleccionar el tramite -- CONFIRMADO por diagnostico real:
            # el campo que el usuario realmente ve e interactua es
            # "servicios" (visible=True) -- "tramite" es un campo
            # SECUNDARIO/oculto (visible=False) que probablemente se
            # sincroniza solo, seleccionarlo directamente no funcionaba.
            # Se cambia el valor por JavaScript y se dispara el evento
            # "change" nativo, que es lo que AngularJS espera para
            # actualizar su modelo.
            page.evaluate("""() => {
                var el = document.querySelector('#servicios');
                if (!el) return false;
                el.value = '90';  // "Comprador/Vendedor - Traspaso"
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }""")
            page.wait_for_timeout(1500)  # dar tiempo a que aparezca el campo de placa (se agrega dinamicamente)

            if _envigado_hay_aviso_sin_agenda(page):
                print("Aviso 'sin agenda disponible' detectado despues de elegir el tramite -- confirmado, no hay citas.", flush=True)
                return []

            # DIAGNOSTICO TEMPORAL -- el campo de Placa no aparecia en el
            # listado inicial de campos, seguramente porque se agrega
            # recien despues de elegir el tramite. Se vuelve a listar
            # todos los campos en este punto para encontrar su nombre
            # real, en vez de seguir adivinando. Quitar despues.
            try:
                campos_tras_tramite = page.evaluate("""() => {
                    var els = document.querySelectorAll('input, select, textarea');
                    var resultado = [];
                    els.forEach(function(el){
                        resultado.push({
                            tag: el.tagName, name: el.name || null, id: el.id || null,
                            type: el.type || null, placeholder: el.placeholder || null,
                            visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                        });
                    });
                    return resultado;
                }""")
                print("=== DIAGNOSTICO: campos reales DESPUES de elegir tramite ===", flush=True)
                for c in campos_tras_tramite:
                    print(c, flush=True)
                print("=== FIN DIAGNOSTICO campos tras tramite ===", flush=True)
            except Exception as e_diag2:
                print(f"No se pudo listar los campos tras el tramite: {e_diag2}", flush=True)

            # El campo de Placa no aparecia en el formulario inicial --
            # probablemente se agrega recien despues de elegir el tramite.
            # Se prueban varios selectores, ya que "get_by_label" no fue
            # confiable en este formulario (AngularJS clasico).
            placa_ok = False
            for selector_placa in ['input[name="placa"]', 'input[name="placaVehiculo"]', 'input#placa']:
                try:
                    page.locator(selector_placa).fill(ENVIGADO_CITAS_PLACA, timeout=5000)
                    placa_ok = True
                    break
                except Exception:
                    continue

            if not placa_ok:
                # ACLARADO: no es una lista de vehiculos para elegir -- es
                # un campo de "etiquetas" (Select2 en modo tags). Se
                # escribe la placa libremente y se presiona Enter, y
                # queda como una etiqueta con una X para quitarla.
                # El envoltorio visual de Select2 bloquea los clics
                # normales de Playwright ("intercepts pointer events"),
                # asi que se usa force=True para saltar esa validacion.
                try:
                    campo_placa = page.locator('input[type="search"]').first
                    campo_placa.click(timeout=5000, force=True)
                    campo_placa.fill(ENVIGADO_CITAS_PLACA, timeout=5000, force=True)
                    campo_placa.press("Enter")
                    page.wait_for_timeout(1000)
                    placa_ok = True
                    print(f"Placa {ENVIGADO_CITAS_PLACA} escrita y confirmada con Enter.", flush=True)
                except Exception as e_placa_tag:
                    print(f"No se pudo escribir la placa como etiqueta: {e_placa_tag}", flush=True)

            if not placa_ok:
                print("No se pudo llenar/elegir la placa con ningun metodo conocido (revisar diagnostico de arriba).", flush=True)

            # Boton "Agregar servicio" -- avanza al paso donde el sitio
            # consulta la disponibilidad real. A veces el aviso de "sin
            # agenda" ya tapa este boton o lo bloquea -- si el clic falla,
            # se revisa el aviso antes de reportarlo como error.
            try:
                page.get_by_text("Agregar servicio", exact=True).click(timeout=8000)
            except Exception as e_clic_agregar:
                if _envigado_hay_aviso_sin_agenda(page):
                    print("Aviso 'sin agenda disponible' detectado al intentar hacer clic en 'Agregar servicio' -- confirmado, no hay citas.", flush=True)
                    return []
                raise e_clic_agregar

            if _envigado_hay_aviso_sin_agenda(page):
                print("Aviso 'sin agenda disponible' detectado justo despues de 'Agregar servicio' -- confirmado, no hay citas.", flush=True)
                return []

            # Esperar a que la peticion de puntos se dispare y la
            # respuesta llegue. Se revisa el aviso de "sin agenda" en
            # cada vuelta tambien, por si aparece justo en este momento
            # (a veces tarda un poco en mostrarse despues del clic).
            for _ in range(10):
                if "getPuntosAtencionServiciosLowcode" in respuestas_capturadas:
                    break
                if _envigado_hay_aviso_sin_agenda(page):
                    print("Aviso 'sin agenda disponible' detectado mientras se esperaba la respuesta -- confirmado, no hay citas.", flush=True)
                    return []
                page.wait_for_timeout(1000)

            puntos = respuestas_capturadas.get("getPuntosAtencionServiciosLowcode")
            if not puntos:
                # Diagnostico: se imprime un resumen del HTML visible para
                # poder ajustar los selectores si algo no coincidio.
                print("=== DIAGNOSTICO citas Envigado: no se capturo la respuesta esperada ===", flush=True)
                print("URL actual:", page.url, flush=True)
                try:
                    print("Texto visible de la pagina (primeros 2000 caracteres):", flush=True)
                    print(page.inner_text("body")[:2000], flush=True)
                except Exception as e_txt:
                    print("No se pudo leer el texto de la pagina:", e_txt, flush=True)
                print("=== FIN DIAGNOSTICO ===", flush=True)
                context.close(); browser.close()
                return []

            print(f"Puntos de atencion encontrados (revisando fecha objetivo {fecha_objetivo}): {puntos}", flush=True)

            # Por cada punto/sede que ofrece el tramite, se elige esa sede
            # y se revisa si el PROXIMO DIA HABIL especificamente esta en
            # su lista de fechas disponibles -- esto es lo que realmente
            # confirma que hay cupo, a diferencia de solo listar el
            # tramite como ofrecido.
            for punto in puntos:
                nombre_sede = punto.get("nombreSubsede") or "Sede desconocida"
                id_subsede = punto.get("idSubsede")
                if id_subsede is None:
                    continue

                respuestas_capturadas.pop("getFechasDisponibles", None)
                try:
                    # Se selecciona por el TEXTO VISIBLE de la opcion
                    # (nombre de la sede), no por su "value" interno --
                    # se confirmo con diagnostico real que el <select> de
                    # Angular no usa el idSubsede puro como value (nunca
                    # se encontraba una opcion con ese valor exacto,
                    # aunque el <select> si tenia opciones renderizadas).
                    # select_option con "label" busca por el texto que ve
                    # el usuario, evitando ese problema. Se reintenta por
                    # si las opciones aun no han terminado de renderizarse.
                    seleccionado_ok = False
                    for intento_sede in range(5):
                        try:
                            page.select_option('#seleccione_punto_atencion', label=nombre_sede, timeout=2000)
                            seleccionado_ok = True
                            break
                        except Exception:
                            page.wait_for_timeout(500)
                    if not seleccionado_ok:
                        # Respaldo: el texto exacto no coincidio (puede
                        # tener espacios/mayusculas distintas) -- se busca
                        # cualquier <option> cuyo texto CONTENGA el nombre
                        # de la sede, sin importar mayusculas/espacios.
                        seleccionado_ok = page.evaluate(f"""() => {{
                            var el = document.querySelector('#seleccione_punto_atencion');
                            if (!el) return false;
                            var buscado = {json.dumps(nombre_sede)}.trim().toLowerCase();
                            for (var i = 0; i < el.options.length; i++) {{
                                var texto = (el.options[i].textContent || '').trim().toLowerCase();
                                if (texto.indexOf(buscado) >= 0 || buscado.indexOf(texto) >= 0) {{
                                    el.selectedIndex = i;
                                    el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                    return true;
                                }}
                            }}
                            return false;
                        }}""")
                    print(f"Sede '{nombre_sede}' (idSubsede={id_subsede}) -- se pudo seleccionar por texto: {seleccionado_ok}", flush=True)
                    if seleccionado_ok:
                        page.evaluate("""() => {
                            var el = document.querySelector('#seleccione_punto_atencion');
                            if (el) el.dispatchEvent(new Event('change', { bubbles: true }));
                        }""")
                except Exception as e_sede:
                    print(f"No se pudo seleccionar la sede '{nombre_sede}': {e_sede}", flush=True)
                    continue
                page.wait_for_timeout(1500)

                for _ in range(10):
                    if "getFechasDisponibles" in respuestas_capturadas:
                        break
                    page.wait_for_timeout(1000)

                if "getFechasDisponibles" not in respuestas_capturadas:
                    print(f"Sede '{nombre_sede}': NUNCA se capturo la respuesta de getFechasDisponibles (la peticion no se disparo, o tardo mas de 10 segundos).", flush=True)
                    print(f"Todas las URLs de red relacionadas a citas vistas hasta ahora: {todas_las_urls_vistas}", flush=True)

                fechas_disponibles = respuestas_capturadas.get("getFechasDisponibles") or []
                dias_atencion = [f.get("diaAtencion") for f in fechas_disponibles if isinstance(f, dict)]
                print(f"Fechas disponibles en '{nombre_sede}': {dias_atencion} (respuesta cruda: {fechas_disponibles})", flush=True)

                if fecha_objetivo in dias_atencion:
                    resultado_final.append({"sede": nombre_sede, "fecha": fecha_objetivo})

        except Exception as e:
            print(f"Error en el flujo de Playwright para citas Envigado: {e}", flush=True)
            try:
                print("=== DIAGNOSTICO citas Envigado: texto visible al momento del error ===", flush=True)
                print(page.inner_text("body")[:2000], flush=True)
                print("=== FIN DIAGNOSTICO ===", flush=True)
            except Exception:
                pass
            return None
        finally:
            context.close(); browser.close()

    return resultado_final


def envigado_revisar_citas_disponibles(dias_adelante=14):
    """Revisa si el proximo dia habil tiene cupo en alguna sede, y guarda
    el resultado en la base de datos. 'dias_adelante' ya no se usa (se
    revisa unicamente el proximo dia habil, no un rango de dias) -- se
    deja como parametro por compatibilidad con quien ya llama esta
    funcion. Devuelve una tupla (resultados, hubo_error):
    - resultados: lista de {sede, fecha, cantidad_horarios}
    - hubo_error: True si la consulta fallo por dentro (para NO
      confundirlo con un "confirmado, sin citas ahora mismo")."""
    puntos = envigado_hay_puntos_disponibles()
    if puntos is None:
        print("Error consultando disponibilidad de citas Envigado (ver logs arriba).", flush=True)
        return [], True

    resultados = []
    conn = get_db_conn()
    cur = conn.cursor()

    # Limpieza general: se borra CUALQUIER registro cuya fecha de cita ya
    # paso, sin importar el resultado de esta consulta -- antes solo se
    # limpiaban los resultados vacios del dia de HOY, asi que un positivo
    # de dias anteriores que nunca se volvio a consultar se quedaba
    # mostrandose para siempre.
    cur.execute("""
        DELETE FROM envigado_citas_disponibles
        WHERE TO_DATE(fecha_dia, 'DD/MM/YYYY') < CURRENT_DATE
    """)

    if isinstance(puntos, list) and len(puntos) > 0:
        # 'puntos' ya viene filtrado -- cada elemento es una sede que SI
        # tiene el proximo dia habil disponible de verdad (no solo que
        # ofrece el tramite).
        print("=== CITAS ENVIGADO: hay cupo para el proximo dia habil ===", flush=True)
        print(puntos, flush=True)
        print("=== FIN ===", flush=True)
        for p in puntos:
            resultados.append({
                "sede": p["sede"],
                "fecha": p["fecha"],
                "cantidad_horarios": 1  # se confirma que hay cupo; el conteo exacto de horas se ve al reservar
            })
            cur.execute("""
                INSERT INTO envigado_citas_disponibles (sede, id_subsede, fecha_dia, cantidad_horarios, verificado_en)
                VALUES (%s, %s, %s, %s, NOW())
                ON CONFLICT (sede, fecha_dia) DO UPDATE SET
                    cantidad_horarios=EXCLUDED.cantidad_horarios, verificado_en=NOW()
            """, (p["sede"], 0, p["fecha"], 1))

    else:
        # Sin citas en este momento -- se limpia cualquier resultado
        # positivo anterior guardado hoy, para no mostrar un aviso viejo
        # que ya no es cierto.
        cur.execute("""
            DELETE FROM envigado_citas_disponibles
            WHERE verificado_en::date = CURRENT_DATE
        """)

    conn.commit()
    cur.close(); conn.close()
    return resultados, False


ENVIGADO_RECAPTCHA_SITEKEY = "6LdZ-WUsAAAAAEEs0_PbIzNhEoDTBqV1CwBEE8B-"  # confirmado con un HAR real

# Carpeta temporal donde se guardan las capturas de pantalla del flujo de
# reserva de citas de Envigado -- se sirven despues via /envigado-captura.
CAPTURAS_ENVIGADO_DIR = "/tmp/capturas_envigado"
os.makedirs(CAPTURAS_ENVIGADO_DIR, exist_ok=True)


def envigado_reservar_cita(solicitud):
    """Completa el flujo ENTERO de agendar una cita real en Envigado --
    llena el formulario con los datos de la persona (no datos de
    prueba), elige la primera sede/fecha/hora disponible que coincida
    con la hora aproximada pedida, resuelve el reCAPTCHA con 2captcha,
    y confirma la cita.

    'solicitud' es un dict con: nombres, apellidos, tipo_documento,
    numero_documento, correo, celular, placa, id_servicio,
    sede_preferida (puede ser None/vacio = cualquiera), hora_aproximada
    (entero 0-23).

    NOTA IMPORTANTE: los pasos de elegir sede/fecha/hora e interactuar
    con el calendario del sitio son dificiles de programar a ciegas sin
    ver la pantalla real (a diferencia del formulario inicial, que ya
    esta confirmado). Por eso esta funcion imprime diagnosticos
    detallados en cada paso -- es posible que los primeros intentos
    reales necesiten ajustes una vez veamos esos diagnosticos, igual
    que paso con el registro de usuarios de Medellin.

    Devuelve un dict: {"exito": bool, "nro_atencion": str|None,
    "mensaje": str, "detalle": dict|None}."""
    etiqueta = f"[ENVIGADO-RESERVAR-{uuid.uuid4().hex[:6]}]"
    resultado = {"exito": False, "nro_atencion": None, "mensaje": "", "detalle": None}

    respuestas_capturadas = {}

    def _capturar_respuesta(response):
        for nombre_endpoint in ["getPuntosAtencionServiciosLowcode", "getFechasDisponibles", "getHorasDisponibles", "agendarCitaGAComponentes"]:
            if nombre_endpoint in response.url:
                try:
                    respuestas_capturadas[nombre_endpoint] = response.json()
                except Exception:
                    pass

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True, args=[
            "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
            "--single-process", "--no-zygote", "--disable-setuid-sandbox"
        ])
        proxy_config = _dataimpulse_proxy_config(etiqueta)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Linux; Android 15; Pixel 9) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Mobile Safari/537.36",
            viewport={"width": 390, "height": 844},
            proxy=proxy_config,
        )
        page = context.new_page()
        page.on("response", _capturar_respuesta)

        try:
            print(f"{etiqueta} Iniciando reserva para {solicitud['nombres']} {solicitud['apellidos']}, placa {solicitud['placa']}, hora aproximada {solicitud['hora_aproximada']}:00", flush=True)
            page.goto("https://movilidad.envigado.gov.co/portal-servicios/#/agendar-cita-publica",
                      wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(3000)

            if _envigado_hay_aviso_sin_agenda(page):
                resultado["mensaje"] = "No hay agenda disponible en este momento."
                return resultado

            # --- Llenar el formulario con los datos REALES de la solicitud ---
            try:
                page.select_option("#tipoDocumento", label="CC", timeout=5000)
            except Exception:
                pass
            page.locator('input[name="numeroDocumento"]').fill(solicitud["numero_documento"], timeout=10000)
            page.locator('input[name="nombres"]').fill(solicitud["nombres"], timeout=8000)
            page.locator('input[name="apellidos"]').fill(solicitud["apellidos"], timeout=8000)
            page.locator('input[name="emailSolicitante"]').fill(solicitud["correo"], timeout=8000)
            page.locator('input[name="confirmarEmail"]').fill(solicitud["correo"], timeout=8000)
            page.locator('input[name="phone"]').fill(solicitud["celular"], timeout=8000)

            id_servicio = solicitud.get("id_servicio") or ENVIGADO_CITAS_ID_SERVICIO
            page.evaluate(f"""() => {{
                var el = document.querySelector('#servicios');
                if (!el) return false;
                el.value = '{id_servicio}';
                el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                return true;
            }}""")
            page.wait_for_timeout(1500)

            if _envigado_hay_aviso_sin_agenda(page):
                resultado["mensaje"] = "No hay agenda disponible para este trámite."
                return resultado

            placa_ok = False
            for selector_placa in ['input[name="placa"]', 'input[name="placaVehiculo"]', 'input#placa']:
                try:
                    page.locator(selector_placa).fill(solicitud["placa"], timeout=5000)
                    placa_ok = True
                    break
                except Exception:
                    continue
            if not placa_ok:
                for intento_placa in range(3):
                    try:
                        campo_placa = page.locator('input[type="search"]').first
                        campo_placa.wait_for(state="attached", timeout=5000)
                        page.wait_for_timeout(500)
                        campo_placa.click(timeout=5000, force=True)
                        campo_placa.fill(solicitud["placa"], timeout=5000, force=True)
                        campo_placa.press("Enter")
                        page.wait_for_timeout(1000)
                        placa_ok = True
                        break
                    except Exception as e_placa_tag:
                        print(f"{etiqueta} No se pudo escribir la placa (intento {intento_placa+1}/3): {e_placa_tag}", flush=True)
                        page.wait_for_timeout(1500)
            if not placa_ok:
                resultado["mensaje"] = "No se pudo llenar el campo de placa."
                return resultado

            try:
                page.get_by_text("Agregar servicio", exact=True).click(timeout=8000)
            except Exception as e_clic_agregar:
                if _envigado_hay_aviso_sin_agenda(page):
                    resultado["mensaje"] = "No hay agenda disponible."
                    return resultado
                raise e_clic_agregar

            if _envigado_hay_aviso_sin_agenda(page):
                resultado["mensaje"] = "No hay agenda disponible."
                return resultado

            for _ in range(10):
                if "getPuntosAtencionServiciosLowcode" in respuestas_capturadas:
                    break
                if _envigado_hay_aviso_sin_agenda(page):
                    resultado["mensaje"] = "No hay agenda disponible."
                    return resultado
                page.wait_for_timeout(1000)

            puntos = respuestas_capturadas.get("getPuntosAtencionServiciosLowcode") or []
            print(f"{etiqueta} Puntos de atencion encontrados: {puntos}", flush=True)
            if not puntos:
                # Diagnostico enriquecido -- antes esta funcion solo
                # reportaba "vacio" sin mas detalle. Se imprime el texto
                # visible y los campos reales del formulario en este
                # punto, para distinguir si el sitio de verdad respondio
                # "sin puntos" o si la respuesta nunca se capturo (ej. el
                # clic en "Agregar servicio" no disparo la peticion).
                print(f"{etiqueta} === DIAGNOSTICO: puntos vacio -- revisando estado de la pagina ===", flush=True)
                print(f"{etiqueta} URL actual: {page.url}", flush=True)
                print(f"{etiqueta} Se capturo getPuntosAtencionServiciosLowcode en absoluto: {'getPuntosAtencionServiciosLowcode' in respuestas_capturadas}", flush=True)
                try:
                    print(f"{etiqueta} Texto visible de la pagina (primeros 2000 caracteres):", flush=True)
                    print(page.inner_text("body")[:2000], flush=True)
                except Exception as e_txt:
                    print(f"{etiqueta} No se pudo leer el texto de la pagina: {e_txt}", flush=True)
                try:
                    campos_diag = page.evaluate("""() => {
                        var els = document.querySelectorAll('input, select, textarea');
                        var resultado = [];
                        els.forEach(function(el){
                            resultado.push({
                                tag: el.tagName, name: el.name || null, id: el.id || null,
                                value: (el.value || '').substring(0, 40),
                                visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                            });
                        });
                        return resultado;
                    }""")
                    print(f"{etiqueta} Campos reales (con su valor actual) en este punto:", flush=True)
                    for c in campos_diag:
                        print(f"{etiqueta}   {c}", flush=True)
                except Exception as e_campos:
                    print(f"{etiqueta} No se pudieron listar los campos: {e_campos}", flush=True)
                print(f"{etiqueta} === FIN DIAGNOSTICO ===", flush=True)
                resultado["mensaje"] = "No se encontraron puntos de atención con este trámite."
                return resultado

            # Elegir la sede -- preferida si se pidio y esta en la lista,
            # si no la primera disponible.
            sede_elegida = None
            if solicitud.get("sede_preferida"):
                for p in puntos:
                    if solicitud["sede_preferida"].strip().lower() in (p.get("nombreSubsede") or "").lower():
                        sede_elegida = p
                        break
            if not sede_elegida:
                sede_elegida = puntos[0]
            print(f"{etiqueta} Sede elegida: {sede_elegida}", flush=True)

            # --- Elegir la sede en el <select> real ---
            # Se selecciona por el TEXTO VISIBLE de la opcion (nombre de
            # la sede), no por su "value" interno -- se confirmo con
            # diagnostico real que el <select> de Angular no usa el
            # idSubsede puro como value.
            id_subsede_elegida = sede_elegida.get("idSubsede")
            nombre_sede_elegida = sede_elegida.get("nombreSubsede") or ""
            seleccionado_ok = False
            for intento_sede in range(5):
                try:
                    page.select_option('#seleccione_punto_atencion', label=nombre_sede_elegida, timeout=2000)
                    seleccionado_ok = True
                    break
                except Exception:
                    page.wait_for_timeout(500)
            if not seleccionado_ok:
                seleccionado_ok = page.evaluate(f"""() => {{
                    var el = document.querySelector('#seleccione_punto_atencion');
                    if (!el) return false;
                    var buscado = {json.dumps(nombre_sede_elegida)}.trim().toLowerCase();
                    for (var i = 0; i < el.options.length; i++) {{
                        var texto = (el.options[i].textContent || '').trim().toLowerCase();
                        if (texto.indexOf(buscado) >= 0 || buscado.indexOf(texto) >= 0) {{
                            el.selectedIndex = i;
                            el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            return true;
                        }}
                    }}
                    return false;
                }}""")
            if seleccionado_ok:
                page.evaluate("""() => {
                    var el = document.querySelector('#seleccione_punto_atencion');
                    if (el) el.dispatchEvent(new Event('change', { bubbles: true }));
                }""")
            print(f"{etiqueta} Sede '{nombre_sede_elegida}' (idSubsede={id_subsede_elegida}) -- se pudo seleccionar por texto: {seleccionado_ok}", flush=True)
            page.wait_for_timeout(1500)

            for _ in range(10):
                if "getFechasDisponibles" in respuestas_capturadas:
                    break
                page.wait_for_timeout(1000)
            fechas_disponibles = respuestas_capturadas.get("getFechasDisponibles") or []
            print(f"{etiqueta} Fechas disponibles en '{sede_elegida.get('nombreSubsede')}': {fechas_disponibles}", flush=True)
            if not fechas_disponibles:
                resultado["mensaje"] = f"La sede '{sede_elegida.get('nombreSubsede')}' no tiene fechas disponibles en este momento."
                return resultado

            # Se toma la PRIMERA fecha disponible (la mas cercana).
            fecha_elegida = fechas_disponibles[0].get("diaAtencion")  # formato "DD/MM/YYYY", confirmado con el HAR
            print(f"{etiqueta} Fecha elegida: {fecha_elegida}", flush=True)

            # --- Elegir la fecha en el datepicker ---
            respuestas_capturadas.pop("getHorasDisponibles", None)

            # Confirmado por diagnostico real: el sitio usa la libreria
            # "Air Datepicker" -- cada dia es un <div> con
            # data-date/data-month/data-year (mes de 0 a 11), y hay que
            # hacer CLIC en la celda del dia exacto para que se dispare
            # la consulta de horas disponibles (escribir texto en el
            # input no basta).
            dia_num, mes_num, anio_num = fecha_elegida.split("/")
            mes_datepicker = str(int(mes_num) - 1)  # Air Datepicker usa el mes de 0 a 11
            dia_datepicker = str(int(dia_num))  # sin cero a la izquierda, ej "19" no "019"

            page.click('#agendarCitaDatePicker', timeout=5000)
            page.wait_for_timeout(800)

            selector_dia = (
                f'.datepicker--cell-day[data-date="{dia_datepicker}"]'
                f'[data-month="{mes_datepicker}"][data-year="{anio_num}"]:not(.-other-month-)'
            )
            fecha_click_ok = False
            try:
                page.click(selector_dia, timeout=5000)
                fecha_click_ok = True
            except Exception as e_click_dia:
                print(f"{etiqueta} No se pudo hacer clic en la celda del dia ({selector_dia}): {e_click_dia}", flush=True)

            print(f"{etiqueta} Clic en el dia {fecha_elegida} del calendario -- exitoso: {fecha_click_ok}", flush=True)

            # Respaldo (por si la libreria cambia o el selector no
            # coincide): escribir el texto directamente en el input.
            if not fecha_click_ok:
                try:
                    page.fill('#agendarCitaDatePicker', fecha_elegida, timeout=5000)
                    page.locator('#agendarCitaDatePicker').press('Tab')
                except Exception:
                    page.evaluate(f"""() => {{
                        var el = document.querySelector('#agendarCitaDatePicker');
                        if (!el) return false;
                        el.value = '{fecha_elegida}';
                        el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                        el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                        return true;
                    }}""")
            page.wait_for_timeout(1500)

            for _ in range(10):
                if "getHorasDisponibles" in respuestas_capturadas:
                    break
                page.wait_for_timeout(1000)
            horas_disponibles = respuestas_capturadas.get("getHorasDisponibles") or []
            print(f"{etiqueta} Horas disponibles para {fecha_elegida}: {horas_disponibles}", flush=True)
            if not horas_disponibles:
                # Diagnostico enriquecido -- igual que se hizo antes con
                # la sede, para ver que esta pasando realmente en la
                # pagina cuando esto falla.
                print(f"{etiqueta} === DIAGNOSTICO: horas vacias -- revisando estado de la pagina ===", flush=True)
                try:
                    valor_datepicker_actual = page.evaluate("""() => {
                        var el = document.querySelector('#agendarCitaDatePicker');
                        return el ? el.value : null;
                    }""")
                    print(f"{etiqueta} Valor actual del datepicker: {valor_datepicker_actual!r}", flush=True)
                except Exception:
                    pass
                try:
                    print(f"{etiqueta} Texto visible de la pagina (primeros 2000 caracteres):", flush=True)
                    print(page.inner_text("body")[:2000], flush=True)
                except Exception as e_txt:
                    print(f"{etiqueta} No se pudo leer el texto de la pagina: {e_txt}", flush=True)
                print(f"{etiqueta} === FIN DIAGNOSTICO ===", flush=True)
                resultado["mensaje"] = f"No se encontraron horas disponibles para el {fecha_elegida} (puede que el datepicker no haya respondido como se esperaba -- revisar diagnostico)."
                return resultado

            # 'horas_disponibles' viene AGRUPADA por franja (ej. "7:00 AM
            # - 12:00 PM", "1:00 PM - 7:00 PM"), cada grupo con su propia
            # lista interna "horarios" -- se aplanan todos los horarios
            # individuales de todos los grupos en una sola lista antes de
            # buscar el mas cercano (si no, min() comparaba los GRUPOS
            # completos entre si, que no tienen "horaIni" propio, y el
            # resultado quedaba con idControlCapacidad/idTaquilla/horaIni
            # vacios).
            horarios_individuales = []
            for grupo in horas_disponibles:
                if isinstance(grupo, dict) and isinstance(grupo.get("horarios"), list):
                    horarios_individuales.extend(grupo["horarios"])
                elif isinstance(grupo, dict) and "horaIni" in grupo:
                    horarios_individuales.append(grupo)  # por si algun dia SI viene plano
            if not horarios_individuales:
                resultado["mensaje"] = f"Se encontraron franjas para el {fecha_elegida}, pero ninguna tenia horarios individuales dentro."
                return resultado

            # Se busca la hora mas cercana a la hora aproximada pedida.
            # 'horaIni' viene como numero tipo 1027 = 10:27 -- se compara
            # solo la parte de la hora (// 100) contra la hora pedida, y
            # si no hay coincidencia exacta se toma la mas cercana.
            hora_pedida = int(solicitud["hora_aproximada"])
            mejor_horario = min(
                horarios_individuales,
                key=lambda h: abs((h.get("horaIni", 0) // 100) - hora_pedida)
            )
            print(f"{etiqueta} Horario elegido (mas cercano a las {hora_pedida}:00): {mejor_horario}", flush=True)

            # --- Elegir la hora en el <select> real ---
            id_control_capacidad = mejor_horario.get("idControlCapacidad")
            id_taquilla = mejor_horario.get("idTaquilla")
            hora_ini_valor = mejor_horario.get("horaIni")
            page.evaluate(f"""() => {{
                var el = document.querySelector('#agendarCitaHoraIniSelect');
                if (!el) return false;
                el.value = '{hora_ini_valor}';
                el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                return true;
            }}""")
            page.wait_for_timeout(1000)

            # --- Confirmado por pruebas reales que sede/fecha/hora se
            # eligen bien -- se procede a resolver el reCAPTCHA. Se
            # SIGUE deteniendo justo antes del envio final (el clic que
            # si reservaria la cita de verdad), para confirmar primero
            # que el captcha se acepta correctamente. ---
            print(f"{etiqueta} Resolviendo reCAPTCHA con 2captcha (puede tardar 15-40 segundos)...", flush=True)
            try:
                token_captcha = resolver_recaptcha_2captcha(ENVIGADO_RECAPTCHA_SITEKEY, page.url)
                print(f"{etiqueta} Token de 2captcha obtenido (primeros 30 caracteres): {token_captcha[:30]}...", flush=True)
            except Exception as e_captcha:
                resultado["mensaje"] = f"No se pudo resolver el reCAPTCHA: {e_captcha}"
                return resultado

            # Se inyecta el token en el textarea estandar de reCAPTCHA v2.
            # Se confirmo con diagnostico real que el sitio tiene un
            # <div id="widgetReCaptcha"> vacio -- el widget de Google
            # nunca termino de renderizarse ahi (posiblemente por el
            # entorno headless/automatizado). No hace falta el widget
            # visual para que el formulario acepte la respuesta: si el
            # textarea "g-recaptcha-response" no existe, se CREA
            # manualmente dentro de ese div (es exactamente lo que el
            # script de Google crea normalmente al renderizar), y se le
            # pone el token ahi.
            resultado_inyeccion = page.evaluate(f"""() => {{
                var resultado = {{ textarea_encontrado: false, textarea_creado: false, callback_llamado: false, callback_nombre: null }};
                var textarea = document.getElementById('g-recaptcha-response');
                if (!textarea) {{
                    var contenedor = document.getElementById('widgetReCaptcha') || document.body;
                    textarea = document.createElement('textarea');
                    textarea.id = 'g-recaptcha-response';
                    textarea.name = 'g-recaptcha-response';
                    textarea.style.width = '250px';
                    textarea.style.height = '40px';
                    textarea.style.border = '1px solid #c1c1c1';
                    textarea.style.margin = '10px 25px';
                    textarea.style.padding = '0px';
                    textarea.style.resize = 'none';
                    textarea.style.display = 'none';
                    contenedor.appendChild(textarea);
                    resultado.textarea_creado = true;
                }}
                textarea.style.display = 'block';
                textarea.value = {json.dumps(token_captcha)};
                textarea.dispatchEvent(new Event('change', {{ bubbles: true }}));
                textarea.dispatchEvent(new Event('input', {{ bubbles: true }}));
                resultado.textarea_encontrado = true;

                var widget = document.querySelector('.g-recaptcha[data-callback], div[data-callback]');
                if (widget) {{
                    var nombreCallback = widget.getAttribute('data-callback');
                    resultado.callback_nombre = nombreCallback;
                    if (nombreCallback && typeof window[nombreCallback] === 'function') {{
                        window[nombreCallback]({json.dumps(token_captcha)});
                        resultado.callback_llamado = true;
                    }}
                }}
                return resultado;
            }}""")
            print(f"{etiqueta} Resultado de inyectar el token: {resultado_inyeccion}", flush=True)

            # El sitio usa AngularJS (clasico) -- el boton "Confirmar cita"
            # probablemente esta deshabilitado por una expresion tipo
            # ng-disabled que Angular solo revisa de nuevo cuando corre su
            # "digest cycle" (no se entera de cambios hechos por fuera de
            # su propio framework, como crear un textarea a mano). Se
            # busca el scope de Angular mas cercano al boton y se le
            # fuerza un $apply() para que reevalue todas sus condiciones,
            # incluida la del boton.
            #
            # ADEMAS: si la condicion del boton llama directamente a
            # grecaptcha.getResponse() (la funcion real de Google, en vez
            # de leer el textarea), esa funcion no sabe nada de nuestro
            # token porque nunca se llamo grecaptcha.render() de verdad
            # -- se SOBRESCRIBE esa funcion para que siempre devuelva
            # nuestro token ya resuelto, sin importar el widget id.
            resultado_angular = page.evaluate(f"""() => {{
                var resultado = {{ angular_encontrado: false, apply_ok: false, boton_disabled_despues: null, grecaptcha_sobreescrito: false }};

                if (typeof grecaptcha !== 'undefined') {{
                    try {{
                        grecaptcha.getResponse = function(id) {{ return {json.dumps(token_captcha)}; }};
                        resultado.grecaptcha_sobreescrito = true;
                    }} catch (e) {{
                        resultado.error_grecaptcha = String(e);
                    }}
                }}

                if (typeof angular === 'undefined') return resultado;
                resultado.angular_encontrado = true;
                var boton = document.getElementById('btnGuardarCita');
                var el = boton || document.querySelector('[ng-app]') || document.body;
                try {{
                    var scope = angular.element(el).scope();
                    if (!scope) {{
                        // A veces el elemento exacto no tiene scope propio -- se busca hacia arriba.
                        var actual = el;
                        while (actual && !scope) {{
                            scope = angular.element(actual).scope();
                            actual = actual.parentElement;
                        }}
                    }}
                    if (scope) {{
                        scope.$apply();
                        resultado.apply_ok = true;
                    }}
                }} catch (e) {{
                    resultado.error = String(e);
                }}
                if (boton) resultado.boton_disabled_despues = boton.disabled;
                return resultado;
            }}""")
            print(f"{etiqueta} Resultado de forzar el digest de Angular: {resultado_angular}", flush=True)

            # El boton sigue deshabilitado a pesar de forzar el digest y
            # sobrescribir grecaptcha.getResponse -- en vez de seguir
            # adivinando el mecanismo, se lee DIRECTAMENTE del boton (y
            # de su scope de Angular) que atributos/variables controlan
            # su estado, para saber con certeza que falta.
            diagnostico_boton = page.evaluate("""() => {
                var boton = document.getElementById('btnGuardarCita');
                if (!boton) return { error: 'boton no encontrado' };
                var atributos = {};
                for (var i = 0; i < boton.attributes.length; i++) {
                    var a = boton.attributes[i];
                    atributos[a.name] = a.value;
                }
                var resultado = { atributos: atributos };
                try {
                    var scope = angular.element(boton).scope();
                    if (!scope) {
                        var actual = boton;
                        while (actual && !scope) {
                            scope = angular.element(actual).scope();
                            actual = actual.parentElement;
                        }
                    }
                    if (scope) {
                        // Se listan las propiedades del scope que parezcan
                        // relacionadas a captcha, formulario, o validez.
                        var propsRelevantes = {};
                        for (var key in scope) {
                            if (scope.hasOwnProperty(key) && /captcha|valid|form|disable|recaptcha/i.test(key)) {
                                try { propsRelevantes[key] = JSON.stringify(scope[key]).substring(0, 200); }
                                catch (e2) { propsRelevantes[key] = '(no se pudo convertir a texto)'; }
                            }
                        }
                        resultado.scope_propiedades_relevantes = propsRelevantes;
                    }
                } catch (e) {
                    resultado.error_scope = String(e);
                }
                return resultado;
            }""")
            print(f"{etiqueta} Diagnostico directo del boton (atributos y scope): {diagnostico_boton}", flush=True)

            # Confirmado por diagnostico real: el boton usa
            # ng-disabled="ctrl.isBloquearAgendarCita" -- es una
            # propiedad del CONTROLADOR (sintaxis "controller as ctrl"),
            # no del scope raiz directamente. Se pone en false y se
            # fuerza el digest de nuevo para que Angular actualice el
            # atributo "disabled" real del boton en el DOM.
            resultado_forzar_boton = page.evaluate("""() => {
                var resultado = { ctrl_encontrado: false, valor_anterior: null, valor_nuevo: null, boton_disabled_final: null };
                var boton = document.getElementById('btnGuardarCita');
                if (!boton) return resultado;
                try {
                    var scope = angular.element(boton).scope();
                    if (!scope) {
                        var actual = boton;
                        while (actual && !scope) {
                            scope = angular.element(actual).scope();
                            actual = actual.parentElement;
                        }
                    }
                    if (scope && scope.ctrl) {
                        resultado.ctrl_encontrado = true;
                        resultado.valor_anterior = scope.ctrl.isBloquearAgendarCita;
                        scope.$apply(function() {
                            scope.ctrl.isBloquearAgendarCita = false;
                        });
                        resultado.valor_nuevo = scope.ctrl.isBloquearAgendarCita;
                    }
                } catch (e) {
                    resultado.error = String(e);
                }
                resultado.boton_disabled_final = boton.disabled;
                return resultado;
            }""")
            print(f"{etiqueta} Resultado de forzar ctrl.isBloquearAgendarCita a false: {resultado_forzar_boton}", flush=True)

            # Si no se encontro el textarea en la pagina principal, puede
            # estar dentro de un iframe (asi renderiza normalmente
            # reCAPTCHA) -- se busca en TODOS los frames de la pagina, y
            # tambien se lista cualquier elemento relacionado a captcha
            # (iframe con "recaptcha" en el src, cualquier id/clase que
            # contenga "captcha") para entender la estructura real.
            if not resultado_inyeccion.get("textarea_encontrado"):
                print(f"{etiqueta} === DIAGNOSTICO: buscando captcha en iframes ===", flush=True)
                try:
                    todos_los_frames = page.frames
                    print(f"{etiqueta} Cantidad de frames en la pagina: {len(todos_los_frames)}", flush=True)
                    for f in todos_los_frames:
                        print(f"{etiqueta} Frame URL: {f.url}", flush=True)
                except Exception as e_frames:
                    print(f"{etiqueta} No se pudieron listar los frames: {e_frames}", flush=True)
                try:
                    elementos_captcha = page.evaluate("""() => {
                        var resultado = [];
                        document.querySelectorAll('[id*="captcha" i], [class*="captcha" i], iframe').forEach(function(el){
                            resultado.push({
                                tag: el.tagName, id: el.id || null, clase: el.className || null,
                                src: el.src || null,
                                visible: !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length)
                            });
                        });
                        return resultado;
                    }""")
                    print(f"{etiqueta} Elementos relacionados a captcha/iframes en la pagina principal: {elementos_captcha}", flush=True)
                except Exception as e_elcap:
                    print(f"{etiqueta} No se pudieron listar elementos de captcha: {e_elcap}", flush=True)
                print(f"{etiqueta} === FIN DIAGNOSTICO captcha ===", flush=True)

            page.wait_for_timeout(1500)

            # --- DIAGNOSTICO: se detiene ANTES de confirmar la cita de
            # verdad -- primero hay que revisar en estos logs que el
            # captcha se haya aceptado (ej. que no aparezca un mensaje de
            # error de captcha en la pagina), antes de dar el ultimo paso
            # que si aparta una cita real. ---
            print(f"{etiqueta} === DIAGNOSTICO: estado antes de confirmar (aun NO se ha reservado nada) ===", flush=True)
            print(f"{etiqueta} idControlCapacidad={id_control_capacidad}, idTaquilla={id_taquilla}, horaIni={hora_ini_valor}, fecha={fecha_elegida}", flush=True)
            try:
                print(f"{etiqueta} Texto visible de la pagina:", page.inner_text("body")[:2000], flush=True)
            except Exception:
                pass
            try:
                boton_confirmar = page.evaluate("""() => {
                    var botones = document.querySelectorAll('button, input[type="button"], input[type="submit"]');
                    var resultado = [];
                    botones.forEach(function(b){
                        if (b.offsetWidth || b.offsetHeight || b.getClientRects().length) {
                            resultado.push({tag: b.tagName, texto: (b.innerText||b.value||'').trim(), id: b.id||null, disabled: !!b.disabled});
                        }
                    });
                    return resultado;
                }""")
                print(f"{etiqueta} Botones visibles en este punto: {boton_confirmar}", flush=True)
            except Exception:
                pass

            # --- PASO FINAL: clic real en "Confirmar cita" -- esto SI
            # reserva la cita de verdad. Se captura la respuesta de
            # agendarCitaGAComponentes (el endpoint real de confirmacion,
            # ya venia previsto en _capturar_respuesta) para saber si
            # salio bien y obtener el numero de atencion. ---
            resultado["capturas"] = {}

            captura_antes = f"{etiqueta.strip('[]')}_1_antes_confirmar.png"
            try:
                page.screenshot(path=os.path.join(CAPTURAS_ENVIGADO_DIR, captura_antes), full_page=False, timeout=8000)
                resultado["capturas"]["antes_confirmar"] = captura_antes
                print(f"{etiqueta} Captura guardada: {captura_antes}", flush=True)
            except Exception as e_cap1:
                print(f"{etiqueta} No se pudo tomar captura antes de confirmar: {e_cap1}", flush=True)

            respuestas_capturadas.pop("agendarCitaGAComponentes", None)
            print(f"{etiqueta} Haciendo clic en 'Confirmar cita' -- ESTO RESERVA LA CITA DE VERDAD...", flush=True)
            try:
                page.click('#btnGuardarCita', timeout=8000)
            except Exception as e_clic_confirmar:
                resultado["mensaje"] = f"No se pudo hacer clic en 'Confirmar cita': {e_clic_confirmar}"
                return resultado

            for _ in range(15):
                if "agendarCitaGAComponentes" in respuestas_capturadas:
                    break
                page.wait_for_timeout(1000)

            respuesta_confirmacion = respuestas_capturadas.get("agendarCitaGAComponentes")
            print(f"{etiqueta} Respuesta de agendarCitaGAComponentes: {respuesta_confirmacion}", flush=True)

            captura_despues = f"{etiqueta.strip('[]')}_2_despues_confirmar.png"
            try:
                page.screenshot(path=os.path.join(CAPTURAS_ENVIGADO_DIR, captura_despues), full_page=False, timeout=8000)
                resultado["capturas"]["despues_confirmar"] = captura_despues
                print(f"{etiqueta} Captura guardada: {captura_despues}", flush=True)
            except Exception as e_cap2:
                print(f"{etiqueta} No se pudo tomar captura despues de confirmar: {e_cap2}", flush=True)

            try:
                print(f"{etiqueta} Texto visible de la pagina despues del clic final:", page.inner_text("body")[:2000], flush=True)
            except Exception:
                pass

            # Si no se capturo la respuesta esperada, se listan TODAS las
            # peticiones de red vistas justo despues del clic (por si el
            # endpoint real de confirmacion tiene otro nombre), y
            # cualquier mensaje de error visible en la pagina.
            if not respuesta_confirmacion:
                print(f"{etiqueta} === DIAGNOSTICO: no se capturo la confirmacion -- revisando alternativas ===", flush=True)
                print(f"{etiqueta} Todas las respuestas capturadas hasta ahora: {list(respuestas_capturadas.keys())}", flush=True)
                try:
                    posibles_errores = page.evaluate("""() => {
                        var resultado = [];
                        document.querySelectorAll('[class*="error" i], [class*="alert" i], .toast, .swal2-popup, .modal').forEach(function(el){
                            if (el.offsetWidth || el.offsetHeight || el.getClientRects().length) {
                                resultado.push({tag: el.tagName, clase: el.className, texto: (el.innerText||'').trim().substring(0, 300)});
                            }
                        });
                        return resultado;
                    }""")
                    print(f"{etiqueta} Elementos de error/alerta/modal visibles: {posibles_errores}", flush=True)
                except Exception as e_err:
                    print(f"{etiqueta} No se pudieron listar posibles errores: {e_err}", flush=True)
                print(f"{etiqueta} === FIN DIAGNOSTICO confirmacion ===", flush=True)

            if respuesta_confirmacion:
                # La estructura exacta se confirma con el resultado real
                # -- se buscan las claves mas probables para el numero de
                # atencion, sin asumir un unico nombre de campo.
                nro_atencion = None
                if isinstance(respuesta_confirmacion, dict):
                    for clave_posible in ("nroAtencion", "numeroAtencion", "nro_atencion", "codigo", "id"):
                        if respuesta_confirmacion.get(clave_posible):
                            nro_atencion = respuesta_confirmacion[clave_posible]
                            break
                resultado["exito"] = True
                resultado["nro_atencion"] = nro_atencion
                resultado["mensaje"] = f"Cita reservada exitosamente para el {fecha_elegida} a las {hora_ini_valor}." + (f" Nro. atencion: {nro_atencion}" if nro_atencion else "")
                resultado["detalle"] = respuesta_confirmacion
            else:
                resultado["mensaje"] = "Se hizo clic en 'Confirmar cita' pero no se pudo confirmar el resultado (revisar diagnostico en los logs y las capturas -- puede que si se haya reservado)."
            return resultado

        except Exception as e:
            print(f"{etiqueta} Error en el flujo de reserva: {e}", flush=True)
            resultado["mensaje"] = str(e)
            try:
                print(f"{etiqueta} Texto visible al momento del error:", page.inner_text("body")[:2000], flush=True)
            except Exception:
                pass
            return resultado
        finally:
            context.close(); browser.close()




ENVIGADO_TURNOS_API = "https://gacomponentes.envigado.gov.co/backga/back-ga/turnos/findAtencionesMonitor"

# Bandera simple en memoria para saber si ya hay una sesion de monitoreo
# corriendo (evita que se disparen varias sesiones encimadas por error).
_envigado_monitoreo_estado = {
    "activo": False, "inicio": None, "fin_esperado": None,
    "numeros_vigilados": [], "detener": False
}

# Misma idea pero para el monitoreo CONSTANTE de citas disponibles
# (revisa cada 30 segundos, separado del monitoreo de turnos llamados).
_envigado_citas_monitoreo_estado = {
    "activo": False, "inicio": None, "fin_esperado": None, "detener": False
}


def _envigado_procesar_cola_solicitudes():
    """Revisa la cola de solicitudes de citas PENDIENTES, e intenta
    reservar cada una (llamando al flujo completo de reserva). Se llama
    solo cuando el monitoreo YA confirmo que hay citas disponibles en
    algun lado -- asi no se intenta el flujo completo (mas pesado) en
    cada ciclo de 30 segundos sin necesidad."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombres, apellidos, tipo_documento, numero_documento, correo, celular,
                   placa, id_servicio, sede_preferida, hora_aproximada
            FROM envigado_citas_solicitudes WHERE estado = 'pendiente' ORDER BY id ASC
        """)
        pendientes = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error leyendo cola de solicitudes de citas: {e}", flush=True)
        return

    if not pendientes:
        return

    print(f"Hay {len(pendientes)} solicitud(es) pendiente(s) en la cola -- intentando reservar...", flush=True)
    for fila in pendientes:
        solicitud = {
            "id": fila[0], "nombres": fila[1], "apellidos": fila[2], "tipo_documento": fila[3],
            "numero_documento": fila[4], "correo": fila[5], "celular": fila[6], "placa": fila[7],
            "id_servicio": fila[8], "sede_preferida": fila[9], "hora_aproximada": fila[10],
        }
        try:
            resultado = envigado_reservar_cita(solicitud)
        except Exception as e:
            resultado = {"exito": False, "mensaje": str(e), "nro_atencion": None}

        try:
            conn2 = get_db_conn()
            cur2 = conn2.cursor()
            if resultado.get("exito"):
                cur2.execute("""
                    UPDATE envigado_citas_solicitudes
                    SET estado = 'reservada', nro_atencion = %s, detalle_reserva = %s, reservado_en = NOW()
                    WHERE id = %s
                """, (resultado.get("nro_atencion"), json.dumps(resultado.get("detalle") or {}), solicitud["id"]))
                enviar_notificacion_push(
                    "¡Cita reservada en Envigado!",
                    f"{solicitud['nombres']} {solicitud['apellidos']} -- Nro. atención: {resultado.get('nro_atencion')}",
                    "/ejecucion.html"
                )
            else:
                # Se queda en "pendiente" (no se marca error) si el
                # problema fue simplemente que no habia sede/fecha/hora
                # que coincidiera todavia -- se reintenta en el proximo
                # ciclo. Solo se marca "error" si fue un problema real
                # de datos/tecnico persistente.
                print(f"No se pudo reservar la solicitud {solicitud['id']} en este intento: {resultado.get('mensaje')}", flush=True)
            conn2.commit()
            cur2.close(); conn2.close()
        except Exception as e:
            print(f"Error actualizando el estado de la solicitud {solicitud['id']}: {e}", flush=True)


def _envigado_polling_citas(duracion_segundos, intervalo_segundos=30):
    """Revisa si hay citas disponibles en Envigado cada 'intervalo_segundos'
    (30 por defecto), durante 'duracion_segundos'. Usa la misma logica que
    ya existe (envigado_revisar_citas_disponibles), que guarda el
    resultado en la base de datos -- el aviso en Liquidacion y el boton
    de revision manual en Ejecucion ya leen de ahi, asi que no hace falta
    nada adicional para que se vea el resultado. Ademas, manda una
    notificacion push -- pero solo quando PASA de "sin citas" a "con
    citas" (no en cada ciclo de 30 segundos mientras sigan disponibles,
    para no saturar de notificaciones repetidas). Si hay citas, tambien
    intenta procesar la cola de solicitudes pendientes."""
    tenia_citas_antes = False
    fin = time.time() + duracion_segundos
    while time.time() < fin:
        if _envigado_citas_monitoreo_estado["detener"]:
            break
        try:
            resultados, hubo_error = envigado_revisar_citas_disponibles()
            hay_citas_ahora = bool(resultados) and not hubo_error
            if hay_citas_ahora and not tenia_citas_antes:
                enviar_notificacion_push(
                    "¡Hay citas disponibles en Envigado!",
                    "Se encontró disponibilidad para agendar. Revisa Tramy para más detalles.",
                    "/liquidacion.html"
                )
            tenia_citas_antes = hay_citas_ahora
            if hay_citas_ahora:
                _envigado_procesar_cola_solicitudes()
        except Exception as e:
            print(f"Error en monitoreo constante de citas Envigado: {e}", flush=True)
        time.sleep(intervalo_segundos)

    _envigado_citas_monitoreo_estado["activo"] = False
    _envigado_citas_monitoreo_estado["detener"] = False


# Igual que el de Envigado, pero para Medellin -- se guarda el ultimo
# hallazgo EN MEMORIA (se pierde si el servidor se reinicia, pero eso ya
# se acepta igual para el resto de estos monitoreos en vivo) para que el
# aviso en Ejecucion pueda mostrar el resultado sin tener que esperar el
# siguiente ciclo.
_medellin_citas_monitoreo_estado = {
    "activo": False, "inicio": None, "fin_esperado": None, "detener": False,
    "ultimo_hallazgo": None,
}

# Estado del monitor "espejo" que SIEMPRE usa el proxy de IPRoyal --
# separado por completo del monitor normal de arriba, para que el
# usuario sepa con certeza cual de los dos esta gastando saldo del
# proxy y cual no.
_medellin_citas_proxy_monitoreo_estado = {
    "activo": False, "inicio": None, "fin_esperado": None, "detener": False,
    "ultimo_hallazgo": None,
}


def _medellin_polling_citas(usuario, password, placa, id_servicio, duracion_segundos, intervalo_segundos=30, sede_deseada=None):
    """Revisa si hay citas disponibles en Medellin cada 'intervalo_segundos'
    (30 por defecto), durante 'duracion_segundos'. A diferencia de
    Envigado, aqui hay que volver a iniciar sesion en cada ciclo (cada
    revision abre su propia sesion de Playwright). Manda notificacion
    push solo la PRIMERA vez que encuentra citas (no en cada ciclo)."""
    fin = time.time() + duracion_segundos
    while time.time() < fin:
        if _medellin_citas_monitoreo_estado["detener"]:
            break
        try:
            hay_citas, detalle = medellin_hay_citas_disponibles(usuario, password, placa, id_servicio, sede_deseada=sede_deseada)
            if hay_citas:
                ya_habia_hallazgo = _medellin_citas_monitoreo_estado["ultimo_hallazgo"] is not None
                _medellin_citas_monitoreo_estado["ultimo_hallazgo"] = {
                    "detalle": detalle,
                    "encontrado_en": datetime.now().isoformat() + "Z",  # UTC -- el navegador lo convierte solo a hora local
                }
                if not ya_habia_hallazgo:
                    sede = (detalle or {}).get("sede", "")
                    enviar_notificacion_push(
                        "¡Hay citas disponibles en Medellín!",
                        f"Sede: {sede}. Revisa Tramy para más detalles." if sede else "Revisa Tramy para más detalles.",
                        "/ejecucion.html"
                    )
        except Exception as e:
            print(f"Error en monitoreo constante de citas Medellin: {e}", flush=True)
        time.sleep(intervalo_segundos)

    _medellin_citas_monitoreo_estado["activo"] = False
    _medellin_citas_monitoreo_estado["detener"] = False


# ============================================================
# PROGRAMADOR AUTOMATICO 24/7 -- corre solo, sin que nadie tenga
# que darle "Iniciar" cada dia. Revisa cada monitor SOLO dentro
# de su horario configurado, con el intervalo que se haya puesto
# en el panel de Ejecucion (se puede cambiar en caliente, sin
# reiniciar el servidor -- se lee de la base de datos en cada
# vuelta).
# ============================================================

def _monitoreo_config_leer(monitor):
    """Lee la configuracion actual de un monitor (activo, intervalo,
    horario, credenciales) desde la base de datos."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT activo, intervalo_segundos, hora_inicio, hora_fin, usuario, password, placa, sede
            FROM monitoreo_config WHERE monitor = %s
        """, (monitor,))
        fila = cur.fetchone()
        cur.close(); conn.close()
        if not fila:
            return None
        return {
            "activo": fila[0], "intervalo_segundos": fila[1],
            "hora_inicio": fila[2], "hora_fin": fila[3],
            "usuario": fila[4], "password": fila[5], "placa": fila[6],
            "sede": fila[7],
        }
    except Exception as e:
        print(f"Error leyendo config de monitoreo ({monitor}): {e}", flush=True)
        return None


def _monitoreo_config_guardar(monitor, **campos):
    """Actualiza uno o varios campos de la configuracion de un monitor."""
    if not campos:
        return False
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        sets = ", ".join(f"{k} = %s" for k in campos.keys())
        valores = list(campos.values()) + [monitor]
        cur.execute(f"UPDATE monitoreo_config SET {sets}, actualizado_en = NOW() WHERE monitor = %s", valores)
        conn.commit()
        cur.close(); conn.close()
        return True
    except Exception as e:
        print(f"Error guardando config de monitoreo ({monitor}): {e}", flush=True)
        return False


def _dentro_de_horario(hora_inicio_str, hora_fin_str):
    """Revisa si la hora actual EN COLOMBIA esta dentro del rango dado
    (formato 'HH:MM'). Soporta rangos normales (ej 07:00-17:00)."""
    ahora = (datetime.utcnow() - timedelta(hours=5)).time()
    try:
        h_ini = datetime.strptime(hora_inicio_str, "%H:%M").time()
        h_fin = datetime.strptime(hora_fin_str, "%H:%M").time()
    except Exception:
        return False
    if h_ini <= h_fin:
        return h_ini <= ahora <= h_fin
    return ahora >= h_ini or ahora <= h_fin  # rango que cruza la medianoche


# Estado en vivo del programador automatico -- lo que el panel de
# Ejecucion consulta para saber si de verdad esta corriendo ahora mismo
# (no solo si esta "activado" en la configuracion).
_programador_automatico_estado = {
    "envigado_citas": {"ultima_revision": None, "dentro_de_horario": False, "ultimo_error": None},
    "medellin_citas": {"ultima_revision": None, "dentro_de_horario": False, "ultimo_error": None},
    "medellin_citas_proxy": {"ultima_revision": None, "dentro_de_horario": False, "ultimo_error": None},
}


def _programador_automatico_loop():
    """Hilo unico que corre para siempre desde que arranca el servidor.
    Cada 10 segundos revisa la configuracion de cada monitor (activo,
    horario, intervalo) y decide si le toca consultar en esta vuelta."""
    ultimo_chequeo = {"envigado_citas": 0, "medellin_citas": 0, "medellin_citas_proxy": 0}

    while True:
        try:
            ahora_ts = time.time()

            # --- Envigado ---
            cfg = _monitoreo_config_leer("envigado_citas")
            dentro_horario_env = bool(cfg) and _dentro_de_horario(cfg["hora_inicio"], cfg["hora_fin"])
            _programador_automatico_estado["envigado_citas"]["dentro_de_horario"] = dentro_horario_env
            if cfg and cfg["activo"] and dentro_horario_env:
                if ahora_ts - ultimo_chequeo["envigado_citas"] >= cfg["intervalo_segundos"]:
                    ultimo_chequeo["envigado_citas"] = ahora_ts
                    try:
                        resultados, hubo_error = envigado_revisar_citas_disponibles()
                        hay_citas_ahora = bool(resultados) and not hubo_error
                        ya_habia = _envigado_citas_monitoreo_estado.get("_auto_tenia_citas", False)
                        if hay_citas_ahora and not ya_habia:
                            enviar_notificacion_push(
                                "¡Hay citas disponibles en Envigado!",
                                "Se encontró disponibilidad para agendar. Revisa Tramy para más detalles.",
                                "/liquidacion.html"
                            )
                        _envigado_citas_monitoreo_estado["_auto_tenia_citas"] = hay_citas_ahora
                        if hay_citas_ahora:
                            _envigado_procesar_cola_solicitudes()
                        _programador_automatico_estado["envigado_citas"]["ultima_revision"] = datetime.now().isoformat() + "Z"
                        _programador_automatico_estado["envigado_citas"]["ultimo_error"] = None
                    except Exception as e:
                        print(f"Error en programador automatico (Envigado): {e}", flush=True)
                        _programador_automatico_estado["envigado_citas"]["ultimo_error"] = str(e)

            # --- Medellin ---
            cfg = _monitoreo_config_leer("medellin_citas")
            dentro_horario_med = bool(cfg) and _dentro_de_horario(cfg["hora_inicio"], cfg["hora_fin"])
            _programador_automatico_estado["medellin_citas"]["dentro_de_horario"] = dentro_horario_med
            if cfg and cfg["activo"] and dentro_horario_med:
                if ahora_ts - ultimo_chequeo["medellin_citas"] >= cfg["intervalo_segundos"]:
                    ultimo_chequeo["medellin_citas"] = ahora_ts
                    if cfg["usuario"] and cfg["password"] and cfg["placa"]:
                        try:
                            hay_citas, detalle = medellin_hay_citas_disponibles(cfg["usuario"], cfg["password"], cfg["placa"], sede_deseada=cfg.get("sede"))
                            if hay_citas:
                                ya_habia = _medellin_citas_monitoreo_estado["ultimo_hallazgo"] is not None
                                _medellin_citas_monitoreo_estado["ultimo_hallazgo"] = {
                                    "detalle": detalle,
                                    "encontrado_en": datetime.now().isoformat() + "Z",
                                }
                                if not ya_habia:
                                    sede = (detalle or {}).get("sede", "")
                                    enviar_notificacion_push(
                                        "¡Hay citas disponibles en Medellín!",
                                        f"Sede: {sede}. Revisa Tramy para más detalles." if sede else "Revisa Tramy para más detalles.",
                                        "/ejecucion.html"
                                    )
                            else:
                                _medellin_citas_monitoreo_estado["ultimo_hallazgo"] = None
                            _programador_automatico_estado["medellin_citas"]["ultima_revision"] = datetime.now().isoformat() + "Z"
                            _programador_automatico_estado["medellin_citas"]["ultimo_error"] = None
                        except Exception as e:
                            print(f"Error en programador automatico (Medellin): {e}", flush=True)
                            _programador_automatico_estado["medellin_citas"]["ultimo_error"] = str(e)
                    else:
                        print("Programador automatico Medellin activo pero faltan usuario/password/placa en la configuracion.", flush=True)

            # --- Medellin (ESPEJO, siempre con proxy de IPRoyal) ---
            # Modulo identico al de arriba, pero completamente separado
            # -- este SIEMPRE gasta saldo del proxy, mientras el de
            # arriba nunca lo toca. Asi el usuario sabe con certeza cual
            # de los dos esta usando el proxy en cada momento.
            cfg_p = _monitoreo_config_leer("medellin_citas_proxy")
            dentro_horario_med_p = bool(cfg_p) and _dentro_de_horario(cfg_p["hora_inicio"], cfg_p["hora_fin"])
            _programador_automatico_estado["medellin_citas_proxy"]["dentro_de_horario"] = dentro_horario_med_p
            if cfg_p and cfg_p["activo"] and dentro_horario_med_p:
                if ahora_ts - ultimo_chequeo["medellin_citas_proxy"] >= cfg_p["intervalo_segundos"]:
                    ultimo_chequeo["medellin_citas_proxy"] = ahora_ts
                    if cfg_p["usuario"] and cfg_p["password"] and cfg_p["placa"]:
                        try:
                            hay_citas_p, detalle_p = medellin_hay_citas_disponibles(cfg_p["usuario"], cfg_p["password"], cfg_p["placa"], sede_deseada=cfg_p.get("sede"), usar_proxy=True)
                            if hay_citas_p:
                                ya_habia_p = _medellin_citas_proxy_monitoreo_estado["ultimo_hallazgo"] is not None
                                _medellin_citas_proxy_monitoreo_estado["ultimo_hallazgo"] = {
                                    "detalle": detalle_p,
                                    "encontrado_en": datetime.now().isoformat() + "Z",
                                }
                                if not ya_habia_p:
                                    sede_p = (detalle_p or {}).get("sede", "")
                                    enviar_notificacion_push(
                                        "¡Hay citas disponibles en Medellín! (proxy)",
                                        f"Sede: {sede_p}. Revisa Tramy para más detalles." if sede_p else "Revisa Tramy para más detalles.",
                                        "/ejecucion.html"
                                    )
                            else:
                                _medellin_citas_proxy_monitoreo_estado["ultimo_hallazgo"] = None
                            _programador_automatico_estado["medellin_citas_proxy"]["ultima_revision"] = datetime.now().isoformat() + "Z"
                            _programador_automatico_estado["medellin_citas_proxy"]["ultimo_error"] = None
                        except Exception as e:
                            print(f"Error en programador automatico (Medellin proxy): {e}", flush=True)
                            _programador_automatico_estado["medellin_citas_proxy"]["ultimo_error"] = str(e)
                    else:
                        print("Programador automatico Medellin (proxy) activo pero faltan usuario/password/placa en la configuracion.", flush=True)

        except Exception as e:
            print(f"Error general en el programador automatico: {e}", flush=True)

        time.sleep(10)


# Se arranca UNA sola vez, apenas se carga la aplicacion (no depende de
# que nadie visite ninguna pagina ni le de clic a nada).
threading.Thread(target=_programador_automatico_loop, daemon=True).start()


def _envigado_polling_turnos(duracion_segundos=7200, intervalo_segundos=8, id_monitor=3, numeros_vigilados=None, placas_por_numero=None):
    """Revisa el "monitor de turnos" de Envigado cada pocos segundos,
    durante 'duracion_segundos'. Cada vez que aparece un idGestionAtencion
    que no habiamos visto, lo guarda con la hora en que Tramy lo detecto
    (la plataforma no entrega un timestamp exacto propio del llamado --
    el campo 'fechaInicial' viene vacio).
    Si se indica 'numeros_vigilados' (lista, ej. ["C-89", "G-78"]), en
    cuanto aparezca CUALQUIERA de esos numeros se agrega a
    _envigado_monitoreo_estado['encontrados'] para que el frontend pueda
    mostrar una alerta destacada (con sonido y vibracion).
    'placas_por_numero' (dict opcional, ej. {"C-89": "ABC123"}) -- si el
    usuario indico la placa de esa cita al agregarla a vigilar, se guarda
    junto con el turno capturado para mostrarla en la alerta."""
    fin = time.time() + duracion_segundos
    ids_vistos = set()
    hoy_str = datetime.now(TZ_COLOMBIA).strftime("%d/%m/%Y")
    numeros_vigilados_norm = set((n or "").strip().upper() for n in (numeros_vigilados or []))
    placas_por_numero = placas_por_numero or {}

    while time.time() < fin:
        if _envigado_monitoreo_estado["detener"]:
            break
        try:
            params = {
                "idMonitor": id_monitor,
                "cantidadTurnos": 20,
                "fechaInicio": f"{hoy_str} 00:00:00",
                "fechaFin": f"{hoy_str} 23:59:59",
                "_": str(int(time.time() * 1000)),
            }
            r = requests.get(ENVIGADO_TURNOS_API, params=params, timeout=15)
            r.raise_for_status()
            data = r.json()
            if isinstance(data, list) and data:
                conn = get_db_conn()
                cur = conn.cursor()
                for item in data:
                    idg = item.get("idGestionAtencion")
                    if idg is None or idg in ids_vistos:
                        continue
                    ids_vistos.add(idg)
                    nro_norm = (item.get("nroAtencion") or "").strip().upper()
                    es_vigilado = bool(numeros_vigilados_norm and nro_norm in numeros_vigilados_norm)
                    placa_asociada = placas_por_numero.get(nro_norm, "")

                    cur.execute("""
                        INSERT INTO envigado_turnos_llamados
                            (id_gestion_atencion, nro_atencion, nombre_usuario, nombre_taquilla,
                             nombre_servicio, id_estado, fue_vigilado, placa, detectado_en)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
                        ON CONFLICT (id_gestion_atencion) DO NOTHING
                    """, (idg, item.get("nroAtencion"), item.get("nombreUsuario"),
                          item.get("nombreTaquilla"), item.get("nombreServicio"),
                          item.get("idEstadoGestionAtencion"), es_vigilado, placa_asociada))

                    if es_vigilado:
                        cur.execute("""
                            UPDATE envigado_citas_vigiladas_historial
                            SET encontrado = TRUE, taquilla = %s, nombre_usuario = %s, detectado_en = NOW()
                            WHERE id = (
                                SELECT id FROM envigado_citas_vigiladas_historial
                                WHERE numero = %s AND encontrado = FALSE AND fecha_cita = CURRENT_DATE
                                ORDER BY creado_en DESC LIMIT 1
                            )
                        """, (item.get("nombreTaquilla"), item.get("nombreUsuario"), nro_norm))
                conn.commit()
                cur.close(); conn.close()
        except Exception as e:
            print(f"Error en monitoreo de turnos Envigado: {e}", flush=True)
        time.sleep(intervalo_segundos)

    _envigado_monitoreo_estado["activo"] = False
    _envigado_monitoreo_estado["detener"] = False


def _envigado_polling_turnos_con_espera(espera_segundos, duracion_segundos, **kwargs_polling):
    """Espera 'espera_segundos' antes de arrancar el monitoreo real -- para
    poder programar un inicio en el futuro (ej. 5 minutos antes de la
    hora de una cita), sin necesitar un servicio de tareas programadas
    aparte. El estado 'activo' ya queda en True desde que se programa
    (no solo cuando arranca de verdad), para que el boton de iniciar se
    bloquee de una vez y no se pueda programar dos veces por error."""
    if espera_segundos > 0:
        time.sleep(espera_segundos)
    if _envigado_monitoreo_estado["detener"]:
        _envigado_monitoreo_estado["activo"] = False
        _envigado_monitoreo_estado["detener"] = False
        return
    _envigado_polling_turnos(duracion_segundos=duracion_segundos, **kwargs_polling)


def cache_antioquia_guardar_paz_salvo(placa, avaluo, estado_veh):
    """Guarda en caché que la placa está a paz y salvo hasta fin de año."""
    try:
        anio_actual = datetime.now().year
        expira = f"{anio_actual}-12-31"
        retefuente = round(avaluo / 100) if avaluo else 0
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO cache_impuestos_antioquia
                (placa, vigencia, total_pagar, avaluo_comercial, retefuente, estado, expira_en, creado_en)
            VALUES (%s, %s, 0, %s, %s, 'PAZ_Y_SALVO', %s, NOW())
            ON CONFLICT (placa, vigencia) DO UPDATE SET
                total_pagar=0, avaluo_comercial=EXCLUDED.avaluo_comercial,
                retefuente=EXCLUDED.retefuente, estado='PAZ_Y_SALVO',
                expira_en=EXCLUDED.expira_en, actualizado_en=NOW()
        """, (placa.upper(), str(anio_actual), avaluo or 0, retefuente, expira))
        conn.commit()
        cur.close(); conn.close()
        print(f"  → Caché guardado PAZ_Y_SALVO para {placa}")
    except Exception as e:
        print(f"Error cache guardar paz y salvo: {e}")


def guardar_estado_cuenta_antioquia(placa, data3):
    """Guarda TODOS los datos de la consulta (estadoCuenta, historial de
    declaraciones, procesos fiscales, bloqueos, novedades) para poder
    generar despues el documento Estado de Cuenta sin tener que volver a
    consultar. Se llama cada vez que un vehiculo sale a paz y salvo."""
    try:
        estado_veh = data3.get("estadoCuenta", {}) or {}
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO estado_cuenta_antioquia
                (placa, numero_certificado, valor_certificado, periodo_inicio, periodo_fin,
                 estado_cuenta_json, lista_detalle_pagos, lista_proceso_fiscal, lista_bloqueo,
                 novedades, actualizado_en)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (placa) DO UPDATE SET
                numero_certificado=EXCLUDED.numero_certificado,
                valor_certificado=EXCLUDED.valor_certificado,
                periodo_inicio=EXCLUDED.periodo_inicio,
                periodo_fin=EXCLUDED.periodo_fin,
                estado_cuenta_json=EXCLUDED.estado_cuenta_json,
                lista_detalle_pagos=EXCLUDED.lista_detalle_pagos,
                lista_proceso_fiscal=EXCLUDED.lista_proceso_fiscal,
                lista_bloqueo=EXCLUDED.lista_bloqueo,
                novedades=EXCLUDED.novedades,
                actualizado_en=NOW()
        """, (
            placa.upper(),
            estado_veh.get("numeroCertificadoSap"),
            estado_veh.get("valorEstadoCuenta"),
            estado_veh.get("periodoInicioCertificacion"),
            estado_veh.get("periodoFinCertificacion"),
            json.dumps(estado_veh, ensure_ascii=False, default=str),
            json.dumps(data3.get("listaDetallePagos", []), ensure_ascii=False, default=str),
            json.dumps(data3.get("listaProcesoFiscal", []), ensure_ascii=False, default=str),
            json.dumps(data3.get("listaBloqueo", []), ensure_ascii=False, default=str),
            json.dumps(data3.get("novedades", []), ensure_ascii=False, default=str),
        ))
        conn.commit()
        cur.close(); conn.close()
        print(f"  → Estado de Cuenta guardado para {placa}")
    except Exception as e:
        print(f"Error guardando estado de cuenta: {e}")


def cache_antioquia_guardar_deuda(placa, vigencias_data, avaluo):
    """Guarda en caché vigencias con deuda.
    - Vigencia año actual antes del 1 agosto: expira el 31 de julio
    - Vigencia año actual desde 1 agosto, y vigencias anteriores: expira en 2 meses
    """
    try:
        ahora       = datetime.now()
        anio_actual = ahora.year
        conn = get_db_conn()
        cur  = conn.cursor()
        retefuente = round(avaluo / 100) if avaluo else 0

        for vig in vigencias_data:
            anio_vig   = int(vig.get('vigencia', 0))
            total      = vig.get('total_pagar', 0) or 0

            # Calcular expiración según reglas
            es_anio_actual    = (anio_vig == anio_actual)
            antes_de_agosto   = ahora.month < 8  # antes del 1 de agosto

            if es_anio_actual and antes_de_agosto:
                expira = f"{anio_actual}-07-31 23:59:59"
            else:
                # 2 meses desde ahora
                mes_exp  = ahora.month + 2
                anio_exp = anio_actual
                if mes_exp > 12:
                    mes_exp  -= 12
                    anio_exp += 1
                expira = f"{anio_exp}-{mes_exp:02d}-{ahora.day:02d} 23:59:59"

            expira_date = expira[:10]  # solo YYYY-MM-DD para columna date
            cur.execute("""
                INSERT INTO cache_impuestos_antioquia
                    (placa, vigencia, total_pagar, avaluo_comercial, retefuente, estado, expira_en, creado_en)
                VALUES (%s, %s, %s, %s, %s, 'CON_DEUDA', %s, NOW())
                ON CONFLICT (placa, vigencia) DO UPDATE SET
                    total_pagar=EXCLUDED.total_pagar,
                    avaluo_comercial=EXCLUDED.avaluo_comercial,
                    retefuente=EXCLUDED.retefuente,
                    estado='CON_DEUDA',
                    expira_en=EXCLUDED.expira_en,
                    actualizado_en=NOW()
            """, (placa.upper(), int(anio_vig), total, avaluo or 0, retefuente, expira_date))

        conn.commit()
        cur.close(); conn.close()
        print(f"  → Caché CON_DEUDA guardado para {placa}: {[v.get('vigencia') for v in vigencias_data]}")
    except Exception as e:
        print(f"Error cache guardar deuda: {e}")


# ============================================================
#  CACHE IMPUESTOS MUNICIPALES (Envigado, Sabaneta, Itagui, Bello,
#  La Estrella, Medellin, etc.) -- mismo principio que el cache de
#  Antioquia: si una placa esta a paz y salvo, lo esta hasta fin de
#  año, asi que no hace falta volver a consultar la pagina del
#  municipio (que es una consulta lenta via Playwright).
# ============================================================

def cache_municipal_buscar(placa, municipio):
    """Busca PAZ_Y_SALVO en cache municipal para el año actual."""
    try:
        anio_actual = datetime.now().year
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT fecha_pago, marca_pago, valor_pago, placa_vista
            FROM cache_impuestos_municipales
            WHERE placa = %s AND municipio = %s AND vigencia = %s AND estado = 'PAZ_Y_SALVO'
              AND (expira_en IS NULL OR expira_en >= NOW())
            ORDER BY creado_en DESC LIMIT 1
        """, (placa.upper(), municipio.lower(), str(anio_actual)))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return {"fecha_pago": row[0] or "", "marca": row[1] or "", "valor_pago": row[2] or "", "placa_vista": row[3] or ""}
        return None
    except Exception as e:
        print(f"Error cache municipal buscar: {e}")
        return None


def cache_municipal_guardar_paz_salvo(placa, municipio, fecha_pago, marca, valor_pago, placa_vista):
    """Guarda en cache que la placa esta a paz y salvo en ese municipio hasta fin de año."""
    try:
        anio_actual = datetime.now().year
        expira = f"{anio_actual}-12-31"
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO cache_impuestos_municipales
                (placa, municipio, vigencia, estado, fecha_pago, marca_pago, valor_pago, placa_vista, expira_en, creado_en)
            VALUES (%s, %s, %s, 'PAZ_Y_SALVO', %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (placa, municipio, vigencia) DO UPDATE SET
                estado='PAZ_Y_SALVO', fecha_pago=EXCLUDED.fecha_pago,
                marca_pago=EXCLUDED.marca_pago, valor_pago=EXCLUDED.valor_pago,
                placa_vista=EXCLUDED.placa_vista, expira_en=EXCLUDED.expira_en,
                actualizado_en=NOW()
        """, (placa.upper(), municipio.lower(), str(anio_actual), fecha_pago or '', marca or '', valor_pago or '', placa_vista or '', expira))
        conn.commit()
        cur.close(); conn.close()
        print(f"  → Cache municipal guardado PAZ_Y_SALVO para {placa} en {municipio}")
    except Exception as e:
        print(f"Error cache municipal guardar: {e}")


def resolver_captcha_imagen_2captcha(imagen_base64, intentos=3):
    """Resuelve un captcha de imagen simple (texto distorsionado) con 2Captcha.
    A diferencia de resolver_recaptcha_2captcha (que usa 'userrecaptcha'/'turnstile'
    con un sitekey), esto manda la imagen directamente y 2Captcha devuelve el texto
    que un humano leeria en ella. Se usa para el captcha del RUNT."""
    ultimo_error = None
    for intento in range(intentos):
        try:
            resp = requests.post("https://2captcha.com/in.php", data={
                "key": TWOCAPTCHA_API_KEY, "method": "base64",
                "body": imagen_base64, "json": 1,
            }, timeout=15)
            data = resp.json()
            if data.get("status") != 1:
                raise Exception(f"2captcha error: {data.get('request')}")
            captcha_id = data["request"]

            for _ in range(24):  # hasta 2 minutos (24 x 5s)
                time.sleep(5)
                resp2 = requests.get("https://2captcha.com/res.php", params={
                    "key": TWOCAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1,
                }, timeout=15)
                data2 = resp2.json()
                if data2.get("status") == 1:
                    return data2["request"]
                if data2.get("request") != "CAPCHA_NOT_READY":
                    raise Exception(f"2captcha error: {data2.get('request')}")

            raise Exception("2captcha timeout esperando solucion")
        except Exception as e:
            ultimo_error = e
            print(f"  → Intento {intento+1} de captcha imagen fallo: {e}")
    raise ultimo_error


RUNT_URL = "https://portalpublico.runt.gov.co/#/consulta-vehiculo/consulta/consulta-ciudadana"
RUNT_TIPO_DOC_MAP = {
    "CC": "Cédula Ciudadanía", "CE": "Cédula Extranjería", "NIT": "NIT",
    "PASAP": "Pasaporte", "TI": "Tarjeta Identidad", "PPT": "Permiso por Protección Temporal",
}


def _runt_seleccionar_mat_select(page, form_control, texto_opcion):
    """Los <select> de Angular Material no son <select> normales -- hay que
    hacer click para abrir el desplegable y luego click en la opcion deseada."""
    page.click(f'mat-select[formcontrolname="{form_control}"]')
    page.click(f'mat-option:has-text("{texto_opcion}")')


def consultar_runt_vehiculo(page, placa, cedula, tipo_documento="CC", job_id=None):
    """Consulta 'Placa y Propietario' en el RUNT. Devuelve un dict con los
    campos ya organizados segun el esquema de la tabla 'vehiculos', mas un
    sub-dict 'persona' si el RUNT tambien confirmo datos basicos del
    propietario en esta misma consulta."""
    if job_id:
        job_actualizar(job_id, "Abriendo el RUNT...", "procesando")

    # "networkidle" esperaba a que la red quedara completamente en reposo,
    # pero el portal del RUNT tiene actividad de fondo constante que nunca
    # la deja quieta del todo -- causaba timeouts frecuentes aunque la
    # pagina ya hubiera cargado bien. "domcontentloaded" resulto ser
    # demasiado temprano (la pagina apenas empieza a construirse), asi
    # que se usa "load" (espera a que terminen de cargar los recursos,
    # un punto medio) y se le da al formulario mas tiempo real de sobra
    # para aparecer -- la idea es reducir cuantas veces hace falta
    # reintentar, no solo agregar mas intentos.
    page.goto(RUNT_URL, wait_until="load", timeout=60000)
    try:
        page.wait_for_selector('input[formcontrolname="placa"]', timeout=45000)
    except Exception as e_primer_intento:
        # El portal del RUNT a veces tarda mas de lo normal en terminar de
        # cargar (o se queda a medias) -- antes de rendirse del todo, se
        # intenta una vez mas con una recarga completa, en vez de fallar
        # de una con el primer timeout.
        print(f"Timeout esperando el formulario del RUNT, reintentando con recarga completa: {e_primer_intento}", flush=True)
        if job_id:
            job_actualizar(job_id, "El RUNT tardó más de lo normal, reintentando...", "procesando")
        try:
            page.goto(RUNT_URL, wait_until="load", timeout=60000)
            page.wait_for_selector('input[formcontrolname="placa"]', timeout=45000)
        except Exception as e_segundo_intento:
            # Diagnostico: se imprime la URL actual y el texto visible de
            # la pagina para entender que esta pasando cuando esto ocurre
            # (por ejemplo, un aviso de mantenimiento, una pantalla de
            # error, o simplemente que el sitio esta caido).
            print("=== DIAGNOSTICO RUNT: el formulario no aparecio tras 2 intentos ===", flush=True)
            print("URL actual:", page.url, flush=True)
            try:
                print("Texto visible de la pagina (primeros 1500 caracteres):", flush=True)
                print(page.inner_text("body")[:1500], flush=True)
            except Exception:
                pass
            print("=== FIN DIAGNOSTICO RUNT ===", flush=True)
            raise e_segundo_intento

    # Procedencia (Nacional) y Consultar Por (Placa y Propietario) ya vienen
    # seleccionados por defecto -- no hace falta tocarlos.
    page.fill('input[formcontrolname="placa"]', placa.upper())

    if tipo_documento != "CC":
        # "Cedula Ciudadania" es el default, solo se cambia si es otro tipo
        texto = RUNT_TIPO_DOC_MAP.get(tipo_documento, "Cédula Ciudadanía")
        _runt_seleccionar_mat_select(page, "tipoDocumento", texto)

    page.fill('input[formcontrolname="documento"]', cedula)

    if job_id:
        job_actualizar(job_id, "Resolviendo captcha...", "procesando")

    # Reintenta hasta 3 veces si el captcha resulta incorrecto (el RUNT
    # regenera la imagen cada vez que falla).
    for intento_captcha in range(3):
        # Se espera explicitamente a que la imagen del captcha este
        # realmente cargada (no solo presente en el DOM) -- si se lee el
        # atributo "src" demasiado pronto puede venir vacio, y de ahi
        # salia el error generico "list index out of range" al intentar
        # separar el prefijo "data:image/png;base64,".
        page.wait_for_selector('img.img-responsive.img-fluid[src]', timeout=15000)
        img_src = page.get_attribute('img.img-responsive.img-fluid', 'src')
        if not img_src or ',' not in img_src:
            raise Exception(f"No se pudo leer la imagen del captcha del RUNT (src {'vacio' if not img_src else 'con formato inesperado'}). Intenta consultar de nuevo.")
        imagen_base64 = img_src.split(',', 1)[1]  # quitar el prefijo "data:image/png;base64,"

        texto_captcha = resolver_captcha_imagen_2captcha(imagen_base64)

        page.fill('input[formcontrolname="captcha"]', texto_captcha)

        if job_id:
            job_actualizar(job_id, "Consultando informacion...", "procesando")

        page.click('button[type="submit"]')

        try:
            page.wait_for_selector(
                'cyrconsultavehiculo-info-vehiculo-detallada, .mat-error, .swal2-popup',
                timeout=20000
            )
        except Exception:
            pass

        # Si el captcha estaba mal, el RUNT normalmente muestra un mensaje
        # de error (swal2) y limpia el campo -- reintentamos con una imagen nueva.
        error_captcha = page.query_selector('.swal2-popup:has-text("captcha")') \
                     or page.query_selector('.swal2-popup:has-text("Captcha")')
        if error_captcha:
            page.click('.swal2-confirm') if page.query_selector('.swal2-confirm') else None
            continue

        break

    # Si el RUNT muestra cualquier otro error (ej. "los datos registrados
    # no corresponden con los propietarios activos"), se propaga el mensaje
    # real en vez de seguir intentando adivinar por que fallo.
    error_popup = page.query_selector('.swal2-popup')
    if error_popup:
        texto_error = error_popup.inner_text().strip()
        if texto_error:
            if page.query_selector('.swal2-confirm'):
                page.click('.swal2-confirm')
            raise Exception(texto_error)

    # Los paneles de SOAT, RTM, garantias, etc. cargan sus datos con
    # peticiones asincronas separadas, despues de que aparece el bloque
    # principal. Si leemos el texto antes de que esas peticiones terminen,
    # esos campos salen vacios aunque el panel ya este expandido. Se espera
    # a que la red quede inactiva (sin peticiones pendientes) antes de seguir.
    if job_id:
        job_actualizar(job_id, "Esperando a que carguen todas las secciones...", "procesando")

    try:
        page.wait_for_load_state("networkidle", timeout=25000)
    except Exception:
        pass

    # Por si ademas hay paneles genuinamente colapsados (no solo cargando),
    # los desplegamos tambien.
    if job_id:
        job_actualizar(job_id, "Desplegando secciones del resultado...", "procesando")

    for _ in range(3):
        headers_colapsados = page.query_selector_all('mat-expansion-panel-header[aria-expanded="false"]')
        for header in headers_colapsados:
            try:
                header.click()
                page.wait_for_timeout(300)
            except Exception:
                pass
        page.wait_for_timeout(400)

    if job_id:
        job_actualizar(job_id, "Extrayendo datos...", "procesando")

    return _parsear_resultado_runt_vehiculo(page)


def _extraer_tarjetas_runt(page):
    """Lee directamente el HTML (no el texto visual) de cada <mat-card> de
    la pagina de resultados, devolviendo un diccionario {etiqueta: valor}
    por cada tarjeta. Esto es inmune a que el CSS reordene visualmente las
    etiquetas y los valores. Revisa tanto <p> como <div> como contenedor,
    y tanto <strong> como <b> como marcador de etiqueta, porque el RUNT usa
    una combinacion distinta segun la seccion (Info General usa <p><strong>,
    Datos Tecnicos usa <p><b>, RTM usa <div><strong>)."""
    return page.evaluate("""
        () => {
            const tarjetas = [];
            document.querySelectorAll('mat-card').forEach(card => {
                const campos = {};
                card.querySelectorAll('p, div').forEach(el => {
                    const marcador = el.querySelector(':scope > strong, :scope > b');
                    if (marcador) {
                        const label = marcador.innerText.replace(/:\\s*$/, '').trim();
                        const value = el.innerText.slice(marcador.innerText.length).trim();
                        if (label) campos[label] = value;
                    }
                });
                card.querySelectorAll('div.col-12').forEach(div => {
                    const labs = div.querySelectorAll('label');
                    // Se procesan de a PARES (0,1) (2,3) (4,5)... -- antes
                    // solo se tomaba el primer par (indices 0 y 1) y se
                    // ignoraban los demas, lo que mezclaba los valores
                    // cuando habia varios campos parecidos juntos en el
                    // mismo bloque (ej. 'Capacidad Pasajeros Sentados' y
                    // 'Capacidad de Pasajeros' uno al lado del otro).
                    for (let i = 0; i + 1 < labs.length; i += 2) {
                        const label = labs[i].innerText.replace(/:\\s*$/, '').trim();
                        const value = labs[i + 1].innerText.trim();
                        if (label) campos[label] = value;
                    }
                });
                const titulo = card.querySelector('mat-card-title');
                if (titulo) campos['_titulo'] = titulo.innerText.trim();
                if (Object.keys(campos).length > 0) tarjetas.push(campos);
            });
            return tarjetas;
        }
    """)


def _extraer_resumen_runt(page):
    """La franja superior (placa, estado del vehiculo, tipo de servicio,
    clase de vehiculo) no vive dentro de ninguna <mat-card> -- son pares de
    <label>Etiqueta:</label> y <b>Valor</b> como hermanos dentro de un
    mismo '.row'. Se emparejan por posicion dentro de cada fila."""
    return page.evaluate("""
        () => {
            const resumen = {};
            document.querySelectorAll('.row').forEach(row => {
                const labels = Array.from(row.querySelectorAll(':scope > div > label'));
                const valores = Array.from(row.querySelectorAll(':scope > div.show-grande > b'));
                if (labels.length > 0 && labels.length === valores.length) {
                    labels.forEach((lab, i) => {
                        const key = lab.innerText.replace(/:\\s*$/, '').trim();
                        resumen[key] = valores[i].innerText.trim();
                    });
                }
            });
            return resumen;
        }
    """) or {}


def _parsear_resultado_runt_vehiculo(page):
    tarjetas = _extraer_tarjetas_runt(page)
    resumen = _extraer_resumen_runt(page)

    # Se combinan todos los campos de todas las tarjetas en un diccionario
    # plano para leerlos facil. Algunas tarjetas (Info General) tienen un
    # solo campo cada una; otras (Datos Tecnicos) traen varios campos juntos
    # en una sola tarjeta. Las tarjetas que se repiten (SOAT, RTM, cada
    # poliza historica) tambien quedan aqui, pero no importa que se
    # sobreescriban entre si porque esos campos se leen aparte, directo de
    # la lista `tarjetas`, no de este diccionario plano.
    plano = dict(resumen)
    for t in tarjetas:
        for k, v in t.items():
            if k != "_titulo":
                plano[k] = v

    plano_lower = {k.lower(): v for k, v in plano.items()}

    print(f"=== DIAGNOSTICO RUNT: campos relacionados a capacidad/pasajeros ===", flush=True)
    for k, v in plano.items():
        if "capacidad" in k.lower() or "pasajer" in k.lower() or "pax" in k.lower():
            print(f"  '{k}' = '{v}'", flush=True)
    print(f"=== FIN DIAGNOSTICO capacidad ===", flush=True)

    def campo(nombre):
        return plano_lower.get(nombre.lower(), "")

    datos = {
        "marca": campo("Marca"),
        "linea": campo("Línea"),
        "modelo": campo("Modelo"),
        "color": campo("Color"),
        "clase": campo("Clase de vehículo"),
        "servicio": campo("Tipo de servicio"),
        "numero_serie": campo("Número de serie"),
        "numero_motor": campo("Número de motor"),
        "numero_chasis": campo("Número de chasis"),
        "vin": campo("Número de VIN"),
        "cilindrada": campo("Cilindraje"),
        "carroceria": campo("Tipo de carrocería"),
        "combustible": campo("Tipo Combustible"),
        "autoridad_transito": campo("Autoridad de tránsito"),
        "puertas": campo("Puertas"),
        "capacidad_carga": campo("Capacidad de Carga"),
        "peso_bruto_vehicular": campo("Peso Bruto Vehicular"),
        # Son DOS datos reales y distintos en el RUNT (confirmado con un
        # caso real) -- antes se mezclaban con una logica de "el primero
        # que no sea 0", pero eso a veces devolvia el dato de UN campo
        # guardado bajo el nombre del OTRO. Ahora cada uno se lee por su
        # propia etiqueta exacta, sin adivinar ni mezclar.
        "capacidad_pasajeros": campo("Capacidad de Pasajeros"),
        "pasajeros_sentados": campo("Capacidad Pasajeros Sentados") or campo("Pasajeros Sentados"),
        "numero_ejes": campo("Número de Ejes"),
        "estado_vehiculo": campo("Estado del vehículo"),
        "gravamenes_propiedad": campo("Gravámenes a la propiedad").upper() == "SI",
        "fecha_matricula_inicial": _convertir_fecha_ddmmyyyy(campo("Fecha de Matricula Inicial")),
    }

    # SOAT vigente: primera tarjeta con "Número de póliza" cuyo Estado diga VIGENTE
    datos["soat_vigente"] = False
    for t in tarjetas:
        if "Número de póliza" in t:
            estado = t.get("Estado", "").upper()
            if "VIGENTE" in estado and "NO VIGENTE" not in estado:
                datos["soat_vigente"] = True
                datos["soat_fecha_fin"] = _convertir_fecha_ddmmyyyy(t.get("Fecha fin de vigencia", ""))
                break

    # RTM vigente: tarjeta "REVISION TECNICO-MECANICO" con Vigente = SI
    datos["rtm_vigente"] = False
    for t in tarjetas:
        if t.get("_titulo", "").upper().startswith("REVISION TECNICO"):
            if t.get("Vigente", "").upper() == "SI":
                datos["rtm_vigente"] = True
                datos["rtm_fecha_fin"] = _convertir_fecha_ddmmyyyy(t.get("Fecha Vigencia", ""))
                break

    # Ultimo tramite relevante (no SOAT ni RTM) -- primera tarjeta "Solicitud NNN"
    for t in tarjetas:
        if t.get("_titulo", "").startswith("Solicitud"):
            tramites = t.get("Trámites Realizados", "")
            if tramites and "revision tecnico mecanica" not in tramites.lower() and "soat" not in tramites.lower():
                datos["ultimo_tramite_tipo"] = tramites.strip(", ")
                datos["ultimo_tramite_fecha"] = _convertir_fecha_ddmmyyyy(t.get("Fecha de Solicitud", ""))
                datos["ultimo_tramite_estado"] = t.get("Estado", "")
                datos["ultimo_tramite_entidad"] = t.get("Entidad", "")
                break

    # Garantias a Favor De -- solo si el acreedor esta afiliado a Confecamaras
    for t in tarjetas:
        if "Acreedor" in t and "Identificación Acreedor" in t:
            datos["garantia_favor_acreedor"] = t.get("Acreedor", "")
            datos["garantia_favor_entidad_nit"] = t.get("Identificación Acreedor", "").replace("NIT", "").strip()
            datos["garantia_favor_fecha_inscripcion"] = _convertir_fecha_ddmmyyyy(t.get("Fecha Inscripción", ""))
            break

    # Garantias Mobiliarias -- hasta 2 registros (inscripcion / levantamiento),
    # se distinguen por el texto libre del campo "Estado".
    for t in tarjetas:
        if "ID Prenda" in t:
            estado_texto = t.get("Estado", "").lower()
            prefijo = "garantia_levantamiento_" if "levantamiento" in estado_texto else "garantia_inscripcion_"
            datos[prefijo + "id_prenda"] = t.get("ID Prenda", "")
            datos[prefijo + "entidad"] = t.get("Entidad", "")
            datos[prefijo + "entidad_nit"] = t.get("Identificación Entidad", "").replace("NIT", "").strip()
            datos[prefijo + "fecha"] = _convertir_fecha_ddmmyyyy(t.get("Fecha de Registro", ""))

    # Limitaciones a la Propiedad (embargo, hurto, etc.) -- distinto de
    # Prenda/Garantias Mobiliarias. Solo se guarda la mas reciente/vigente.
    # Nombres de columnas confirmados por el CSS real del componente
    # (fechaExpedicion, fechaRadicacion, noDocumento, departamento,
    # municipio) -- se prueban varias variantes de texto por si acaso.
    for t in tarjetas:
        titulo_lower = t.get("_titulo", "").lower()
        if "limitacion" in titulo_lower:
            datos["limitacion_tipo"] = t.get("Tipo de Limitación", "") or t.get("Tipo", "")
            datos["limitacion_numero_oficio"] = (
                t.get("Número de Documento", "") or t.get("No. de Documento", "")
                or t.get("Número de Oficio", "") or t.get("No. Oficio", ""))
            datos["limitacion_entidad"] = t.get("Entidad", "")
            datos["limitacion_departamento"] = t.get("Departamento", "")
            datos["limitacion_municipio"] = t.get("Municipio", "")
            datos["limitacion_fecha_oficio"] = _convertir_fecha_ddmmyyyy(
                t.get("Fecha de Expedición", "") or t.get("Fecha Expedición", "") or t.get("Fecha de Expedición del Oficio", ""))
            datos["limitacion_fecha_registro"] = _convertir_fecha_ddmmyyyy(
                t.get("Fecha de Radicación", "") or t.get("Fecha Radicación", "") or t.get("Fecha de Registro", ""))
            break


    datos["placa"] = campo("PLACA DEL VEHÍCULO").upper()

    # --- DIAGNOSTICO TEMPORAL (quitar despues de confirmar los nombres reales) ---
    datos["_debug_tarjetas_limitacion"] = [t for t in tarjetas if "limitacion" in t.get("_titulo", "").lower()]
    datos["_debug_titulos_tarjetas"] = [t.get("_titulo", "") for t in tarjetas if t.get("_titulo")]
    # --- FIN DIAGNOSTICO TEMPORAL ---

    return datos


def _convertir_fecha_ddmmyyyy(fecha_str):
    """Convierte 'dd/mm/yyyy' (formato del RUNT) a 'yyyy-mm-dd' (formato de
    Postgres), o None si el texto viene vacio."""
    fecha_str = (fecha_str or "").strip()
    if not fecha_str:
        return None
    try:
        dd, mm, yyyy = fecha_str.split("/")
        return f"{yyyy}-{mm}-{dd}"
    except Exception:
        return None


def guardar_vehiculo_runt(datos):
    """Guarda (o actualiza) los datos de un vehiculo consultado en el RUNT.
    El RUNT siempre marca fuente='RUNT' y sobrescribe cualquier dato previo
    que hubiera venido solo de una lectura por OCR."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        # Los campos "_debug_*" son solo para diagnostico en pantalla, no
        # corresponden a columnas reales de la tabla.
        columnas = [k for k in datos.keys() if k != "placa" and not k.startswith("_debug")]
        columnas.append("fuente")
        set_clause = ", ".join(f"{c}=EXCLUDED.{c}" for c in columnas)
        cols_sql = ", ".join(["placa"] + columnas + ["leido_en"])
        vals_sql = ", ".join(["%s"] * (len(columnas) + 2))
        valores = [datos["placa"]] + [datos.get(c) for c in columnas[:-1]] + ["RUNT"] + [datetime.now()]
        cur.execute(f"""
            INSERT INTO vehiculos ({cols_sql})
            VALUES ({vals_sql})
            ON CONFLICT (placa) DO UPDATE SET {set_clause}, leido_en=EXCLUDED.leido_en
        """, valores)
        conn.commit()
        cur.close(); conn.close()
        print(f"  → Vehiculo RUNT guardado: {datos['placa']}")
    except Exception as e:
        print(f"Error guardando vehiculo RUNT: {e}")


import unicodedata
import subprocess
import shutil
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.cell import MergedCell
import openpyxl as _openpyxl
import copy

# --- Generador de FUN (Formulario Unico Nacional) ---
# La plantilla debe subirse al repositorio junto a app.py, con este mismo
# nombre exacto ("AppJX.xlsm"), en el mismo directorio.
FUN_PLANTILLA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "AppJX.xlsm")
DECLARACION_MANUAL_PLANTILLA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DECLARACION_MANUAL_DE_IMPUESTOS_DEPARTAMENTALES V2.xlsx")

VERDE_MARCA = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# Cada opcion marca DOS celdas: el numero/casilla y la etiqueta de texto,
# para que la seleccion se vea claramente (no solo el numero).
CELDAS_TRAMITE = {
    "MATRICULA/ REGISTRO": ("A7", "B7"), "TRASPASO": ("E7", "F7"),
    "TRASLADO MATRICULA / REGISTRO": ("I7", "J7"), "RADICADO  MATRICULA / REGISTRO": ("N7", "O7"),
    "CAMBIO DE COLOR": ("Q7", "R7"), "CAMBIO DE SERVICIO": ("T7", "U7"),
    "REGRABAR MOTOR": ("A9", "B9"), "REGRABAR CHASIS": ("E9", "F9"), "TRANSFORMACION": ("I9", "J9"),
    "DUPLICADO LICENCIA TRANSITO": ("N9", "O9"), "INSCRIPC. PRENDA": ("Q9", "R9"), "LEVANTA PRENDA": ("T9", "U9"),
    "CANCELACION MATRICULA / REGISTRO": ("A12", "B12"), "CAMBIO DE PLACAS": ("E12", "F12"),
    "DUPLICADO DE PLACAS": ("I12", "J12"), "REMATRICULA": ("N12", "O12"),
    "CAMBIO DE CARROCERIA": ("Q12", "R12"),
}
# "OTROS" ya no vive aqui -- se marca aparte, solo cuando hay traslado (ver mas abajo)
CELDA_OTROS_TRAMITE = ("T12", "U12")

CELDAS_CLASE = {
    "AUTOMOVIL": ("A17", "A16"), "BUS": ("D17", "D16"), "BUSETA": ("H17", "H16"),
    "CAMION": ("L17", "L16"), "CAMIONETA": ("O17", "O16"), "CAMPERO": ("P17", "P16"),
    "MICROBUS": ("S17", "S16"), "TRACTOCAMION": ("A19", "A18"), "MOTOCICLETA": ("D19", "D18"),
    "MOTOCARRO": ("H19", "H18"), "MOTOTRICICLO": ("L19", "L18"), "CUATRIMOTO": ("O19", "O18"),
    "VOLQUETA": ("P19", "P18"), "OTRO": ("S19", "S18"),
}
CELDAS_COMBUSTIBLE = {
    "GASOLINA": ("AC8", "AC7"), "DIESEL": ("AE8", "AE7"), "GAS": ("AF8", "AF7"),
    "MIXTO": ("AG8", "AG7"), "ELECTRICO": ("AH8", "AH7"), "HIDROGENO": ("AI8", "AI7"),
    "ETANOL": ("AJ8", "AJ7"), "BIODIESEL": ("AK8", "AK7"),
}
CELDAS_SERVICIO = {
    "PARTICULAR": ("AE29", "AE28"), "PUBLICO": ("AF29", "AF28"), "DIPLOMATICO": ("AG29", "AG28"),
    "OFICIAL": ("AH29", "AH28"), "ESPECIAL": ("AI29", "AI28"), "OTROS": ("AJ29", "AJ28"),
}
# Datos del VEHICULO -- confirmado revisando la plantilla que estas
# coordenadas son IDENTICAS en las 3 hojas de Formulario, asi que este
# bloque aplica a las 3 por igual.
CELDAS_REFERENCIA_SIMPLE_VEHICULO = {
    "AJ3": "placa", "W7": "marca", "Z7": "linea", "W10": "color",
    "AG10": "modelo", "AI10": "cilindrada",
    # NOTA: "capacidad" (W13) NO va en este diccionario a proposito -- ya
    # se escribe aparte (ver APPJX_CELDA_CAPACIDAD mas abajo), con una
    # regla especial que trata "0" como vacio (0 pasajeros no es un dato
    # real, es la ausencia del dato). Si se agregara aqui tambien, este
    # bloque genérico volvia a escribir "0" encima de esa correccion.
    "AE17": "numero_motor", "W19": "carroceria", "AE19": "numero_chasis",
    "AE22": "numero_serie", "AE24": "vin", "AC2": "autoridad_transito",
}
# Datos de PERSONAS (propietario/comprador) -- estas coordenadas SI son
# especificas del layout de la hoja BASE (en "(2)"/"(3)" caen en celdas
# distintas por el espacio de la segunda persona), asi que este bloque
# solo aplica a la hoja base.
CELDAS_REFERENCIA_SIMPLE_PERSONAS = {
    "A24": "propietario_primer_apellido", "I24": "propietario_segundo_apellido",
    "P24": "propietario_nombres", "S26": "propietario_documento",
    "A29": "propietario_direccion", "M29": "propietario_ciudad", "S29": "propietario_telefono",
    "A37": "comprador_primer_apellido", "I37": "comprador_segundo_apellido",
    "P37": "comprador_nombres", "S41": "comprador_documento",
    "A44": "comprador_direccion", "M44": "comprador_ciudad", "S44": "comprador_telefono",
    "AG41": "traslado_municipio",
}
# Se mantiene el nombre viejo (union de ambos) por compatibilidad con
# generar_fun, que SI aplica solo a un unico documento (el FUN clasico,
# no las hojas de AppJX) y no tiene este problema de coordenadas
# distintas entre variantes.
CELDAS_REFERENCIA_SIMPLE = {**CELDAS_REFERENCIA_SIMPLE_VEHICULO, **CELDAS_REFERENCIA_SIMPLE_PERSONAS}



def _fun_normalizar(texto):
    if not texto:
        return ""
    texto = str(texto).upper().strip()
    return "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")


def _fun_coincide(valor_tramy, etiqueta_formulario):
    """Compara por PALABRA COMPLETA (no por 'contiene'), para que por ejemplo
    'GAS' no haga match por accidente dentro de 'GASOLINA'."""
    a, b = _fun_normalizar(valor_tramy), _fun_normalizar(etiqueta_formulario)
    if not a or not b:
        return False
    if a == b:
        return True
    return re.search(r"\b" + re.escape(b) + r"\b", a) is not None


def _fun_marcar_checkboxes(ws, mapa_celdas, valor_tramy):
    """Marca TODAS las opciones que coincidan (no solo la primera) -- un
    vehiculo de combustible dual (ej. 'GASOLINA Y GAS') debe marcar ambas."""
    if not valor_tramy:
        return
    for etiqueta, celdas in mapa_celdas.items():
        if _fun_coincide(valor_tramy, etiqueta):
            for celda in celdas:
                ws[celda].fill = VERDE_MARCA


def generar_fun(datos, ruta_salida_pdf):
    """Genera el FUN diligenciado en PDF a partir de la plantilla Excel real
    (FORMULARIO + EXPORTAR), con las casillas marcadas en verde."""
    wb = _openpyxl.load_workbook(FUN_PLANTILLA, data_only=False, keep_vba=True)
    exportar = wb["EXPORTAR"]
    formulario = wb["FORMULARIO"]

    for row in formulario.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "[1]EXPORTAR!" in cell.value:
                cell.value = cell.value.replace("[1]EXPORTAR!", "EXPORTAR!")
            elif isinstance(cell.value, str) and "[1]DATOS!" in cell.value:
                cell.value = cell.value.replace("[1]DATOS!", "DATOS!")

    exportar["D8"] = datos.get("propietario_nombres", "")
    exportar["D9"] = datos.get("propietario_primer_apellido", "")
    exportar["D10"] = datos.get("propietario_segundo_apellido", "")
    exportar["D11"] = datos.get("propietario_documento", "")
    exportar["D12"] = datos.get("propietario_direccion", "")
    exportar["D13"] = datos.get("propietario_ciudad", "")
    exportar["D14"] = datos.get("propietario_telefono", "")
    exportar["D16"] = datos.get("comprador_nombres", "")
    exportar["D17"] = datos.get("comprador_primer_apellido", "")
    exportar["D18"] = datos.get("comprador_segundo_apellido", "")
    exportar["D19"] = datos.get("comprador_documento", "")
    exportar["D20"] = datos.get("comprador_direccion", "")
    exportar["D21"] = datos.get("comprador_ciudad", "")
    exportar["D22"] = datos.get("comprador_telefono", "")
    exportar["D27"] = datos.get("placa", "")
    exportar["D28"] = datos.get("servicio", "")
    exportar["D29"] = datos.get("clase", "")
    exportar["D30"] = datos.get("marca", "")
    exportar["D31"] = datos.get("linea", "")
    exportar["D32"] = datos.get("modelo", "")
    exportar["D33"] = datos.get("color", "")
    exportar["D34"] = datos.get("numero_serie", "")
    exportar["D35"] = datos.get("numero_motor", "")
    exportar["D36"] = datos.get("numero_chasis", "")
    exportar["D37"] = datos.get("cilindrada", "")
    exportar["D38"] = datos.get("carroceria", "")
    exportar["D39"] = datos.get("combustible", "")
    exportar["D40"] = datos.get("autoridad_transito", "")
    exportar["D41"] = datos.get("capacidad", "")
    exportar["D42"] = datos.get("vin", "")
    exportar["D43"] = "SI" if datos.get("gravamenes_propiedad") else "NO"
    exportar["D44"] = datos.get("fecha_matricula_inicial", "")
    exportar["D47"] = datos.get("tramite", "")
    exportar["D51"] = datos.get("traslado_municipio", "")

    formulario["AC2"] = datos.get("autoridad_transito", "")
    formulario["AC2"].alignment = Alignment(horizontal="center", vertical="center")
    formulario["AA4"] = datos.get("municipio", "")

    for celda, clave in CELDAS_REFERENCIA_SIMPLE.items():
        if not datos.get(clave):
            formulario[celda] = ""

    _fun_marcar_checkboxes(formulario, CELDAS_TRAMITE, datos.get("tramite", ""))
    _fun_marcar_checkboxes(formulario, CELDAS_CLASE, datos.get("clase", ""))
    _fun_marcar_checkboxes(formulario, CELDAS_COMBUSTIBLE, datos.get("combustible", ""))
    _fun_marcar_checkboxes(formulario, CELDAS_SERVICIO, datos.get("servicio", ""))

    # "OTROS" (en Tramite Solicitado) NO depende del tramite elegido -- se
    # marca unica y exclusivamente cuando hay un traslado de cuenta.
    if datos.get("traslado_municipio"):
        for celda in CELDA_OTROS_TRAMITE:
            formulario[celda].fill = VERDE_MARCA

    formulario["A51"] = ""  # pie de pagina "Juridicox.com..." -- se quita

    hojas_a_conservar = {"FORMULARIO", "EXPORTAR", "DATOS"}
    for nombre in list(wb.sheetnames):
        if nombre not in hojas_a_conservar:
            del wb[nombre]
    formulario.sheet_state = "visible"
    wb.active = wb.sheetnames.index("FORMULARIO")

    formulario.print_area = "A1:AY47"
    formulario.page_setup.orientation = "landscape"
    formulario.page_setup.paperSize = formulario.PAPERSIZE_LETTER
    formulario.page_setup.scale = 100

    id_temp = str(uuid.uuid4())[:8]
    ruta_xlsm_temp = f"/tmp/_fun_{id_temp}.xlsm"
    wb.save(ruta_xlsm_temp)

    subprocess.run([
        "soffice", "--headless", "--convert-to", "pdf",
        "--outdir", os.path.dirname(ruta_salida_pdf), ruta_xlsm_temp
    ], check=True, timeout=90)

    generado = os.path.join(os.path.dirname(ruta_salida_pdf), f"_fun_{id_temp}.pdf")
    shutil.move(generado, ruta_salida_pdf)
    os.remove(ruta_xlsm_temp)


# --- Documentos de AppJX.xlsm para la pestaña Revision -- SOLO con datos
# del vehiculo (no personas: comprador/vendedor/mandante/mandatario
# quedan sin llenar por ahora). Mapa: clave usada en la URL -> (nombre
# visible en el boton, nombre real de la hoja en AppJX.xlsm). El nombre
# de hoja se confirmo comparando el contenido real de cada plantilla
# (ej. "FORMULARIO (2)" menciona "DATOS DE LOS PROPIETARIOS" en plural y
# un solo comprador -> es la de "dos vendedores").
APPJX_DOCUMENTOS = {
    "formulario":                    ("Formulario",                              "FORMULARIO"),
    "formulario_dos_vendedores":     ("Formulario (dos vendedores)",             "FORMULARIO (2)"),
    "formulario_dos_compradores":    ("Formulario (dos compradores)",            "FORMULARIO (3)"),
    "compraventa":                   ("Compraventa",                             "COMPRA VENTA"),
    "compraventa_dos_vendedores":    ("Compraventa (dos vendedores)",            "COMPRA VENTA (2)"),
    "compraventa_dos_compradores":   ("Compraventa (dos compradores)",           "COMPRA VENTA (3)"),
    "compraventa_persona_juridica":  ("Compraventa (persona jurídica vende)",    "COMPRA VENTA NIT"),
    "mandato":                       ("Mandato",                                 "MANDATO"),
    "mandato_persona_juridica":      ("Mandato (persona jurídica)",              "MANDATO NIT"),
    "mandato_dos_vendedores":        ("Mandato (dos vendedores)",                "MANDATO (2)"),
    "mandato_comprador_vendedor":    ("Mandato (comprador y vendedor)",          "MANDATO (3)"),
    "mandato_4":                     ("Mandato (dos mandatarios)",               "MANDATO (4)"),
    "traspaso_indeterminado":        ("Traspaso indeterminado",                  "INDETERMINADO"),
    "revocatoria_indeterminado":     ("Traspaso a Favor del interesado",         "REVOCATORIA"),
    "afirmacion_traspaso":           ("Afirmación de traspaso",                  "AFIRMACION"),
    "levantamiento_prenda":          ("Levantamiento de prenda",                 "LEVANTAMIENTO PRENDA"),
    "inscripcion_prenda":            ("Inscripción de prenda",                   "INSCRIPCION PRENDA"),
    "acta_responsabilidad":          ("Acta de responsabilidad civil",           "ACTA RESPONSABILIDAD"),
}

# Celda dentro de cada documento que muestra la linea de datos de la
# empresa (referencia "=DATOS!W2" en la plantilla original) -- se
# confirmo revisando las 17 hojas una por una. Se usa para escribir el
# valor de forma directa (en vez de depender de que se recalcule la
# formula, que no siempre pasa de forma confiable al convertir a PDF).
APPJX_CELDA_LINEA_EMPRESA = {
    "formulario": "A51", "formulario_dos_vendedores": "A52", "formulario_dos_compradores": "A48",
    "compraventa": "A35", "compraventa_dos_vendedores": "A36", "compraventa_dos_compradores": "A37",
    "compraventa_persona_juridica": "A35",
    "mandato": "A49", "mandato_persona_juridica": "A49",
    "mandato_dos_vendedores": "A46", "mandato_comprador_vendedor": "A46",
    "traspaso_indeterminado": "A47", "revocatoria_indeterminado": "A46",
    "afirmacion_traspaso": "A46",
    "levantamiento_prenda": "A42", "inscripcion_prenda": "A42",
    "acta_responsabilidad": "A44",
}

# Filas de los 4 documentos de MANDATO que tienen un salto de linea
# ("\n") dentro del texto de la celda, pero se quedaron con la altura de
# fila de una sola linea -- por eso el texto se veia apeñuscado (la
# segunda linea se montaba encima de la fila de abajo). Se les da mas
# alto para que las 2 lineas quepan bien.
APPJX_FILAS_ALTURA_EXTRA = {
    "MANDATO": [3, 7, 21],
    "MANDATO NIT": [3, 8, 22],
    "MANDATO (2)": [3, 5, 21],
    "MANDATO (3)": [5, 6],
}

# Celda dentro de cada documento que muestra la capacidad de pasajeros
# (referencia "=EXPORTAR!D41" en la plantilla) -- solo los 7 documentos
# listados aqui la muestran. Igual que con la linea de empresa, se
# escribe directo por el mismo problema de recalculo de formulas.
APPJX_CELDA_CAPACIDAD = {
    "formulario": "W13", "formulario_dos_vendedores": "W13", "formulario_dos_compradores": "W13",
    "compraventa": "B20", "compraventa_dos_vendedores": "B20",
    "compraventa_dos_compradores": "B20", "compraventa_persona_juridica": "B20",
}


# Celdas de TELEFONO en cada documento que dependen de una formula
# (=EXPORTAR!D14 o similar) que, cuando la celda de origen esta vacia,
# Excel/LibreOffice la evalua como el NUMERO 0 (asi es como Excel trata
# SIEMPRE una referencia a una celda vacia, sin importar el formato de
# la celda que muestra el resultado -- cambiar el numero_format NO
# alcanza a arreglar esto). La solucion real es escribir el valor de
# telefono DIRECTO en la celda del documento (no depender de la
# formula), igual que ya se hace con otros datos fragiles como la linea
# de empresa o la capacidad. Cada entrada dice que ROL de persona
# corresponde a esa celda.
# Casillas de TRAMITE en las 3 variantes de Formulario -- confirmado
# revisando la plantilla real que la cuadrilla de 18 casillas (filas
# 7, 9 y 12) es IDENTICA en las 3 hojas (Formulario, Formulario (2),
# Formulario (3)), asi que un solo mapeo aplica a las tres. Cada entrada
# tiene DOS celdas (numero + etiqueta) que se pintan de verde juntas.
# Los 14 tramites "canonicos" -- estos MISMOS nombres se usan en los 3
# lugares donde se elige un tramite (el catalogo real de Liquidacion se
# normaliza a estos, y Preparacion + el modulo de documentos de
# Liquidacion los muestran identicos). Cada uno mapea a su casilla en
# Formulario cuando existe (algunos, como "CAMBIO DE MOTOR" o
# "REGRABACION DE SERIE", no tienen casilla propia en la plantilla, asi
# que simplemente no resaltan nada ahi -- solo aparecen en el texto de
# Mandato).
CELDAS_TRAMITE_FORMULARIO = {
    "MATRICULA INICIAL": ("A7", "B7"),
    "TRASPASO DE PROPIEDAD": ("E7", "F7"),
    "TRASLADO DE CUENTA": ("I7", "J7"),
    "RADICADO DE CUENTA": ("N7", "O7"),
    "CAMBIO DE COLOR": ("Q7", "R7"),
    "REGRABACION DE MOTOR": ("A9", "B9"),
    "REGRABACION DE CHASIS": ("E9", "F9"),
    "DUPLICADO DE LICENCIA DE TRANSITO": ("N9", "O9"),
    "INSCRIPCION DE PRENDA": ("Q9", "R9"),
    "LEVANTAMIENTO DE PRENDA": ("T9", "U9"),
    "CANCELACION DE CUENTA": ("A12", "B12"),
    "DUPLICADO DE PLACAS": ("I12", "J12"),
}
CELDA_OTROS_TRAMITE_FORMULARIO = ("T12", "U12")

# Celdas de "Tipo de Servicio" (Particular/Publico/Diplomatico) -- a
# diferencia de la cuadricula de TRAMITES (identica en las 3 hojas), esta
# seccion SI cambia de posicion entre variantes (confirmado revisando la
# plantilla real: en "FORMULARIO" el texto esta en la fila 28 y el
# recuadro que se resalta en las filas 29-30; en "FORMULARIO (2)" todo
# baja una fila (texto en 29, recuadro en 31-32); en "FORMULARIO (3)" el
# texto vuelve a la fila 28 pero el recuadro queda en las filas 30-31).
# Cada entrada son TODAS las celdas que hay que pintar de verde juntas.
CELDAS_SERVICIO_POR_DOCUMENTO = {
    "formulario": {
        "PARTICULAR": ["AE28", "AE29", "AE30"], "PUBLICO": ["AF28", "AF29", "AF30"], "DIPLOMATICO": ["AG29", "AG30"],
    },
    "formulario_dos_vendedores": {
        "PARTICULAR": ["AE29", "AE30", "AE31"], "PUBLICO": ["AF29", "AF30", "AF31"], "DIPLOMATICO": ["AG29", "AG30", "AG31"],
    },
    "formulario_dos_compradores": {
        "PARTICULAR": ["AE28", "AE29", "AE30", "AE31"], "PUBLICO": ["AF28", "AF29", "AF30", "AF31"], "DIPLOMATICO": ["AG28", "AG29", "AG30", "AG31"],
    },
}

# Celdas de "Tipo de documento" (C.C / NIT / N.N / Pasaporte / C.Extranj. /
# T.Identi.) del propietario y del comprador -- SOLO se resaltan C.C y
# NIT (los demas tipos no se marcan, segun se pidio explicitamente).
# Confirmado revisando la plantilla real: cada casilla tiene una celda de
# ETIQUETA (fila de arriba) y, quiza, una celda de CODIGO (fila de abajo,
# ej. "C"/"N") -- se pintan ambas cuando existen las dos.
CELDAS_TIPO_DOC_FORMULARIO = {
    "formulario": {
        "propietario": {"CC": ["A25", "A26"], "NIT": ["C25", "C26"]},
        "comprador": {"CC": ["A40", "A41"], "NIT": ["C40", "C41"]},
    },
    "formulario_dos_vendedores": {
        "propietario": {"CC": ["A26", "A27"], "NIT": ["C26", "C27"]},
        "comprador": {"CC": ["A42", "A43"], "NIT": ["C42", "C43"]},
    },
    "formulario_dos_compradores": {
        "propietario": {"CC": ["A26"], "NIT": ["C26"]},
        "comprador": {"CC": ["A40", "A41"], "NIT": ["C40", "C41"]},
    },
}

# Celda donde se escribe DIRECTO (no por formula) el texto de traslado de
# cuenta cuando se elige el tramite "TRASLADO MATRICULA / REGISTRO" --
# cada variante de Formulario tiene esta celda en una fila distinta
# (confirmado revisando la plantilla real).
APPJX_CELDA_TRASLADO_TEXTO = {
    "formulario": "W41",
    "formulario_dos_vendedores": "W43",
    "formulario_dos_compradores": "W41",
}

# Celdas de DIRECCION/CIUDAD/TELEFONO por documento y rol -- se escriben
# DIRECTO (no por formula) porque una formula que apunta a una celda
# vacia se evalua como 0 en Excel/LibreOffice, sin importar el formato
# de la celda de destino (confirmado con casos reales: paso primero con
# telefono, luego se confirmo que direccion/ciudad tienen el mismo
# problema en "FORMULARIO (2)"). No se incluye la hoja base "FORMULARIO"
# aqui porque esa ya tiene su propio manejo (CELDAS_REFERENCIA_SIMPLE_
# PERSONAS, mas abajo), que ya escribe vacio correctamente.
APPJX_CELDAS_PERSONA_A_CORREGIR = {
    "formulario_dos_vendedores": {
        "propietario": {"direccion": "A30", "ciudad": "M30", "telefono": "S30"},
        "otro_propietario": {"direccion": "A31", "ciudad": "M31", "telefono": "S31"},
        "comprador": {"direccion": "A46", "ciudad": "M46", "telefono": "S46"},
    },
    "formulario_dos_compradores": {
        "propietario": {"direccion": "A29", "ciudad": "M29", "telefono": "S29"},
        "comprador": {"direccion": "A45", "ciudad": "M45", "telefono": "S45"},
        "otro_comprador": {"direccion": "A44", "ciudad": "M44", "telefono": "S44"},
    },
    "compraventa": {
        "propietario": {"telefono": "B7"},
        "comprador": {"telefono": "B14"},
    },
    "compraventa_dos_vendedores": {
        "propietario": {"telefono": "B7"},
        "otro_propietario": {"telefono": "G7"},
        "comprador": {"telefono": "B14"},
    },
    "compraventa_dos_compradores": {
        "propietario": {"telefono": "B7"},
        "otro_comprador": {"telefono": "G7"},
        "comprador": {"telefono": "B14"},
    },
    "compraventa_persona_juridica": {
        "propietario": {"telefono": "B7"},
        "comprador": {"telefono": "B14"},
    },
}


import random

# Mismas listas que ya existen en Preparacion (boton "Datos Falsos") --
# se guardan tambien aqui para poder rellenar automaticamente los datos
# del propietario cuando falten, al generar un documento desde
# Liquidacion (que no tiene ese boton en su interfaz).
TRAMY_PREFIJOS_TELEFONO_FALSO = ["310508", "313205", "301528", "320854", "300633", "323787", "315325", "314458", "333477", "316968"]
TRAMY_DIRECCIONES_FALSAS = [
    "Cra 80  # 50 - 52 apto (201)", "Cra 69 # 32 - 25 (401) Palomares",
    "Calle 34  # 22 - 38 201 Urb calle larga", "Calle 21 # 18 - 26 apto 301",
    "Diag  77  # 32 - 40 Ed el bosque (501)", "Calle 58 # 70 - 25 granero la palma",
    "Diag 30  82 - 48 edificio puente verde (908)", "Cl 98a  #65-122",
    "Calle 50 # 42-54", "Cr 36  #10 B-38", "Calle 81a #52a-60 (piso 4)",
    "Cra 45 # 42-42 (apto 501)", "Calle 12 # 31-185 edificio la cigala",
    "Diagonal 49 # 34-92 (urb casa verde casa 18)", "Calle 155b # 8C-22 apto 502",
    "Carrera 69A #93-20 torre 2 apto 1010", "Carrera 87 N° 46 - 33",
    "Cra 59 No 36- 56 casa 3", "Calle 54 #85-40 apto 201", "Cr 55 #69-07 esquina apto 501",
    "Cra 52 # 1-81 la pola", "Calle 51 #49-11 Of. 603", "Transversal 34 A Sur No 32 D -18",
    "Calle 38 Sur 43-85 Cons. 201", "Carrera 50 A # 33-74", "Calle 18 #58-06",
    "Calle 39B Sur # 38-9", "Carrera 42 #14-74", "Calle 30A # 79-117",
    "Calle 78 SUR # 57-83 Local 110", "CRA 65 Nº 43-10", "Carrera 22 # 80 Sur - 32",
    "Calle 60 sur # 20-16 Diagonal a Andar", "Carrera 43 B #12-157", "Calle 65 # 87 - 59",
    "Calle 46 N. 54-48 Almacén AYACUCHO", "Cr 42 16 A sur 41 - Mall Aerocentro, local 1",
]


def _rellenar_datos_falsos_si_faltan(persona, municipio_vehiculo):
    """Si la persona no tiene telefono/direccion/ciudad, se rellenan con
    datos de prueba (igual que el boton 'Datos Falsos' de Preparacion) --
    si YA tiene algun dato puesto, ese se respeta y no se toca. Se usa
    para que los documentos generados desde Liquidacion (que no tiene
    ese boton) igual salgan completos, ya que las secretarias de
    transito no reciben documentos con campos en blanco."""
    if not persona:
        return persona
    persona = dict(persona)  # no modificar el original
    if not persona.get("telefono"):
        prefijo = random.choice(TRAMY_PREFIJOS_TELEFONO_FALSO)
        resto = str(random.randint(1000, 9999))
        persona["telefono"] = prefijo + resto
    if not persona.get("direccion"):
        persona["direccion"] = random.choice(TRAMY_DIRECCIONES_FALSAS) + " *"
    if not persona.get("ciudad") and municipio_vehiculo:
        persona["ciudad"] = municipio_vehiculo.strip().upper()
    return persona


def generar_documento_vehiculo_appjx(clave_documento, datos_vehiculo, ruta_salida_pdf):
    """Genera en PDF cualquiera de los documentos de AppJX.xlsm listados
    en APPJX_DOCUMENTOS, llenando SOLO los datos del vehiculo (placa,
    marca, linea, modelo, etc. -- los mismos campos que ya usa el FUN).
    Los datos de personas (comprador/vendedor/mandante/mandatario) se
    dejan sin llenar por ahora."""
    if clave_documento not in APPJX_DOCUMENTOS:
        raise ValueError(f"Documento desconocido: {clave_documento}")
    _, nombre_hoja = APPJX_DOCUMENTOS[clave_documento]

    # Si el propietario (vendedor) no tiene telefono/direccion/ciudad, se
    # rellenan con datos de prueba (igual que el boton "Datos Falsos" de
    # Preparacion) -- respeta cualquier dato que YA tenga puesto, solo
    # llena lo que falte. Esto hace que los documentos generados desde
    # Liquidacion (que no tiene ese boton) tambien salgan completos.
    # NO se aplica al comprador -- esos datos se ponen a mano cuando se
    # necesiten, a proposito.
    _municipio_para_datos_falsos = datos_vehiculo.get("municipio", "")
    for _rol_relleno in ("propietario", "otro_propietario"):
        if datos_vehiculo.get(_rol_relleno):
            datos_vehiculo[_rol_relleno] = _rellenar_datos_falsos_si_faltan(
                datos_vehiculo[_rol_relleno], _municipio_para_datos_falsos
            )

    wb = _openpyxl.load_workbook(FUN_PLANTILLA, data_only=False, keep_vba=True)
    hoja = wb[nombre_hoja]

    # Igual que en generar_fun: algunas celdas quedan con una referencia
    # externa rota ("[1]EXPORTAR!...") en vez de la referencia normal a
    # la misma hoja -- se corrige antes de guardar.
    for row in hoja.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "[1]EXPORTAR!" in cell.value:
                cell.value = cell.value.replace("[1]EXPORTAR!", "EXPORTAR!")
            elif isinstance(cell.value, str) and "[1]DATOS!" in cell.value:
                cell.value = cell.value.replace("[1]DATOS!", "DATOS!")

    exportar = wb["EXPORTAR"]
    exportar["D27"] = datos_vehiculo.get("placa", "")
    exportar["D28"] = datos_vehiculo.get("servicio", "")
    exportar["D29"] = datos_vehiculo.get("clase", "")
    exportar["D30"] = datos_vehiculo.get("marca", "")
    exportar["D31"] = datos_vehiculo.get("linea", "")
    exportar["D32"] = datos_vehiculo.get("modelo", "")
    exportar["D33"] = datos_vehiculo.get("color", "")
    exportar["D34"] = datos_vehiculo.get("numero_serie", "")
    exportar["D35"] = datos_vehiculo.get("numero_motor", "")
    exportar["D36"] = datos_vehiculo.get("numero_chasis", "")
    exportar["D37"] = datos_vehiculo.get("cilindrada", "")
    exportar["D38"] = datos_vehiculo.get("carroceria", "")
    exportar["D39"] = datos_vehiculo.get("combustible", "")
    exportar["D40"] = datos_vehiculo.get("autoridad_transito", "")
    _capacidad_pax = str(datos_vehiculo.get("capacidad", "") or "").strip()
    _capacidad_pax_final = "" if _capacidad_pax in ("", "0") else _capacidad_pax
    exportar["D41"] = _capacidad_pax_final
    celda_capacidad = APPJX_CELDA_CAPACIDAD.get(clave_documento)
    if celda_capacidad:
        try:
            hoja[celda_capacidad] = _capacidad_pax_final
        except AttributeError:
            pass  # celda combinada -- no se puede escribir directo
        hoja[celda_capacidad].number_format = "General"
    exportar["D42"] = datos_vehiculo.get("vin", "")
    exportar["D43"] = "SI" if datos_vehiculo.get("gravamenes_propiedad") else "NO"
    exportar["D44"] = datos_vehiculo.get("fecha_matricula_inicial", "")

    # Casilla "1. ORGANISMO DE TRANSITO" / "NOMBRE" -- solo existe en los
    # 3 Formularios, en la celda AC2 (junto a la etiqueta "NOMBRE" en
    # AA2). Se escribe directo (no via EXPORTAR) porque esta celda no
    # tenia ninguna formula/referencia en la plantilla original.
    if nombre_hoja in ("FORMULARIO", "FORMULARIO (2)", "FORMULARIO (3)"):
        hoja["AC2"] = datos_vehiculo.get("autoridad_transito", "")

    # Afirmacion de Traspaso tiene una celda con la fecha de hoy
    # (=TODAY()) que LibreOffice muestra en INGLES (ej. "15-August-2026")
    # porque no usa la configuracion regional en español -- se escribe
    # directo el texto ya formateado en español, y se centra (la
    # plantilla la traia alineada a la izquierda).
    if nombre_hoja == "AFIRMACION":
        _meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                     "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        _hoy = datetime.now()
        hoja["F8"] = f"{_hoy.day}-{_meses_es[_hoy.month - 1]}-{_hoy.year}"
        hoja["F8"].alignment = Alignment(horizontal="center", vertical=hoja["F8"].alignment.vertical)

    # Las 4 hojas de MANDATO tienen varias celdas sin alineacion vertical
    # definida -- heredan la del tema de la plantilla, que en algunos
    # casos las deja "pegadas al piso" de una fila mas alta de lo normal,
    # dando la impresion de que el dato esta en la fila de abajo. Se
    # centran verticalmente para que siempre se vean en su fila correcta
    # sin importar la altura.
    _celdas_alinear_mandato = {
        "MANDATO": ["C1"],
        "MANDATO NIT": ["D3"],
        "MANDATO (2)": ["C1"],
        "MANDATO (3)": ["C1", "F5"],
    }.get(nombre_hoja, [])
    for _celda_alinear in _celdas_alinear_mandato:
        hoja[_celda_alinear].alignment = Alignment(
            horizontal=hoja[_celda_alinear].alignment.horizontal,
            vertical="center", wrap_text=False,
        )

    exportar["D51"] = ""  # traslado_municipio -- vacio explicito, si no la formula '=EXPORTAR!D51' en FORMULARIO muestra "0" (una celda totalmente vacia, sin ni siquiera comillas vacias, se lee como cero en una referencia directa)
    exportar["D24"] = datos_vehiculo.get("precio_venta") or ""  # precio -- viene del campo "Precio de venta" en Tramites (Preparacion/Liquidacion); si no se indica, queda vacio para llenarlo a mano en el documento impreso
    exportar["D24"].number_format = "General"

    # Tramites seleccionados en Preparacion, conectados con los 4
    # contratos de MANDATO -- a diferencia de Formulario (que resalta
    # casillas), aqui cada tramite elegido se escribe como texto, uno por
    # linea, en EXPORTAR!D47/D48/D49 (la plantilla original ya tenia esta
    # conexion prevista con esas 3 celdas -- "MANDATO (3)" solo muestra
    # las primeras 2, las demas variantes muestran las 3). Los nombres
    # que llegan aqui ya son los 14 nombres "canonicos" (los mismos que
    # se eligen en Preparacion/Liquidacion), asi que se usan tal cual,
    # sin necesidad de normalizarlos de nuevo.
    _tramites_mandato = (datos_vehiculo.get("tramites_seleccionados") or [])[:3]
    _lineas_mandato = list(_tramites_mandato)
    exportar["D47"] = _lineas_mandato[0] if len(_lineas_mandato) > 0 else ""
    exportar["D48"] = _lineas_mandato[1] if len(_lineas_mandato) > 1 else ""
    exportar["D49"] = _lineas_mandato[2] if len(_lineas_mandato) > 2 else ""

    # "MANDATO (3)" solo tenia 2 lineas conectadas (A14/A15, con formula
    # a D47/D48) -- A13 estaba vacia y sin usar, pero hay espacio real
    # ahi para una 3ra linea. Se escribe DIRECTO (no hay formula previa
    # que reutilizar), copiando la fuente de A14 para que se vea igual.
    if nombre_hoja == "MANDATO (3)":
        hoja["A13"].value = _lineas_mandato[2] if len(_lineas_mandato) > 2 else ""
        hoja["A13"].font = copy.copy(hoja["A14"].font)

    # Linea de datos de la empresa (nombre, telefono, correo, etc.) que
    # aparece al pie de cada documento -- se escribe en DATOS!W2 (por si
    # algun otro lado de la plantilla la usa) Y TAMBIEN directo en la
    # celda del documento actual que la muestra, ya que esa celda
    # depende de que se recalcule "=DATOS!W2" y eso no siempre pasa de
    # forma confiable al convertir a PDF (el mismo problema visto con
    # las demas referencias directas).
    linea_empresa = datos_vehiculo.get("linea_empresa", "")
    if "DATOS" in wb.sheetnames:
        wb["DATOS"]["W2"] = linea_empresa
    celda_linea_empresa = APPJX_CELDA_LINEA_EMPRESA.get(clave_documento)
    if celda_linea_empresa:
        try:
            hoja[celda_linea_empresa] = linea_empresa
        except AttributeError:
            pass  # celda combinada -- no se puede escribir directo
        hoja[celda_linea_empresa].number_format = "General"

    # En los 4 documentos de MANDATO, algunas filas tienen texto con
    # salto de linea interno pero se quedaron con altura de una sola
    # linea -- se les da mas espacio para que no se vean apeñuscadas.
    for fila in APPJX_FILAS_ALTURA_EXTRA.get(nombre_hoja, []):
        alto_actual = hoja.row_dimensions[fila].height or 15
        hoja.row_dimensions[fila].height = max(alto_actual, 30)

    # Datos de personas (asesor/propietario/comprador) -- si se
    # proporcionan (como dict con nombres/apellido/segundo_apellido/
    # numero_documento/direccion/ciudad/telefono), se escriben DIRECTO en
    # las celdas del documento -- igual que con los datos del vehiculo,
    # se evita depender de la formula VLOOKUP (que ademas en la
    # plantilla busca por NOMBRE, no por documento, lo cual es fragil).
    # Si no se proporciona una persona para un rol, ese bloque queda en
    # blanco (comportamiento anterior).
    def _escribir_bloque_persona(fila_nombres, con_direccion, persona, columna=4):
        p = persona or {}
        filas_campos = [
            (fila_nombres,     (p.get("nombres") or "")),
            (fila_nombres + 1, (p.get("apellido") or "")),
            (fila_nombres + 2, (p.get("segundo_apellido") or "")),
            (fila_nombres + 3, (p.get("numero_documento") or "")),
        ]
        if con_direccion:
            filas_campos += [
                (fila_nombres + 4, (p.get("direccion") or "")),
                (fila_nombres + 5, (p.get("ciudad") or "")),
                (fila_nombres + 6, (p.get("telefono") or "")),
            ]
        for fila, valor in filas_campos:
            celda = exportar.cell(row=fila, column=columna)
            celda.value = valor
            celda.number_format = "General"

    _escribir_bloque_persona(3, con_direccion=False, persona=datos_vehiculo.get("asesor"), columna=4)
    _escribir_bloque_persona(8, con_direccion=True, persona=datos_vehiculo.get("propietario"), columna=4)
    _escribir_bloque_persona(8, con_direccion=True, persona=datos_vehiculo.get("otro_propietario"), columna=7)
    _escribir_bloque_persona(16, con_direccion=True, persona=datos_vehiculo.get("comprador"), columna=4)
    _escribir_bloque_persona(16, con_direccion=True, persona=datos_vehiculo.get("otro_comprador"), columna=7)

    # Se escriben direccion/ciudad/telefono TAMBIEN directo en la celda
    # del documento (ademas de en EXPORTAR) -- una formula que apunta a
    # una celda vacia se evalua como 0 en Excel/LibreOffice sin importar
    # el formato de la celda de destino, asi que la unica forma confiable
    # de que un dato vacio se vea en blanco es no depender de la formula.
    for _rol_doc, _campos_doc in APPJX_CELDAS_PERSONA_A_CORREGIR.get(clave_documento, {}).items():
        _persona_rol_doc = datos_vehiculo.get(_rol_doc) or {}
        for _campo_doc, _celda_doc in _campos_doc.items():
            _celda_obj_doc = hoja[_celda_doc]
            _celda_obj_doc.value = _persona_rol_doc.get(_campo_doc) or ""
            _celda_obj_doc.number_format = "General"

    # "COMPRA VENTA NIT" -- I9 traia una formula rota (hacia referencia a
    # una celda de OTRA hoja, "FORMULARIO!P37", que no corresponde a nada
    # en este documento -- error de copia en la plantilla original). Se
    # escribe directo el nombre completo del comprador, e I10 (numero de
    # documento) tambien se escribe directo por seguridad, aunque su
    # formula original (=B11) si apuntaba al lugar correcto.
    if clave_documento == "compraventa_persona_juridica":
        _comprador_nit = datos_vehiculo.get("comprador") or {}
        _nombre_completo_comprador = " ".join(filter(None, [
            _comprador_nit.get("nombres"), _comprador_nit.get("apellido"), _comprador_nit.get("segundo_apellido"),
        ]))
        hoja["I9"] = _nombre_completo_comprador
        hoja["I9"].number_format = "General"
        hoja["I10"] = _comprador_nit.get("numero_documento") or ""
        hoja["I10"].number_format = "General"

    # Rol "OTRO" -- se conecta con estas 6 hojas especificas. En
    # "COMPRA VENTA NIT" y "MANDATO NIT" las celdas estaban vacias (sin
    # formula previa); en las otras 4 SI habia una formula que mostraba
    # al propietario por defecto (=EXPORTAR!D8/D9/D10 para el nombre,
    # =EXPORTAR!D11 para el documento) -- se sobrescribe con los datos
    # de "otro" en todos los casos, escribiendo directo (no por formula)
    # para que tambien funcione bien si "otro" queda vacio.
    _CELDAS_ROL_OTRO = {
        "compraventa_persona_juridica": {"nombre": "H7", "documento": "J8"},
        "mandato_persona_juridica": {"nombre": "C1", "documento": "G2"},
        "revocatoria_indeterminado": {"nombre": "B7", "documento": "C8"},
        "levantamiento_prenda": {"nombre": "A8", "documento": "C9"},
        "inscripcion_prenda": {"nombre": "D3", "documento": "C4"},
        "acta_responsabilidad": {"nombre": "D3", "documento": "C4"},
    }
    if clave_documento in _CELDAS_ROL_OTRO:
        _otro_persona = datos_vehiculo.get("otro") or {}
        _nombre_completo_otro = " ".join(filter(None, [
            _otro_persona.get("nombres"), _otro_persona.get("apellido"), _otro_persona.get("segundo_apellido"),
        ]))
        _celdas_otro_doc = _CELDAS_ROL_OTRO[clave_documento]
        hoja[_celdas_otro_doc["nombre"]] = _nombre_completo_otro
        hoja[_celdas_otro_doc["nombre"]].number_format = "General"
        hoja[_celdas_otro_doc["documento"]] = _otro_persona.get("numero_documento") or ""
        hoja[_celdas_otro_doc["documento"]].number_format = "General"

    # Rol "MANDATARIO" -- se conecta con las 5 hojas de Mandato. Estas
    # celdas ya tenian una formula que apuntaba a EXPORTAR!D3-D6 (el rol
    # "asesor", que nunca se conecto desde la interfaz, asi que siempre
    # quedaban vacias) -- se sobrescriben directo con los datos de
    # "mandatario" en su lugar. "MANDATO (4)" es la unica variante con
    # espacio para un SEGUNDO mandatario (otro_mandatario).
    _CELDAS_ROL_MANDATARIO = {
        "mandato": {"nombre": "A4", "documento": "A6"},
        "mandato_dos_vendedores": {"nombre": "B7", "documento": "F8"},
        "mandato_comprador_vendedor": {"nombre": "A7", "documento": "G8"},
        "mandato_persona_juridica": {"nombre": "D5", "documento": "F6"},
        "mandato_4": {"nombre": "C4", "documento": "A6"},
    }
    if clave_documento in _CELDAS_ROL_MANDATARIO:
        _mandatario_persona = datos_vehiculo.get("mandatario") or {}
        _nombre_completo_mandatario = " ".join(filter(None, [
            _mandatario_persona.get("nombres"), _mandatario_persona.get("apellido"), _mandatario_persona.get("segundo_apellido"),
        ]))
        _celdas_mandatario_doc = _CELDAS_ROL_MANDATARIO[clave_documento]
        hoja[_celdas_mandatario_doc["nombre"]] = _nombre_completo_mandatario
        hoja[_celdas_mandatario_doc["nombre"]].number_format = "General"
        hoja[_celdas_mandatario_doc["documento"]] = _mandatario_persona.get("numero_documento") or ""
        hoja[_celdas_mandatario_doc["documento"]].number_format = "General"

    # Segundo mandatario -- SOLO existe en "MANDATO (4)" por ahora.
    if clave_documento == "mandato_4":
        _otro_mandatario_persona = datos_vehiculo.get("otro_mandatario") or {}
        _nombre_completo_otro_mandatario = " ".join(filter(None, [
            _otro_mandatario_persona.get("nombres"), _otro_mandatario_persona.get("apellido"), _otro_mandatario_persona.get("segundo_apellido"),
        ]))
        hoja["B7"] = _nombre_completo_otro_mandatario
        hoja["B7"].number_format = "General"
        hoja["F8"] = _otro_mandatario_persona.get("numero_documento") or ""
        hoja["F8"].number_format = "General"

    # Precio de venta -- se escribe directo en la celda del documento
    # (ademas de en EXPORTAR) por el mismo motivo de siempre: una formula
    # que apunta a una celda vacia se evalua como 0, sin importar el
    # formato de la celda de destino.
    _CELDA_PRECIO_VENTA = {
        "compraventa": "G3",
        "compraventa_dos_vendedores": "G10",
        "compraventa_dos_compradores": "G10",
    }
    if clave_documento in _CELDA_PRECIO_VENTA:
        _celda_precio = _CELDA_PRECIO_VENTA[clave_documento]
        _precio_venta_raw = datos_vehiculo.get("precio_venta")
        try:
            _precio_venta_valor = float(str(_precio_venta_raw).replace(",", "").replace(".", "").strip()) if _precio_venta_raw else ""
        except (ValueError, TypeError):
            _precio_venta_valor = _precio_venta_raw or ""
        hoja[_celda_precio] = _precio_venta_valor
        hoja[_celda_precio].number_format = '"$"#,##0'

    # Las celdas DENTRO del documento (no en EXPORTAR) que muestran estos
    # datos de personas dependen de que LibreOffice recalcule su formula
    # al convertir a PDF -- eso no siempre pasa de forma confiable (igual
    # que con el caso de "traslado" en FORMULARIO). Para los roles que
    # SI tienen persona asignada, se confia en el recalculo normal (la
    # celda de EXPORTAR ya tiene texto real, no vacio, asi que no debería
    # dar el problema del "0"). Para los roles que quedaron SIN persona,
    # se sigue escribiendo vacio directo en la celda que se ve, como
    # antes, para que no aparezca "0".
    rangos_por_rol = [
        (3, 6, [4], datos_vehiculo.get("asesor")),
        (8, 14, [4], datos_vehiculo.get("propietario")),
        (8, 14, [7], datos_vehiculo.get("otro_propietario")),
        (16, 22, [4], datos_vehiculo.get("comprador")),
        (16, 22, [7], datos_vehiculo.get("otro_comprador")),
    ]
    filas_sin_persona = set()
    for fila_ini, fila_fin, columnas, persona in rangos_por_rol:
        if not persona:  # rol sin persona asignada -- sus filas quedan candidatas a vaciar
            for f in range(fila_ini, fila_fin + 1):
                for c in columnas:
                    filas_sin_persona.add((f, c))

    def _fila_columna_de_referencia_exportar(texto_formula):
        """Extrae (fila, columna) de una referencia tipo 'EXPORTAR!D9' o
        'EXPORTAR!$G$14' -- para saber si esa referencia cae en un rol
        sin persona asignada."""
        m = re.search(r"EXPORTAR!\$?([A-Z]+)\$?([0-9]+)", texto_formula)
        if not m:
            return None
        col_letra, fila_texto = m.group(1), m.group(2)
        return (int(fila_texto), _openpyxl.utils.column_index_from_string(col_letra))

    filas_persona_regex = re.compile(r"EXPORTAR!\$?D\$?([3-9]|1[0-9]|2[0-2]|24|51)\b")
    celdas_vaciadas = set()  # coordenadas (ej. "D5") que se dejaron vacias en esta pasada
    for row in hoja.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            ref = _fila_columna_de_referencia_exportar(cell.value)
            es_referencia_persona = ref is not None and (
                (3 <= ref[0] <= 6) or (8 <= ref[0] <= 14) or (16 <= ref[0] <= 22)
            )
            if es_referencia_persona and ref not in filas_sin_persona:
                continue  # este rol SI tiene persona -- se deja que la formula muestre el dato real
            if es_referencia_persona or filas_persona_regex.search(cell.value):
                celdas_vaciadas.add(cell.coordinate)
                try:
                    cell.value = ""
                except AttributeError:
                    pass  # celda combinada -- no se puede escribir directo, se deja como esta
                cell.number_format = "General"

    # Algunas celdas no referencian EXPORTAR directamente, sino OTRA
    # celda de la misma hoja que a su vez si referencia EXPORTAR (ej.
    # "=D5", donde D5 es una de las celdas que se acaba de vaciar arriba)
    # -- se repite unas cuantas vueltas para seguir esas cadenas, por si
    # hay mas de un nivel (una celda que referencia a otra que
    # referencia a otra).
    referencia_local_regex = re.compile(r"^=\$?([A-Z]{1,3})\$?([0-9]+)$")
    for _ in range(3):
        nuevas = set()
        for row in hoja.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    m = referencia_local_regex.match(cell.value.strip())
                    if m and m.group(1) + m.group(2) in celdas_vaciadas and cell.coordinate not in celdas_vaciadas:
                        try:
                            cell.value = ""
                        except AttributeError:
                            pass
                        cell.number_format = "General"
                        nuevas.add(cell.coordinate)
        if not nuevas:
            break
        celdas_vaciadas |= nuevas

    # Las 3 hojas de FORMULARIO tienen casillas que se resaltan en verde
    # segun el tramite/clase/combustible/servicio del vehiculo. Primero
    # se limpia CUALQUIER resaltado que pudiera haber (por si la plantilla
    # trae algo marcado de una prueba anterior), y despues se vuelve a
    # aplicar SOLO segun los datos reales de este vehiculo. El resaltado
    # de TRAMITE (unico, campo "tramite") no aplica todavia (esta
    # herramienta aun no deja elegir un tramite asi) -- si se agrega mas
    # adelante, ya queda listo: basta con mandar datos_vehiculo["tramite"].
    if nombre_hoja in ("FORMULARIO", "FORMULARIO (2)", "FORMULARIO (3)"):
        sin_relleno = PatternFill(fill_type=None)
        for mapa_celdas in (CELDAS_TRAMITE, CELDAS_CLASE, CELDAS_COMBUSTIBLE):
            for celdas in mapa_celdas.values():
                for celda in celdas:
                    hoja[celda].fill = sin_relleno
        # El servicio SI cambia de celdas entre variantes -- se limpian
        # las 3 posibles ubicaciones conocidas para curar cualquier
        # resto, sin importar cual le corresponde a esta hoja en concreto.
        for _mapa_serv in CELDAS_SERVICIO_POR_DOCUMENTO.values():
            for _celdas_serv in _mapa_serv.values():
                for _celda_serv in _celdas_serv:
                    try:
                        hoja[_celda_serv].fill = sin_relleno
                    except Exception:
                        pass
        for celda in CELDA_OTROS_TRAMITE:
            hoja[celda].fill = sin_relleno
        hoja["W38"].fill = sin_relleno  # bloque "ESPECIFIQUE LA PALABRA OTRO..." (combinado W38:AK40)
        # W41/AG41 muestran el texto de "traslado de cuenta" via formula
        # (=EXPORTAR!D51) -- LibreOffice no siempre recalcula esa formula
        # en la conversion a PDF, asi que se escribe vacio DIRECTAMENTE en
        # la celda visible en vez de depender de la formula. Protegido por
        # si acaso en "(2)"/"(3)" esa celda resulta combinada distinto.
        for celda_fija in ("W41", "AG41"):
            try:
                hoja[celda_fija].value = ""
            except AttributeError:
                pass
            hoja[celda_fija].fill = sin_relleno

        # Datos del VEHICULO (placa, marca, VIN, capacidad, etc.) -- estas
        # coordenadas SI son identicas en las 3 hojas (confirmado
        # revisando la plantilla), asi que este bloque corre siempre,
        # para las 3 variantes. Sin esto, un campo vacio (ej. sin VIN)
        # aparecia como "0" en vez de blanco -- una formula que apunta a
        # una celda vacia se evalua como 0 en Excel/LibreOffice, sin
        # importar el formato de la celda que muestra el resultado.
        for celda, clave in CELDAS_REFERENCIA_SIMPLE_VEHICULO.items():
            try:
                hoja[celda] = datos_vehiculo.get(clave) or ""
            except Exception:
                pass

        # CELDAS_REFERENCIA_SIMPLE_PERSONAS usa coordenadas de celda FIJAS
        # que solo coinciden con el layout de la hoja BASE "FORMULARIO" --
        # en "FORMULARIO (2)"/"(3)" esas mismas coordenadas caen en celdas
        # distintas (por el layout de dos personas), asi que escribir ahi
        # SOBREESCRIBIA datos de otro campo (esto causaba que el segundo
        # comprador mostrara los mismos datos que el primero). Se
        # restringe este bloque a que SOLO corra en la hoja base.
        if nombre_hoja == "FORMULARIO":
            # CELDAS_REFERENCIA_SIMPLE_PERSONAS espera claves "planas" (ej.
            # "propietario_nombres"), mientras que el resto de esta funcion
            # recibe los datos de personas como diccionarios (ej.
            # datos_vehiculo["propietario"] = {"nombres": ..., ...}) -- se
            # traduce de un formato al otro aqui, solo para FORMULARIO.
            for _rol, _prefijo in (("propietario", "propietario"), ("comprador", "comprador")):
                _persona_rol = datos_vehiculo.get(_rol)
                if _persona_rol:
                    datos_vehiculo[f"{_prefijo}_nombres"] = _persona_rol.get("nombres", "")
                    datos_vehiculo[f"{_prefijo}_primer_apellido"] = _persona_rol.get("apellido", "")
                    datos_vehiculo[f"{_prefijo}_segundo_apellido"] = _persona_rol.get("segundo_apellido", "")
                    datos_vehiculo[f"{_prefijo}_documento"] = _persona_rol.get("numero_documento", "")
                    datos_vehiculo[f"{_prefijo}_direccion"] = _persona_rol.get("direccion", "")
                    datos_vehiculo[f"{_prefijo}_ciudad"] = _persona_rol.get("ciudad", "")
                    datos_vehiculo[f"{_prefijo}_telefono"] = _persona_rol.get("telefono", "")

            # Igual que generar_fun: las celdas de "referencia simple" (que
            # no son formulas de EXPORTAR, sino texto/numero directo, como
            # documento/telefono del propietario y comprador) tambien deben
            # quedar en blanco cuando no hay ese dato -- si no, algunas
            # aparecen como "0" en vez de vacio.
            for celda, clave in CELDAS_REFERENCIA_SIMPLE_PERSONAS.items():
                if datos_vehiculo.get(clave):
                    try:
                        hoja[celda] = datos_vehiculo[clave]
                    except Exception:
                        pass
                else:
                    try:
                        hoja[celda] = ""
                    except Exception:
                        pass


        _fun_marcar_checkboxes(hoja, CELDAS_CLASE, datos_vehiculo.get("clase", ""))
        _fun_marcar_checkboxes(hoja, CELDAS_COMBUSTIBLE, datos_vehiculo.get("combustible", ""))

        # Servicio -- usa el mapeo especifico de ESTE documento (las
        # celdas cambian de posicion entre variantes, a diferencia de
        # clase/combustible que si son iguales en las 3).
        _servicio_normalizado = _fun_normalizar(datos_vehiculo.get("servicio", ""))
        _mapa_servicio_doc = CELDAS_SERVICIO_POR_DOCUMENTO.get(clave_documento, {})
        for _etiqueta_serv, _celdas_serv in _mapa_servicio_doc.items():
            if _fun_coincide(datos_vehiculo.get("servicio", ""), _etiqueta_serv):
                for _celda_serv in _celdas_serv:
                    hoja[_celda_serv].fill = VERDE_MARCA

        # Tipo de documento (C.C / NIT) del propietario y del comprador --
        # solo se resalta si es exactamente uno de esos dos tipos (los
        # demas, como Pasaporte o T.I., no se marcan).
        _mapa_tipodoc_doc = CELDAS_TIPO_DOC_FORMULARIO.get(clave_documento, {})
        for _rol_td, _opciones_td in _mapa_tipodoc_doc.items():
            _persona_td = datos_vehiculo.get(_rol_td) or {}
            _tipo_doc_persona = (_persona_td.get("tipo_documento") or "").strip().upper()
            _celdas_td = _opciones_td.get(_tipo_doc_persona)
            if _celdas_td:
                for _celda_td in _celdas_td:
                    hoja[_celda_td].fill = VERDE_MARCA

        if datos_vehiculo.get("tramite"):
            _fun_marcar_checkboxes(hoja, CELDAS_TRAMITE, datos_vehiculo["tramite"])
            if "TRASLADO" in _fun_normalizar(datos_vehiculo["tramite"]):
                hoja["W38"].fill = VERDE_MARCA

    hojas_a_conservar = {nombre_hoja, "EXPORTAR", "DATOS"}
    for nombre in list(wb.sheetnames):
        if nombre not in hojas_a_conservar:
            del wb[nombre]
    hoja.sheet_state = "visible"
    # EXPORTAR y DATOS deben seguir EXISTIENDO (las formulas de la hoja
    # del documento las necesitan), pero OCULTAS -- si no, LibreOffice
    # las imprime tambien como paginas propias ademas del documento real
    # (EXPORTAR sola ya son decenas de paginas de "MENU").
    exportar.sheet_state = "hidden"
    if "DATOS" in wb.sheetnames:
        wb["DATOS"].sheet_state = "hidden"

    # Todos estos documentos deben salir en UNA sola pagina -- ninguno
    # tenia "ajustar a una pagina" configurado en la plantilla, asi que
    # el area usada se desbordaba a una segunda pagina casi vacia (con
    # el titulo y un pie de pagina de marca repetidos -- ese pie
    # resulto ser un elemento flotante, no una celda, asi que no se pudo
    # recortar con precision; en vez de eso se encoge todo para que
    # quepa siempre en una pagina).
    # NOTA: se evita la propiedad "hoja.page_setup.fitToPage" porque
    # tiene un bug interno en openpyxl (intenta usar el worksheet padre,
    # que a veces no esta enlazado) -- se marca directo en
    # sheet_properties.pageSetUpPr en su lugar.
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 1
    if hoja.sheet_properties.pageSetUpPr is None:
        from openpyxl.worksheet.properties import PageSetupProperties
        hoja.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        hoja.sheet_properties.pageSetUpPr.fitToPage = True
    wb.active = wb.sheetnames.index(nombre_hoja)

    # Resaltado de tramites (solo aplica a las 3 variantes de
    # Formulario -- son las unicas que tienen esta cuadricula de
    # casillas). Se normaliza el texto (mayusculas, espacios multiples
    # colapsados) para que coincida sin importar variaciones pequeñas de
    # como se escribio en el modulo de tramites.
    if clave_documento in APPJX_CELDA_TRASLADO_TEXTO:
        tramites_elegidos = datos_vehiculo.get("tramites_seleccionados") or []
        for tramite_texto in tramites_elegidos:
            tramite_normalizado = re.sub(r"\s+", " ", (tramite_texto or "").strip().upper())
            if tramite_normalizado == "OTROS":
                celdas_marcar = [CELDA_OTROS_TRAMITE_FORMULARIO]
            else:
                # Se busca por coincidencia normalizada contra las claves
                # del mapeo (tambien normalizadas), en vez de exigir un
                # match exacto de texto.
                celdas_marcar = []
                for clave_mapa, celdas in CELDAS_TRAMITE_FORMULARIO.items():
                    if re.sub(r"\s+", " ", clave_mapa.strip().upper()) == tramite_normalizado:
                        celdas_marcar = [celdas]
                        break
            for celda_num, celda_etq in celdas_marcar:
                hoja[celda_num].fill = VERDE_MARCA
                hoja[celda_etq].fill = VERDE_MARCA

            # Caso especial: "TRASLADO MATRICULA / REGISTRO" ademas
            # escribe el texto DIRECTO (no por formula) en la celda de
            # abajo del parrafo "ESPECIFIQUE LA PALABRA OTRO..." -- igual
            # que con el telefono, una formula que depende de una celda
            # vacia puede fallar al convertir a PDF, asi que se escribe
            # el valor ya armado. Usa el MUNICIPIO DE DESTINO elegido a
            # mano en el frontend (no el municipio del vehiculo -- ese es
            # de donde SALE el tramite, no hacia donde se traslada).
            if tramite_normalizado == "TRASLADO DE CUENTA":
                celda_traslado = APPJX_CELDA_TRASLADO_TEXTO[clave_documento]
                municipio_destino = (datos_vehiculo.get("traslado_municipio_destino") or "").strip()
                celda_obj = hoja[celda_traslado]
                celda_obj.value = f"Traslado de Cuenta hacia la secretaria de transito de {municipio_destino}".strip()
                celda_obj.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                celda_obj.fill = VERDE_MARCA

    id_temp = str(uuid.uuid4())[:8]
    ruta_xlsm_temp = f"/tmp/_appjxdoc_{id_temp}.xlsm"
    wb.save(ruta_xlsm_temp)

    subprocess.run([
        "soffice", "--headless", "--convert-to", "pdf",
        "--outdir", os.path.dirname(ruta_salida_pdf), ruta_xlsm_temp
    ], check=True, timeout=90)

    generado = os.path.join(os.path.dirname(ruta_salida_pdf), f"_appjxdoc_{id_temp}.pdf")
    shutil.move(generado, ruta_salida_pdf)
    os.remove(ruta_xlsm_temp)


def _moneda(valor):
    """Formatea un valor numerico como texto de moneda '$ X,XXX,XXX' para
    la declaracion manual."""
    try:
        return "$ {:,.0f}".format(float(valor))
    except (TypeError, ValueError):
        return valor


def generar_declaracion_manual_pdf(datos, ruta_salida_pdf):
    """Genera la Declaracion Manual (formulario FO-M8-P6-008) diligenciada,
    a partir de la plantilla Excel real (hojas CONTRIBUYENTE y ENTIDAD
    RECAUDADORA). 'datos' es un dict con todos los campos ya resueltos:
    datos del vehiculo, del propietario, y la liquidacion privada."""
    wb = _openpyxl.load_workbook(DECLARACION_MANUAL_PLANTILLA)

    # FECHA GENERACIÓN / FECHA LÍMITE PAGO -- el PDF de la Declaracion
    # Sugerida (que viene directo de la Gobernacion) siempre trae estas
    # dos fechas. Se calculan aqui con la MISMA logica de validez que ya
    # se usa para el cache (_antioquia_calcular_validez_pdf): una
    # vigencia vencida solo sirve el mismo dia en que se genero (los
    # intereses suben a diario); la del año en curso sirve durante toda
    # la ventana de pronto pago (hasta el 30 de abril o el 31 de julio,
    # segun cuando se genere), y desde el 1 de agosto se comporta igual
    # que una vencida. Las etiquetas fijas ("FECHA LÍMITE PAGO" / "FECHA
    # GENERACIÓN") ya estan en la plantilla (E68/E69) -- aqui solo se
    # escribe el valor de cada fecha, en las celdas AE68/AE69.
    hoy = datetime.utcnow() - timedelta(hours=5)  # Colombia = UTC-5 todo el año, sin horario de verano
    fecha_generacion_str = hoy.strftime("%d/%m/%Y")
    fecha_limite = _antioquia_calcular_validez_pdf(datos.get("vigencia", hoy.year))
    fecha_limite_str = fecha_limite.strftime("%d/%m/%Y")

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]

        # A. Periodo
        ws["H12"] = datos.get("vigencia", "")

        ws["AE68"] = fecha_limite_str
        ws["AE69"] = fecha_generacion_str

        # Las etiquetas fijas "FECHA LÍMITE PAGO" / "FECHA GENERACIÓN"
        # (E68/E69) usaban color de tema (theme=1) en vez de un color
        # fijo -- LibreOffice (el motor que convierte a PDF en el
        # servidor) a veces no resuelve bien ese color de tema y el
        # texto sale invisible, aunque en Excel se vea negro normal.
        # Mismo problema que ya se habia resuelto antes en otra parte de
        # este documento (D37/D39/D41-43) -- se usa el MISMO patron que
        # ya funciona ahi: copiar la fuente completa con copy.copy() y
        # solo cambiar el color (reconstruir un Font() nuevo desde cero
        # perdia algun atributo que hacia que LibreOffice lo siguiera
        # tratando como color de tema).
        for celda_etiqueta in ["E68", "E69"]:
            celda = ws[celda_etiqueta]
            nueva_fuente = copy.copy(celda.font)
            nueva_fuente.color = "FF000000"
            celda.font = nueva_fuente

        # C. Declarante
        ws["D15"] = datos.get("nombre_completo", "")
        ws["D18"] = datos.get("apellidos", "")
        ws["AZ18"] = datos.get("celular", "")
        ws["CA18"] = datos.get("telefono", "")
        ws["CO18"] = datos.get("email", "")
        ws["D21"] = datos.get("direccion", "")
        ws["BE21"] = datos.get("municipio_residencia", "")
        ws["CI21"] = datos.get("departamento_residencia", "ANTIOQUIA")
        ws["CR15"] = datos.get("numero_documento", "")
        tipo_doc = (datos.get("tipo_documento") or "").upper()
        casillas_tipo_doc = {"CC": "BM15", "NIT": "BV15", "TI": "CC15", "CE": "CH15", "OTRO": "CN15"}
        if tipo_doc in casillas_tipo_doc:
            ws[casillas_tipo_doc[tipo_doc]] = "X"

        # D. Vehiculo
        ws["D24"] = datos.get("placa", "")
        ws["Z24"] = datos.get("marca", "")
        ws["BB24"] = datos.get("linea", "")
        ws["CP24"] = datos.get("modelo", "")
        ws["D27"] = datos.get("clase", "")
        ws["AJ27"] = datos.get("carroceria", "")
        # D7 GRUPO se deja en blanco (instructivo oficial)
        ws["BZ27"] = datos.get("puertas", "")
        ws["CL27"] = datos.get("cilindraje", "")
        ws["D30"] = datos.get("capacidad_carga", "")
        ws["AJ30"] = datos.get("capacidad_pasajeros", "")
        ws["BP30"] = datos.get("municipio_matricula", "")
        ws["CL30"] = datos.get("departamento_matricula", "ANTIOQUIA")

        if datos.get("blindado"):
            ws["S33"] = "X"
        if datos.get("importado"):
            ws["AN33"] = "X"

        ws["CI35"] = datos.get("caja", "")
        ws["CY35"] = datos.get("traccion", "")

        # E. Liquidacion privada -- con formato de moneda
        ws["AF38"] = _moneda(datos.get("avaluo", 0))
        ws["AF40"] = _moneda(datos.get("impuesto", 0))
        ws["AF41"] = _moneda(datos.get("sanciones", 0))
        ws["AF42"] = _moneda(datos.get("descuentos", 0))
        ws["AF43"] = _moneda(datos.get("total_cargo_5", 0))
        ws["CJ37"] = _moneda(datos.get("total_cargo_6", 0))
        ws["CJ38"] = _moneda(datos.get("intereses_mora", 0))
        ws["CJ39"] = _moneda(datos.get("pagos_anteriores", 0))
        ws["CJ40"] = _moneda(datos.get("descuento_interes", 0))
        ws["CJ41"] = _moneda(datos.get("saldo_favor", 0))
        ws["CJ42"] = _moneda(datos.get("total_pagar", 0))

        # J. Distribucion del recaudo -- 20% Municipio, 80% Departamento,
        # calculado sobre el total a pagar (instructivo oficial, seccion J).
        total_pagar_num = datos.get("total_pagar", 0) or 0
        try:
            total_pagar_num = float(total_pagar_num)
        except (TypeError, ValueError):
            total_pagar_num = 0
        valor_municipio = round(total_pagar_num * 0.20)
        valor_departamento = total_pagar_num - valor_municipio
        ws["AE64"] = _moneda(valor_municipio)
        ws["AE66"] = _moneda(valor_departamento)

        # G. Declarante -- se deja en blanco a proposito. Se firma y se
        # diligencia a lapicero de forma manual, no se prellena.

        # Configuracion de pagina: una sola pagina por hoja, vertical.
        ws.print_area = "B3:DI82"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.gridLines = False
        ws.sheet_view.showGridLines = False

        # Limpiar bordes sueltos (restos de la plantilla) que generaban
        # lineas verticales gruesas sin pertenecer a ninguna casilla real.
        sin_borde = Border()
        for coord in ["I7", "K7", "G35", "H35", "M49", "B50", "B64", "B66", "J66", "L66", "N66", "V66"]:
            ws[coord].border = sin_borde

        # D50 (G.2 Nombres y Apellidos) SI es la celda real de esa casilla
        # (no un borde suelto) -- se le restaura el contorno completo.
        borde_fino = Side(style="thin")
        ws["D50"].border = Border(top=borde_fino, bottom=borde_fino, left=borde_fino, right=borde_fino)

        # F40 ("VEHICULOS AUTOMOTORES") y F38 ("DEL VEHICULO") tenian
        # alineacion vertical sin fijar -- al agrandar sus filas el texto
        # se hundia y se veia separado de la linea de arriba.
        celda_f40 = ws["F40"]
        celda_f40.value = "VEHICULOS AUTOMOTORES"
        celda_f40.alignment = Alignment(vertical="top", horizontal=celda_f40.alignment.horizontal)
        celda_f38 = ws["F38"]
        celda_f38.alignment = Alignment(vertical="top", horizontal=celda_f38.alignment.horizontal)

        ws.row_dimensions[38].height = 14
        ws.row_dimensions[40].height = 14
        ws.row_dimensions[28].height = 13.5
        ws.row_dimensions[29].height = 1.5
        ws.row_dimensions[27].height = 16.5

        # Los numeros de renglon (1 al 5) mostraban "###" en el PDF sin
        # importar el ancho de columna -- convertirlos de numero a TEXTO
        # elimina el problema de raiz (el "###" solo le pasa a numeros).
        ws["D37"] = "1"
        ws["D39"] = "2"
        ws["D41"] = "3"
        ws["D42"] = "4"
        ws["D43"] = "5"

        # Ademas usan un color de TEMA que LibreOffice a veces interpreta
        # mal al convertir (texto invisible) -- se fija a negro explicito.
        for coord in ["D37", "D39", "D41", "D42", "D43"]:
            celda = ws[coord]
            nueva_fuente = copy.copy(celda.font)
            nueva_fuente.color = "FF000000"
            celda.font = nueva_fuente

    id_temp = str(uuid.uuid4())[:8]
    ruta_xlsx_temp = f"/tmp/_decl_manual_{id_temp}.xlsx"
    wb.save(ruta_xlsx_temp)

    subprocess.run([
        "soffice", "--headless", "--convert-to", "pdf",
        "--outdir", os.path.dirname(ruta_salida_pdf), ruta_xlsx_temp
    ], check=True, timeout=90)

    generado = os.path.join(os.path.dirname(ruta_salida_pdf), f"_decl_manual_{id_temp}.pdf")
    shutil.move(generado, ruta_salida_pdf)
    os.remove(ruta_xlsx_temp)


def _moneda_pys(valor):
    """Formatea un valor como '$\\xa0X.XXX.XXX', igual al formato que ya
    usa la plantilla PYS de AppJX.xlsm (con espacio duro y punto de miles)."""
    try:
        return "$\xa0" + "{:,.0f}".format(float(valor)).replace(",", ".")
    except (TypeError, ValueError):
        return "$\xa00"


def _desfusionar_zona_tabla(ws, fila_inicio, fila_fin, columnas):
    """Quita cualquier fusion de celdas que se cruce con el rango de
    filas/columnas indicado. Se usa antes de llenar una tabla fila por
    fila (como la de declaraciones), porque si una celda combinada
    abarca VARIAS filas de esa tabla, escribir en cada fila por separado
    terminaba pisando el valor anterior (todas las filas de una misma
    fusion apuntan a la misma celda 'ancla') -- en vez de eso, se separa
    la fusion de una vez para que cada fila quede independiente.

    Ademas, vuelve a aplicar un borde fino explicito a CADA celda del
    ANCHO COMPLETO de las fusiones que se separan (no solo a la columna
    que se estaba buscando) -- las celdas combinadas en Excel suelen
    usar la propiedad 'outline' (que solo se ve bien mientras siguen
    fusionadas) en vez de un borde normal en cada celda individual, asi
    que al separar una fusion de varias columnas de ancho (ej. A:O),
    solo la celda ancla (A) tenia borde propio -- las demas (B a O)
    quedaban sin ninguno, dando la apariencia de seguir combinadas."""
    columnas = set(columnas)
    rangos_a_quitar = []
    columnas_con_borde = set(columnas)  # se va ampliando con el ancho real de cada fusion encontrada
    for rango in list(ws.merged_cells.ranges):
        si_se_cruza = (rango.min_row <= fila_fin and rango.max_row >= fila_inicio
                       and any(c in columnas for c in range(rango.min_col, rango.max_col + 1)))
        if si_se_cruza:
            rangos_a_quitar.append(str(rango))
            columnas_con_borde.update(range(rango.min_col, rango.max_col + 1))
    for rango_texto in rangos_a_quitar:
        ws.unmerge_cells(rango_texto)

    borde_fino = Side(style="thin", color="FF000000")
    columna_izq = min(columnas_con_borde)
    columna_der = max(columnas_con_borde)
    for fila in range(fila_inicio, fila_fin + 1):
        for columna in columnas_con_borde:
            ws.cell(row=fila, column=columna).border = Border(
                left=borde_fino if columna == columna_izq else None,
                right=borde_fino if columna == columna_der else None,
                top=borde_fino, bottom=borde_fino,
            )


def generar_estado_cuenta_pdf(datos, ruta_salida_pdf):
    """Genera el documento Estado de Cuenta (certificado de paz y salvo),
    a partir de la plantilla AppJX.xlsm real (hojas PYS + ESTADO DE
    CUENTA). 'datos' es un dict con: estado_veh (dict de estadoCuenta),
    lista_detalle_pagos, lista_proceso_fiscal, lista_bloqueo, novedades."""
    wb = _openpyxl.load_workbook(FUN_PLANTILLA, data_only=False, keep_vba=True)
    pys = wb["PYS"]
    edc = wb["ESTADO DE CUENTA"]

    estado_veh = datos.get("estado_veh", {}) or {}
    inicio = estado_veh.get("periodoInicioCertificacion", "")
    fin    = estado_veh.get("periodoFinCertificacion", "")

    # A22:B31 -- informacion general
    # A22 es la ETIQUETA "PERIODO DE CERTIFICACION" -- la plantilla la
    # deja vacia esperando que el codigo la escriba (junto con el valor
    # en B22); si se deja vacia, la formula de la hoja ESTADO DE CUENTA
    # que la usa (UPPER(A22)) la interpreta como "0" en vez de texto
    # vacio, y por eso aparecia "0" en vez de la etiqueta real.
    pys["A22"] = "PERÍODO DE CERTIFICACIÓN"
    pys["B22"] = f"{inicio} a {fin}" if inicio and fin else ""
    pys["B23"] = estado_veh.get("placa", "")
    pys["B24"] = estado_veh.get("modelo", "")
    pys["B25"] = estado_veh.get("municipioMatricula", "")
    pys["B26"] = estado_veh.get("departamentoMatricula", "")
    # Fecha de expedicion -- debe ser la fecha en que REALMENTE se
    # consulto y se obtuvo este numero de certificado (guardada en
    # 'fecha_consulta'), no la fecha en que se genera el documento --
    # el numero de certificado es unico de esa consulta puntual.
    # OJO: el servidor corre en UTC, pero la fecha debe ser la de
    # Colombia (UTC-5) -- sin este ajuste, de 7pm a 12am hora Colombia
    # el servidor ya cree que es el dia siguiente.
    fecha_consulta = datos.get("fecha_consulta")
    if fecha_consulta:
        pys["B27"] = fecha_consulta - timedelta(hours=5)

    # "FECHA DE EXPEDICIÓN" (la que se ve en el certificado final) en
    # realidad NO sale de B27 -- sale de PYS!AA26, que en la plantilla
    # traia la formula "=TODAY()". Esa formula la calcula LibreOffice al
    # momento de convertir a PDF usando la hora DEL SERVIDOR (UTC), sin
    # ningun ajuste posible desde Python -- por eso mostraba un dia
    # adelantado en las noches. Se reemplaza esa formula por el valor ya
    # calculado (con el ajuste de zona horaria de Colombia aplicado),
    # sin mover la celda de lugar. SIEMPRE se usa la fecha de HOY (no la
    # fecha en que se hizo la consulta original), como se pidio.
    # IMPORTANTE: se escribe SOLO la fecha (.date(), sin hora/minutos),
    # porque AA26 tambien alimenta un VLOOKUP (via U24) que necesita
    # coincidencia EXACTA contra una tabla de fechas -- escribir un
    # datetime completo (con hora) rompia esa busqueda exacta, lo cual
    # en cascada causaba que el documento saliera con muchas paginas de
    # mas en vez de las 1-3 normales.
    hoy_colombia = (datetime.utcnow() - timedelta(hours=5)).date()
    pys["AA26"] = hoy_colombia
    pys["B28"] = estado_veh.get("marca", "")
    pys["B29"] = estado_veh.get("cilindraje", "")
    pys["B30"] = estado_veh.get("linea", "")
    pys["B31"] = estado_veh.get("capacidadCarga", "")

    # A35:J.. -- declaraciones presentadas (una fila por cada elemento)
    declaraciones = datos.get("lista_detalle_pagos", []) or []
    if declaraciones:
        _desfusionar_zona_tabla(pys, 35, 35 + len(declaraciones) - 1, range(1, 11))
    fila = 35
    for d in declaraciones:
        pys.cell(row=fila, column=1,  value=d.get("tipoLiquidacion", ""))
        pys.cell(row=fila, column=2,  value=d.get("formularioLiquidacion", ""))
        fecha_pago = d.get("fechaPago")
        if fecha_pago:
            try:
                pys.cell(row=fila, column=3, value=datetime.utcfromtimestamp(fecha_pago / 1000))
            except (TypeError, ValueError, OSError):
                pys.cell(row=fila, column=3, value="")
        pys.cell(row=fila, column=4,  value=_moneda_pys(d.get("impuesto", 0)))
        pys.cell(row=fila, column=5,  value=_moneda_pys(d.get("sancion", 0)))
        pys.cell(row=fila, column=6,  value=_moneda_pys(d.get("descuento", 0)))
        pys.cell(row=fila, column=7,  value=_moneda_pys(d.get("interesMora", 0)))
        pys.cell(row=fila, column=8,  value=_moneda_pys(d.get("totalPagar", 0)))
        pys.cell(row=fila, column=9,  value=_moneda_pys(d.get("avaluoComercial", 0)))
        pys.cell(row=fila, column=10, value=d.get("vigencia", ""))
        fila += 1

    # M22.. / O22.. / P22.. -- procesos fiscales, bloqueos, novedades
    procesos = datos.get("lista_proceso_fiscal", []) or []
    bloqueos = datos.get("lista_bloqueo", []) or []
    novedades = datos.get("novedades", []) or []

    if procesos:
        _desfusionar_zona_tabla(pys, 22, 22 + len(procesos) - 1, [13])
    if bloqueos:
        _desfusionar_zona_tabla(pys, 22, 22 + len(bloqueos) - 1, [15])
    if novedades:
        _desfusionar_zona_tabla(pys, 22, 22 + len(novedades) - 1, [16])

    for i, p in enumerate(procesos):
        texto = f"{p.get('descripcionProcesoFiscal', '')} ({p.get('vigencia', '')})"
        pys.cell(row=22 + i, column=13, value=texto)  # M
    for i, b in enumerate(bloqueos):
        texto = f"{b.get('descripcionBloqueo', '')} ({b.get('vigencia', '')})"
        pys.cell(row=22 + i, column=15, value=texto)  # O
    for i, n in enumerate(novedades):
        if isinstance(n, dict):
            descripcion = n.get("descripcionNovedad", "")
            fecha_raw = (n.get("fechaNovedad") or "")[:10]  # "YYYY-MM-DD"
            fecha_fmt = fecha_raw
            if fecha_raw:
                try:
                    fecha_fmt = datetime.strptime(fecha_raw, "%Y-%m-%d").strftime("%d/%m/%Y")
                except ValueError:
                    fecha_fmt = fecha_raw
            texto = f"{descripcion} - {fecha_fmt}" if fecha_fmt else descripcion
        else:
            texto = str(n)
        pys.cell(row=22 + i, column=16, value=texto)  # P

    # "El vehiculo no presenta observaciones" solo si las 3 listas estan
    # vacias -- si hay cualquier cosa, se quita ese aviso.
    if procesos or bloqueos or novedades:
        edc["A74"] = ""

    # Columna "VIGENCIAS ADEUDADAS" (A53:A72) siempre se deja en blanco --
    # este documento solo se genera cuando el vehiculo esta a paz y salvo,
    # nunca hay vigencias adeudadas que mostrar aqui. Se sobrescribe
    # directo porque la formula original (=IF(AND(...),"")) no tiene rama
    # para cuando la condicion es falsa, y en ese caso Excel/LibreOffice
    # muestra el texto literal "FALSE" en vez de dejarlo vacio.
    _desfusionar_zona_tabla(edc, 53, 72, [1])
    for r in range(53, 73):
        edc.cell(row=r, column=1, value="")

    # Borde suelto (resto de la plantilla) que cortaba visualmente el
    # texto "Tipos de declaraciones..." justo en la palabra "Corrección".
    edc["O47"].border = Border()

    # Certificado No. -- se usa el numero real que entrega la Gobernacion
    # en esta consulta puntual (cambia cada vez que se consulta).
    edc["AG5"] = estado_veh.get("numeroCertificadoSap", "")

    # NOTA: se habia reducido la altura de esta fila (de 27.75 a 14) para
    # acercar el "CERTIFICADO No." al titulo, pero eso dejo apeñuscado el
    # texto "El suscrito funcionario..." con el logo. Se prueba un punto
    # intermedio (20) que de un poco de aire sin comerse tanta capacidad
    # de la tabla de declaraciones (con la altura original de 27.75, un
    # caso de 18 declaraciones ya no cabia en una sola pagina).
    edc.row_dimensions[5].height = 18

    # NOTA: se habia reducido la altura de esta fila (de 27.75 a 14) para
    # acercar el "CERTIFICADO No." al titulo, pero eso dejo apeñuscado el
    # texto "El suscrito funcionario..." con el logo justo debajo -- se
    # revierte, la altura original ya tenia un espacio aceptable.

    # "Avaluo para la vigencia" -- CORREGIDO: se usa el avaluo de la
    # declaracion MAS RECIENTE (la de mayor vigencia dentro de la tabla
    # de declaraciones), no estadoCuenta.avaluoComercial -- se confirmo
    # con un caso real que ese campo general de la Gobernacion puede
    # traer un valor distinto (de otra referencia) al que realmente
    # aparece declarado para el año en curso en la propia tabla de este
    # mismo documento.
    avaluo_vigencia_actual = None
    mejor_vigencia_pdf = -1
    for d in declaraciones:
        try:
            vig_d = int(d.get("vigencia", 0) or 0)
        except (TypeError, ValueError):
            continue
        if vig_d > mejor_vigencia_pdf:
            mejor_vigencia_pdf = vig_d
            avaluo_vigencia_actual = d.get("avaluoComercial", 0)
    if not avaluo_vigencia_actual:
        avaluo_vigencia_actual = estado_veh.get("avaluoComercial", 0)  # respaldo si no hay declaraciones
    edc["AE76"] = _moneda_pys(avaluo_vigencia_actual)

    # Ocultar las filas vacias sobrantes de ambas tablas (no todas las
    # placas tienen 30 declaraciones ni observaciones). Se OCULTAN en vez
    # de borrarlas (LibreOffice no imprime filas ocultas) porque borrar
    # filas con openpyxl en una hoja con tantas celdas fusionadas como
    # esta corrompe las fusiones y daña el diseno.
    def _ocultar_filas(hoja, desde, hasta):
        for r in range(desde, hasta + 1):
            hoja.row_dimensions[r].hidden = True

    FILA_OBS_INICIO, FILA_OBS_FIN = 53, 72
    max_obs = max(len(procesos), len(bloqueos), len(novedades))
    if max_obs == 0:
        _ocultar_filas(edc, FILA_OBS_INICIO, FILA_OBS_FIN)
    else:
        ultima_fila_obs_usada = FILA_OBS_INICIO + max_obs - 1
        if ultima_fila_obs_usada < FILA_OBS_FIN:
            _ocultar_filas(edc, ultima_fila_obs_usada + 1, FILA_OBS_FIN)

    # Se libera una fila mas para la tabla de declaraciones (para que
    # quepa una vigencia adicional): la leyenda "* Tipos de declaraciones
    # ..." que vivia en la fila 47 se traslada a la fila 48 (que era un
    # espaciador chico y quedaba libre), y la fila 47 pasa a ser una fila
    # mas de la tabla -- copiando las FORMULAS reales de la fila 46 (que
    # trae los datos de PYS!*65) pero apuntando a PYS!*66, para que la
    # fila nueva si traiga datos de verdad y no quede vacia.
    celda_leyenda = edc["A47"]
    edc["A48"] = celda_leyenda.value
    edc["A48"].font = copy.copy(celda_leyenda.font)
    edc["A48"].alignment = Alignment(
        horizontal=celda_leyenda.alignment.horizontal,
        vertical=celda_leyenda.alignment.vertical,
        wrapText=False
    )
    # Fusion ancha (igual que el parrafo legal de la fila 6, que va de A a
    # BH) para que el texto quede en UNA sola linea, sin partirse en dos.
    edc.merge_cells("A48:BH48")

    for col in range(1, 60):
        celda_origen = edc.cell(row=46, column=col)
        if isinstance(celda_origen.value, str) and celda_origen.value.startswith("="):
            celda_destino = edc.cell(row=47, column=col)
            celda_destino.value = celda_origen.value.replace("65", "66")
            celda_destino.font = copy.copy(celda_origen.font)
            celda_destino.alignment = copy.copy(celda_origen.alignment)
            celda_destino.border = copy.copy(celda_origen.border)
            celda_destino.number_format = celda_origen.number_format
        elif col == 1:
            # A47 tenia la leyenda -- se limpia para dejarla lista como
            # celda de datos (ya se copio arriba a A48).
            celda_leyenda.value = None

    edc.row_dimensions[47].height = 21.0   # misma altura que las demas filas de declaraciones
    edc.row_dimensions[48].height = 20.25  # altura que antes tenia la leyenda, para que quepa completa

    FILA_DECL_INICIO, FILA_DECL_FIN = 16, 47
    filas_declaraciones = len(declaraciones)
    if filas_declaraciones == 0:
        _ocultar_filas(edc, FILA_DECL_INICIO, FILA_DECL_FIN)
    else:
        ultima_fila_decl_usada = FILA_DECL_INICIO + filas_declaraciones - 1
        if ultima_fila_decl_usada < FILA_DECL_FIN:
            _ocultar_filas(edc, ultima_fila_decl_usada + 1, FILA_DECL_FIN)

    # "Observaciones" y "Avaluo para la vigencia" deben quedar SIEMPRE
    # juntos en la misma pagina, y nunca deben cortarse a la mitad -- se
    # fuerza un salto de pagina justo antes de "OBSERVACIONES" solo
    # cuando de verdad hace falta. IMPORTANTE: esto NO se basa en
    # "cuantas filas caben", sino en el ESPACIO REAL en puntos que queda
    # en la pagina justo despues de donde termino la tabla de
    # Declaraciones -- si la tabla de Declaraciones ya se desbordo por
    # su cuenta a una segunda pagina, puede quedar espacio de sobra en
    # ESA misma pagina para Observaciones (tablas 2 y 3), y en ese caso
    # no hay que forzar otro salto (si no, Observaciones salta a una
    # TERCERA pagina sin necesidad, dejando espacio vacio en la
    # segunda). Se usan las alturas REALES de cada fila (en puntos) de
    # la hoja VISIBLE (ESTADO DE CUENTA -- la hoja PYS esta oculta en el
    # PDF final, asi que sus alturas no afectan el diseño impreso).
    def _altura_filas_pts(ws, fila_inicio, fila_fin, defecto=15.0):
        total = 0.0
        for r in range(fila_inicio, fila_fin + 1):
            h = ws.row_dimensions[r].height
            total += h if h is not None else defecto
        return total

    # Alto util real de una pagina carta con los margenes de esta
    # plantilla (792pt de alto - margenes superior/inferior en puntos).
    # Calibrado ademas con un caso limite real: "18 declaraciones + 0
    # observaciones" cabia por muy poco en una sola pagina.
    ALTO_PAGINA_PT = 792.0
    ALTO_UTIL_PT = ALTO_PAGINA_PT - (edc.page_margins.top * 72) - (edc.page_margins.bottom * 72)

    altura_antes_declaraciones = _altura_filas_pts(edc, 1, FILA_DECL_INICIO - 1)
    altura_una_fila_declaracion = edc.row_dimensions[FILA_DECL_INICIO].height or 21.0
    altura_gap_y_titulo_obs = _altura_filas_pts(edc, FILA_DECL_FIN + 1, 52)  # filas 48..52: espacio + titulo "OBSERVACIONES" + encabezados de columna (SIN incluir las filas de datos 53+, esas se cuentan aparte con max_obs)
    altura_una_fila_obs = edc.row_dimensions[53].height or 21.0
    altura_tabla4 = _altura_filas_pts(edc, 77, 79)  # Periodo/Valor/Fecha/Firma

    altura_declaraciones_total = filas_declaraciones * altura_una_fila_declaracion
    # La PAGINA 1 tiene MENOS espacio disponible para Declaraciones que
    # las paginas siguientes, porque el encabezado ("INFORMACION
    # GENERAL", etc.) ocupa espacio ahi y NO se repite en la pagina 2 en
    # adelante -- por eso no se puede usar un modulo simple sobre el
    # total acumulado desde el principio (eso asumia que todas las
    # paginas tienen la misma capacidad, lo cual fallaba cuando
    # Declaraciones era tan larga que se desbordaba mas alla de la
    # pagina 2).
    capacidad_pagina1_para_declaraciones = ALTO_UTIL_PT - altura_antes_declaraciones
    if altura_declaraciones_total <= capacidad_pagina1_para_declaraciones:
        # Todas las declaraciones caben en la pagina 1 -- el espacio
        # libre que queda ahi es el resto despues del encabezado + tabla.
        espacio_libre_pagina_actual = ALTO_UTIL_PT - altura_antes_declaraciones - altura_declaraciones_total
    else:
        # Se desbordo mas alla de la pagina 1 -- de ahi en adelante cada
        # pagina nueva arranca con capacidad COMPLETA (sin el encabezado
        # de la pagina 1), asi que se calcula cuanto sobra despues de
        # llenar paginas completas con el resto.
        resto_tras_pagina1 = altura_declaraciones_total - capacidad_pagina1_para_declaraciones
        resto_en_pagina_actual = resto_tras_pagina1 % ALTO_UTIL_PT
        espacio_libre_pagina_actual = ALTO_UTIL_PT - resto_en_pagina_actual

    altura_necesaria_obs = altura_gap_y_titulo_obs + (max_obs * altura_una_fila_obs) + altura_tabla4
    if altura_necesaria_obs > espacio_libre_pagina_actual:
        edc.row_breaks.append(Break(id=50))  # quiebre despues de la fila 50 -> "OBSERVACIONES" (fila 51, con su titulo) arranca en pagina nueva

    hojas_a_conservar = {"PYS", "ESTADO DE CUENTA"}
    for nombre in list(wb.sheetnames):
        if nombre not in hojas_a_conservar:
            del wb[nombre]
    edc.sheet_state = "visible"
    # PYS es una hoja de "datos de fondo" (solo alimenta formulas para
    # ESTADO DE CUENTA) -- si queda visible, LibreOffice la imprime como
    # paginas propias ademas de las de ESTADO DE CUENTA (confirmado: en
    # un libro con muchas hojas/formulas, esto por si solo agregaba
    # hasta 15 paginas de mas). Ocultarla no afecta ninguna formula (las
    # hojas ocultas se siguen calculando igual), solo evita que se
    # impriman sus propias paginas.
    pys.sheet_state = "hidden"
    wb.active = wb.sheetnames.index("ESTADO DE CUENTA")

    # El texto se veia "gris suave" en vez de negro solido -- mismo
    # problema de color de TEMA que ya resolvimos en la Declaracion
    # Manual (LibreOffice a veces interpreta mal el indice del tema al
    # convertir). Se fuerza negro explicito en todas las celdas con texto.
    for hoja in (pys, edc):
        for fila_celdas in hoja.iter_rows():
            for celda in fila_celdas:
                if celda.value is not None:
                    nueva_fuente = copy.copy(celda.font)
                    nueva_fuente.color = "FF000000"
                    celda.font = nueva_fuente

    id_temp = str(uuid.uuid4())[:8]
    ruta_xlsm_temp = f"/tmp/_edc_{id_temp}.xlsm"
    wb.save(ruta_xlsm_temp)

    subprocess.run([
        "soffice", "--headless", "--convert-to", "pdf",
        "--outdir", os.path.dirname(ruta_salida_pdf), ruta_xlsm_temp
    ], check=True, timeout=90)

    generado = os.path.join(os.path.dirname(ruta_salida_pdf), f"_edc_{id_temp}.pdf")
    shutil.move(generado, ruta_salida_pdf)
    os.remove(ruta_xlsm_temp)


def bloquear_recursos(page):
    page.route("**/*", lambda route: route.abort()
               if route.request.resource_type in ["image", "stylesheet", "font", "media", "other"]
               else route.continue_())


def resolver_recaptcha_2captcha(site_key, page_url, intentos=3):
    ultimo_error = None
    for intento in range(intentos):
        try:
            resp = requests.post("https://2captcha.com/in.php", data={
                "key": TWOCAPTCHA_API_KEY, "method": "userrecaptcha",
                "googlekey": site_key, "pageurl": page_url, "json": 1,
            }, timeout=15)
            data = resp.json()
            if data.get("status") != 1:
                raise Exception(f"2captcha error: {data.get('request')}")
            captcha_id = data["request"]
            for _ in range(24):
                time.sleep(5)
                resultado = requests.get("https://2captcha.com/res.php", params={
                    "key": TWOCAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1,
                }, timeout=10).json()
                if resultado.get("status") == 1:
                    return resultado["request"]
                if resultado.get("request") not in ("CAPCHA_NOT_READY", "CAPTCHA_NOT_READY"):
                    raise Exception(f"2captcha error: {resultado.get('request')}")
            raise Exception("2captcha tardo demasiado.")
        except Exception as e:
            ultimo_error = e
            if "IP_BANNED" in str(e) and intento < intentos - 1:
                time.sleep(3)
                continue
            raise
    raise ultimo_error


def resolver_turnstile_2captcha(site_key, page_url, intentos=3):
    ultimo_error = None
    for intento in range(intentos):
        try:
            resp = requests.post("https://2captcha.com/in.php", data={
                "key": TWOCAPTCHA_API_KEY, "method": "turnstile",
                "sitekey": site_key, "pageurl": page_url, "json": 1,
            }, timeout=15)
            data = resp.json()
            if data.get("status") != 1:
                raise Exception(f"2captcha error: {data.get('request')}")
            captcha_id = data["request"]
            for _ in range(24):
                time.sleep(5)
                resultado = requests.get("https://2captcha.com/res.php", params={
                    "key": TWOCAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1,
                }, timeout=10).json()
                if resultado.get("status") == 1:
                    return resultado["request"]
                if resultado.get("request") not in ("CAPCHA_NOT_READY", "CAPTCHA_NOT_READY"):
                    raise Exception(f"2captcha error: {resultado.get('request')}")
            raise Exception("2captcha tardo demasiado.")
        except Exception as e:
            ultimo_error = e
            if "IP_BANNED" in str(e) and intento < intentos - 1:
                time.sleep(3)
                continue
            raise
    raise ultimo_error


# ============================================================
#  MUNICIPIOS (sin tocar)
# ============================================================
def consultar_envigado(page, placa):
    url = "https://movilidad.envigado.gov.co/portal-servicios/#/impuesto-local"
    page.goto(url, wait_until="domcontentloaded")
    page.get_by_role("textbox", name="Placa").fill(placa)
    page.get_by_role("button", name="Buscar").click()
    page.wait_for_function("""() => {
        const texto = document.body.innerText;
        const tabla = document.querySelector('#tablaCollapseVigencias');
        const noMatriculado = texto.includes('El vehiculo no se encuentra matriculado en la Secretaria de Movilidad');
        // Paz y salvo: esperar que la tabla tenga al menos una fila con dato real
        const pazYSalvoHeader = texto.includes('Último pago realizado');
        const pazYSalvoConDatos = pazYSalvoHeader && document.querySelectorAll('table tr td').length >= 3;
        return tabla || noMatriculado || pazYSalvoConDatos;
    }""", timeout=TIMEOUT)
    if page.get_by_text(MSG_NO_MATRICULADO).is_visible():
        return [], 0

    # Espera extra para que Angular termine de renderizar
    page.wait_for_timeout(1500)
    texto_pagina = page.inner_text("body")

    # Verificar en el DOM real si existe Y ES VISIBLE la tabla de vigencias
    # pendientes. La tabla #tablaCollapseVigencias SIEMPRE existe en el DOM
    # (Angular la renderiza vacía), solo se oculta con ng-hide en el div
    # contenedor cuando no hay deuda — por eso hay que chequear visibilidad
    # real (is_visible), no solo presencia (.count() > 0).
    try:
        tiene_vigencias_pendientes = page.locator("#tablaCollapseVigencias").is_visible()
    except Exception:
        tiene_vigencias_pendientes = False

    # Paz y salvo — extraer datos de la tabla #tablaUltimosPagos
    if 'Último pago realizado' in texto_pagina and not tiene_vigencias_pendientes:
        try:
            page.wait_for_selector("#tablaUltimosPagos tbody tr td", timeout=5000)
        except Exception:
            pass
        fecha_pago = ""
        marca_veh  = ""
        placa_veh  = ""
        try:
            fila = page.locator("#tablaUltimosPagos tbody tr").first
            placa_veh  = (fila.locator("td[data-label='Placa']").inner_text() or "").strip()
            marca_veh  = (fila.locator("td[data-label='Marca']").inner_text() or "").strip()
            fecha_pago = (fila.locator("td[data-label='Fecha pago']").inner_text() or "").strip()
        except Exception:
            pass
        return [{
            "vigencia":       "PAZ Y SALVO",
            "estado":         f"Vehículo a paz y salvo en el Tránsito de Envigado. Último pago: {fecha_pago}".strip(". "),
            "total_vigencia": 0,
            "paz_y_salvo":    True,
            "fecha_pago":     fecha_pago,
            "marca":          marca_veh,
            "placa_info":     placa_veh,
        }], 0

    if page.locator("#selectall").is_visible():
        page.locator("#selectall").check()

    # Extraer datos de último pago aunque haya deuda (verificación anti-falso-positivo)
    placa_ult = ""; marca_ult = ""; fecha_ult = ""; valor_ult = ""
    try:
        fila_ult = page.locator("#tablaUltimosPagos tbody tr").first
        placa_ult  = (fila_ult.locator("td[data-label='Placa']").inner_text() or "").strip()
        marca_ult  = (fila_ult.locator("td[data-label='Marca']").inner_text() or "").strip()
        fecha_ult  = (fila_ult.locator("td[data-label='Fecha pago']").inner_text() or "").strip()
        valor_ult  = (fila_ult.locator("td[data-label='Valor pago']").inner_text() or "").strip()
    except Exception:
        pass

    registros = []
    filas = page.locator("#tablaCollapseVigencias tr").all()
    for fila in filas:
        texto_fila = fila.inner_text().strip()
        if not texto_fila:
            continue
        año = re.search(r'\b(20\d{2})\b', texto_fila)
        montos = re.findall(r'\$\s*[\d.]+', texto_fila)
        if año and montos:
            valor_str = montos[-1].replace('$', '').replace(' ', '').replace('.', '')
            try:
                registros.append({
                    'vigencia': año.group(), 'estado': 'Pendiente de pago',
                    'total_vigencia': int(valor_str),
                    'placa_ultimo_pago': placa_ult,
                    'marca_ultimo_pago': marca_ult,
                    'fecha_ultimo_pago': fecha_ult,
                    'valor_ultimo_pago': valor_ult,
                })
            except ValueError:
                pass
    total = sum(r['total_vigencia'] for r in registros)
    return registros, total


def consultar_sabaneta(page, placa):
    url = "https://transitosabaneta.utsetsa.com/#/impuesto-local"
    page.goto(url, wait_until="domcontentloaded", timeout=30000)
    page.locator("#placa").wait_for(state="visible", timeout=15000)
    page.locator("#placa").fill(placa)
    page.get_by_role("button", name="Buscar").click()
    page.wait_for_timeout(20000)
    texto_pagina = page.inner_text("body")
    html_pagina  = page.content()
    if MSG_NO_MATRICULADO in texto_pagina:
        return [], 0
    if 'Último pago realizado' in texto_pagina and 'Vigencias pendientes' not in texto_pagina:
        placa_sab = ""; marca_sab = ""; fecha_sab = ""; valor_sab = ""
        try:
            fila = page.locator("#tablaUltimosPagos tbody tr").first
            celdas = fila.locator("td").all()
            texts = [c.inner_text().strip() for c in celdas]
            # Orden: Placa, Marca, Fecha pago, Valor pago
            if len(texts) > 0: placa_sab = texts[0]
            if len(texts) > 1: marca_sab = texts[1]
            if len(texts) > 2: fecha_sab = texts[2]
            if len(texts) > 3: valor_sab = texts[3]
        except Exception:
            pass
        return [{
            "vigencia":       "PAZ Y SALVO",
            "estado":         "Vehículo a paz y salvo en el Tránsito de Sabaneta.",
            "total_vigencia": 0,
            "paz_y_salvo":    True,
            "placa_info":     placa_sab,
            "marca":          marca_sab,
            "fecha_pago":     fecha_sab,
            "valor_pago":     valor_sab,
        }], 0
    if 'Vigencias pendientes' not in texto_pagina:
        return [], 0
    page.locator("#tablaCollapseVigencias").wait_for(state="visible", timeout=15000)
    checkbox = page.locator("#selectall")
    checkbox.wait_for(state="visible", timeout=15000)
    if checkbox.is_enabled():
        checkbox.check()
    page.wait_for_timeout(5000)
    spans_cop = page.locator("span.fs-16.ng-binding").all()
    total = 0
    for span in spans_cop[::-1]:
        texto = span.inner_text().strip()
        if "COP" in texto and texto != "COP 0":
            valor_str = texto.replace("COP", "").replace(".", "").strip()
            try:
                total = int(valor_str)
                break
            except ValueError:
                pass
    # Extraer datos de último pago aunque haya deuda (sirven para verificar
    # que el scraper realmente consultó ESTE vehículo y no un falso positivo)
    placa_ult = ""; marca_ult = ""; fecha_ult = ""; valor_ult = ""
    try:
        fila_ult = page.locator("#tablaUltimosPagos tbody tr").first
        celdas_ult = fila_ult.locator("td").all()
        texts_ult = [c.inner_text().strip() for c in celdas_ult]
        if len(texts_ult) > 0: placa_ult = texts_ult[0]
        if len(texts_ult) > 1: marca_ult = texts_ult[1]
        if len(texts_ult) > 2: fecha_ult = texts_ult[2]
        if len(texts_ult) > 3: valor_ult = texts_ult[3]
    except Exception:
        pass

    registros = []
    filas = page.locator("#tablaCollapseVigencias tr").all()
    for fila in filas:
        texto_fila = fila.inner_text().strip()
        if not texto_fila:
            continue
        año = re.search(r'\b(20\d{2})\b', texto_fila)
        montos = re.findall(r'COP\s*[\d.]+', texto_fila)
        if año and montos:
            valor_fila = montos[-1].replace('COP', '').replace(' ', '').replace('.', '')
            try:
                registros.append({
                    'vigencia': año.group(), 'estado': 'Pendiente de pago',
                    'total_vigencia': int(valor_fila),
                    'placa_ultimo_pago': placa_ult,
                    'marca_ultimo_pago': marca_ult,
                    'fecha_ultimo_pago': fecha_ult,
                    'valor_ultimo_pago': valor_ult,
                })
            except ValueError:
                pass
    return registros, total


def consultar_itagui(page, placa):
    url = "https://movilidad.transitoitagui.gov.co/portal-servicios/#/impuesto-local"
    page.goto(url, wait_until="domcontentloaded")
    page.get_by_role("textbox", name="Placa").fill(placa)
    page.get_by_role("button", name="Buscar").click()
    page.wait_for_function("""() => {
        const texto = document.body.innerText;
        const noMatriculado = texto.includes('El vehiculo no se encuentra matriculado en la Secretaria de Movilidad');
        const conDeuda = texto.includes('Vigencias pendientes');
        const pazYSalvo = texto.includes('Último pago realizado') && !texto.includes('Vigencias pendientes');
        return noMatriculado || conDeuda || pazYSalvo;
    }""", timeout=20000)
    texto_pagina = page.inner_text("body")
    if MSG_NO_MATRICULADO in texto_pagina:
        return [], 0

    # Paz y salvo — extraer datos de verificación (placa/marca/fecha) igual que Envigado
    if 'Vigencias pendientes' not in texto_pagina and AÑO_ACTUAL in texto_pagina:
        try:
            page.wait_for_selector("#tablaUltimosPagos tbody tr td", timeout=5000)
        except Exception:
            pass
        placa_veh = ""; marca_veh = ""; fecha_pago = ""
        try:
            fila = page.locator("#tablaUltimosPagos tbody tr").first
            placa_veh  = (fila.locator("td[data-label='Placa']").inner_text() or "").strip()
            marca_veh  = (fila.locator("td[data-label='Marca']").inner_text() or "").strip()
            fecha_pago = (fila.locator("td[data-label='Fecha pago']").inner_text() or "").strip()
        except Exception:
            pass
        return [{
            "vigencia":       "PAZ Y SALVO",
            "estado":         f"Vehículo a paz y salvo en el Tránsito de Itagüí. Último pago: {fecha_pago}".strip(". "),
            "total_vigencia": 0,
            "paz_y_salvo":    True,
            "fecha_pago":     fecha_pago,
            "marca":          marca_veh,
            "placa_info":     placa_veh,
        }], 0

    page.locator("#tablaCollapseVigencias").wait_for(state="visible", timeout=15000)
    checkbox = page.locator("#selectall")
    checkbox.wait_for(state="visible", timeout=15000)
    if checkbox.is_enabled():
        checkbox.check()
    page.wait_for_timeout(3000)
    spans_cop = page.locator("span.fs-16.ng-binding").all()
    total = 0
    for span in spans_cop[::-1]:
        texto = span.inner_text().strip()
        if "COP" in texto and texto != "COP 0":
            valor_str = texto.replace("COP", "").replace(".", "").strip()
            try:
                total = int(valor_str)
                break
            except ValueError:
                pass

    # Extraer datos de último pago aunque haya deuda (verificación anti-falso-positivo)
    placa_ult = ""; marca_ult = ""; fecha_ult = ""; valor_ult = ""
    try:
        fila_ult = page.locator("#tablaUltimosPagos tbody tr").first
        placa_ult = (fila_ult.locator("td[data-label='Placa']").inner_text() or "").strip()
        marca_ult = (fila_ult.locator("td[data-label='Marca']").inner_text() or "").strip()
        fecha_ult = (fila_ult.locator("td[data-label='Fecha pago']").inner_text() or "").strip()
        valor_ult = (fila_ult.locator("td[data-label='Valor pago']").inner_text() or "").strip()
    except Exception:
        pass

    registros = []
    filas = page.locator("#tablaCollapseVigencias tr").all()
    for fila in filas:
        texto_fila = fila.inner_text().strip()
        if not texto_fila:
            continue
        año = re.search(r'\b(20\d{2})\b', texto_fila)
        montos = re.findall(r'COP\s*[\d.]+', texto_fila)
        if año and montos:
            valor_fila = montos[-1].replace('COP', '').replace(' ', '').replace('.', '')
            try:
                registros.append({
                    'vigencia': año.group(), 'estado': 'Pendiente de pago',
                    'total_vigencia': int(valor_fila),
                    'placa_ultimo_pago': placa_ult,
                    'marca_ultimo_pago': marca_ult,
                    'fecha_ultimo_pago': fecha_ult,
                    'valor_ultimo_pago': valor_ult,
                })
            except ValueError:
                pass
    return registros, total


def consultar_bello(page, placa):
    url = "https://serviciosdigitales.movilidadavanzadabello.com.co/portal-servicios/#/public"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_function("""() => { return document.querySelectorAll('input[type="search"]').length > 0; }""", timeout=30000)
    try:
        page.get_by_role("button", name="Close").click(timeout=5000)
    except:
        pass
    page.get_by_role("searchbox", name="Placa").nth(3).fill(placa)
    page.get_by_role("button").nth(5).click()
    try:
        page.wait_for_url("**/impuesto-local", timeout=15000)
    except:
        return [], 0
    page.wait_for_timeout(10000)
    texto_pagina = page.inner_text("body")

    def _extraer_verificacion():
        """Intenta extraer placa/marca/fecha/valor del último pago para verificar
        que el sistema consultó el vehículo real (anti falso-positivo)."""
        placa_v = ""; marca_v = ""; fecha_v = ""; valor_v = ""
        try:
            fila = page.locator("#tablaUltimosPagos tbody tr").first
            if fila.count() > 0:
                placa_v = (fila.locator("td[data-label='Placa']").inner_text() or "").strip()
                marca_v = (fila.locator("td[data-label='Marca']").inner_text() or "").strip()
                fecha_v = (fila.locator("td[data-label='Fecha pago']").inner_text() or "").strip()
                valor_v = (fila.locator("td[data-label='Valor pago (COP)']").inner_text() or "").strip()
        except Exception:
            pass
        return placa_v, marca_v, fecha_v, valor_v

    if 'paz y salvo' in texto_pagina or 'No se encontraron registros' in texto_pagina:
        placa_v, marca_v, fecha_v, _ = _extraer_verificacion()
        if placa_v or marca_v or fecha_v:
            return [{
                "vigencia":       "PAZ Y SALVO",
                "estado":         f"Vehículo a paz y salvo en el Tránsito de Bello. Último pago: {fecha_v}".strip(". "),
                "total_vigencia": 0,
                "paz_y_salvo":    True,
                "fecha_pago":     fecha_v,
                "marca":          marca_v,
                "placa_info":     placa_v,
            }], 0
        return [], 0

    # La sección "Vigencias pendientes" se oculta con ng-hide (display:none)
    # cuando NO hay deuda, y Playwright's inner_text() no incluye texto oculto.
    # Por eso hay que extraer la verificación (tabla de últimos pagos, que SÍ
    # es visible siempre) ANTES de decidir si hay o no vigencias pendientes.
    placa_ult, marca_ult, fecha_ult, valor_ult = _extraer_verificacion()

    if 'Vigencias pendientes' not in texto_pagina:
        if placa_ult or marca_ult or fecha_ult:
            return [{
                "vigencia":       "PAZ Y SALVO",
                "estado":         f"Vehículo a paz y salvo en el Tránsito de Bello. Último pago: {fecha_ult}".strip(". "),
                "total_vigencia": 0,
                "paz_y_salvo":    True,
                "fecha_pago":     fecha_ult,
                "marca":          marca_ult,
                "placa_info":     placa_ult,
            }], 0
        return [], 0

    registros = []
    filas_vig = page.locator("#tablaCollapseVigencias tr").all()
    for fila in filas_vig:
        texto = fila.inner_text().strip()
        if not texto:
            continue
        año = re.search(r'\b(20\d{2})\b', texto)
        montos = re.findall(r'COP\s*[\d.]+', texto)
        if año and montos:
            valor_fila = montos[-1].replace('COP', '').replace(' ', '').replace('.', '')
            try:
                registros.append({
                    'vigencia': año.group(), 'estado': 'Pendiente de pago' if 'Pendiente' in texto else 'Desconocido',
                    'total_vigencia': int(valor_fila),
                    'placa_ultimo_pago': placa_ult,
                    'marca_ultimo_pago': marca_ult,
                    'fecha_ultimo_pago': fecha_ult,
                    'valor_ultimo_pago': valor_ult,
                })
            except ValueError:
                pass
    match_total = re.search(r'Total a pagar:\s*COP\s*([\d.]+)', texto_pagina)
    total = int(match_total.group(1).replace('.', '')) if match_total else sum(r['total_vigencia'] for r in registros)

    # Si no hay vigencias reales con deuda (registros vacío / total 0) pero sí
    # se logró extraer placa/marca/fecha del último pago, es un paz y salvo
    # verificado (esto es lo que realmente pasa en Bello 27.1: la página no
    # muestra el texto "paz y salvo", solo "Vigencias pendientes ()" vacío).
    if not registros and total == 0 and (placa_ult or marca_ult or fecha_ult):
        return [{
            "vigencia":       "PAZ Y SALVO",
            "estado":         f"Vehículo a paz y salvo en el Tránsito de Bello. Último pago: {fecha_ult}".strip(". "),
            "total_vigencia": 0,
            "paz_y_salvo":    True,
            "fecha_pago":     fecha_ult,
            "marca":          marca_ult,
            "placa_info":     placa_ult,
        }], 0

    return registros, total


def _parsear_emtrasur(data):
    registros = []
    for r in data:
        registros.append({
            "vigencia": str(r.get("AnioNoFacturado", "")),
            "estado": "Pendiente de pago",
            "total_vigencia": r.get("ValorPorFacturar", 0),
            "tipo_vehiculo": r.get("TipoVehiculo", ""),
            "ultimo_pago": r.get("AnioPagado", ""),
            "descripcion": r.get("DescripcionNoFacturada", "").strip(),
        })
    total = sum(r["total_vigencia"] for r in registros)
    return registros, total


def consultar_laestrella(page, placa):
    token = resolver_recaptcha_2captcha(EMTRASUR_SITE_KEY, EMTRASUR_URL)
    api_url = f"https://sistematizacion.emtrasur.com.co/api/Sistematizacion/{placa}"
    resp = requests.get(api_url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Referer": EMTRASUR_URL,
        "Origin": "https://sistematizacion.emtrasur.com.co",
        "X-Captcha-Token": token,
    }, timeout=15)
    if resp.status_code == 200:
        data = resp.json()
        if data.get("Success"):
            return _parsear_emtrasur(data.get("Data", []))
    raise Exception(f"EMTRASUR respondio {resp.status_code}: {resp.text[:200]}")


# ============================================================
#  ANTIOQUIA — MÓDULO NUEVO
# ============================================================
def _calcular_digito_nit(nit):
    """Calcula el dígito de verificación de un NIT colombiano (algoritmo DIAN)."""
    factores = [3, 7, 13, 17, 19, 23, 29, 37, 41, 43]
    n = str(nit).strip().replace("-", "").replace(".", "").zfill(10)
    suma = sum(int(n[::-1][i]) * factores[i] for i in range(10))
    r = suma % 11
    return 0 if r == 0 else (1 if r == 1 else 11 - r)


def _sesion_antioquia(placa, identificacion, tipo_documento_id,
                      modelo, organismo_transito, apellidos_propietario):
    """
    Abre sesión completa en Antioquia y retorna (session, token_cuestionario, data3).
    Costo: 2 Turnstiles.
    """
    try:
        token_captcha = resolver_turnstile_2captcha(ANTIOQUIA_SITE_KEY, ANTIOQUIA_URL)
    except Exception as e:
        raise Exception(f"Error resolviendo captcha inicial: {e}")

    session = requests.Session()
    session.headers.update({
        "Accept": "*/*",
        "Content-Type": "application/json",
        "captcha": token_captcha,
        "Referer": "https://www.vehiculosantioquia.com.co/impuestosweb/",
        "X-Requested-With": "XMLHttpRequest",
        "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Mobile Safari/537.36",
    })

    r1 = session.post(
        f"{ANTIOQUIA_API}/ConsultarEstadoCuentaImpAntioquia/obtenerCuestionarioEstadoCuenta",
        json={"placa": placa, "idTipoIdentificacion": tipo_documento_id, "identificacion": identificacion},
        timeout=60
    )
    try:
        data1 = r1.json()
    except Exception:
        data1 = None

    if not data1 or not isinstance(data1, dict):
        raise Exception("La placa ingresada no coincide con la identificacion del propietario. Verifica los datos e intenta de nuevo.")

    if data1.get("codigo") == 0 or (not data1.get("referencia") and data1.get("mensaje")):
        mensaje = data1.get("mensaje") or data1.get("descripcion") or "La placa ingresada no coincide con la identificacion del propietario."
        raise Exception(mensaje)

    referencia = data1.get("referencia")
    if not referencia:
        raise Exception("La placa ingresada no coincide con la identificacion del propietario. Verifica los datos e intenta de nuevo.")

    opciones_nombre = (data1.get("preguntaNombrePropietario") or {}).get("opcionesPregunta", [])
    primer_apellido = apellidos_propietario.upper().split()[0] if apellidos_propietario.strip() else ""
    nombre_encontrado = next(
        (n for n in opciones_nombre if primer_apellido in n.upper()), None
    )
    if not nombre_encontrado:
        raise Exception(f"No se encontró propietario con apellidos '{apellidos_propietario}'. Opciones: {opciones_nombre}")

    r2 = session.post(
        f"{ANTIOQUIA_API}/ConsultarEstadoCuentaImpAntioquia/validarCuestionarioEstadoCuenta",
        json={
            "placa": placa, "tipoDocumento": tipo_documento_id, "numeroDocumento": identificacion,
            "idEstadoCuenta": referencia,
            "respuestas": {
                "respuestaModelo": modelo,
                "respuestaOrganismoTransito": organismo_transito,
                "respuestaNombrePropietario": nombre_encontrado
            }
        },
        timeout=60
    )
    validacion = r2.json()
    if validacion.get("codigo") != 1:
        raise Exception(f"Cuestionario inválido: {validacion.get('descripcion')}")

    try:
        token_captcha2 = resolver_turnstile_2captcha(ANTIOQUIA_SITE_KEY, ANTIOQUIA_URL)
    except Exception as e:
        raise Exception(f"Error resolviendo segundo captcha: {e}")
    session.headers.update({"captcha": token_captcha2})

    token_cuestionario = session.cookies.get("token_cuestionario")
    if not token_cuestionario:
        raise Exception("No se pudo obtener el token de sesión.")

    r3 = session.post(
        f"{ANTIOQUIA_API}/ConsultarEstadoCuentaImpAntioquia/consultarEstadoCuentaVehiculoHomePublico",
        json={"placa": placa, "informacionDeclarante": {
            "idsolicitante": identificacion, "idTipoIdentificacion": tipo_documento_id
        }},
        headers={"Cookie": f"token_cuestionario={token_cuestionario}"},
        timeout=60
    )
    return session, token_cuestionario, r3.json()


def _consultar_vigencia_antioquia(vigencia, session, token_cuestionario,
                                   placa, identificacion, tipo_documento_id,
                                   doc_abreviatura, doc_nombre,
                                   celular, email, direccion, municipio, municipio_cod, departamento_cod):
    """
    Consulta el costo de una vigencia específica.
    Costo: 2 Turnstiles adicionales.
    """
    try:
        token_prop = resolver_turnstile_2captcha(ANTIOQUIA_SITE_KEY, ANTIOQUIA_URL)
    except Exception as e:
        raise Exception(f"Error resolviendo captcha vigencia {vigencia}: {e}")
    session.headers.update({"captcha": token_prop})

    r4 = session.post(
        f"{ANTIOQUIA_API}/UsuariosPortalAntioquia/consultarPropietarioVehiculo",
        json={"tipoDoc": doc_abreviatura, "nroDoc": identificacion, "placa": placa, "vigencia": vigencia},
        headers={"Cookie": f"token_cuestionario={token_cuestionario}"},
        timeout=60
    )
    propietario = r4.json().get("propietario", {})

    session.post(f"{ANTIOQUIA_API}/TablasTipo/obtenerTablasPropietario", json={},
                 headers={"Cookie": f"token_cuestionario={token_cuestionario}"}, timeout=60)
    session.get(f"{ANTIOQUIA_API}/UtilImpuestos/obtenerDescripcionPPST",
                headers={"Cookie": f"token_cuestionario={token_cuestionario}"}, timeout=60)
    session.post(f"{ANTIOQUIA_API}/Pagos/parametrosPago", json={},
                 headers={"Cookie": f"token_cuestionario={token_cuestionario}"}, timeout=60)
    session.get(f"{ANTIOQUIA_API}/UtilImpuestos/obtenerVigenciaMinimaAutodeclarar",
                headers={"Cookie": f"token_cuestionario={token_cuestionario}"}, timeout=60)

    try:
        token_decl = resolver_turnstile_2captcha(ANTIOQUIA_SITE_KEY, ANTIOQUIA_URL)
    except Exception as e:
        raise Exception(f"Error resolviendo captcha declaración vigencia {vigencia}: {e}")
    session.headers.update({"captcha": token_decl})
    session.cookies.clear()

    es_nit = (str(tipo_documento_id) == "2")
    if es_nit:
        declarante = {
            "idsolicitante": identificacion,
            "idtipodocumento": doc_abreviatura,
            "desctipodocument": doc_nombre,
            "nombres": propietario.get("nameOrg1", ""),
            "apellidos": "",
            "celular": celular,
            "telefono": propietario.get("celphone", celular),
            "email": email,
            "direccion": direccion, "municipio": municipio,
            "departamento": "ANTIOQUIA", "nivreclamacion": 0, "procedimiento": ""
        }
    else:
        declarante = {
            "idsolicitante": identificacion,
            "idtipodocumento": doc_abreviatura,
            "desctipodocument": doc_nombre,
            "nombres": propietario.get("nameFirst", ""),
            "apellidos": propietario.get("nameLast", ""),
            "celular": celular, "telefono": celular, "email": email,
            "direccion": direccion, "municipio": municipio,
            "departamento": "ANTIOQUIA", "nivreclamacion": 0, "procedimiento": ""
        }

    r5 = session.post(
        f"{ANTIOQUIA_API}/LiquidacionAntioquia/crearDeclaracionImpuestoAnt",
        json={
            "formularioLiquidacion": "",
            "declarante": declarante,
            "iIdliqIm": 0,
            "informacionComplementaria": {
                "idTipoDocumento": int(tipo_documento_id),
                "distribucionDepartamento": departamento_cod,
                "distribucionMunicipio": municipio_cod,
                "direccionCompleta": direccion,
                "nombreDistribucionDepartamento": "ANTIOQUIA",
                "nombreDistribucionMunicipio": municipio,
                "tipoCanalLiquidacion": 2, "tipoOpcionLiquidacion": 1
            },
            "placa": placa,
            "vigencia": [{"persl": vigencia}]
        },
        timeout=60
    )
    return r5.json()


def _antioquia_construir_documento(tipo_documento_id, doc_abreviatura, doc_nombre):
    """Arma el objeto 'documento' que pide el endpoint de aceptacion de
    terminos, con la misma forma que se ve en el request real capturado."""
    es_nit = (str(tipo_documento_id) == "2")
    return {
        "idDocumentoIdentidad": int(tipo_documento_id),
        "tipoPersona": "J" if es_nit else "N",
        "abreviatura": doc_abreviatura,
        "nombreDocumento": doc_nombre,
    }


def _antioquia_aceptar_terminos_liquidacion(session, identificacion, tipo_documento_id,
                                              doc_abreviatura, doc_nombre):
    """Acepta las 3 casillas (tratamiento de datos, terminos y condiciones,
    firma digital) que en el sitio real hay que marcar antes de que se
    habilite el boton de imprimir/descargar la declaracion sugerida."""
    body = {
        "numeroDocumento": str(identificacion),
        "documento": _antioquia_construir_documento(tipo_documento_id, doc_abreviatura, doc_nombre),
    }
    r = session.post(
        f"{ANTIOQUIA_API}/AceptacionTerminoCondiciones/insertAceptaTerminosLiquidacion",
        json=body, timeout=60
    )
    if r.status_code not in (200, 204):
        raise Exception(f"Error aceptando terminos de liquidacion: {r.status_code} {r.text[:300]}")


def _antioquia_descargar_pdf_liquidacion(session, formulario_liquidacion):
    """Pide el PDF de la declaracion sugerida ya generada. El servidor lo
    devuelve codificado en base64 dentro de un campo 'archivo'."""
    r = session.post(
        f"{ANTIOQUIA_API}/LiquidacionAntioquia/gestionarImprimirLiquidacion",
        json=formulario_liquidacion, timeout=60
    )
    if r.status_code != 200:
        raise Exception(f"Error descargando PDF de liquidacion: {r.status_code} {r.text[:300]}")
    data = r.json()
    archivo_b64 = data.get("archivo")
    if not archivo_b64:
        raise Exception(f"La respuesta no trajo el campo 'archivo': {json.dumps(data)[:300]}")
    return base64.b64decode(archivo_b64)


def _extraer_nombre_apellidos_declaracion(pdf_bytes):
    """Extrae el NOMBRE (C.1) y los APELLIDOS (C.3) directamente del texto
    del PDF de la Declaracion Sugerida -- estos son los datos OFICIALES
    que la Gobernacion tiene registrados para el propietario (confirmado
    con la ficha de seguridad que responde el propio sistema de la
    Gobernacion), asi que son mas confiables que cualquier nombre que el
    usuario haya escrito a mano. Se usan para que la Declaracion Manual
    quede con el mismo nombre exacto que la Declaracion Sugerida.

    El orden de extraccion de pypdf para este PDF en particular pone el
    valor de "C.1 NOMBRE..." ANTES de su propia etiqueta, y el valor de
    "C.3 APELLIDOS" DESPUES de su etiqueta (confirmado con un PDF real) --
    por eso se buscan con patrones distintos para cada uno.
    Si no los encuentra (formato distinto), se devuelven vacios -- el
    llamador debe usar como respaldo lo que el usuario haya escrito."""
    try:
        import io, re
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        texto = ""
        for pagina in reader.pages:
            texto += (pagina.extract_text() or "") + "\n"

        match_nombres = re.search(r'^([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]*)$\nC\.1\s+NOMBRE', texto, re.MULTILINE)
        nombres = match_nombres.group(1).strip() if match_nombres else ""

        match_apellidos = re.search(r'C\.3\s+APELLIDOS\s*\n?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]*?)\s+\d', texto)
        apellidos = match_apellidos.group(1).strip() if match_apellidos else ""

        return {"nombres": nombres, "apellidos": apellidos}
    except Exception as e:
        print(f"No se pudo extraer nombre/apellidos del PDF de declaracion: {e}", flush=True)
    return {"nombres": "", "apellidos": ""}


def _extraer_caja_traccion_declaracion(pdf_bytes):
    """Extrae 'Caja' (transmision) y 'Traccion' directamente del texto del
    PDF de la Declaracion Sugerida -- la Gobernacion los incluye ahi
    aunque no vengan en la respuesta JSON de la consulta (crearDeclaracion
    ImpuestoAnt), asi que en vez de perseguirlos por la API los leemos del
    mismo documento que ya generamos.

    IMPORTANTE: el orden en que aparecen las etiquetas D.21/D.22 varia
    segun la plantilla (a veces "D.21 CAJA ... D.22 TRACCION", otras veces
    al reves, y hasta el nombre de la etiqueta cambia). Por eso, en vez de
    depender de la posicion, se buscan directamente los VALORES conocidos
    (MT/AT/CVT para caja, 4X2/4X4 para traccion) dentro de una ventana de
    texto alrededor de esas etiquetas.

    Si no los encuentra (formato distinto, o el vehiculo no tiene esos
    datos), se devuelven vacios -- el instructivo oficial permite dejarlos
    en blanco de todas formas."""
    try:
        import io, re
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        texto = ""
        for pagina in reader.pages:
            texto += (pagina.extract_text() or "") + "\n"
        texto_norm = texto.upper()

        idx21 = texto_norm.find("D.21")
        idx22 = texto_norm.find("D.22")
        indices = [i for i in (idx21, idx22) if i != -1]
        ventana = texto_norm[max(0, min(indices) - 60): min(indices) + 100] if indices else texto_norm

        match_traccion = re.search(r'\b(4X[24])', ventana)
        match_caja = re.search(r'\b(MT|AT|CVT|TM|TA)\b', ventana)

        return {
            "traccion": match_traccion.group(1) if match_traccion else "",
            "caja": match_caja.group(1) if match_caja else ""
        }
    except Exception as e:
        print(f"No se pudo extraer caja/traccion del PDF de declaracion: {e}", flush=True)
    return {"traccion": "", "caja": ""}


def antioquia_generar_pdf_declaracion(placa, identificacion, tipo_documento_abrev, vigencia,
                                       modelo, municipio_transito, apellidos_propietario,
                                       celular="3000000000", email="consulta@consulta.com",
                                       direccion="CRA", municipio="MEDELLIN",
                                       municipio_cod=5001000, departamento_cod=5):
    """Flujo completo para obtener el PDF de la declaracion sugerida (el que
    se lleva al banco), TODO dentro de la misma sesion (el numero de
    liquidacion que genera el sitio solo es valido dentro de la sesion que
    lo creo, no se puede reutilizar en una sesion nueva).
    Devuelve una tupla (pdf_bytes, data_vig) -- data_vig es la respuesta
    cruda de la liquidacion (avaluo, impuesto, sanciones, intereses, etc.),
    util para reutilizar esos mismos datos en la declaracion manual sin
    tener que volver a consultar."""
    tipo_documento_id = ANTIOQUIA_TIPO_DOC_MAP.get(tipo_documento_abrev.upper(), "1")
    tipo_doc_info = ANTIOQUIA_TIPOS_DOCUMENTO.get(tipo_documento_id, ANTIOQUIA_TIPOS_DOCUMENTO["1"])
    doc_abreviatura = tipo_doc_info["abreviatura"]
    doc_nombre = tipo_doc_info["nombre"]

    if tipo_documento_id == "2":
        identificacion = str(identificacion) + str(_calcular_digito_nit(identificacion))

    # 1. Nueva sesion (igual que cualquier consulta normal)
    session, token_cuestionario, _data3 = _sesion_antioquia(
        placa, identificacion, tipo_documento_id,
        modelo, municipio_transito, apellidos_propietario
    )

    # 2. Crear la declaracion sugerida para la vigencia pedida (misma sesion)
    data_vig = _consultar_vigencia_antioquia(
        vigencia, session, token_cuestionario,
        placa, identificacion, tipo_documento_id,
        doc_abreviatura, doc_nombre,
        celular, email, direccion, municipio, municipio_cod, departamento_cod
    )
    formulario_liquidacion = data_vig.get("formularioLiquidacion")
    if not formulario_liquidacion:
        raise Exception(f"No se pudo generar la declaracion: {json.dumps(data_vig, ensure_ascii=False)[:300]}")

    # 3. Aceptar las 3 casillas (misma sesion)
    _antioquia_aceptar_terminos_liquidacion(session, identificacion, tipo_documento_id,
                                              doc_abreviatura, doc_nombre)

    # 4. Descargar el PDF ya generado (misma sesion)
    # El sitio espera este valor como NUMERO en el JSON, no como texto
    # entre comillas -- por eso se convierte antes de enviarlo.
    pdf_bytes = _antioquia_descargar_pdf_liquidacion(session, int(formulario_liquidacion))
    return pdf_bytes, data_vig


def _antioquia_calcular_validez_pdf(vigencia):
    """La declaracion sugerida de una vigencia VENCIDA solo sirve el mismo
    dia en que se genero (los intereses suben a diario). La de la vigencia
    ACTUAL (el año en curso) es distinta: sirve durante toda la ventana de
    pronto pago -- del 1 de enero al 30 de abril, o del 1 de mayo al 31 de
    julio -- y desde el 1 de agosto en adelante se comporta igual que una
    vigencia vencida (valida solo el mismo dia).
    OJO: el servidor corre en UTC, pero "el mismo dia" debe ser el dia en
    Colombia -- sin este ajuste, de 7pm a 12am hora Colombia el servidor
    ya cree que es el dia siguiente (UTC va 5 horas adelante)."""
    hoy = (datetime.utcnow() - timedelta(hours=5)).date()  # Colombia = UTC-5 todo el año, sin horario de verano
    anio_actual = hoy.year

    if int(vigencia) == anio_actual:
        if hoy.month <= 4:
            return date(anio_actual, 4, 30)
        elif hoy.month <= 7:
            return date(anio_actual, 7, 31)
    return hoy  # vigencias vencidas, o vigencia actual desde agosto en adelante


def _limpiar_declaraciones_vencidas():
    """Borra de R2 y de la base de datos los PDFs de declaraciones que ya
    vencieron (valido_hasta < hoy en Colombia) -- las de una vigencia
    vencida solo sirven el mismo dia en que se generaron, asi que no
    tiene sentido dejarlas guardadas para siempre. Se llama de forma
    'perezosa' (cada vez que se hace una consulta de Impuesto
    Departamental), no con una tarea programada aparte."""
    try:
        hoy_colombia = (datetime.utcnow() - timedelta(hours=5)).date()
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, url, url_manual FROM cache_declaraciones_antioquia
            WHERE valido_hasta < %s
        """, (hoy_colombia,))
        vencidas = cur.fetchall()
        if not vencidas:
            cur.close(); conn.close()
            return
        for fila_id, url, url_manual in vencidas:
            if url:
                borrar_de_r2(url)
            if url_manual:
                borrar_de_r2(url_manual)
        cur.execute("DELETE FROM cache_declaraciones_antioquia WHERE valido_hasta < %s", (hoy_colombia,))
        conn.commit()
        cur.close(); conn.close()
        print(f"Limpieza de declaraciones vencidas: se borraron {len(vencidas)} registro(s).", flush=True)
    except Exception as e:
        print(f"Error limpiando declaraciones vencidas: {e}")


def _cache_declaracion_buscar(placa, vigencia):
    """Busca lo que ya se haya generado hoy (o dentro de su ventana de
    validez) para esta placa/vigencia. Devuelve un dict {url, url_manual,
    datos} si existe (url/url_manual/datos pueden venir en None si nunca
    se genero ese PDF en particular), o None si no hay nada cacheado."""
    try:
        # OJO: no se usa CURRENT_DATE de Postgres porque el servidor de
        # base de datos corre en su propia zona horaria (UTC), distinta
        # a la de Colombia -- eso causaba que, en la noche (cuando UTC ya
        # va en el dia siguiente), el cache pareciera "vencido" aunque
        # todavia fuera valido segun la hora real de Colombia. Se manda
        # la fecha de Colombia calculada en Python, para que ambos lados
        # de la comparacion usen la MISMA referencia de "hoy".
        hoy_colombia = (datetime.utcnow() - timedelta(hours=5)).date()
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT url, datos_json, url_manual FROM cache_declaraciones_antioquia
            WHERE placa = %s AND vigencia = %s AND valido_hasta >= %s
        """, (placa.upper(), int(vigencia), hoy_colombia))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            return None
        return {"url": row[0], "datos": row[1], "url_manual": row[2]}
    except Exception as e:
        print(f"Error buscando cache de declaracion: {e}")
        return None


def _cache_declaracion_guardar(placa, vigencia, url, datos_extra=None, url_manual=None):
    """Guarda el PDF generado en cache. 'datos_extra', si se da, guarda
    ademas la liquidacion completa (avaluo, impuesto, sanciones, etc.) y
    caja/traccion, para que otras herramientas (como la Declaracion
    Manual) puedan reutilizarlos sin tener que consultar de nuevo en
    vivo a la Gobernacion. 'url_manual', si se da, guarda la URL del PDF
    de la Declaracion Manual ya generado (para poder saltarla si se
    vuelve a pedir la misma vigencia el mismo dia)."""
    try:
        valido_hasta = _antioquia_calcular_validez_pdf(vigencia)
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cache_declaraciones_antioquia (placa, vigencia, url, valido_hasta, datos_json, url_manual)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (placa, vigencia) DO UPDATE SET
                url = COALESCE(EXCLUDED.url, cache_declaraciones_antioquia.url),
                valido_hasta = EXCLUDED.valido_hasta,
                datos_json = COALESCE(EXCLUDED.datos_json, cache_declaraciones_antioquia.datos_json),
                url_manual = COALESCE(EXCLUDED.url_manual, cache_declaraciones_antioquia.url_manual),
                creado_en = NOW()
        """, (placa.upper(), int(vigencia), url, valido_hasta,
              json.dumps(datos_extra) if datos_extra else None, url_manual))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error guardando cache de declaracion: {e}")


def antioquia_generar_todas_declaraciones(placa, identificacion, tipo_documento_abrev, vigencias,
                                           modelo, municipio_transito, apellidos_propietario,
                                           celular="3000000000", email="consulta@consulta.com",
                                           direccion="CRA", municipio="MEDELLIN",
                                           municipio_cod=5001000, departamento_cod=5, job_id=None,
                                           ignorar_cache=False):
    """Genera (o reutiliza del cache, si sigue vigente) el PDF de cada
    vigencia adeudada. Cada vigencia es INDEPENDIENTE: si una falla (a
    veces la Gobernacion tiene problemas puntuales con alguna), las demas
    igual se entregan, y solo habria que reintentar la que fallo.
    Si 'ignorar_cache' es True (ej. porque se dieron datos reales del
    cliente para el PDF final), se ignora cualquier PDF ya cacheado y se
    genera uno nuevo -- para no entregar por error una version vieja con
    datos de relleno.
    Devuelve una lista de dicts: {vigencia, ok, url, error}."""
    placa = placa.upper()
    resultados = []

    for vigencia in vigencias:
        cache = None if ignorar_cache else _cache_declaracion_buscar(placa, vigencia)
        # Solo se reutiliza directo si el cache tiene TANTO la url del
        # PDF COMO los datos completos -- si le faltan los datos (fila
        # "a medias", ej. de antes de que existiera esta funcionalidad),
        # no se salta: se hace la consulta en vivo igual, para completar
        # lo que falta en vez de quedarse con un hueco para siempre.
        if cache and cache.get("url") and cache.get("datos"):
            resultados.append({"vigencia": vigencia, "ok": True, "url": cache["url"]})
            if job_id:
                job_actualizar(job_id, f"Vigencia {vigencia}: usando declaración ya generada hoy...",
                                datos_parciales=resultados)
            continue

        if job_id:
            job_actualizar(job_id, f"Generando declaración de la vigencia {vigencia}...",
                            datos_parciales=resultados)
        try:
            pdf_bytes, data_vig = antioquia_generar_pdf_declaracion(
                placa, identificacion, tipo_documento_abrev, vigencia,
                modelo, municipio_transito, apellidos_propietario,
                celular, email, direccion, municipio, municipio_cod, departamento_cod
            )
            id_unico = uuid.uuid4().hex[:8]
            ruta = f"/tmp/decl_{placa}_{vigencia}_{id_unico}.pdf"
            with open(ruta, "wb") as f:
                f.write(pdf_bytes)

            url = subir_a_r2(ruta, f"declaraciones/{placa}_{vigencia}_{id_unico}.pdf",
                              nombre_descarga=f"Declaracion_Sugerida_{placa}_{vigencia}.pdf")
            os.remove(ruta)

            # Se extrae caja/traccion (y se guarda la liquidacion completa)
            # en cache -- asi otras herramientas (ej. la Declaracion
            # Manual) pueden reutilizar estos datos sin volver a consultar
            # en vivo a la Gobernacion.
            try:
                caja_traccion = _extraer_caja_traccion_declaracion(pdf_bytes)
            except Exception:
                caja_traccion = {}
            datos_extra = dict(data_vig or {})
            datos_extra["caja"] = caja_traccion.get("caja", "")
            datos_extra["traccion"] = caja_traccion.get("traccion", "")

            _cache_declaracion_guardar(placa, vigencia, url, datos_extra=datos_extra)
            resultados.append({"vigencia": vigencia, "ok": True, "url": url})
        except Exception as e:
            print(f"Error generando declaracion vigencia {vigencia} para {placa}: {e}", flush=True)
            resultados.append({"vigencia": vigencia, "ok": False, "error": str(e)})

        if job_id:
            job_actualizar(job_id, f"Vigencia {vigencia} lista.", datos_parciales=resultados)

    return resultados


def consultar_antioquia(page, placa, identificacion, tipo_documento_abrev,
                        modelo, municipio_transito, apellidos_propietario,
                        celular="3000000000", email="consulta@consulta.com",
                        direccion="CRA", municipio="MEDELLIN",
                        municipio_cod=5001000, departamento_cod=5, job_id=None):
    """
    Proceso completo para Antioquia.
    Retorna (registros, total, avaluo, estado_vehiculo, excede_limite).
    """
    LIMITE = ANTIOQUIA_LIMITE_VIGENCIAS

    # Limpieza perezosa: se aprovecha cada consulta de Impuesto
    # Departamental para borrar de paso los PDFs de declaraciones que ya
    # vencieron (no sirven para otro dia, no tiene sentido dejarlos
    # ocupando espacio).
    _limpiar_declaraciones_vencidas()


    # Resolver tipo de documento
    tipo_documento_id = ANTIOQUIA_TIPO_DOC_MAP.get(tipo_documento_abrev.upper(), "1")
    tipo_doc_info     = ANTIOQUIA_TIPOS_DOCUMENTO.get(tipo_documento_id, ANTIOQUIA_TIPOS_DOCUMENTO["1"])
    doc_abreviatura   = tipo_doc_info["abreviatura"]
    doc_nombre        = tipo_doc_info["nombre"]

    # Si es NIT, calcular y agregar dígito de verificación
    if tipo_documento_id == "2":
        identificacion = str(identificacion) + str(_calcular_digito_nit(identificacion))

    if job_id:
        job_actualizar(job_id, "Estoy ingresando a la página de la Gobernación de Antioquia...")
    print(f"\n  → Consultando primer bloque de datos ({placa})...")
    session0, token0, data3 = _sesion_antioquia(
        placa, identificacion, tipo_documento_id,
        modelo, municipio_transito, apellidos_propietario
    )

    estado_veh          = data3.get("estadoCuenta", {})
    vigencias_adeudadas = data3.get("listaVigenciasAdeudas", [])
    # Se usa el avaluo de la DECLARACION MAS RECIENTE (ej. la de 2026),
    # no el campo general estadoCuenta.avaluoComercial -- se confirmo con
    # un caso real que ese campo general puede traer un valor distinto
    # (mas viejo o de otra referencia) al avaluo que realmente aparece
    # declarado para el año en curso.
    avaluo              = _avaluo_declaracion_mas_reciente(data3)
    print(f"  → Vigencias adeudadas encontradas: {len(vigencias_adeudadas)}")
    if job_id:
        if not vigencias_adeudadas:
            job_actualizar(job_id, "Este vehículo está a paz y salvo con la Gobernación de Antioquia.")
        else:
            job_actualizar(job_id, f"Encontré {len(vigencias_adeudadas)} año(s) con impuesto pendiente. Consultando valores...")

    # Paz y salvo — solo retornar si el avaluo es confiable (> 0)
    if not vigencias_adeudadas:
        if not avaluo or avaluo == 0:
            raise Exception("No se pudo obtener información completa del vehículo. Por favor intente de nuevo.")
        return [], 0, avaluo, estado_veh, False

    total_vigencias       = len(vigencias_adeudadas)
    vigencias_a_consultar = sorted(vigencias_adeudadas, key=lambda x: x["vigencia"], reverse=True)
    excede_limite         = total_vigencias > LIMITE
    if excede_limite:
        vigencias_a_consultar = vigencias_a_consultar[:LIMITE]

    registros         = []
    total_suma        = 0
    avaluo_actual     = 0
    retefuente_actual = 0
    MAX_INTENTOS      = 2

    # Vigencias actualmente adeudadas según el portal
    anios_adeudados = set(str(v.get("vigencia")) for v in vigencias_a_consultar)

    # Limpiar del caché las vigencias que ya fueron pagadas
    try:
        conn_c = get_db_conn()
        cur_c  = conn_c.cursor()
        cur_c.execute("""
            SELECT vigencia FROM cache_impuestos_antioquia
            WHERE placa = %s AND estado = 'CON_DEUDA'
              AND (expira_en IS NULL OR expira_en >= NOW())
        """, (placa.upper(),))
        anios_en_cache = set(str(r[0]) for r in cur_c.fetchall())
        cur_c.close(); conn_c.close()
        for anio_pagado in (anios_en_cache - anios_adeudados):
            cache_antioquia_eliminar_vigencia(placa, anio_pagado)
    except Exception as e:
        print(f"  → Error limpiando caché: {e}")

    for v in vigencias_a_consultar:
        anio = v.get("vigencia")
        if job_id:
            job_actualizar(job_id, f"Estoy consultando el impuesto del año {anio}...")
        print(f"\n  → Consultando vigencia {anio}...")

        total_pagar  = None
        avaluo_vig   = 0

        # Intentar desde caché primero
        cache_vig = cache_antioquia_buscar_vigencia(placa, anio)
        if cache_vig:
            total_pagar = cache_vig['total_pagar']
            avaluo_vig  = cache_vig['avaluo']
            print(f"  ✔ Vigencia {anio} desde caché: ${total_pagar:,}")
        else:
            for intento in range(1, MAX_INTENTOS + 1):
                if intento > 1:
                    print(f"  ↺ Reintentando vigencia {anio}...")
                try:
                    session_v, token_v, _ = _sesion_antioquia(
                        placa, identificacion, tipo_documento_id,
                        modelo, municipio_transito, apellidos_propietario
                    )
                    data_vig = _consultar_vigencia_antioquia(
                        anio, session_v, token_v,
                        placa, identificacion, tipo_documento_id,
                        doc_abreviatura, doc_nombre,
                        celular, email, direccion, municipio, municipio_cod, departamento_cod
                    )
                    _msg    = data_vig.get("mensaje") or data_vig.get("descripcion")
                    _codigo = data_vig.get("codigo")
                    if _codigo and _codigo != 1 and _msg:
                        print(f"  ✖ Error servidor vigencia {anio}: {_msg}")

                    total_pagar = data_vig.get("totalPagar")
                    avaluo_vig  = data_vig.get("avaluoComercial", 0) or 0
                    if total_pagar is not None:
                        print(f"  ✔ Vigencia {anio}: ${total_pagar:,}")
                        # Guardar en caché (costo de la vigencia -- esta
                        # parte NO se toca, sigue igual que siempre)
                        try:
                            cache_antioquia_guardar_deuda(placa, [{
                                'vigencia': anio,
                                'total_pagar': total_pagar,
                            }], avaluo_vig or avaluo)
                            print(f"  → Caché guardado exitosamente para {placa} vigencia {anio}")
                        except Exception as ce:
                            print(f"  ✖ Error guardando caché vigencia {anio}: {ce}")

                        # Se aprovecha la MISMA sesion (sin gastar captcha
                        # adicional -- aceptar terminos y descargar el PDF
                        # no lo necesitan) para tambien generar y guardar
                        # el PDF de la declaracion de esta vigencia, listo
                        # por si luego se necesita descargar. Se guarda
                        # con la misma validez que ya se usa en otras
                        # partes: la vigencia actual sigue la ventana de
                        # pronto pago (enero-abril / mayo-julio), y las
                        # demas solo sirven el mismo dia.
                        try:
                            formulario_liquidacion = data_vig.get("formularioLiquidacion")
                            if formulario_liquidacion:
                                _antioquia_aceptar_terminos_liquidacion(
                                    session_v, identificacion, tipo_documento_id,
                                    doc_abreviatura, doc_nombre
                                )
                                pdf_bytes_vig = _antioquia_descargar_pdf_liquidacion(
                                    session_v, int(formulario_liquidacion)
                                )
                                id_unico_vig = uuid.uuid4().hex[:8]
                                ruta_pdf_vig = f"/tmp/decl_{placa}_{anio}_{id_unico_vig}.pdf"
                                with open(ruta_pdf_vig, "wb") as f_pdf:
                                    f_pdf.write(pdf_bytes_vig)
                                url_pdf_vig = subir_a_r2(
                                    ruta_pdf_vig, f"declaraciones/{placa}_{anio}_{id_unico_vig}.pdf",
                                    nombre_descarga=f"Declaracion_Sugerida_{placa}_{anio}.pdf"
                                )
                                os.remove(ruta_pdf_vig)

                                try:
                                    caja_traccion_vig = _extraer_caja_traccion_declaracion(pdf_bytes_vig)
                                except Exception:
                                    caja_traccion_vig = {}
                                datos_extra_vig = dict(data_vig or {})
                                datos_extra_vig["caja"] = caja_traccion_vig.get("caja", "")
                                datos_extra_vig["traccion"] = caja_traccion_vig.get("traccion", "")
                                _cache_declaracion_guardar(placa, anio, url_pdf_vig, datos_extra=datos_extra_vig)
                                print(f"  → PDF de la vigencia {anio} guardado (aprovechando la misma consulta).")
                        except Exception as e_pdf:
                            # Si falla la parte del PDF, no se interrumpe
                            # la consulta principal -- el dato del costo
                            # ya se obtuvo bien, que es lo importante.
                            print(f"  ✖ No se pudo generar/guardar el PDF de la vigencia {anio}: {e_pdf}")

                        break
                except Exception as e:
                    print(f"  ✖ Error vigencia {anio} intento {intento}: {e}")

        if not avaluo_actual and avaluo_vig:
            avaluo_actual     = avaluo_vig
            retefuente_actual = round(avaluo_vig / 100)

        if total_pagar is not None:
            total_suma += total_pagar

        registros.append({
            "vigencia":       str(anio),
            "estado":         "Pendiente de pago",
            "total_vigencia": total_pagar,
        })

        if job_id and total_pagar is not None:
            job_actualizar(job_id,
                f"Año {anio}: impuesto es ${total_pagar:,}. Continuando...",
                datos_parciales=list(registros))

    print(f"\n  ✔ ¡Consulta Antioquia finalizada!")
    return registros, total_suma, avaluo_actual or avaluo, estado_veh, excede_limite



def consultar_medellin(page, placa, identificacion, modelo, apellidos_propietario,
                       celular="3208578787", email="consulta@juridicox.com",
                       direccion="CRA 20 20 20"):
    """Consulta impuesto municipal de Medellín (servicio público) con valores reales incluyendo intereses."""
    import re as _re

    url = "https://www.medellin.gov.co/irj/portal/medellin/pago-impuesto-circulacion-transito"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)

    # Esperar popup de validación
    page.wait_for_selector("#popupValidacion", timeout=30000)

    # Cerrar popup de imagen si aparece
    try:
        cerrar = page.locator(".divCerrarPopup")
        if cerrar.is_visible(timeout=3000):
            cerrar.click()
            page.wait_for_timeout(500)
    except Exception:
        pass

    # Paso 0 — popup validación: servicio público + Medellín
    page.locator("input[name='tipoVehiculo'][value='publico']").check()
    page.wait_for_timeout(500)
    page.wait_for_selector("#matriculaLugar", timeout=5000)
    page.locator("input[name='lugarMatricula'][value='medellin']").check()
    page.wait_for_timeout(500)
    page.wait_for_function("() => !document.getElementById('btnContinuar').disabled", timeout=5000)
    page.locator("#btnContinuar").click()

    # Paso 1 — llenar placa y documento
    page.wait_for_selector("#placa", timeout=15000)
    page.locator("#placa").fill(placa.upper())
    page.locator("#id").fill(identificacion)
    page.locator("button.boton_consulta").click()

    # Esperar tabla de vigencias
    page.wait_for_selector("#cont_paso1 table tbody tr", timeout=30000)

    body_text = page.inner_text("body").lower()
    if "no se encuentra matriculado" in body_text or "no está matriculado" in body_text:
        raise Exception("Este vehículo no está matriculado en la Secretaría de Movilidad de Medellín.")
    if "no presenta deuda" in body_text or "no adeuda" in body_text or "paz y salvo" in body_text:
        return [], 0

    # Seleccionar todos los checkboxes de vigencias
    try:
        page.locator(".sel_todo").click(timeout=3000)
    except Exception:
        pass
    page.wait_for_timeout(500)
    # Si no quedaron todos marcados, marcarlos uno a uno
    checkboxes = page.locator("#cont_paso1 input[type='checkbox']").all()
    for cb in checkboxes:
        try:
            if not cb.is_checked():
                cb.check()
        except Exception:
            pass
    page.wait_for_timeout(500)
    # Verificar que hay checkboxes marcados antes de continuar
    marcados = page.locator("#cont_paso1 input[type='checkbox']:checked").count()
    btn = page.locator("button.boton_continuar").first
    disabled = btn.get_attribute("disabled")
    # Forzar click aunque esté deshabilitado
    btn.evaluate("el => el.click()")
    page.wait_for_timeout(1000)

    # Paso 2a — modelo y propietario
    page.wait_for_selector("#modelo_veh", timeout=15000)
    modelo_str = str(modelo).strip()[:4].zfill(4)
    page.locator("#modelo_veh").fill(modelo_str)

    # Seleccionar propietario — primera opción disponible si no hay match por apellido
    try:
        opciones = page.locator("#nombres_props option.valorSel").all()
        valor_sel = opciones[0].get_attribute("value") if opciones else None
        for op in opciones:
            texto = (op.inner_text() or "").upper()
            if apellidos_propietario and apellidos_propietario.split()[0].upper() in texto:
                valor_sel = op.get_attribute("value")
                break
        if valor_sel:
            page.locator("#nombres_props").select_option(valor_sel)
    except Exception:
        pass

    # Forzar click en boton_validar ignorando validación HTML5
    page.locator("button.boton_validar").evaluate("el => el.click()")
    page.wait_for_timeout(1500)

    # Paso 2b — datos de contacto del propietario
    page.wait_for_selector("#correo", timeout=15000)
    page.locator("#correo").fill(email)
    page.locator("#celular").fill(celular)
    try:
        page.locator("#telefono").fill("6042379933")
    except Exception:
        pass

    # Dirección — abrir popup y llenar
    try:
        page.locator("#direccion").click()
        page.wait_for_selector("#tipo_via", state="visible", timeout=5000)
        page.locator("#tipo_via").select_option("CARRERA")
        page.locator("#numero1").fill("20")
        page.locator("#numero2").fill("20")
        page.locator("#numero3").fill("20")
        # Guardar dirección con evaluate para bypass validación
        page.locator("button.boton_dir").evaluate("el => el.click()")
        page.wait_for_timeout(1000)
        # Verificar que la dirección quedó guardada
        dir_val = page.locator("#direccion").input_value()
    except Exception as e:
        # Si falla el popup, inyectar la dirección directamente
        try:
            page.evaluate("document.getElementById('direccion').removeAttribute('readonly')")
            page.locator("#direccion").fill("CARRERA 20 20 20")
        except Exception:
            pass

    # Departamento y municipio
    try:
        page.locator("#departamento").select_option("05")
        page.wait_for_timeout(500)
        page.locator("#municipio").select_option("000000005001")
        page.wait_for_timeout(500)
    except Exception as e:
        pass

    # Guardar datos del propietario — bypass validación HTML5
    # Intentar el botón de guardar del form info_propietario
    try:
        page.locator("button[form='info_propietario']").first.evaluate("el => el.click()")
    except Exception:
        # Si no existe, buscar boton_continuar o cualquier submit visible
        try:
            page.locator(".divContBotones button:not(.boton_cancelar)").first.evaluate("el => el.click()")
        except Exception as e2:
            pass
    page.wait_for_timeout(1500)

    # Esperar tabla del paso 3 con valores reales
    page.wait_for_selector("#cont_paso3 table tbody tr", timeout=30000)

    # Extraer total general del tfoot
    total = 0
    try:
        tfoot_text = page.locator("#cont_paso3 table tfoot").inner_text()
        total_match = _re.search(r'\$([\d\.]+)', tfoot_text)
        if total_match:
            total = int(total_match.group(1).replace('.', ''))
    except Exception:
        pass

    # Extraer vigencias con valores reales (impuesto + intereses + total por vigencia)
    registros = []
    filas = page.locator("#cont_paso3 table tbody tr").all()
    for fila in filas:
        texto = fila.inner_text().strip()
        if not texto:
            continue
        anio = _re.search(r'\b(20\d{2}|19\d{2})\b', texto)
        # Buscar "Total a pagar" que es la última columna
        valores = _re.findall(r'\$([\d\.]+)', texto)
        if anio and valores:
            try:
                # El último valor es "Total a pagar" por vigencia
                total_vigencia = int(valores[-1].replace('.', ''))
                impuesto = int(valores[-3].replace('.', '')) if len(valores) >= 3 else 0
                interes = int(valores[-2].replace('.', '')) if len(valores) >= 2 else 0
                if total_vigencia > 0:
                    registros.append({
                        'vigencia': anio.group(),
                        'estado': 'Pendiente de pago',
                        'impuesto_base': impuesto,
                        'interes_mora': interes,
                        'total_vigencia': total_vigencia
                    })
            except ValueError:
                pass

    if not total and registros:
        total = sum(r['total_vigencia'] for r in registros)

    return registros, total

# ============================================================
#  MAPA DE MUNICIPIOS
# ============================================================
MUNICIPIOS = {
    "envigado":    consultar_envigado,
    "sabaneta":    consultar_sabaneta,
    "itagui":      consultar_itagui,
    "bello":       consultar_bello,
    "laestrella":  consultar_laestrella,
    "la estrella": consultar_laestrella,
    "medellin":    consultar_medellin,
    "medellín":    consultar_medellin,
}


@app.route("/consultar", methods=["GET"])
def consultar():
    import traceback
    placa     = request.args.get("placa", "").upper().strip()
    municipio = request.args.get("municipio", "").lower().strip()
    if not placa or not municipio:
        return jsonify({"error": "Debes proporcionar placa y municipio."}), 400
    if municipio not in MUNICIPIOS and municipio != "antioquia":
        return jsonify({"error": f"Municipio '{municipio}' no reconocido.", "opciones": list(MUNICIPIOS.keys()) + ["antioquia"]}), 400

    identificacion     = request.args.get("identificacion", "").strip()
    tipo_documento     = request.args.get("tipo_documento", "CC").strip().upper() or "CC"
    modelo             = request.args.get("modelo", "").strip()
    municipio_transito = request.args.get("municipio_transito", "").upper().strip()
    apellidos          = request.args.get("apellidos_propietario", "").upper().strip()
    celular            = request.args.get("celular", "3000000000").strip()
    email              = request.args.get("email", "consulta@consulta.com").strip()
    direccion          = request.args.get("direccion", "CRA").strip()
    mun_declarante     = request.args.get("municipio_declarante", "MEDELLIN").strip().upper()
    municipio_cod      = int(request.args.get("municipio_cod", 5001000))
    departamento_cod   = int(request.args.get("departamento_cod", 5))

    if municipio == "antioquia":
        if not identificacion or not modelo or not municipio_transito or not apellidos:
            return jsonify({"error": "Para Antioquia debes proporcionar: identificacion, modelo, municipio_transito, apellidos_propietario."}), 400

    # Municipios síncronos
    if municipio != "antioquia":
        # Verificar cache antes de lanzar Playwright -- si ya sabemos que esta
        # a paz y salvo este año, no hace falta volver a consultar la pagina
        # del municipio (evita una consulta lenta e innecesaria).
        cache_hit_mun = cache_municipal_buscar(placa, municipio)
        if cache_hit_mun:
            print(f"  → Cache hit municipal para {placa} en {municipio} — respondiendo sin Playwright")
            return jsonify({
                "placa":       placa,
                "municipio":   municipio,
                "registros":   [],
                "total":       0,
                "sin_deuda":   True,
                "verificado":  True,
                "placa_vista": cache_hit_mun["placa_vista"],
                "fecha_pago":  cache_hit_mun["fecha_pago"],
                "marca":       cache_hit_mun["marca"],
                "valor_pago":  cache_hit_mun["valor_pago"],
                "desde_cache": True,
            })

        resultado       = {}
        error_container = {}

        def ejecutar_mun():
            try:
                with sync_playwright() as playwright:
                    browser = playwright.chromium.launch(headless=True, args=[
                        "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
                        "--single-process", "--no-zygote", "--disable-setuid-sandbox"
                    ])
                    context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
                    page = context.new_page()
                    if municipio not in ["bello", "sabaneta", "laestrella"]:
                        bloquear_recursos(page)
                    funcion = MUNICIPIOS[municipio]
                    if municipio == "medellin":
                        registros, total = funcion(page, placa,
                            identificacion=identificacion,
                            modelo=modelo,
                            apellidos_propietario=apellidos,
                            celular=celular,
                            email=email)
                    else:
                        registros, total = funcion(page, placa)
                    resultado['registros'] = registros
                    resultado['total']     = total
                    context.close(); browser.close()
            except Exception as e:
                error_container['error'] = str(e)
                print(traceback.format_exc(), flush=True)

        hilo = threading.Thread(target=ejecutar_mun)
        hilo.start()
        hilo.join(timeout=620)

        if hilo.is_alive():
            return jsonify({"error": "La consulta tardo demasiado. Intenta de nuevo."}), 504
        if error_container:
            return jsonify({"error": error_container['error']}), 500
        registros_mun   = resultado.get('registros', [])
        total_mun       = resultado.get('total', 0)
        fecha_pago_mun  = ""
        marca_pago_mun  = ""
        valor_pago_mun  = ""
        placa_vista_mun = ""

        # Extraer datos de paz y salvo si el municipio los devuelve
        if registros_mun and registros_mun[0].get('paz_y_salvo'):
            r0              = registros_mun[0]
            fecha_pago_mun  = r0.get('fecha_pago', '')
            marca_pago_mun  = r0.get('marca', '')
            valor_pago_mun  = r0.get('valor_pago', '')
            placa_vista_mun = r0.get('placa_info', '')
            registros_mun   = []
            total_mun       = 0
        # Extraer último pago de registros con deuda (si viene en el primer registro)
        elif registros_mun and registros_mun[0].get('fecha_ultimo_pago'):
            r0              = registros_mun[0]
            fecha_pago_mun  = r0.get('fecha_ultimo_pago', '')
            marca_pago_mun  = r0.get('marca_ultimo_pago', '')
            valor_pago_mun  = r0.get('valor_ultimo_pago', '')
            placa_vista_mun = r0.get('placa_ultimo_pago', '')

        # Reintento automático si paz y salvo sin evidencia de verificación
        # (posible falso positivo: la página no cargó los datos reales del vehículo).
        # Se considera "verificado" cuando trajo placa y/o marca vistas en la página.
        verificado_mun = bool(placa_vista_mun or marca_pago_mun)
        if total_mun == 0 and not verificado_mun and municipio in ("envigado", "sabaneta", "itagui", "bello"):
            resultado2    = {}
            error2        = {}
            funcion_reint = MUNICIPIOS[municipio]
            def _reintento():
                try:
                    with sync_playwright() as pw2:
                        b2 = pw2.chromium.launch(headless=True, args=[
                            "--no-sandbox","--disable-dev-shm-usage","--disable-gpu",
                            "--single-process","--no-zygote","--disable-setuid-sandbox"
                        ])
                        ctx2  = b2.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
                        pg2   = ctx2.new_page()
                        bloquear_recursos(pg2)
                        r2, t2 = funcion_reint(pg2, placa)
                        resultado2['registros'] = r2
                        resultado2['total']     = t2
                        ctx2.close(); b2.close()
                except Exception as e2:
                    error2['error'] = str(e2)
            import threading as _th2
            hilo2 = _th2.Thread(target=_reintento)
            hilo2.start()
            hilo2.join(timeout=120)
            if not error2 and resultado2:
                r2 = resultado2.get('registros', [])
                t2 = resultado2.get('total', 0)
                fp2 = ""; mp2 = ""; vp2 = ""; pv2 = ""
                if r2 and r2[0].get('paz_y_salvo'):
                    fp2 = r2[0].get('fecha_pago', '')
                    mp2 = r2[0].get('marca', '')
                    vp2 = r2[0].get('valor_pago', '')
                    pv2 = r2[0].get('placa_info', '')
                    r2  = []
                    t2  = 0
                elif r2 and r2[0].get('fecha_ultimo_pago'):
                    fp2 = r2[0].get('fecha_ultimo_pago', '')
                    mp2 = r2[0].get('marca_ultimo_pago', '')
                    vp2 = r2[0].get('valor_ultimo_pago', '')
                    pv2 = r2[0].get('placa_ultimo_pago', '')
                # Si el reintento también da paz y salvo verificado → confirmado
                # Si el reintento da deuda → el primero era falso positivo
                registros_mun   = r2
                total_mun       = t2
                fecha_pago_mun  = fp2
                marca_pago_mun  = mp2
                valor_pago_mun  = vp2
                placa_vista_mun = pv2
                verificado_mun  = bool(pv2 or mp2)

        # Guardar en cache si quedo confirmado a paz y salvo -- asi no se
        # vuelve a consultar este municipio para esta placa el resto del año.
        if verificado_mun and total_mun == 0 and not registros_mun:
            cache_municipal_guardar_paz_salvo(
                placa, municipio, fecha_pago_mun, marca_pago_mun, valor_pago_mun, placa_vista_mun
            )

        return jsonify({
            "placa":       placa,
            "municipio":   municipio,
            "registros":   registros_mun,
            "total":       total_mun,
            "sin_deuda":   total_mun == 0 and not registros_mun,
            "verificado":  verificado_mun,
            "placa_vista": placa_vista_mun,
            "fecha_pago":  fecha_pago_mun,
            "marca":       marca_pago_mun,
            "valor_pago":  valor_pago_mun,
        })

    # Antioquia — verificar caché de vigencias antes de lanzar Playwright
    # El snippet pasa las vigencias adeudadas que ya conoce del paso 1
    vigencias_param = request.args.get("vigencias", "").strip()
    if vigencias_param:
        anios_solicitados = [a.strip() for a in vigencias_param.split(",") if a.strip()]
        registros_cache = []
        avaluo_cache    = 0
        total_cache     = 0
        todos_cacheados = True

        for anio in anios_solicitados:
            cv = cache_antioquia_buscar_vigencia(placa, anio)
            if cv:
                registros_cache.append({
                    "vigencia":       str(anio),
                    "estado":         "Pendiente de pago",
                    "total_vigencia": cv['total_pagar'],
                })
                total_cache  += cv['total_pagar']
                if not avaluo_cache:
                    avaluo_cache = cv['avaluo']
            else:
                todos_cacheados = False
                break

        if todos_cacheados and registros_cache:
            print(f"  → Cache hit completo para {placa} — respondiendo sin Playwright")
            return jsonify({
                "placa":      placa,
                "municipio":  "antioquia",
                "placa_info": {},
                "registros":  registros_cache,
                "total":      total_cache,
                "avaluo":     avaluo_cache,
                "retefuente": round(avaluo_cache / 100) if avaluo_cache else 0,
                "sin_deuda":  False,
                "desde_cache": True,
            })

    # Antioquia — sistema asíncrono
    job_id = str(uuid.uuid4())[:12]
    job_actualizar(job_id, "Iniciando consulta...", "procesando")

    def ejecutar_antioquia():
        try:
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(headless=True, args=[
                    "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
                    "--single-process", "--no-zygote", "--disable-setuid-sandbox"
                ])
                context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
                page = context.new_page()
                registros, total, avaluo, estado_veh, excede = consultar_antioquia(
                    page, placa, identificacion, tipo_documento,
                    modelo, municipio_transito, apellidos,
                    celular, email, direccion, mun_declarante,
                    municipio_cod, departamento_cod, job_id=job_id
                )
                context.close(); browser.close()

            respuesta = {
                "placa":      placa,
                "municipio":  "antioquia",
                "placa_info": {
                    "marca":       estado_veh.get("marca", ""),
                    "linea":       estado_veh.get("linea", ""),
                    "modelo":      estado_veh.get("modelo", ""),
                    "propietario": estado_veh.get("nombrePropietario", ""),
                },
                "registros":  registros,
                "total":      total,
                "avaluo":     avaluo,
                "retefuente": round(avaluo / 100) if avaluo else 0,
                "sin_deuda":  len(registros) == 0,
            }
            if excede:
                respuesta["excede_limite"]  = True
                respuesta["mensaje_limite"] = f"El límite de consulta es de {ANTIOQUIA_LIMITE_VIGENCIAS} vigencias. Comunícate con un asesor de la Gobernación de Antioquia al 6044444666."
            job_terminar(job_id, respuesta)
        except Exception as e:
            print(traceback.format_exc(), flush=True)
            msg = str(e)
            if any(x in msg.lower() for x in ["net::err", "connection"]):
                msg = "No se pudo conectar al portal de Antioquia. Intenta más tarde."
            job_error(job_id, msg)

    threading.Thread(target=ejecutar_antioquia, daemon=True).start()
    return jsonify({"job_id": job_id, "estado": "procesando"})


@app.route("/consultar/antioquia/vigencias", methods=["GET"])
def consultar_antioquia_vigencias():
    """PASO 1 — Rápido (2 captchas). Devuelve lista de vigencias sin valores."""
    import traceback
    placa              = request.args.get("placa", "").upper().strip()
    identificacion     = request.args.get("identificacion", "").strip()
    tipo_documento     = request.args.get("tipo_documento", "CC").strip().upper()
    modelo             = request.args.get("modelo", "").strip()
    municipio_transito = request.args.get("municipio_transito", "").upper().strip()
    apellidos          = request.args.get("apellidos_propietario", "").upper().strip()

    if not all([placa, identificacion, modelo, municipio_transito, apellidos]):
        return jsonify({"error": "Faltan datos requeridos"}), 400

    # Verificar caché primero
    cache = cache_antioquia_buscar(placa)
    if cache and cache['estado'] == 'PAZ_Y_SALVO':
        print(f"  → Cache hit PAZ_Y_SALVO para {placa}")
        return jsonify({
            "placa":       placa,
            "sin_deuda":   True,
            "avaluo":      cache.get('avaluo', 0),
            "retefuente":  cache.get('retefuente', 0),
            "vigencias":   [],
            "placa_info":  {},
            "desde_cache": True,
        })


    resultado       = {}
    error_container = {}

    def ejecutar():
        try:
            tipo_documento_id = ANTIOQUIA_TIPO_DOC_MAP.get(tipo_documento, "1")
            if tipo_documento_id == "2":
                ident = str(identificacion) + str(_calcular_digito_nit(identificacion))
            else:
                ident = identificacion
            session0, token0, data3 = _sesion_antioquia(
                placa, ident, tipo_documento_id,
                modelo, municipio_transito, apellidos
            )
            estado_veh          = data3.get("estadoCuenta", {})
            vigencias_adeudadas = data3.get("listaVigenciasAdeudas", [])
            # Se usa el avaluo de la DECLARACION MAS RECIENTE, no el campo
            # general estadoCuenta.avaluoComercial (ver comentario en
            # _avaluo_declaracion_mas_reciente para el detalle del caso
            # real que confirmo esta discrepancia).
            avaluo              = _avaluo_declaracion_mas_reciente(data3)
            resultado['vigencias']  = vigencias_adeudadas
            resultado['avaluo']     = avaluo
            resultado['estado_veh'] = estado_veh
            resultado['sin_deuda']  = len(vigencias_adeudadas) == 0
            # Guardar en caché si está a paz y salvo
            if not vigencias_adeudadas and avaluo and avaluo > 0:
                cache_antioquia_guardar_paz_salvo(placa, avaluo, estado_veh)
                # Se guardan tambien TODOS los datos (historial de
                # declaraciones, procesos fiscales, bloqueos, novedades)
                # para poder generar el documento Estado de Cuenta despues.
                guardar_estado_cuenta_antioquia(placa, data3)
        except Exception as e:
            error_container['error'] = str(e)
            print(traceback.format_exc(), flush=True)

    hilo = threading.Thread(target=ejecutar)
    hilo.start()
    hilo.join(timeout=120)

    if hilo.is_alive():
        return jsonify({"error": "La consulta tardó demasiado. Intenta de nuevo."}), 504
    if error_container:
        return jsonify({"error": error_container['error']}), 500

    estado_veh = resultado.get('estado_veh', {})
    avaluo     = resultado.get('avaluo', 0)
    vigencias  = resultado.get('vigencias', [])

    return jsonify({
        "placa":      placa,
        "sin_deuda":  resultado.get('sin_deuda', True),
        "avaluo":     avaluo,
        "retefuente": round(avaluo / 100) if avaluo else 0,
        "vigencias":  vigencias,
        "placa_info": {
            "marca":       estado_veh.get("marca", ""),
            "linea":       estado_veh.get("linea", ""),
            "modelo":      estado_veh.get("modelo", ""),
            "propietario": estado_veh.get("nombrePropietario", ""),
        }
    })


@app.route("/consultar-runt-vehiculo", methods=["GET"])
def consultar_runt_vehiculo_endpoint():
    """Consulta el RUNT (Placa y Propietario) para una placa + cedula.
    Es una consulta con su propio costo de 2Captcha, independiente de las
    consultas de impuestos -- por eso se guarda directo en la tabla
    'vehiculos' cada vez que se llama, sin verificar cache primero (los
    datos del RUNT cambian con cada tramite, a diferencia del estado de
    paz y salvo que dura todo el año)."""
    placa  = request.args.get("placa", "").upper().strip()
    cedula = request.args.get("cedula", "").strip()
    tipo_documento = request.args.get("tipo_documento", "CC").strip().upper() or "CC"
    user_id = request.args.get("user_id", "").strip()  # id del usuario en Supabase (opcional)

    if not placa or not cedula:
        return jsonify({"error": "Debes proporcionar placa y cedula."}), 400

    # Limite: no se puede forzar una nueva consulta al RUNT para la misma
    # placa si ya se hizo una en los ultimos 3 dias -- en ese caso hay que
    # usar el dato ya guardado en cache (vehiculo-runt-guardado).
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT leido_en FROM vehiculos WHERE placa = %s AND fuente = 'RUNT'
        """, (placa,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row and row[0]:
            transcurrido = datetime.now() - row[0]
            if transcurrido < timedelta(days=3):
                faltan = timedelta(days=3) - transcurrido
                horas_faltantes = int(faltan.total_seconds() // 3600)
                return jsonify({
                    "error": f"Esta placa ya se consultó en el RUNT hace menos de 3 días. "
                             f"Debes usar el dato ya guardado; podrás forzar una nueva consulta "
                             f"en aproximadamente {horas_faltantes} horas.",
                    "limite_activo": True,
                    "horas_restantes": horas_faltantes
                }), 429
    except Exception as e:
        print(f"Error verificando limite de consulta RUNT: {e}", flush=True)
        # Si falla la verificacion, se deja continuar (no se bloquea por un
        # problema tecnico ajeno al limite en si).

    job_id = str(uuid.uuid4())[:12]
    job_actualizar(job_id, "Iniciando consulta RUNT...", "procesando")

    def ejecutar():
        try:
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(headless=True, args=[
                    "--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu",
                    "--single-process", "--no-zygote", "--disable-setuid-sandbox"
                ])
                context = browser.new_context(
                    user_agent="Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
                    viewport={"width": 390, "height": 844},
                )
                page = context.new_page()
                datos = consultar_runt_vehiculo(page, placa, cedula, tipo_documento, job_id=job_id)
                context.close(); browser.close()

            if not datos.get("placa"):
                job_error(job_id, "No se pudo leer la placa en el resultado. Verifica los datos o intenta de nuevo.")
                return

            guardar_vehiculo_runt(datos)
            if user_id:
                guardar_mi_consulta(user_id, datos["placa"], cedula)
            datos["leido_en"] = (datetime.now() - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M")
            job_terminar(job_id, datos)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    hilo = threading.Thread(target=ejecutar)
    hilo.start()

    return jsonify({"job_id": job_id})


def guardar_mi_consulta(user_id, placa, cedula):
    """Registra que este usuario en particular consulto esta placa (y
    cedula), para el historial personal de 'Mis vehiculos consultados'.
    La restriccion unica es solo (user_id, placa) -- si la cedula viene
    distinta a una consulta anterior de esa misma placa (ej. por
    diferencias de formato), se ACTUALIZA la fila existente en vez de
    crear una copia nueva."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO mis_consultas (user_id, placa, cedula, actualizado_en)
            VALUES (%s, %s, %s, NOW())
            ON CONFLICT (user_id, placa) DO UPDATE SET cedula = EXCLUDED.cedula, actualizado_en = NOW()
        """, (user_id, placa, cedula))
        conn.commit()
        cur.close(); conn.close()
        return True, None
    except Exception as e:
        print(f"Error guardando mi_consulta: {e}")
        return False, str(e)


@app.route("/guardar-vehiculo-ocr", methods=["POST"])
def guardar_vehiculo_ocr():
    """Guarda lo leido por OCR de una tarjeta de propiedad. Si esa placa ya
    tenia datos de una consulta al RUNT, el RUNT prevalece: no se
    sobrescriben marca/linea/modelo/etc, ni la fuente ni la fecha de
    lectura -- solo se actualizan los campos que el RUNT NUNCA trae
    (tipo de documento, cedula, nombre del propietario, municipio, y la
    limitacion a la propiedad tal como la leyo el OCR)."""
    data = request.get_json(silent=True) or {}
    placa = (data.get("placa") or "").upper().strip()
    if not placa:
        return jsonify({"error": "Debes proporcionar placa."}), 400

    user_id = (data.get("user_id") or "").strip()
    cedula  = (data.get("cedula") or "").strip()

    campos_ocr_siempre = ["municipio", "propietario_tipo_documento", "propietario_cedula", "propietario_nombre", "ocr_limitacion_propiedad"]
    campos_ocr_si_no_hay_runt = ["clase", "marca", "linea", "modelo", "cilindrada", "servicio", "carroceria", "capacidad_pasajeros"]

    valores = {
        "municipio": data.get("municipio"),
        "propietario_tipo_documento": data.get("tipo_documento"),
        "propietario_cedula": cedula,
        "propietario_nombre": data.get("apellidos"),
        "ocr_limitacion_propiedad": data.get("limitacion_propiedad"),
        "clase": data.get("clase"),
        "marca": data.get("marca"),
        "linea": data.get("linea"),
        "modelo": data.get("modelo"),
        "cilindrada": data.get("cilindrada"),
        "servicio": data.get("servicio"),
        "carroceria": data.get("carroceria"),
        "capacidad_pasajeros": data.get("capacidad"),
    }

    try:
        conn = get_db_conn()
        cur = conn.cursor()

        todas_columnas = campos_ocr_siempre + campos_ocr_si_no_hay_runt
        cols_sql = ", ".join(["placa"] + todas_columnas + ["fuente", "leido_en"])
        vals_sql = ", ".join(["%s"] * (len(todas_columnas) + 3))
        params = [placa] + [valores[c] for c in todas_columnas] + ["OCR", datetime.now()]

        set_siempre = ", ".join(f"{c}=EXCLUDED.{c}" for c in campos_ocr_siempre)
        set_condicional = ", ".join(
            f"{c}=CASE WHEN vehiculos.fuente='RUNT' THEN vehiculos.{c} ELSE EXCLUDED.{c} END"
            for c in campos_ocr_si_no_hay_runt
        )

        cur.execute(f"""
            INSERT INTO vehiculos ({cols_sql})
            VALUES ({vals_sql})
            ON CONFLICT (placa) DO UPDATE SET
                {set_siempre}, {set_condicional},
                fuente = CASE WHEN vehiculos.fuente='RUNT' THEN 'RUNT' ELSE 'OCR' END,
                leido_en = CASE WHEN vehiculos.fuente='RUNT' THEN vehiculos.leido_en ELSE EXCLUDED.leido_en END
        """, params)
        conn.commit()
        cur.close(); conn.close()

        if user_id:
            guardar_mi_consulta(user_id, placa, cedula)

        return jsonify({"ok": True})
    except Exception as e:
        print(f"Error guardando vehiculo OCR: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/combinar-pdfs", methods=["POST"])
def combinar_pdfs_endpoint():
    """Toma varias URLs de PDFs YA generados (ej. de /generar-pdf-declaracion)
    y las une en un solo PDF de varias paginas -- pensado para el boton
    'Combinar PDFs' en el frontend, una vez ya se generaron los individuales.
    No vuelve a consultar nada ni gasta captchas, solo descarga y pega."""
    datos = request.get_json(silent=True) or {}
    urls = datos.get("urls", [])
    placa = (datos.get("placa") or "declaraciones").upper().strip()
    # El nombre de archivo es configurable -- este endpoint se reutiliza
    # tanto para combinar Declaraciones Sugeridas (su uso original) como
    # para los combos de documentos de Preparacion/Liquidacion (Combo
    # Traspaso, Combo FUN-Mandato, etc.), que necesitan su propio nombre
    # en vez de que todo diga "Declaracion".
    nombre_archivo = (datos.get("nombre_archivo") or "Declaracion_Sugerida").strip()
    nombre_archivo = re.sub(r"[^\w\s-]", "", nombre_archivo).replace(" ", "_")

    if not urls or len(urls) < 2:
        return jsonify({"error": "Se necesitan al menos 2 URLs para combinar"}), 400

    rutas_temp = []
    try:
        writer = PdfWriter()
        for i, url in enumerate(urls):
            resp = requests.get(url, timeout=30)
            if resp.status_code != 200:
                raise Exception(f"No se pudo descargar {url}")
            ruta = f"/tmp/combinar_{i}_{uuid.uuid4().hex[:6]}.pdf"
            with open(ruta, "wb") as f:
                f.write(resp.content)
            rutas_temp.append(ruta)
            writer.append(ruta)

        id_unico = uuid.uuid4().hex[:8]
        ruta_combinado = f"/tmp/combinado_{placa}_{id_unico}.pdf"
        with open(ruta_combinado, "wb") as f:
            writer.write(f)
        writer.close()

        url_final = subir_a_r2(ruta_combinado, f"declaraciones/combinado_{placa}_{id_unico}.pdf",
                                nombre_descarga=f"{nombre_archivo}_{placa}_combinado.pdf")
        os.remove(ruta_combinado)
        for ruta in rutas_temp:
            os.remove(ruta)

        return jsonify({"ok": True, "url": url_final})
    except Exception as e:
        for ruta in rutas_temp:
            if os.path.exists(ruta):
                os.remove(ruta)
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/generar-pdf-declaracion", methods=["GET"])
def generar_pdf_declaracion_endpoint():
    """Genera el/los PDF(s) de la declaracion sugerida (pago en banco) EN
    SEGUNDO PLANO -- con varias vigencias (cada una con sus propios
    captchas) el proceso puede tardar varios minutos, y una sola peticion
    tan larga corre el riesgo de que el navegador (o algun proxy en el
    camino) la corte con 'Failed to fetch' aunque el servidor siga
    trabajando bien. Por eso, igual que la consulta de impuestos, esto
    responde de inmediato con un job_id para consultar el avance en
    /consultar/estado."""
    placa = request.args.get("placa", "").upper().strip()
    identificacion = request.args.get("identificacion", "").strip()
    tipo_documento = request.args.get("tipo_documento", "CC").strip()
    vigencias_raw = request.args.get("vigencia", "").strip()
    modelo = request.args.get("modelo", "").strip()
    municipio_transito = request.args.get("municipio_transito", "").strip()
    apellidos_propietario = request.args.get("apellidos_propietario", "").strip()
    # Opcionales -- para que el PDF quede diligenciado con los datos reales
    # del cliente que va a pagar, en vez de los datos de relleno. Si no se
    # envian, se usan los valores por defecto de siempre.
    celular_raw = request.args.get("celular", "").strip()
    email_raw = request.args.get("email", "").strip()
    direccion_raw = request.args.get("direccion", "").strip()
    municipio_raw = request.args.get("municipio", "").strip()
    municipio_cod_raw = request.args.get("municipio_cod", "").strip()
    departamento_cod_raw = request.args.get("departamento_cod", "").strip()

    celular = celular_raw or "3000000000"
    email = email_raw or "consulta@consulta.com"
    direccion = direccion_raw or "CRA"
    municipio_residencia = municipio_raw or "MEDELLIN"
    municipio_cod = int(municipio_cod_raw) if municipio_cod_raw.isdigit() else 5001000
    departamento_cod = int(departamento_cod_raw) if departamento_cod_raw.isdigit() else 5

    # Si se dieron datos reales, no se reutiliza un PDF cacheado (podria
    # tener los datos de relleno de una consulta anterior).
    tiene_datos_reales = bool(celular_raw or email_raw or direccion_raw or municipio_raw)

    if not all([placa, identificacion, vigencias_raw, modelo, municipio_transito, apellidos_propietario]):
        return jsonify({"error": "Faltan parametros: placa, identificacion, vigencia, modelo, municipio_transito, apellidos_propietario"}), 400

    vigencias = [v.strip() for v in vigencias_raw.split(",") if v.strip()]
    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Iniciando...", "procesando")

    def ejecutar_declaraciones():
        try:
            resultados = antioquia_generar_todas_declaraciones(
                placa, identificacion, tipo_documento, vigencias,
                modelo, municipio_transito, apellidos_propietario,
                celular=celular, email=email, direccion=direccion,
                municipio=municipio_residencia, municipio_cod=municipio_cod,
                departamento_cod=departamento_cod,
                ignorar_cache=tiene_datos_reales,
                job_id=job_id
            )
            job_terminar(job_id, {"resultados": resultados})
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar_declaraciones, daemon=True).start()
    return jsonify({"job_id": job_id, "estado": "procesando"})


@app.route("/generar-declaracion-manual", methods=["GET"])
def generar_declaracion_manual_endpoint():
    """Genera la Declaracion Manual (formulario FO-M8-P6-008) diligenciada,
    combinando: datos del vehiculo ya guardados en Tramy (tabla vehiculos),
    los datos del propietario que escribe el usuario, y la liquidacion
    privada + caja/traccion que se obtienen de una consulta real a la
    Gobernacion (la misma consulta que usa la Declaracion Sugerida).
    Acepta varias vigencias separadas por coma (igual que el Generador de
    Declaraciones Sugeridas) -- cada una es independiente, si una falla
    las demas igual se entregan. Se ejecuta en segundo plano por el mismo
    motivo que las demas consultas a la Gobernacion: puede tardar por los
    captchas."""
    placa = request.args.get("placa", "").upper().strip()
    identificacion = request.args.get("identificacion", "").strip()
    tipo_documento = request.args.get("tipo_documento", "CC").strip()
    vigencias_raw = request.args.get("vigencia", "").strip()
    modelo = request.args.get("modelo", "").strip()
    municipio_transito = request.args.get("municipio_transito", "").strip()
    apellidos_propietario = request.args.get("apellidos_propietario", "").strip()
    nombres_propietario = request.args.get("nombres_propietario", "").strip()

    # Estos 4 datos son siempre los mismos, sin importar el cliente --
    # se dejan fijos aqui, ignorando cualquier valor que llegue del
    # formulario para estos campos especificos.
    celular = "3107208784"
    telefono_fijo = "2379933"
    email = "Dilydocs@gmail.com"
    direccion = "cra 80  50 - 52"
    municipio_residencia = request.args.get("municipio", "").strip() or "MEDELLIN"
    municipio_cod = request.args.get("municipio_cod", "").strip()
    municipio_cod = int(municipio_cod) if municipio_cod.isdigit() else 5001000
    departamento_cod = request.args.get("departamento_cod", "").strip()
    departamento_cod = int(departamento_cod) if departamento_cod.isdigit() else 5

    if not all([placa, identificacion, vigencias_raw, modelo, municipio_transito, apellidos_propietario]):
        return jsonify({"error": "Faltan parametros: placa, identificacion, vigencia, modelo, municipio_transito, apellidos_propietario"}), 400

    vigencias = [v.strip() for v in vigencias_raw.split(",") if v.strip()]

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Iniciando...", "procesando")

    def ejecutar():
        resultados = []
        try:
            conn = get_db_conn()
            cur = conn.cursor()
            cur.execute("SELECT * FROM vehiculos WHERE placa = %s", (placa,))
            fila = cur.fetchone()
            vehiculo = {}
            if fila:
                columnas = [desc[0] for desc in cur.description]
                vehiculo = dict(zip(columnas, fila))
            cur.close(); conn.close()

            for vigencia in vigencias:
                try:
                    cache = _cache_declaracion_buscar(placa, vigencia)

                    # Si ya se genero un PDF de Declaracion MANUAL hoy para
                    # esta misma placa/vigencia, se reutiliza directo (sin
                    # volver a generar nada) -- por eso se revisa esto
                    # primero, antes de cualquier otra cosa.
                    if cache and cache.get("url_manual"):
                        job_actualizar(job_id, f"Vigencia {vigencia}: ya existe un PDF generado hoy, usando ese...",
                                        datos_parciales=resultados)
                        resultados.append({"vigencia": vigencia, "ok": True, "url": cache["url_manual"]})
                        job_actualizar(job_id, f"Vigencia {vigencia} lista.", datos_parciales=resultados)
                        continue

                    # Si ya se genero una Declaracion Sugerida de esta
                    # misma placa/vigencia hoy (desde esta misma
                    # herramienta o desde el Generador de Declaraciones
                    # Sugeridas), se reutilizan esos datos en vez de
                    # volver a consultar en vivo a la Gobernacion (evita
                    # el captcha y es practicamente instantaneo).
                    if cache and cache.get("datos"):
                        job_actualizar(job_id, f"Vigencia {vigencia}: usando liquidación ya consultada hoy...",
                                        datos_parciales=resultados)
                        data_vig = cache["datos"]
                        caja_traccion = {"caja": data_vig.get("caja", ""), "traccion": data_vig.get("traccion", "")}
                        nombre_real = {"nombres": data_vig.get("nombres_reales", ""), "apellidos": data_vig.get("apellidos_reales", "")}
                    else:
                        job_actualizar(job_id, f"Vigencia {vigencia}: consultando en la Gobernación (puede tardar por el captcha)...",
                                        datos_parciales=resultados)
                        pdf_sugerida_bytes, data_vig = antioquia_generar_pdf_declaracion(
                            placa, identificacion, tipo_documento, vigencia,
                            modelo, municipio_transito, apellidos_propietario,
                            celular=celular, email=email, direccion=direccion,
                            municipio=municipio_residencia, municipio_cod=municipio_cod,
                            departamento_cod=departamento_cod
                        )
                        caja_traccion = _extraer_caja_traccion_declaracion(pdf_sugerida_bytes)
                        # Nombre y apellidos OFICIALES, tal como los tiene
                        # registrados la Gobernacion (leidos del mismo PDF
                        # de la Declaracion Sugerida) -- para que la
                        # Declaracion Manual quede con el mismo nombre
                        # exacto, sin depender de lo que se haya escrito
                        # a mano en Tramy.
                        nombre_real = _extraer_nombre_apellidos_declaracion(pdf_sugerida_bytes)

                        # Se guarda en cache SOLO si ya existia una entrada
                        # con PDF real generado antes (para no crear una
                        # fila sin URL que despues rompa al Generador de
                        # Declaraciones Sugeridas si busca un PDF que no
                        # existe).
                        if cache and cache.get("url"):
                            datos_extra = dict(data_vig or {})
                            datos_extra["caja"] = caja_traccion.get("caja", "")
                            datos_extra["traccion"] = caja_traccion.get("traccion", "")
                            datos_extra["nombres_reales"] = nombre_real.get("nombres", "")
                            datos_extra["apellidos_reales"] = nombre_real.get("apellidos", "")
                            _cache_declaracion_guardar(placa, vigencia, cache["url"], datos_extra=datos_extra)

                    # Si la extraccion del PDF no encontro nada (formato
                    # distinto, PDF fallo, etc.), se usa como respaldo lo
                    # que el usuario haya escrito a mano -- para no dejar
                    # el documento sin nombre en ese caso.
                    nombres_para_pdf = nombre_real.get("nombres") or nombres_propietario
                    apellidos_para_pdf = nombre_real.get("apellidos") or apellidos_propietario

                    datos = {
                        "vigencia": vigencia,
                        "nombre_completo": nombres_para_pdf,
                        "apellidos": apellidos_para_pdf,
                        "celular": celular,
                        "telefono": telefono_fijo,
                        "email": email,
                        "direccion": direccion,
                        "municipio_residencia": municipio_residencia,
                        "departamento_residencia": "ANTIOQUIA",
                        "numero_documento": identificacion,
                        "tipo_documento": tipo_documento,

                        "placa": placa,
                        "marca": vehiculo.get("marca", ""),
                        "linea": vehiculo.get("linea", ""),
                        "modelo": modelo,
                        "clase": vehiculo.get("clase", ""),
                        "carroceria": vehiculo.get("carroceria", ""),
                        "puertas": vehiculo.get("puertas", ""),
                        "cilindraje": vehiculo.get("cilindrada", ""),
                        "capacidad_carga": vehiculo.get("capacidad_carga", ""),
                        "capacidad_pasajeros": vehiculo.get("capacidad_pasajeros", ""),
                        "municipio_matricula": municipio_transito,
                        "departamento_matricula": "ANTIOQUIA",
                        "blindado": bool(vehiculo.get("info_blindaje")),
                        "importado": False,
                        "caja": caja_traccion.get("caja", ""),
                        "traccion": caja_traccion.get("traccion", ""),

                        # Liquidacion privada -- ver nota de mapeo mas abajo
                        "avaluo": data_vig.get("avaluoComercial", 0),
                        "impuesto": data_vig.get("impuesto", 0),
                        "sanciones": data_vig.get("sancion", 0),
                        "descuentos": data_vig.get("descuentoSancion", 0),
                        "total_cargo_5": data_vig.get("totalCargo", 0),
                        "total_cargo_6": data_vig.get("totalCargo", 0),
                        "intereses_mora": data_vig.get("interesesMora", 0),
                        "pagos_anteriores": data_vig.get("pagosAnteriores", 0),
                        "descuento_interes": data_vig.get("descuentoInteresesMora", 0),
                        "saldo_favor": data_vig.get("saldoFavor", 0),
                        # OJO: se usa "saldoPagar", NO "totalPagar" -- este
                        # ultimo incluye el costo de $25.900 del servicio
                        # de declaracion sugerida, que no aplica a la
                        # declaracion manual.
                        "total_pagar": data_vig.get("saldoPagar", 0),
                    }

                    job_actualizar(job_id, f"Vigencia {vigencia}: generando el documento...", datos_parciales=resultados)
                    id_unico = uuid.uuid4().hex[:8]
                    ruta_pdf = f"/tmp/decl_manual_{placa}_{vigencia}_{id_unico}.pdf"
                    generar_declaracion_manual_pdf(datos, ruta_pdf)

                    url = subir_a_r2(ruta_pdf, f"declaraciones-manuales/{placa}_{vigencia}_{id_unico}.pdf",
                                      nombre_descarga=f"Declaracion_Manual_{placa}_{vigencia}.pdf")
                    os.remove(ruta_pdf)

                    # Se guarda para poder saltar esta vigencia si se
                    # vuelve a pedir el mismo dia.
                    _cache_declaracion_guardar(placa, vigencia, None, url_manual=url)

                    resultados.append({"vigencia": vigencia, "ok": True, "url": url})
                except Exception as e_vig:
                    print(f"Error generando declaracion manual vigencia {vigencia} para {placa}: {e_vig}", flush=True)
                    resultados.append({"vigencia": vigencia, "ok": False, "error": str(e_vig)})

                job_actualizar(job_id, f"Vigencia {vigencia} lista.", datos_parciales=resultados)

            job_terminar(job_id, {"resultados": resultados})
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id, "estado": "procesando"})


@app.route("/generar-fun", methods=["POST"])
def generar_fun_endpoint():
    """Genera el FUN diligenciado con los datos recibidos, lo sube a R2, y
    devuelve el enlace de descarga. Recibe todo por POST (no se re-consulta
    la base de datos) porque el frontend ya tiene en pantalla justo los
    datos que el usuario confirmo -- incluyendo cosas que no vivimos en
    'vehiculos' (tramite elegido, datos del comprador, etc.)."""
    datos = request.get_json(silent=True) or {}
    placa = (datos.get("placa") or "SINPLACA").upper().strip()

    if not os.path.exists(FUN_PLANTILLA):
        return jsonify({"error": "No se encontro la plantilla AppJX.xlsm en el servidor."}), 500

    id_doc = str(uuid.uuid4())[:10]
    ruta_pdf_local = f"/tmp/FUN_{placa}_{id_doc}.pdf"

    try:
        generar_fun(datos, ruta_pdf_local)
        nombre_remoto = f"fun/{placa}_{id_doc}.pdf"
        url = subir_a_r2(ruta_pdf_local, nombre_remoto)
        os.remove(ruta_pdf_local)
        return jsonify({"ok": True, "url": url})
    except Exception as e:
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/personas-buscar", methods=["GET"])
def personas_buscar_endpoint():
    """Busca personas por nombre o numero de documento (coincidencia
    parcial) -- para el selector de Asesor/Propietario/Comprador al
    generar un documento. Sin parametro 'q', devuelve las mas recientes."""
    consulta = request.args.get("q", "").strip()
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if consulta:
            patron = f"%{consulta}%"
            cur.execute("""
                SELECT id, nombres, apellido, segundo_apellido, tipo_documento, numero_documento,
                       telefono, direccion, barrio_info, ciudad, email
                FROM personas
                WHERE numero_documento ILIKE %s OR nombres ILIKE %s OR apellido ILIKE %s
                ORDER BY actualizado_en DESC LIMIT 20
            """, (patron, patron, patron))
        else:
            cur.execute("""
                SELECT id, nombres, apellido, segundo_apellido, tipo_documento, numero_documento,
                       telefono, direccion, barrio_info, ciudad, email
                FROM personas ORDER BY actualizado_en DESC LIMIT 20
            """)
        filas = cur.fetchall()
        cur.close(); conn.close()
        personas = [{
            "id": f[0], "nombres": f[1], "apellido": f[2], "segundo_apellido": f[3],
            "tipo_documento": f[4], "numero_documento": f[5], "telefono": f[6],
            "direccion": f[7], "barrio_info": f[8], "ciudad": f[9], "email": f[10],
        } for f in filas]
        return jsonify({"ok": True, "personas": personas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/personas-guardar", methods=["POST"])
def personas_guardar_endpoint():
    """Crea o actualiza una persona -- si ya existe alguien con ese
    numero de documento, se actualizan sus datos (no se duplica)."""
    datos = request.get_json(silent=True) or {}
    numero_documento = (datos.get("numero_documento") or "").strip()
    nombres = (datos.get("nombres") or "").strip()
    if not numero_documento or not nombres:
        return jsonify({"ok": False, "error": "Faltan nombres o número de documento."}), 400

    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO personas (nombres, apellido, segundo_apellido, tipo_documento, numero_documento,
                                   telefono, direccion, barrio_info, ciudad, email, notas, actualizado_en)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (numero_documento) DO UPDATE SET
                nombres = EXCLUDED.nombres, apellido = EXCLUDED.apellido,
                segundo_apellido = EXCLUDED.segundo_apellido, tipo_documento = EXCLUDED.tipo_documento,
                telefono = CASE WHEN EXCLUDED.telefono <> '' THEN EXCLUDED.telefono ELSE personas.telefono END,
                direccion = CASE WHEN EXCLUDED.direccion <> '' THEN EXCLUDED.direccion ELSE personas.direccion END,
                barrio_info = CASE WHEN EXCLUDED.barrio_info <> '' THEN EXCLUDED.barrio_info ELSE personas.barrio_info END,
                ciudad = CASE WHEN EXCLUDED.ciudad <> '' THEN EXCLUDED.ciudad ELSE personas.ciudad END,
                email = CASE WHEN EXCLUDED.email <> '' THEN EXCLUDED.email ELSE personas.email END,
                notas = CASE WHEN EXCLUDED.notas <> '' THEN EXCLUDED.notas ELSE personas.notas END,
                actualizado_en = NOW()
            RETURNING id
        """, (
            nombres.upper(), (datos.get("apellido") or "").strip().upper(),
            (datos.get("segundo_apellido") or "").strip().upper(),
            (datos.get("tipo_documento") or "CC").strip(), numero_documento,
            (datos.get("telefono") or "").strip(), (datos.get("direccion") or "").strip(),
            (datos.get("barrio_info") or "").strip(), (datos.get("ciudad") or "").strip().upper(),
            (datos.get("email") or "").strip(), (datos.get("notas") or "").strip(),
        ))
        persona_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True, "id": persona_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/personas-listar", methods=["GET"])
def personas_listar_endpoint():
    """Lista personas paginadas, con busqueda opcional -- para la tabla
    de gestion (ver/eliminar) en el panel de configuracion."""
    consulta = request.args.get("q", "").strip()
    pagina = max(int(request.args.get("pagina", 1) or 1), 1)
    por_pagina = 30
    offset = (pagina - 1) * por_pagina
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if consulta:
            patron = f"%{consulta}%"
            cur.execute("""
                SELECT id, nombres, apellido, segundo_apellido, tipo_documento, numero_documento,
                       telefono, direccion, barrio_info, ciudad, email
                FROM personas
                WHERE numero_documento ILIKE %s OR nombres ILIKE %s OR apellido ILIKE %s
                ORDER BY nombres ASC LIMIT %s OFFSET %s
            """, (patron, patron, patron, por_pagina, offset))
            filas = cur.fetchall()
            cur.execute("""
                SELECT COUNT(*) FROM personas
                WHERE numero_documento ILIKE %s OR nombres ILIKE %s OR apellido ILIKE %s
            """, (patron, patron, patron))
            total = cur.fetchone()[0]
        else:
            cur.execute("""
                SELECT id, nombres, apellido, segundo_apellido, tipo_documento, numero_documento,
                       telefono, direccion, barrio_info, ciudad, email
                FROM personas ORDER BY nombres ASC LIMIT %s OFFSET %s
            """, (por_pagina, offset))
            filas = cur.fetchall()
            cur.execute("SELECT COUNT(*) FROM personas")
            total = cur.fetchone()[0]
        cur.close(); conn.close()
        personas = [{
            "id": f[0], "nombres": f[1], "apellido": f[2], "segundo_apellido": f[3],
            "tipo_documento": f[4], "numero_documento": f[5], "telefono": f[6],
            "direccion": f[7], "barrio_info": f[8], "ciudad": f[9], "email": f[10],
        } for f in filas]
        return jsonify({"ok": True, "personas": personas, "total": total, "pagina": pagina, "por_pagina": por_pagina})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/personas-eliminar", methods=["POST"])
def personas_eliminar_endpoint():
    """Elimina una persona por su id."""
    datos = request.get_json(silent=True) or {}
    persona_id = datos.get("id")
    if not persona_id:
        return jsonify({"ok": False, "error": "Falta el id de la persona."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM personas WHERE id = %s", (persona_id,))
        eliminada = cur.rowcount > 0
        conn.commit()
        cur.close(); conn.close()
        if not eliminada:
            return jsonify({"ok": False, "error": "No se encontró esa persona."}), 404
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/liquidaciones-guardar", methods=["POST"])
def liquidaciones_guardar_endpoint():
    """Guarda una liquidacion en el historial -- se llama justo cuando el
    usuario da clic en 'Enviar por WhatsApp', con fecha/hora automatica."""
    datos = request.get_json(silent=True) or {}
    placa = (datos.get("placa") or "").strip().upper()
    texto_whatsapp = datos.get("texto_whatsapp") or ""
    if not placa or not texto_whatsapp:
        return jsonify({"ok": False, "error": "Faltan placa o texto_whatsapp."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO liquidaciones_historial
                (placa, municipio, marca, linea, tipo_cliente, tramites, total, texto_whatsapp)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            placa, (datos.get("municipio") or "").strip(),
            (datos.get("marca") or "").strip(), (datos.get("linea") or "").strip(),
            (datos.get("tipo_cliente") or "").strip(), (datos.get("tramites") or "").strip(),
            datos.get("total") or 0, texto_whatsapp,
        ))
        liquidacion_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True, "id": liquidacion_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/liquidaciones-buscar", methods=["GET"])
def liquidaciones_buscar_endpoint():
    """Busca liquidaciones guardadas, principalmente por placa (tambien
    acepta buscar por municipio o tramite). Paginado de 20 en 20."""
    consulta = request.args.get("q", "").strip()
    pagina = max(int(request.args.get("pagina", 1) or 1), 1)
    por_pagina = 20
    offset = (pagina - 1) * por_pagina
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if consulta:
            patron = f"%{consulta}%"
            cur.execute("""
                SELECT id, placa, municipio, marca, linea, tipo_cliente, tramites, total, texto_whatsapp, creado_en
                FROM liquidaciones_historial
                WHERE placa ILIKE %s OR municipio ILIKE %s OR tramites ILIKE %s
                ORDER BY creado_en DESC LIMIT %s OFFSET %s
            """, (patron, patron, patron, por_pagina, offset))
            filas = cur.fetchall()
            cur.execute("""
                SELECT COUNT(*) FROM liquidaciones_historial
                WHERE placa ILIKE %s OR municipio ILIKE %s OR tramites ILIKE %s
            """, (patron, patron, patron))
            total_filas = cur.fetchone()[0]
        else:
            cur.execute("""
                SELECT id, placa, municipio, marca, linea, tipo_cliente, tramites, total, texto_whatsapp, creado_en
                FROM liquidaciones_historial ORDER BY creado_en DESC LIMIT %s OFFSET %s
            """, (por_pagina, offset))
            filas = cur.fetchall()
            cur.execute("SELECT COUNT(*) FROM liquidaciones_historial")
            total_filas = cur.fetchone()[0]
        cur.close(); conn.close()
        liquidaciones = [{
            "id": f[0], "placa": f[1], "municipio": f[2], "marca": f[3], "linea": f[4],
            "tipo_cliente": f[5], "tramites": f[6], "total": float(f[7]) if f[7] is not None else 0,
            "texto_whatsapp": f[8], "creado_en": f[9].isoformat() + "Z",
        } for f in filas]
        return jsonify({"ok": True, "liquidaciones": liquidaciones, "total": total_filas, "pagina": pagina, "por_pagina": por_pagina})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/appjx-documentos-listar", methods=["GET"])
def appjx_documentos_listar_endpoint():
    """Devuelve la lista de documentos disponibles de AppJX.xlsm (clave +
    nombre visible), para armar los botones en la pestaña Revision."""
    return jsonify({
        "ok": True,
        "documentos": [
            {"clave": clave, "nombre": nombre}
            for clave, (nombre, _hoja) in APPJX_DOCUMENTOS.items()
        ]
    })


@app.route("/appjx-generar-documento", methods=["POST"])
def appjx_generar_documento_endpoint():
    """Genera cualquiera de los documentos de AppJX.xlsm, llenando SOLO
    los datos del vehiculo -- recibe los datos por POST igual que
    /generar-fun, para usar justo lo que el usuario tiene en pantalla en
    ese momento (sin depender de que el vehiculo ya este guardado en la
    base de datos)."""
    datos = request.get_json(silent=True) or {}
    clave_documento = datos.get("documento", "")
    if clave_documento not in APPJX_DOCUMENTOS:
        return jsonify({"ok": False, "error": "Documento desconocido."}), 400

    placa = (datos.get("placa") or "SINPLACA").upper().strip()
    if not os.path.exists(FUN_PLANTILLA):
        return jsonify({"ok": False, "error": "No se encontró la plantilla AppJX.xlsm en el servidor."}), 500

    id_doc = str(uuid.uuid4())[:10]
    ruta_pdf_local = f"/tmp/APPJXDOC_{placa}_{id_doc}.pdf"

    try:
        generar_documento_vehiculo_appjx(clave_documento, datos, ruta_pdf_local)
        nombre_remoto = f"appjx-doc/{clave_documento}_{placa}_{id_doc}.pdf"
        url = subir_a_r2(ruta_pdf_local, nombre_remoto)
        os.remove(ruta_pdf_local)
        return jsonify({"ok": True, "url": url})
    except Exception as e:
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-solicitud-agregar", methods=["POST"])
def envigado_citas_solicitud_agregar_endpoint():
    """Agrega una nueva linea a la cola de citas pendientes por reservar
    automaticamente."""
    datos = request.get_json(silent=True) or {}
    campos_obligatorios = ["nombres", "apellidos", "numero_documento", "correo", "celular", "placa", "hora_aproximada"]
    faltantes = [c for c in campos_obligatorios if not datos.get(c)]
    if faltantes:
        return jsonify({"ok": False, "error": f"Faltan datos: {', '.join(faltantes)}"}), 400

    try:
        hora_aproximada = int(datos["hora_aproximada"])
        if not (0 <= hora_aproximada <= 23):
            raise ValueError()
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Hora aproximada invalida (debe ser 0-23)."}), 400

    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO envigado_citas_solicitudes
                (nombres, apellidos, tipo_documento, numero_documento, correo, celular, placa, id_servicio, sede_preferida, hora_aproximada)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            datos["nombres"].strip().upper(), datos["apellidos"].strip().upper(),
            datos.get("tipo_documento", "2").strip(), datos["numero_documento"].strip(),
            datos["correo"].strip(), datos["celular"].strip(), datos["placa"].strip().upper(),
            datos.get("id_servicio", "90").strip(), (datos.get("sede_preferida") or "").strip() or None,
            hora_aproximada,
        ))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True, "id": nuevo_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-solicitud-listar", methods=["GET"])
def envigado_citas_solicitud_listar_endpoint():
    """Devuelve todas las solicitudes de citas de la cola (pendientes,
    reservadas, y con error), mas recientes primero."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombres, apellidos, numero_documento, correo, celular, placa,
                   sede_preferida, hora_aproximada, estado, nro_atencion, error_mensaje,
                   creado_en, reservado_en
            FROM envigado_citas_solicitudes ORDER BY id DESC LIMIT 100
        """)
        filas = cur.fetchall()
        cur.close(); conn.close()
        solicitudes = []
        for f in filas:
            solicitudes.append({
                "id": f[0], "nombres": f[1], "apellidos": f[2], "numero_documento": f[3],
                "correo": f[4], "celular": f[5], "placa": f[6], "sede_preferida": f[7],
                "hora_aproximada": f[8], "estado": f[9], "nro_atencion": f[10], "error_mensaje": f[11],
                "creado_en": f[12].isoformat() if f[12] else None,
                "reservado_en": f[13].isoformat() if f[13] else None,
            })
        return jsonify({"ok": True, "solicitudes": solicitudes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-solicitud-eliminar", methods=["GET"])
def envigado_citas_solicitud_eliminar_endpoint():
    """Elimina/cancela una solicitud de la cola (ya sea porque ya no se
    necesita, o para quitar una que quedo con error)."""
    solicitud_id = request.args.get("id", "")
    if not solicitud_id.isdigit():
        return jsonify({"ok": False, "error": "ID invalido."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM envigado_citas_solicitudes WHERE id = %s", (int(solicitud_id),))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-solicitud-probar-ahora", methods=["GET"])
def envigado_citas_solicitud_probar_ahora_endpoint():
    """Fuerza el intento de reserva de UNA solicitud especifica de la
    cola AHORA MISMO, sin esperar a que el monitoreo automatico detecte
    citas primero -- util para probar/depurar el flujo de reserva sin
    depender de que aparezcan citas reales por casualidad. No cambia el
    estado en la base de datos (la solicitud sigue pendiente despues,
    sin importar el resultado) -- es solo para ver el diagnostico en los
    logs."""
    solicitud_id = request.args.get("id", "")
    if not solicitud_id.isdigit():
        return jsonify({"ok": False, "error": "ID invalido."}), 400

    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombres, apellidos, tipo_documento, numero_documento, correo, celular,
                   placa, id_servicio, sede_preferida, hora_aproximada
            FROM envigado_citas_solicitudes WHERE id = %s
        """, (int(solicitud_id),))
        fila = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    if not fila:
        return jsonify({"ok": False, "error": "No se encontró esa solicitud."}), 404

    solicitud = {
        "id": fila[0], "nombres": fila[1], "apellidos": fila[2], "tipo_documento": fila[3],
        "numero_documento": fila[4], "correo": fila[5], "celular": fila[6], "placa": fila[7],
        "id_servicio": fila[8], "sede_preferida": fila[9], "hora_aproximada": fila[10],
    }

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Probando el flujo de reserva (forzado, sin esperar deteccion)...", "procesando")

    def ejecutar():
        try:
            resultado = envigado_reservar_cita(solicitud)
            job_terminar(job_id, resultado)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/envigado-captura", methods=["GET"])
def envigado_captura_endpoint():
    """Sirve una captura de pantalla guardada por el flujo de reserva de
    citas de Envigado (para descargar o ver en el navegador)."""
    nombre_archivo = request.args.get("nombre", "")
    # Se valida que sea solo un nombre de archivo simple (sin rutas), para
    # que no se pueda pedir ningun otro archivo del servidor con esto.
    if not nombre_archivo or "/" in nombre_archivo or ".." in nombre_archivo:
        return jsonify({"ok": False, "error": "Nombre de archivo inválido."}), 400
    ruta_completa = os.path.join(CAPTURAS_ENVIGADO_DIR, nombre_archivo)
    if not os.path.isfile(ruta_completa):
        return jsonify({"ok": False, "error": "No se encontró esa captura (puede que el servidor se haya reiniciado desde entonces)."}), 404
    return send_file(ruta_completa, mimetype="image/png")


@app.route("/envigado-citas-disponibles", methods=["GET"])
def envigado_citas_disponibles_endpoint():
    """Revisa en vivo las dos sedes de Envigado (Vegas y City Plaza) para
    los proximos dias, y devuelve solo las fechas que SI tienen horarios
    disponibles."""
    dias = request.args.get("dias", "14")
    dias = int(dias) if dias.isdigit() else 14
    try:
        resultados, hubo_error = envigado_revisar_citas_disponibles(dias_adelante=dias)
        if hubo_error:
            return jsonify({
                "ok": False,
                "error": "No se pudo completar la consulta a Envigado (falló la conexión o algún paso previo). Revisa los logs del servidor para más detalle, e intenta de nuevo."
            }), 502
        con_citas = [r for r in resultados if r["cantidad_horarios"] > 0]
        return jsonify({
            "ok": True,
            "hay_citas": len(con_citas) > 0,
            "disponibles": con_citas,
            "verificado_en": datetime.now().isoformat() + "Z"  # UTC -- el navegador lo convierte solo a hora local
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/push-vapid-public-key", methods=["GET"])
def push_vapid_public_key_endpoint():
    """Devuelve la llave publica VAPID -- esta es segura de compartir
    (a diferencia de la privada), el navegador la necesita para crear
    la suscripcion push."""
    return jsonify({"publicKey": VAPID_PUBLIC_KEY})


@app.route("/push-subscribe", methods=["POST"])
def push_subscribe_endpoint():
    """Guarda una suscripcion push nueva (mandada desde el navegador
    despues de que la persona acepta recibir notificaciones)."""
    datos = request.get_json(silent=True) or {}
    endpoint = datos.get("endpoint", "")
    keys = datos.get("keys", {})
    p256dh = keys.get("p256dh", "")
    auth = keys.get("auth", "")
    if not endpoint or not p256dh or not auth:
        return jsonify({"ok": False, "error": "Suscripcion incompleta."}), 400
    ok = guardar_suscripcion_push(endpoint, p256dh, auth)
    return jsonify({"ok": ok})


@app.route("/medellin-citas-disponibles", methods=["GET"])
def medellin_citas_disponibles_endpoint():
    """Revisa si hay citas disponibles en el portal de Medellin para el
    servicio indicado (Traspaso por defecto). Requiere las credenciales
    de una cuenta ya registrada y activa en el portal."""
    usuario = request.args.get("usuario", "").strip()
    password = request.args.get("password", "").strip()
    placa = request.args.get("placa", "").upper().strip()
    id_servicio = request.args.get("id_servicio", MEDELLIN_SERVICIO_TRASPASO).strip()
    sede_deseada = request.args.get("sede", "").strip() or None

    if not all([usuario, password, placa]):
        return jsonify({"error": "Faltan datos: usuario, password, placa"}), 400

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Iniciando sesión en el portal de Medellín...", "procesando")

    def ejecutar():
        try:
            hay_citas, detalle = medellin_hay_citas_disponibles(usuario, password, placa, id_servicio, sede_deseada=sede_deseada)
            job_terminar(job_id, {"hay_citas": hay_citas, "detalle": detalle})
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/medellin-crear-usuario", methods=["GET"])
def medellin_crear_usuario_endpoint():
    """Inicia la creacion de un usuario en el portal de Medellin. Como es
    un proceso que puede tardar (carga de pagina + llenado), se ejecuta
    en segundo plano con el mismo patron de job que las demas consultas
    que usan Playwright."""
    datos = {
        "tipo_sociedad": request.args.get("tipo_sociedad", "").strip(),
        "tipo_identificacion": request.args.get("tipo_identificacion", "").strip(),
        "numero_identificacion": request.args.get("numero_identificacion", "").strip(),
        "nombre": request.args.get("nombre", "").strip(),
        "apellidos": request.args.get("apellidos", "").strip(),
        "genero": request.args.get("genero", "").strip(),
        "email": request.args.get("email", "").strip(),
        "direccion": request.args.get("direccion", "").strip(),
        "telefono": request.args.get("telefono", "").strip(),
    }
    faltantes = [k for k, v in datos.items() if not v]
    if faltantes:
        return jsonify({"error": f"Faltan datos: {', '.join(faltantes)}"}), 400

    # Estos dos campos son opcionales (dependen del perfil), asi que se
    # agregan DESPUES de la revision de campos obligatorios de arriba.
    datos["perfil_usuario"] = request.args.get("perfil_usuario", "").strip().lower()  # "propietario" o "comprador"
    datos["placa_propietario"] = request.args.get("placa_propietario", "").strip().upper()

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Iniciando registro en el portal de Medellín...", "procesando")

    def ejecutar():
        try:
            resultado = medellin_crear_usuario(datos)
            job_terminar(job_id, resultado)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/medellin-activar-cuenta", methods=["GET"])
def medellin_activar_cuenta_endpoint():
    """Segundo paso del registro: lee el correo de activacion que manda
    la Alcaldia de Medellin (con usuario/contraseña temporales), inicia
    sesion con esas credenciales, y deja la cuenta lista para usar.

    Requiere que el correo YA haya llegado (o llegue dentro de los
    minutos de espera) -- normalmente tarda unos minutos despues de
    crear el usuario con /medellin-crear-usuario."""
    cedula = request.args.get("cedula", "").strip()
    email_cuenta = request.args.get("email_cuenta", "").strip()
    password_app_email = request.args.get("password_app_email", "").strip()
    nueva_password = request.args.get("nueva_password", "").strip()

    faltantes = [k for k, v in {
        "cedula": cedula, "email_cuenta": email_cuenta,
        "password_app_email": password_app_email, "nueva_password": nueva_password,
    }.items() if not v]
    if faltantes:
        return jsonify({"error": f"Faltan datos: {', '.join(faltantes)}"}), 400

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Buscando el correo de activación...", "procesando")

    def ejecutar():
        try:
            etiqueta = f"[MEDELLIN-ACTIVAR-{job_id[:6]}]"
            usuario_temp, password_temp = medellin_leer_credenciales_temporales(
                email_cuenta, password_app_email, cedula, minutos_maximo_espera=5, etiqueta=etiqueta
            )
            if not usuario_temp:
                job_error(job_id, "No se encontró el correo de activación a tiempo (esperó 5 minutos).")
                return
            job_actualizar(job_id, "Correo encontrado, iniciando sesión con credenciales temporales...", "procesando")
            resultado = medellin_activar_cuenta(usuario_temp, password_temp, nueva_password, usar_proxy=True)
            job_terminar(job_id, resultado)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/diagnostico-proxy-iproyal", methods=["GET"])
def diagnostico_proxy_iproyal_endpoint():
    """Endpoint TEMPORAL de diagnostico -- corre exactamente el comando
    curl que pidio el soporte de IPRoyal, directo desde el servidor de
    Railway, para descartar si el problema es de Playwright/Chromium o
    de la conexion misma entre Railway y el proxy."""
    if not (IPROYAL_USER and IPROYAL_PASS):
        return jsonify({"error": "Faltan las credenciales de IPRoyal en las variables de entorno."}), 400

    proxy_url = f"http://{IPROYAL_USER}:{IPROYAL_PASS}@{IPROYAL_HOST}:{IPROYAL_PORT}"
    try:
        resultado = subprocess.run(
            ["curl", "-v", "--proxy", proxy_url, "--max-time", "30",
             "https://www.medellin.gov.co/portal-movilidad/index.html"],
            capture_output=True, text=True, timeout=35
        )
        return jsonify({
            "ok": True,
            "codigo_salida": resultado.returncode,
            "salida_estandar": resultado.stdout[-2000:],
            "salida_error_detallada": resultado.stderr[-4000:],
        })
    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "error": "Timeout -- el comando tardo mas de 35 segundos sin responder."}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/diagnostico-proxy-dataimpulse", methods=["GET"])
def diagnostico_proxy_dataimpulse_endpoint():
    """Endpoint de diagnostico -- prueba la conexion real del proxy de
    DataImpulse contra el sitio de citas de Envigado, directo desde el
    servidor de Railway (sin pasar por Playwright), para confirmar que
    las credenciales funcionan y que el sitio .gov no esta bloqueado
    antes de depender del proxy en el flujo completo de citas."""
    if not (DATAIMPULSE_USER and DATAIMPULSE_PASS):
        return jsonify({"error": "Faltan las credenciales de DataImpulse en las variables de entorno (DATAIMPULSE_USER/DATAIMPULSE_PASS)."}), 400

    proxy_url = f"http://{DATAIMPULSE_USER}:{DATAIMPULSE_PASS}@{DATAIMPULSE_HOST}:{DATAIMPULSE_PORT}"
    try:
        resultado = subprocess.run(
            ["curl", "-v", "--proxy", proxy_url, "--max-time", "30",
             "https://movilidad.envigado.gov.co/portal-servicios/#/agendar-cita-publica"],
            capture_output=True, text=True, timeout=35
        )
        return jsonify({
            "ok": True,
            "codigo_salida": resultado.returncode,
            "salida_estandar": resultado.stdout[-2000:],
            "salida_error_detallada": resultado.stderr[-4000:],
        })
    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "error": "Timeout -- el comando tardo mas de 35 segundos sin responder."}), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/medellin-activar-cuenta-directo", methods=["GET"])
def medellin_activar_cuenta_directo_endpoint():
    """Version simplificada de /medellin-activar-cuenta -- en vez de leer
    el correo automaticamente, recibe el usuario y contraseña temporal
    directamente (el usuario los copia del correo y los manda aqui a
    mano). Util mientras la lectura automatica de correo no este
    disponible (ej. si la cuenta de correo quedo bloqueada)."""
    usuario_temporal = request.args.get("usuario_temporal", "").strip()
    password_temporal = request.args.get("password_temporal", "").strip()
    nueva_password = request.args.get("nueva_password", "").strip()

    faltantes = [k for k, v in {
        "usuario_temporal": usuario_temporal, "password_temporal": password_temporal,
        "nueva_password": nueva_password,
    }.items() if not v]
    if faltantes:
        return jsonify({"error": f"Faltan datos: {', '.join(faltantes)}"}), 400

    job_id = str(uuid.uuid4())
    job_actualizar(job_id, "Iniciando sesión con credenciales temporales...", "procesando")

    def ejecutar():
        try:
            resultado = medellin_activar_cuenta(usuario_temporal, password_temporal, nueva_password, usar_proxy=True)
            job_terminar(job_id, resultado)
        except Exception as e:
            import traceback
            print(traceback.format_exc(), flush=True)
            job_error(job_id, str(e))

    threading.Thread(target=ejecutar, daemon=True).start()
    return jsonify({"job_id": job_id})


@app.route("/envigado-citas-iniciar-monitoreo", methods=["GET"])
def envigado_citas_iniciar_monitoreo_endpoint():
    """Arranca una sesion de monitoreo CONSTANTE de citas disponibles en
    Envigado -- revisa cada 30 segundos, durante maximo 2 horas (o hasta
    que se detenga manualmente). Util para la ventana de tiempo del dia
    en la que se sabe que pueden abrir citas nuevas."""
    if _envigado_citas_monitoreo_estado["activo"]:
        return jsonify({
            "ok": False,
            "error": "Ya hay una sesión de monitoreo de citas corriendo.",
            "fin_esperado": _envigado_citas_monitoreo_estado["fin_esperado"]
        }), 409

    duracion_minutos = request.args.get("minutos", "120")
    duracion_minutos = int(duracion_minutos) if duracion_minutos.isdigit() else 120
    duracion_minutos = min(duracion_minutos, 120)  # tope maximo de 2 horas
    duracion_segundos = duracion_minutos * 60

    _envigado_citas_monitoreo_estado["activo"] = True
    _envigado_citas_monitoreo_estado["inicio"] = datetime.now().isoformat() + "Z"  # UTC
    _envigado_citas_monitoreo_estado["fin_esperado"] = (datetime.now() + timedelta(seconds=duracion_segundos)).isoformat() + "Z"  # UTC
    _envigado_citas_monitoreo_estado["detener"] = False

    threading.Thread(
        target=_envigado_polling_citas,
        kwargs={"duracion_segundos": duracion_segundos},
        daemon=True
    ).start()

    return jsonify({
        "ok": True,
        "mensaje": f"Monitoreo de citas iniciado, revisando cada 30 segundos por {duracion_minutos} minutos (o hasta que lo detengas).",
        "fin_esperado": _envigado_citas_monitoreo_estado["fin_esperado"]
    })


@app.route("/envigado-citas-detener-monitoreo", methods=["GET"])
def envigado_citas_detener_monitoreo_endpoint():
    """Detiene la sesion de monitoreo constante de citas antes de que se
    cumpla el tiempo maximo. Puede tardar hasta 30 segundos en detenerse
    del todo (revisa la bandera en cada ciclo)."""
    if not _envigado_citas_monitoreo_estado["activo"]:
        return jsonify({"ok": False, "error": "No hay ningún monitoreo de citas corriendo en este momento."}), 409
    _envigado_citas_monitoreo_estado["detener"] = True
    return jsonify({"ok": True, "mensaje": "Deteniendo el monitoreo de citas..."})


@app.route("/envigado-citas-estado-monitoreo", methods=["GET"])
def envigado_citas_estado_monitoreo_endpoint():
    """Indica si hay una sesion de monitoreo constante de citas activa."""
    return jsonify({
        "ok": True,
        "activo": _envigado_citas_monitoreo_estado["activo"],
        "inicio": _envigado_citas_monitoreo_estado["inicio"],
        "fin_esperado": _envigado_citas_monitoreo_estado["fin_esperado"],
    })


@app.route("/medellin-citas-iniciar-monitoreo", methods=["GET"])
def medellin_citas_iniciar_monitoreo_endpoint():
    """Arranca una sesion de monitoreo CONSTANTE de citas disponibles en
    Medellin -- revisa cada 30 segundos, durante maximo 2 horas (o hasta
    que se detenga manualmente)."""
    if _medellin_citas_monitoreo_estado["activo"]:
        return jsonify({
            "ok": False,
            "error": "Ya hay una sesión de monitoreo de citas de Medellín corriendo.",
            "fin_esperado": _medellin_citas_monitoreo_estado["fin_esperado"]
        }), 409

    usuario = request.args.get("usuario", "").strip()
    password = request.args.get("password", "").strip()
    placa = request.args.get("placa", "").upper().strip()
    id_servicio = request.args.get("id_servicio", MEDELLIN_SERVICIO_TRASPASO).strip()
    sede_deseada = request.args.get("sede", "").strip() or None
    if not all([usuario, password, placa]):
        return jsonify({"error": "Faltan datos: usuario, password, placa"}), 400

    duracion_minutos = request.args.get("minutos", "120")
    duracion_minutos = int(duracion_minutos) if duracion_minutos.isdigit() else 120
    duracion_minutos = min(duracion_minutos, 120)  # tope maximo de 2 horas
    duracion_segundos = duracion_minutos * 60

    _medellin_citas_monitoreo_estado["activo"] = True
    _medellin_citas_monitoreo_estado["inicio"] = datetime.now().isoformat() + "Z"  # UTC
    _medellin_citas_monitoreo_estado["fin_esperado"] = (datetime.now() + timedelta(seconds=duracion_segundos)).isoformat() + "Z"  # UTC
    _medellin_citas_monitoreo_estado["detener"] = False
    _medellin_citas_monitoreo_estado["ultimo_hallazgo"] = None  # se limpia cualquier hallazgo viejo de una sesion anterior

    threading.Thread(
        target=_medellin_polling_citas,
        args=(usuario, password, placa, id_servicio, duracion_segundos),
        kwargs={"sede_deseada": sede_deseada},
        daemon=True
    ).start()

    return jsonify({
        "ok": True,
        "mensaje": f"Monitoreo de citas de Medellín iniciado, revisando cada 30 segundos por {duracion_minutos} minutos (o hasta que lo detengas).",
        "fin_esperado": _medellin_citas_monitoreo_estado["fin_esperado"]
    })


@app.route("/medellin-citas-detener-monitoreo", methods=["GET"])
def medellin_citas_detener_monitoreo_endpoint():
    """Detiene la sesion de monitoreo constante de citas de Medellin
    antes de que se cumpla el tiempo maximo."""
    if not _medellin_citas_monitoreo_estado["activo"]:
        return jsonify({"ok": False, "error": "No hay ningún monitoreo de citas de Medellín corriendo en este momento."}), 409
    _medellin_citas_monitoreo_estado["detener"] = True
    return jsonify({"ok": True, "mensaje": "Deteniendo el monitoreo de citas de Medellín..."})


@app.route("/monitoreo-config", methods=["GET"])
def monitoreo_config_obtener_endpoint():
    """Devuelve la configuracion actual de los monitores automaticos
    (activo, intervalo, horario, y para Medellin ademas si ya tiene
    credenciales guardadas -- NUNCA se devuelve la contraseña real por
    este endpoint, solo si esta configurada o no). Tambien devuelve el
    estado EN VIVO (ultima revision real hecha, si esta dentro del
    horario ahora mismo), para que el panel pueda confirmar que de
    verdad esta corriendo, no solo que esta "activado"."""
    envigado = _monitoreo_config_leer("envigado_citas") or {}
    medellin = _monitoreo_config_leer("medellin_citas") or {}
    medellin_proxy = _monitoreo_config_leer("medellin_citas_proxy") or {}
    return jsonify({
        "ok": True,
        "envigado_citas": {
            "activo": envigado.get("activo", False),
            "intervalo_segundos": envigado.get("intervalo_segundos", 30),
            "hora_inicio": envigado.get("hora_inicio", "11:00"),
            "hora_fin": envigado.get("hora_fin", "16:00"),
            "dentro_de_horario": _programador_automatico_estado["envigado_citas"]["dentro_de_horario"],
            "ultima_revision": _programador_automatico_estado["envigado_citas"]["ultima_revision"],
            "ultimo_error": _programador_automatico_estado["envigado_citas"]["ultimo_error"],
        },
        "medellin_citas": {
            "activo": medellin.get("activo", False),
            "intervalo_segundos": medellin.get("intervalo_segundos", 60),
            "hora_inicio": medellin.get("hora_inicio", "07:00"),
            "hora_fin": medellin.get("hora_fin", "17:00"),
            "usuario": medellin.get("usuario") or "",
            "placa": medellin.get("placa") or "",
            "sede": medellin.get("sede") or "",
            "tiene_password": bool(medellin.get("password")),
            "dentro_de_horario": _programador_automatico_estado["medellin_citas"]["dentro_de_horario"],
            "ultima_revision": _programador_automatico_estado["medellin_citas"]["ultima_revision"],
            "ultimo_error": _programador_automatico_estado["medellin_citas"]["ultimo_error"],
        },
        "medellin_citas_proxy": {
            "activo": medellin_proxy.get("activo", False),
            "intervalo_segundos": medellin_proxy.get("intervalo_segundos", 30),
            "hora_inicio": medellin_proxy.get("hora_inicio", "07:00"),
            "hora_fin": medellin_proxy.get("hora_fin", "17:00"),
            "usuario": medellin_proxy.get("usuario") or "",
            "placa": medellin_proxy.get("placa") or "",
            "sede": medellin_proxy.get("sede") or "",
            "tiene_password": bool(medellin_proxy.get("password")),
            "dentro_de_horario": _programador_automatico_estado["medellin_citas_proxy"]["dentro_de_horario"],
            "ultima_revision": _programador_automatico_estado["medellin_citas_proxy"]["ultima_revision"],
            "ultimo_error": _programador_automatico_estado["medellin_citas_proxy"]["ultimo_error"],
        },
    })


@app.route("/monitoreo-config", methods=["POST"])
def monitoreo_config_guardar_endpoint():
    """Actualiza la configuracion de uno de los monitores automaticos
    desde el panel de Ejecucion. Se manda solo el/los campos que
    cambiaron -- los demas quedan como estaban."""
    datos = request.get_json(silent=True) or {}
    monitor = datos.get("monitor", "")
    if monitor not in ("envigado_citas", "medellin_citas", "medellin_citas_proxy"):
        return jsonify({"ok": False, "error": "Monitor inválido."}), 400

    campos = {}
    if "activo" in datos:
        campos["activo"] = bool(datos["activo"])
    if "intervalo_segundos" in datos:
        try:
            campos["intervalo_segundos"] = max(10, int(datos["intervalo_segundos"]))  # minimo 10s, por seguridad
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "Intervalo inválido."}), 400
    if "hora_inicio" in datos:
        campos["hora_inicio"] = datos["hora_inicio"]
    if "hora_fin" in datos:
        campos["hora_fin"] = datos["hora_fin"]
    if monitor in ("medellin_citas", "medellin_citas_proxy"):
        if "usuario" in datos:
            campos["usuario"] = datos["usuario"]
        if "password" in datos and datos["password"]:  # solo se actualiza si mandaron una nueva, nunca se borra sola
            campos["password"] = datos["password"]
        if "placa" in datos:
            campos["placa"] = (datos["placa"] or "").upper().strip()
        if "sede" in datos:
            campos["sede"] = (datos["sede"] or "").strip()

    ok = _monitoreo_config_guardar(monitor, **campos)
    return jsonify({"ok": ok})


@app.route("/medellin-citas-estado-monitoreo", methods=["GET"])
def medellin_citas_estado_monitoreo_endpoint():
    """Indica si hay una sesion de monitoreo constante de citas de
    Medellin activa, y el ultimo hallazgo (si hay alguno) guardado
    durante esta sesion."""
    return jsonify({
        "ok": True,
        "activo": _medellin_citas_monitoreo_estado["activo"],
        "inicio": _medellin_citas_monitoreo_estado["inicio"],
        "fin_esperado": _medellin_citas_monitoreo_estado["fin_esperado"],
        "ultimo_hallazgo": _medellin_citas_monitoreo_estado["ultimo_hallazgo"],
    })


@app.route("/medellin-citas-proxy-ultimo-hallazgo", methods=["GET"])
def medellin_citas_proxy_ultimo_hallazgo_endpoint():
    """Igual que el de arriba, pero para el monitor ESPEJO (el que
    siempre usa el proxy de IPRoyal) -- estado completamente aparte,
    para que el aviso sonoro/push de cada uno funcione de forma
    independiente."""
    return jsonify({
        "ok": True,
        "ultimo_hallazgo": _medellin_citas_proxy_monitoreo_estado["ultimo_hallazgo"],
    })


@app.route("/medellin-citas-resetear-aviso", methods=["GET"])
def medellin_citas_resetear_aviso_endpoint():
    """Olvida el ultimo hallazgo guardado (tanto del monitoreo manual
    como del programador automatico, que comparten el mismo estado) --
    asi, si la MISMA disponibilidad sigue ahi (nunca desaparecio), la
    proxima revision lo vuelve a contar como 'nuevo' y manda el aviso
    de nuevo (sonido/vibracion/push), en vez de quedarse callado hasta
    que la disponibilidad desaparezca y reaparezca sola."""
    _medellin_citas_monitoreo_estado["ultimo_hallazgo"] = None
    return jsonify({"ok": True, "mensaje": "Aviso reiniciado -- la próxima vez que se detecte disponibilidad, se avisará de nuevo."})


@app.route("/medellin-citas-proxy-resetear-aviso", methods=["GET"])
def medellin_citas_proxy_resetear_aviso_endpoint():
    """Igual que el de arriba, pero para el monitor ESPEJO (proxy)."""
    _medellin_citas_proxy_monitoreo_estado["ultimo_hallazgo"] = None
    return jsonify({"ok": True, "mensaje": "Aviso reiniciado -- la próxima vez que se detecte disponibilidad, se avisará de nuevo."})


@app.route("/envigado-citas-ultimo-resultado", methods=["GET"])
def envigado_citas_ultimo_resultado_endpoint():
    """Devuelve el ultimo resultado GUARDADO (sin volver a consultar la
    API en vivo) -- para el aviso rapido que se muestra al entrar a
    Liquidacion, sin tener que esperar una consulta en vivo cada vez."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT sede, fecha_dia, cantidad_horarios, verificado_en
            FROM envigado_citas_disponibles
            WHERE cantidad_horarios > 0
            ORDER BY verificado_en DESC
        """)
        filas = cur.fetchall()
        cur.close(); conn.close()
        # Se descartan aqui los registros cuya fecha de cita YA PASO --
        # antes se mostraban indefinidamente porque la limpieza automatica
        # solo borra los resultados vacios del DIA EN QUE SE CONSULTO, no
        # los positivos de dias anteriores que quedaron sin revisar de
        # nuevo (ej. si el monitoreo se detuvo antes de volver a
        # consultar esa fecha).
        hoy = datetime.now().date()
        disponibles = []
        for sede, fecha_dia, cantidad_horarios, verificado_en in filas:
            try:
                fecha_cita = datetime.strptime(fecha_dia, "%d/%m/%Y").date()
                if fecha_cita < hoy:
                    continue  # la fecha de la cita ya paso -- se ignora
            except (ValueError, TypeError):
                pass  # si no se puede interpretar la fecha, se muestra igual (mejor prevenir que ocultar por error)
            disponibles.append({
                "sede": sede, "fecha": fecha_dia, "cantidad_horarios": cantidad_horarios,
                "verificado_en": verificado_en.isoformat() + "Z"
            })
        return jsonify({"ok": True, "hay_citas": len(disponibles) > 0, "disponibles": disponibles})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-turnos-iniciar-monitoreo", methods=["GET"])
def envigado_turnos_iniciar_monitoreo_endpoint():
    """Arranca (o programa) una sesion de monitoreo de maximo 2 horas
    (configurable), que revisa el monitor de turnos de Envigado cada
    pocos segundos y va guardando cada llamado nuevo que aparezca -- se
    puede detener antes manualmente. Los DATOS que va capturando (toda
    la lista, y los que coincidan con los numeros vigilados) quedan
    guardados en la base de datos todo el dia, sin importar si la sesion
    de vigilancia ya termino o si se inicia una sesion nueva despues --
    solo se renuevan al dia siguiente. Se puede vigilar hasta 20 numeros
    de cita a la vez.
    Recibe 'citas' como JSON: [{"numero": "C-89", "placa": "ABC123",
    "hora": "14:30", "fecha": "2026-08-25"}, ...] -- placa/hora/fecha son
    opcionales. Si ALGUNA cita trae hora+fecha, el monitoreo se PROGRAMA
    para arrancar 5 minutos antes de la hora mas temprana, y terminar 1
    hora despues de la hora mas tardia (en vez de arrancar de inmediato
    con la duracion fija de 'minutos')."""
    if _envigado_monitoreo_estado["activo"]:
        return jsonify({
            "ok": False,
            "error": "Ya hay una sesión de monitoreo corriendo.",
            "fin_esperado": _envigado_monitoreo_estado["fin_esperado"]
        }), 409

    try:
        citas = json.loads(request.args.get("citas", "[]"))
    except Exception:
        citas = []
    citas = citas[:20]

    numeros_vigilados = [c["numero"].strip().upper() for c in citas if c.get("numero")]
    placas_por_numero = {
        c["numero"].strip().upper(): c["placa"].strip().upper()
        for c in citas if c.get("numero") and c.get("placa")
    }

    # Se guarda cada cita ingresada en el historial -- independiente de
    # si se llega a detectar o no, para poder revisar despues que se
    # dejo vigilando cada dia (no todos los turnos que paso el monitor,
    # solo lo que el usuario pidio vigilar).
    if citas:
        try:
            conn_hist = get_db_conn()
            cur_hist = conn_hist.cursor()
            for c in citas:
                if not c.get("numero"):
                    continue
                cur_hist.execute("""
                    INSERT INTO envigado_citas_vigiladas_historial (numero, placa, hora_cita, fecha_cita)
                    VALUES (%s, %s, %s, %s)
                """, (
                    c["numero"].strip().upper(),
                    (c.get("placa") or "").strip().upper() or None,
                    (c.get("hora") or "").strip() or None,
                    c.get("fecha") or datetime.now(TZ_COLOMBIA).date().isoformat(),
                ))
            conn_hist.commit()
            cur_hist.close(); conn_hist.close()
        except Exception as e:
            print(f"Error guardando historial de citas vigiladas: {e}", flush=True)

    # Si alguna cita trae hora+fecha, se programa el inicio/fin segun eso
    # -- si no, se usa el comportamiento de siempre (duracion fija,
    # arranca de inmediato). Las horas que escribe el usuario son SIEMPRE
    # hora de Colombia (asi se les asigna explicitamente esa zona), sin
    # importar en que zona horaria corra el servidor.
    horas_programadas = []
    for c in citas:
        if c.get("hora") and c.get("fecha"):
            try:
                dt_naive = datetime.strptime(f"{c['fecha']} {c['hora']}", "%Y-%m-%d %H:%M")
                # La oficina solo atiende de 7am a 5pm -- se ignora
                # cualquier hora fuera de ese rango (el frontend ya lo
                # restringe con min/max, pero se valida aqui tambien por
                # si la peticion viene de otro lado).
                hora_valida = 7 <= dt_naive.hour <= 17 and not (dt_naive.hour == 17 and dt_naive.minute > 0)
                if hora_valida:
                    horas_programadas.append(dt_naive.replace(tzinfo=TZ_COLOMBIA))
            except Exception:
                pass

    ahora = datetime.now(TZ_COLOMBIA)
    if horas_programadas:
        inicio_deseado = min(horas_programadas) - timedelta(minutes=5)
        fin_deseado = max(horas_programadas) + timedelta(hours=1)
        espera_segundos = max(0, (inicio_deseado - ahora).total_seconds())
        duracion_segundos = max(60, (fin_deseado - max(ahora, inicio_deseado)).total_seconds())
        duracion_segundos = min(duracion_segundos, 6 * 3600)  # tope de seguridad: 6 horas totales de monitoreo
        inicio_real = max(ahora, inicio_deseado)
        fin_esperado_dt = inicio_real + timedelta(seconds=duracion_segundos)
        mensaje = (
            f"Monitoreo programado para iniciar a las {inicio_deseado.strftime('%H:%M')} "
            f"y terminar a las {fin_esperado_dt.strftime('%H:%M')} (hora Colombia)."
            if espera_segundos > 0 else
            f"Monitoreo iniciado -- corriendo hasta las {fin_esperado_dt.strftime('%H:%M')} (hora Colombia)."
        )
    else:
        duracion_minutos = request.args.get("minutos", "120")
        duracion_minutos = int(duracion_minutos) if duracion_minutos.isdigit() else 120
        duracion_minutos = min(duracion_minutos, 120)  # tope maximo de 2 horas
        duracion_segundos = duracion_minutos * 60
        espera_segundos = 0
        fin_esperado_dt = ahora + timedelta(seconds=duracion_segundos)
        mensaje = f"Monitoreo iniciado por {duracion_minutos} minutos (o hasta que lo detengas)."

    _envigado_monitoreo_estado["activo"] = True
    _envigado_monitoreo_estado["inicio"] = (ahora + timedelta(seconds=espera_segundos)).astimezone(timezone.utc).isoformat()
    _envigado_monitoreo_estado["fin_esperado"] = fin_esperado_dt.astimezone(timezone.utc).isoformat()
    _envigado_monitoreo_estado["numeros_vigilados"] = numeros_vigilados
    _envigado_monitoreo_estado["detener"] = False

    threading.Thread(
        target=_envigado_polling_turnos_con_espera,
        kwargs={
            "espera_segundos": espera_segundos, "duracion_segundos": duracion_segundos,
            "numeros_vigilados": numeros_vigilados, "placas_por_numero": placas_por_numero,
        },
        daemon=True
    ).start()

    return jsonify({
        "ok": True,
        "mensaje": mensaje,
        "programado": espera_segundos > 0,
        "inicio_esperado": _envigado_monitoreo_estado["inicio"],
        "fin_esperado": _envigado_monitoreo_estado["fin_esperado"]
    })


@app.route("/envigado-turnos-detener-monitoreo", methods=["GET"])
def envigado_turnos_detener_monitoreo_endpoint():
    """Detiene la sesion de monitoreo activa antes de que se cumplan las
    2 horas. El hilo revisa esta bandera en cada ciclo (cada 8 segundos),
    asi que puede tardar hasta ese tiempo en detenerse del todo."""
    if not _envigado_monitoreo_estado["activo"]:
        return jsonify({"ok": False, "error": "No hay ningún monitoreo corriendo en este momento."}), 409
    _envigado_monitoreo_estado["detener"] = True
    return jsonify({"ok": True, "mensaje": "Deteniendo el monitoreo..."})


@app.route("/envigado-turnos-estado-monitoreo", methods=["GET"])
def envigado_turnos_estado_monitoreo_endpoint():
    """Indica si hay una sesion de monitoreo activa en este momento."""
    return jsonify({
        "ok": True,
        "activo": _envigado_monitoreo_estado["activo"],
        "inicio": _envigado_monitoreo_estado["inicio"],
        "fin_esperado": _envigado_monitoreo_estado["fin_esperado"],
        "numeros_vigilados": _envigado_monitoreo_estado["numeros_vigilados"],
    })


@app.route("/envigado-turnos-capturados", methods=["GET"])
def envigado_turnos_capturados_endpoint():
    """Devuelve los turnos capturados de un dia en particular, mas
    recientes primero. Por defecto (sin 'fecha') muestra los de hoy --
    'fecha' se recibe en formato YYYY-MM-DD, para poder revisar el
    historial de dias anteriores."""
    limite = request.args.get("limite", "100")
    limite = int(limite) if limite.isdigit() else 100
    fecha = request.args.get("fecha", "").strip()
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if fecha:
            cur.execute("""
                SELECT nro_atencion, nombre_usuario, nombre_taquilla, nombre_servicio, detectado_en, placa
                FROM envigado_turnos_llamados
                WHERE detectado_en::date = %s
                ORDER BY detectado_en DESC
                LIMIT %s
            """, (fecha, limite))
        else:
            cur.execute("""
                SELECT nro_atencion, nombre_usuario, nombre_taquilla, nombre_servicio, detectado_en, placa
                FROM envigado_turnos_llamados
                WHERE detectado_en::date = CURRENT_DATE
                ORDER BY detectado_en DESC
                LIMIT %s
            """, (limite,))
        filas = cur.fetchall()
        cur.close(); conn.close()
        turnos = [
            {"nro_atencion": f[0], "nombre_usuario": f[1], "taquilla": f[2],
             "servicio": f[3], "detectado_en": f[4].isoformat() + "Z", "placa": f[5] or ""}
            for f in filas
        ]
        return jsonify({"ok": True, "turnos": turnos})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-vigiladas-fechas", methods=["GET"])
def envigado_citas_vigiladas_fechas_endpoint():
    """Devuelve la lista de dias (mas reciente primero) en los que se
    dejaron citas vigilando -- para poblar el selector del historial."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT fecha_cita AS dia, COUNT(*) AS total
            FROM envigado_citas_vigiladas_historial
            GROUP BY dia
            ORDER BY dia DESC
            LIMIT 90
        """)
        filas = cur.fetchall()
        cur.close(); conn.close()
        fechas = [{"fecha": f[0].isoformat(), "total": f[1]} for f in filas]
        return jsonify({"ok": True, "fechas": fechas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-citas-vigiladas-historial", methods=["GET"])
def envigado_citas_vigiladas_historial_endpoint():
    """Devuelve las citas que se dejaron vigilando en un dia en
    particular (numero, placa, hora programada), junto con si se llego a
    detectar el llamado y con que datos (taquilla, nombre, hora real)."""
    fecha = request.args.get("fecha", "").strip()
    if not fecha:
        fecha = datetime.now(TZ_COLOMBIA).date().isoformat()
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT numero, placa, hora_cita, encontrado, taquilla, nombre_usuario, detectado_en, creado_en
            FROM envigado_citas_vigiladas_historial
            WHERE fecha_cita = %s
            ORDER BY creado_en ASC
        """, (fecha,))
        filas = cur.fetchall()
        cur.close(); conn.close()
        citas = [{
            "numero": f[0], "placa": f[1] or "", "hora_cita": f[2] or "",
            "encontrado": f[3], "taquilla": f[4] or "", "nombre_usuario": f[5] or "",
            "detectado_en": (f[6].isoformat() + "Z") if f[6] else None,
            "creado_en": f[7].isoformat() + "Z",
        } for f in filas]
        return jsonify({"ok": True, "citas": citas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/envigado-turnos-vigilados-hoy", methods=["GET"])
def envigado_turnos_vigilados_hoy_endpoint():
    """Devuelve los turnos que coincidieron con algun numero vigilado,
    capturados HOY -- persistente en base de datos, asi que sobrevive a
    que se refresque la pagina o se inicie una sesion de monitoreo nueva
    mas tarde en el mismo dia. Se renueva solo al pasar la medianoche
    (deja de aparecer, ya que el filtro es por el dia de hoy)."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT nro_atencion, nombre_usuario, nombre_taquilla, nombre_servicio, detectado_en, placa
            FROM envigado_turnos_llamados
            WHERE fue_vigilado = TRUE AND detectado_en::date = CURRENT_DATE
            ORDER BY detectado_en DESC
        """)
        filas = cur.fetchall()
        cur.close(); conn.close()
        encontrados = [
            {"nro_atencion": f[0], "nombre_usuario": f[1], "taquilla": f[2],
             "servicio": f[3], "detectado_en": f[4].isoformat() + "Z", "placa": f[5] or ""}
            for f in filas
        ]
        return jsonify({"ok": True, "encontrados": encontrados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/generar-estado-cuenta", methods=["GET"])
def generar_estado_cuenta_endpoint():
    """Genera el documento Estado de Cuenta a partir de los datos ya
    guardados (de la ultima vez que esta placa salio a paz y salvo). No
    requiere consulta en vivo ni captcha -- responde directo."""
    placa = request.args.get("placa", "").upper().strip()
    if not placa:
        return jsonify({"error": "Debes proporcionar la placa."}), 400

    if not os.path.exists(FUN_PLANTILLA):
        return jsonify({"error": "No se encontro la plantilla AppJX.xlsm en el servidor."}), 500

    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT estado_cuenta_json, lista_detalle_pagos, lista_proceso_fiscal,
                   lista_bloqueo, novedades, actualizado_en
            FROM estado_cuenta_antioquia WHERE placa = %s
        """, (placa,))
        fila = cur.fetchone()
        cur.close(); conn.close()

        if not fila:
            return jsonify({"error": "No tenemos datos de Estado de Cuenta guardados para esta placa. Debes consultar primero el impuesto departamental (y que el vehículo esté a paz y salvo) antes de poder generar este documento."}), 404

        datos = {
            "estado_veh": fila[0] or {},
            "lista_detalle_pagos": fila[1] or [],
            "lista_proceso_fiscal": fila[2] or [],
            "lista_bloqueo": fila[3] or [],
            "novedades": fila[4] or [],
            # Fecha real en que se consulto y se obtuvo este numero de
            # certificado -- NO la fecha en que se genera el documento,
            # ya que el certificado es unico de esa consulta puntual.
            "fecha_consulta": fila[5],
        }

        id_doc = str(uuid.uuid4())[:10]
        ruta_pdf_local = f"/tmp/EstadoCuenta_{placa}_{id_doc}.pdf"
        generar_estado_cuenta_pdf(datos, ruta_pdf_local)

        nombre_remoto = f"estado-cuenta/{placa}_{id_doc}.pdf"
        url = subir_a_r2(ruta_pdf_local, nombre_remoto,
                          nombre_descarga=f"EstadoCuenta_{placa}.pdf")
        os.remove(ruta_pdf_local)
        return jsonify({"ok": True, "url": url})
    except Exception as e:
        import traceback
        print(traceback.format_exc(), flush=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/registrar-mi-consulta", methods=["GET"])
def registrar_mi_consulta_endpoint():
    """Registra en el historial personal del usuario que consulto esta
    placa, SIN volver a scrapear el RUNT -- se usa cuando el dato vino del
    cache global, para que igual quede en 'Mis vehiculos consultados'."""
    user_id = request.args.get("user_id", "").strip()
    placa   = request.args.get("placa", "").upper().strip()
    cedula  = request.args.get("cedula", "").strip()
    if not user_id or not placa:
        return jsonify({"error": "Debes proporcionar user_id y placa."}), 400
    ok, error = guardar_mi_consulta(user_id, placa, cedula)
    return jsonify({"ok": ok, "error": error})


@app.route("/mis-vehiculos-runt", methods=["GET"])
def mis_vehiculos_runt():
    """Historial personal: solo las placas que ESTE usuario ha consultado
    en el RUNT antes (a diferencia de /vehiculo-runt-guardado, que es
    global para todos los usuarios). Acepta 'q' opcional para filtrar por
    placa mientras el usuario escribe (la lista puede crecer mucho)."""
    user_id = request.args.get("user_id", "").strip()
    texto   = request.args.get("q", "").upper().strip()
    if not user_id:
        return jsonify({"error": "Debes proporcionar user_id."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if texto:
            cur.execute("""
                SELECT v.placa, v.marca, v.linea, v.modelo, mc.actualizado_en, v.fuente
                FROM mis_consultas mc
                JOIN vehiculos v ON v.placa = mc.placa
                WHERE mc.user_id = %s AND v.placa LIKE %s
                ORDER BY mc.actualizado_en DESC LIMIT 8
            """, (user_id, texto + '%'))
        else:
            cur.execute("""
                SELECT v.placa, v.marca, v.linea, v.modelo, mc.actualizado_en, v.fuente
                FROM mis_consultas mc
                JOIN vehiculos v ON v.placa = mc.placa
                WHERE mc.user_id = %s
                ORDER BY mc.actualizado_en DESC LIMIT 8
            """, (user_id,))
        filas = []
        for r in cur.fetchall():
            filas.append({
                "placa": r[0], "marca": r[1], "linea": r[2], "modelo": r[3],
                "actualizado_en": (r[4] - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M") if r[4] else None,
                "fuente": r[5],
            })
        cur.close(); conn.close()
        return jsonify(filas)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/mis-vehiculos-eliminar", methods=["POST"])
def mis_vehiculos_eliminar_endpoint():
    """Elimina una placa del historial personal del usuario (tabla
    mis_consultas) -- NO borra el vehiculo de la tabla global 'vehiculos'
    (otros usuarios que tambien la hayan consultado siguen viendola en su
    propio historial, y la placa se puede volver a consultar despues sin
    problema)."""
    datos = request.get_json(silent=True) or {}
    user_id = (datos.get("user_id") or "").strip()
    placa = (datos.get("placa") or "").strip().upper()
    if not user_id or not placa:
        return jsonify({"ok": False, "error": "Faltan user_id o placa."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM mis_consultas WHERE user_id = %s AND placa = %s", (user_id, placa))
        eliminada = cur.rowcount > 0
        conn.commit()
        cur.close(); conn.close()
        if not eliminada:
            return jsonify({"ok": False, "error": "No se encontró esa placa en tu historial."}), 404
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/vehiculos-buscar", methods=["GET"])
def vehiculos_buscar():
    """Autocompletado: si hay texto ('q'), devuelve placas que empiecen con
    ese texto; si no hay texto, devuelve las mas recientes -- para que el
    desplegable muestre algo util apenas se abre, antes de escribir nada."""
    prefijo = request.args.get("q", "").upper().strip()
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if prefijo:
            cur.execute("""
                SELECT placa, marca, linea FROM vehiculos
                WHERE placa LIKE %s ORDER BY leido_en DESC LIMIT 8
            """, (prefijo + '%',))
        else:
            cur.execute("""
                SELECT placa, marca, linea FROM vehiculos
                ORDER BY leido_en DESC LIMIT 8
            """)
        filas = [{"placa": r[0], "marca": r[1], "linea": r[2]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(filas)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/vigencias-adeudadas-cache", methods=["GET"])
def vigencias_adeudadas_cache_endpoint():
    """Devuelve las vigencias que, segun la ultima consulta guardada en
    cache, estan CON_DEUDA para esta placa -- para autocompletar el
    campo de 'Vigencia(s) a descargar' en el generador de Declaraciones
    Sugeridas, sin tener que volver a consultar."""
    placa = request.args.get("placa", "").upper().strip()
    if not placa:
        return jsonify({"error": "Debes indicar la placa."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT vigencia FROM cache_impuestos_antioquia
            WHERE placa = %s AND estado = 'CON_DEUDA'
              AND (expira_en IS NULL OR expira_en >= NOW())
            ORDER BY vigencia ASC
        """, (placa,))
        vigencias = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"ok": True, "placa": placa, "vigencias": vigencias})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/vehiculo-runt-guardado", methods=["GET"])
def vehiculo_runt_guardado():
    """Trae los datos de RUNT ya guardados para una placa, sin consultar el
    RUNT de nuevo (no tiene costo de 2Captcha). Se usa para que Tramy pueda
    mostrar automaticamente lo que ya se sabe de un vehiculo, y que el
    usuario decida si esta lo bastante reciente o prefiere consultar de nuevo."""
    placa = request.args.get("placa", "").upper().strip()
    if not placa:
        return jsonify({"error": "Debes proporcionar la placa."}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM vehiculos WHERE placa = %s", (placa,))
        fila = cur.fetchone()
        if not fila:
            cur.close(); conn.close()
            return jsonify(None)
        columnas = [desc[0] for desc in cur.description]
        datos = dict(zip(columnas, fila))
        cur.close(); conn.close()
        # Convertir fechas a texto para que se puedan mostrar en JSON.
        # "leido_en" es la unica con hora (las demas son solo fecha), y se
        # guarda en UTC -- se ajusta a hora de Colombia (UTC-5) para mostrar.
        for k, v in datos.items():
            if hasattr(v, "isoformat"):
                if k == "leido_en":
                    datos[k] = (v - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M")
                else:
                    datos[k] = str(v)
        return jsonify(datos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/consultar/estado", methods=["GET"])
def consultar_estado():
    job_id = request.args.get("job_id", "").strip()
    if not job_id:
        return jsonify({"error": "Falta job_id"}), 400
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("SELECT estado, mensaje, resultado FROM consulta_jobs WHERE job_id=%s", (job_id,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row:
            return jsonify({"estado": "no_encontrado"})
        estado, mensaje, resultado = row
        resp = {"estado": estado, "mensaje": mensaje}
        if resultado:
            # Siempre devolver resultado si existe (parcial o final)
            resp["resultado"] = resultado
        if estado == "error":
            resp["error"] = mensaje
        return jsonify(resp)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
#  RETEFUENTE
# ============================================================

# Mapeo clase OCR → tabla retefuente
def _normalizar_marca(cur, marca):
    """Si la marca no existe exacta, busca por ILIKE y devuelve la más cercana."""
    cur.execute("SELECT COUNT(*) FROM retefuente_2026 WHERE marca = %s", (marca,))
    if cur.fetchone()[0] > 0:
        return marca
    cur.execute("SELECT DISTINCT marca FROM retefuente_2026 WHERE marca ILIKE %s LIMIT 1", (f"%{marca}%",))
    row = cur.fetchone()
    return row[0] if row else marca


def _es_carga(capacidad):
    """Devuelve True si la capacidad indica carga (kg) en lugar de pasajeros."""
    if not capacidad:
        return False
    cap = str(capacidad).strip().upper().replace('.','').replace(',','')
    # Si contiene KG o KILO es carga
    if 'KG' in cap or 'KILO' in cap or 'TON' in cap:
        return True
    # Si contiene PAX o PASAJERO es pasajeros
    if 'PAX' in cap or 'PASAJERO' in cap or 'PASAJ' in cap:
        return False
    # Si es número puro para CAMIONETA: >=100 = carga (kg), <100 = pasajeros
    try:
        # Limpiar puntos de miles y comas decimales
        cap_clean = re.sub(r'[^0-9]', '', cap)
        num = int(cap_clean)
        # Si el número original tiene punto como separador de miles (ej: 5.610 -> 5610)
        # ya se limpió arriba. Si era 5 pasajeros sería simplemente "5"
        return num >= 100
    except Exception:
        return False


def _tabla_retefuente(clase, carroceria='', capacidad=''):
    clase      = (clase or '').strip().upper()
    carroceria = (carroceria or '').strip().upper()
    if clase in ('AUTOMOVIL', 'AUTOMÓVIL'):                          return 'T1'
    if clase == 'CAMIONETA CARGA' or clase == 'CAMIONETA ESTACAS':   return 'T7'
    if clase == 'CAMIONETA':
        if carroceria == 'DOBLE CABINA':                             return 'T3'
        if _es_carga(capacidad):                                     return 'T7'
        return 'T2'
    if clase in ('CAMPERO',):                                         return 'T2'
    if clase in ('MOTOCICLETA', 'MOTOCARRO'):                         return 'T5'
    if clase in ('BUS', 'BUSETA', 'MICROBUS', 'MICROBÚS'):           return 'T6'
    if clase in ('CAMION', 'CAMIÓN', 'VOLQUETA', 'TRACTOCAMION'):    return 'T7'
    if clase == 'AMBULANCIA':                                         return 'T8'
    return None

def _col_anio(modelo):
    """Devuelve el nombre de columna según el modelo del vehículo."""
    try:
        anio = int(str(modelo).strip())
    except:
        return 'anio_2001_ant'
    if anio <= 2001:
        return 'anio_2001_ant'
    if anio > 2025:
        return 'anio_2025'
    return f'anio_{anio}'



@app.route("/retefuente/marcas-all", methods=["GET"])
def retefuente_marcas_all():
    """Devuelve todas las marcas para una clase (tabla o clase_bd)."""
    clase      = request.args.get("clase", "").strip().upper()
    clase_bd   = request.args.get("clase_bd", "").strip().upper()
    carroceria = request.args.get("carroceria", "").strip().upper()
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        capacidad_m = request.args.get("capacidad","")
        if clase_bd:
            if clase_bd == 'CAMIONETA':
                cur.execute("SELECT DISTINCT marca FROM retefuente_2026 WHERE clase='CAMIONETA' AND tabla='T7' ORDER BY marca")
            else:
                cur.execute("SELECT DISTINCT marca FROM retefuente_2026 WHERE clase=%s ORDER BY marca", (clase_bd,))
        elif clase:
            tabla = _tabla_retefuente(clase, carroceria, request.args.get('capacidad',''))
            if tabla:
                cur.execute("SELECT DISTINCT marca FROM retefuente_2026 WHERE tabla=%s ORDER BY marca", (tabla,))
            else:
                cur.execute("SELECT DISTINCT marca FROM retefuente_2026 ORDER BY marca")
        else:
            cur.execute("SELECT DISTINCT marca FROM retefuente_2026 ORDER BY marca")
        marcas = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"marcas": marcas})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/retefuente/lineas", methods=["GET"])
def retefuente_lineas():
    """Devuelve las lineas para una marca (y opcionalmente clase o clase_bd)."""
    marca      = request.args.get("marca", "").strip().upper()
    clase      = request.args.get("clase", "").strip().upper()
    clase_bd   = request.args.get("clase_bd", "").strip().upper()
    carroceria = request.args.get("carroceria", "").strip().upper()
    if not marca:
        return jsonify({"error": "Debes enviar marca."}), 400
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        capacidad = request.args.get('capacidad','')
        capacidad_l = request.args.get("capacidad","")
        marca = _normalizar_marca(cur, marca)
        if clase_bd:
            if clase_bd == 'CAMIONETA':
                # clase_bd=CAMIONETA significa explícitamente camioneta de carga → T7
                cur.execute("SELECT DISTINCT linea FROM retefuente_2026 WHERE marca=%s AND clase='CAMIONETA' AND tabla='T7' ORDER BY linea", (marca,))
            else:
                cur.execute("SELECT DISTINCT linea FROM retefuente_2026 WHERE marca=%s AND clase=%s ORDER BY linea", (marca, clase_bd))
        elif clase:
            tabla = _tabla_retefuente(clase, carroceria, capacidad)
            if tabla:
                cur.execute("SELECT DISTINCT linea FROM retefuente_2026 WHERE marca=%s AND tabla=%s ORDER BY linea", (marca, tabla))
            else:
                cur.execute("SELECT DISTINCT linea FROM retefuente_2026 WHERE marca=%s ORDER BY linea", (marca,))
        else:
            cur.execute("SELECT DISTINCT linea FROM retefuente_2026 WHERE marca=%s ORDER BY linea", (marca,))
        lineas = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"lineas": lineas})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/retefuente/modelos", methods=["GET"])
def retefuente_modelos():
    """Devuelve los modelos (anos) disponibles."""
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'retefuente_2026'
              AND column_name LIKE 'anio_%'
            ORDER BY column_name DESC
        """)
        modelos = [r[0].replace('anio_', '').replace('_ant', ' y anteriores') for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"modelos": modelos})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/retefuente/opciones", methods=["GET"])
def retefuente_opciones():
    """
    Devuelve todas las opciones para marca+linea+modelo+cilindraje.
    Cilindraje >= al ingresado. Incluye clase, tonelaje, pasajeros.
    """
    marca      = request.args.get("marca", "").strip().upper()
    linea      = request.args.get("linea", "").strip().upper()
    clase      = request.args.get("clase", "").strip().upper()
    carroceria = request.args.get("carroceria", "").strip().upper()
    modelo     = request.args.get("modelo", "").strip()
    cilindraje = request.args.get("cilindraje", "0").strip()

    if not marca or not modelo:
        return jsonify({"error": "Debes enviar marca y modelo."}), 400

    # Normalizar modelo — acepta año libre, "2001 y anteriores", etc.
    modelo_norm = modelo.replace(" y anteriores", "").replace("_ant", "").strip()
    try:
        anio_int = int(modelo_norm)
    except:
        return jsonify({"error": "Modelo invalido."}), 400
    col_anio = _col_anio(str(anio_int))

    try:
        # Limpiar puntos de miles y comas decimales (ej: 3.760 -> 3760)
        cil = int(re.sub(r'[^0-9]', '', cilindraje)) if cilindraje else 0
    except:
        cil = 0

    try:
        conn = get_db_conn()
        cur  = conn.cursor()

        marca  = _normalizar_marca(cur, marca)
        where  = ["marca = %s", f"{col_anio} > 0"]
        params = [marca]

        if linea:
            palabras = [p for p in linea.split() if len(p) > 2][:3]
            for p in palabras:
                where.append("linea ILIKE %s")
                params.append(f'%{p}%')

        clase_bd   = request.args.get("clase_bd", "").strip().upper()
        capacidad  = request.args.get("capacidad", "")
        if clase_bd:
            if clase_bd == 'CAMIONETA':
                where.append("clase = 'CAMIONETA'")
                where.append("tabla = 'T7'")
            else:
                where.append("clase = %s")
                params.append(clase_bd)
        elif clase:
            tabla = _tabla_retefuente(clase, carroceria, capacidad)
            if tabla:
                where.append("tabla = %s")
                params.append(tabla)

        if cil > 0:
            where.append("cilindraje >= %s")
            params.append(cil)

        sql = f"""
            SELECT marca, linea, cilindraje, tabla, {col_anio} as avaluo,
                   clase, tonelaje, pasajeros
            FROM retefuente_2026
            WHERE {' AND '.join(where)}
            ORDER BY cilindraje ASC
            LIMIT 40
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        # Si no hay resultados con filtro de linea, buscar sin él
        if not rows and linea:
            where2  = ["marca = %s", f"{col_anio} > 0"]
            params2 = [marca]
            if clase_bd:
                if clase_bd == 'CAMIONETA':
                    where2.append("clase = 'CAMIONETA'")
                    where2.append("tabla = 'T7'")
                else:
                    where2.append("clase = %s")
                    params2.append(clase_bd)
            elif clase:
                tabla = _tabla_retefuente(clase, carroceria, capacidad)
                if tabla:
                    where2.append("tabla = %s")
                    params2.append(tabla)
            if cil > 0:
                where2.append("cilindraje >= %s")
                params2.append(cil)
            cil_dist2 = f"ABS(cilindraje - {cil})," if cil > 0 else ""
            sql2 = f"""
                SELECT marca, linea, cilindraje, tabla, {col_anio} as avaluo,
                       clase, tonelaje, pasajeros
                FROM retefuente_2026
                WHERE {' AND '.join(where2)}
                ORDER BY {cil_dist2} cilindraje ASC
                LIMIT 20
            """
            cur.execute(sql2, params2)
            rows = cur.fetchall()

        # Ordenar en Python: 1) cilindraje más cercano, 2) mayor coincidencia con línea del OCR
        linea_words = [w.upper() for w in linea.split() if len(w) > 1][:5] if linea else []
        def score_row(r):
            cil_r    = r[2] or 0
            cil_dist = abs(cil_r - cil) if cil > 0 else cil_r
            lin_score = sum(1 for w in linea_words if w in (r[1] or '').upper())
            return (cil_dist, -lin_score)
        rows = sorted(rows, key=score_row)[:20]

        cur.close(); conn.close()

        TABLA_CLASE = {
            'T1':'Automóvil','T2':'Campero/Camioneta','T3':'Camioneta D.C.',
            'T4':'Eléctrico','T5':'Motocicleta','T6':'Bus/Buseta',
            'T7':'Camión/Volqueta','T8':'Ambulancia','T9':'Híbrido'
        }

        opciones = []
        for r in rows:
            op = {
                "marca":      r[0],
                "linea":      r[1],
                "cilindraje": r[2],
                "tabla":      r[3],
                "avaluo":     r[4],
                "retefuente": round(r[4] / 100) if r[4] else 0,
                "clase_veh":  r[5] or TABLA_CLASE.get(r[3], r[3]),
                "tonelaje":    float(r[6]) if r[6] else None,
                "tonelaje_kg": int(float(r[6]) * 1000) if r[6] else None,
                "pasajeros":   r[7] if r[7] else None,
            }
            opciones.append(op)

        return jsonify({"opciones": opciones})
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route("/retefuente/buscar", methods=["GET"])
def retefuente_buscar():
    """
    Busca opciones de retefuente según marca, clase, carroceria y modelo.
    Devuelve lista de opciones para que el usuario elija.
    """
    marca      = request.args.get("marca", "").strip().upper()
    linea      = request.args.get("linea", "").strip().upper()
    clase      = request.args.get("clase", "").strip().upper()
    carroceria = request.args.get("carroceria", "").strip().upper()
    modelo     = request.args.get("modelo", "").strip()
    cilindraje = request.args.get("cilindraje", "0").strip()

    if not marca or not clase or not modelo:
        return jsonify({"error": "Debes enviar marca, clase y modelo."}), 400

    tabla = _tabla_retefuente(clase, carroceria, request.args.get('capacidad',''))
    if not tabla:
        return jsonify({"error": f"Clase '{clase}' no tiene tabla de retefuente."}), 400

    col_anio = _col_anio(modelo)

    try:
        conn = get_db_conn()
        cur  = conn.cursor()

        # Cilindraje del vehículo para filtrar — mostrar desde (cilindraje - 100) hacia arriba
        try:
            cil_vehiculo = int(cilindraje) if cilindraje else 0
        except:
            cil_vehiculo = 0
        cil_min = max(0, cil_vehiculo - 50) if cil_vehiculo > 0 else 0

        # Función auxiliar para construir query con filtro cilindraje
        def query_retefuente(where_extra, params_extra, cil_desde, limite=8):
            cil_cond = f"AND cilindraje >= {cil_desde}" if cil_desde > 0 else ""
            order = f"CASE WHEN cilindraje >= {cil_vehiculo} THEN cilindraje ELSE cilindraje + 999999 END, linea"
            sql = f"""
                SELECT id, marca, linea, cilindraje, {col_anio} as avaluo
                FROM retefuente_2026
                WHERE tabla = %s AND marca = %s {where_extra} {cil_cond} AND {col_anio} > 0
                ORDER BY {order}
                LIMIT {limite}
            """
            cur.execute(sql, [tabla, marca] + params_extra)
            return cur.fetchall()

        # Buscar solo cilindraje >= vehiculo (exacto o superior)
        rows = []
        palabras = [p for p in linea.split() if len(p) > 2][:3]

        # 1. Con palabras de la línea + cilindraje >= vehiculo
        if palabras and linea:
            like_conds = " AND ".join(["linea ILIKE %s" for _ in palabras])
            rows = query_retefuente(f"AND {like_conds}", [f'%{p}%' for p in palabras], cil_vehiculo)

        # 2. Línea base estándar + cilindraje >= vehiculo
        if not rows:
            rows = query_retefuente(
                "AND (linea ILIKE %s OR linea ILIKE %s)",
                ['%LINEA BASE%', '%BASE ESTANDAR%'], cil_vehiculo
            )

        # 3. Todas las líneas de esa marca + cilindraje >= vehiculo
        if not rows:
            rows = query_retefuente("", [], cil_vehiculo)

        cur.close()
        conn.close()

        opciones = [{
            "id":         r[0],
            "marca":      r[1],
            "linea":      r[2],
            "cilindraje": r[3],
            "avaluo":     r[4],
            "retefuente": round(r[4] / 100) if r[4] else 0
        } for r in rows]

        return jsonify({
            "tabla":   tabla,
            "col_anio": col_anio,
            "opciones": opciones,
            "total":   len(opciones)
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/retefuente/marcas", methods=["GET"])
def retefuente_marcas():
    """Devuelve lista de marcas disponibles para una tabla."""
    clase      = request.args.get("clase", "").strip().upper()
    carroceria = request.args.get("carroceria", "").strip().upper()
    tabla = _tabla_retefuente(clase, carroceria, request.args.get('capacidad',''))
    if not tabla:
        return jsonify({"error": "Clase no reconocida"}), 400
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("SELECT DISTINCT marca FROM retefuente_2026 WHERE tabla=%s ORDER BY marca", (tabla,))
        marcas = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"marcas": marcas})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/tramites/filtros", methods=["GET"])
def tramites_filtros():
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        campo        = request.args.get("campo", "")
        departamento = request.args.get("departamento", "").strip().upper()
        municipio    = request.args.get("municipio", "").strip().upper()
        clase        = request.args.get("clase", "").strip().upper()
        if campo == "departamento" and municipio:
            cur.execute("SELECT DISTINCT departamento FROM tramites_transito WHERE municipio=%s ORDER BY departamento LIMIT 1", (municipio,))
        elif campo == "departamento":
            cur.execute("SELECT DISTINCT departamento FROM tramites_transito ORDER BY departamento")
        elif campo == "municipio" and departamento:
            cur.execute("SELECT DISTINCT municipio FROM tramites_transito WHERE departamento=%s ORDER BY municipio", (departamento,))
        elif campo == "municipio" and not departamento:
            cur.execute("SELECT DISTINCT municipio FROM tramites_transito ORDER BY municipio")
        elif campo == "clase" and municipio:
            cur.execute("SELECT DISTINCT clase FROM tramites_transito WHERE municipio=%s ORDER BY clase", (municipio,))
        elif campo == "tramite" and municipio and clase:
            cur.execute("SELECT DISTINCT tramite FROM tramites_transito WHERE municipio=%s AND clase=%s ORDER BY tramite", (municipio, clase))
        else:
            cur.close(); conn.close()
            return jsonify({"error": "Parametros insuficientes"}), 400
        valores = [row[0] for row in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({"valores": valores})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/tramites/precio", methods=["GET"])
def tramites_precio():
    departamento = request.args.get("departamento", "").strip().upper()
    municipio    = request.args.get("municipio", "").strip().upper()
    clase        = request.args.get("clase", "").strip().upper()
    tramite      = request.args.get("tramite", "").strip().upper()
    if not all([departamento, municipio, clase, tramite]):
        return jsonify({"error": "Debes enviar departamento, municipio, clase y tramite"}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT precio FROM tramites_transito WHERE departamento=%s AND municipio=%s AND clase=%s AND tramite=%s LIMIT 1", (departamento, municipio, clase, tramite))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return jsonify({"departamento": departamento, "municipio": municipio, "clase": clase, "tramite": tramite, "precio": row[0]})
        return jsonify({"error": "No se encontro el tramite"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============ SOAT ============

@app.route("/soat/clases", methods=["GET"])
def soat_clases():
    """Devuelve las clases de vehiculo disponibles, en orden de tarifa (1-9)."""
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT tarifa, clase_vehiculo FROM soat_tarifas
            WHERE periodo = 2026 ORDER BY tarifa
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"tarifa": r[0], "clase_vehiculo": r[1]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/soat/opciones", methods=["GET"])
def soat_opciones():
    """Devuelve las descripciones (cilindraje/toneladas/pasajeros) para una clase dada."""
    clase = request.args.get("clase", "").strip().upper()
    if not clase:
        return jsonify({"error": "Debes enviar clase"}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT codigo, descripcion FROM soat_tarifas
            WHERE periodo = 2026 AND clase_vehiculo = %s ORDER BY codigo
        """, (clase,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"codigo": r[0], "descripcion": r[1]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/soat/modelos", methods=["GET"])
def soat_modelos():
    """Devuelve los rangos de modelo disponibles (si aplica) para clase+descripcion."""
    clase       = request.args.get("clase", "").strip().upper()
    descripcion = request.args.get("descripcion", "").strip()
    if not clase or not descripcion:
        return jsonify({"error": "Debes enviar clase y descripcion"}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT codigo, modelo FROM soat_tarifas
            WHERE periodo = 2026 AND clase_vehiculo = %s AND descripcion = %s
            ORDER BY codigo
        """, (clase, descripcion))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"codigo": r[0], "modelo": r[1]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/soat/precio", methods=["GET"])
def soat_precio():
    """Precio final. 'codigo' es suficiente si ya se conoce (mas directo),
    o se puede armar con clase+descripcion+modelo."""
    codigo = request.args.get("codigo", "").strip()
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        if codigo:
            cur.execute("SELECT clase_vehiculo, descripcion, modelo, valor FROM soat_tarifas WHERE periodo=2026 AND codigo=%s", (codigo,))
        else:
            clase       = request.args.get("clase", "").strip().upper()
            descripcion = request.args.get("descripcion", "").strip()
            modelo      = request.args.get("modelo", "").strip()
            if not clase:
                return jsonify({"error": "Debes enviar codigo, o al menos clase"}), 400
            if descripcion and modelo:
                cur.execute("SELECT clase_vehiculo, descripcion, modelo, valor FROM soat_tarifas WHERE periodo=2026 AND clase_vehiculo=%s AND descripcion=%s AND modelo=%s", (clase, descripcion, modelo))
            elif descripcion:
                cur.execute("SELECT clase_vehiculo, descripcion, modelo, valor FROM soat_tarifas WHERE periodo=2026 AND clase_vehiculo=%s AND descripcion=%s", (clase, descripcion))
            else:
                cur.execute("SELECT clase_vehiculo, descripcion, modelo, valor FROM soat_tarifas WHERE periodo=2026 AND clase_vehiculo=%s", (clase,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return jsonify({"clase_vehiculo": row[0], "descripcion": row[1], "modelo": row[2], "valor": row[3]})
        return jsonify({"error": "No se encontro tarifa SOAT para esos criterios"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============ TECNOMECANICA ============

@app.route("/tecnomecanica/categorias", methods=["GET"])
def tecnomecanica_categorias():
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT categoria, valor FROM tecnomecanica_tarifas WHERE periodo = 2026 ORDER BY id")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"categoria": r[0], "valor": r[1]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/tecnomecanica/precio", methods=["GET"])
def tecnomecanica_precio():
    categoria = request.args.get("categoria", "").strip()
    if not categoria:
        return jsonify({"error": "Debes enviar categoria"}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT valor FROM tecnomecanica_tarifas WHERE periodo = 2026 AND categoria = %s", (categoria,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return jsonify({"categoria": categoria, "valor": row[0]})
        return jsonify({"error": "No se encontro esa categoria"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============ COMPARENDOS ============

@app.route("/comparendos/buscar", methods=["GET"])
def comparendos_buscar():
    """Busca por codigo exacto o por palabra clave dentro de la descripcion."""
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify([])
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT codigo, descripcion, valor, valor_desc_50, valor_desc_25 FROM comparendos_tarifas
            WHERE periodo = 2026 AND (codigo ILIKE %s OR descripcion ILIKE %s)
            ORDER BY codigo LIMIT 20
        """, (f"{q}%", f"%{q}%"))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return jsonify([{"codigo": r[0], "descripcion": r[1], "valor": r[2], "valor_desc_50": r[3], "valor_desc_25": r[4]} for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/comparendos/precio", methods=["GET"])
def comparendos_precio():
    codigo = request.args.get("codigo", "").strip().upper()
    if not codigo:
        return jsonify({"error": "Debes enviar codigo"}), 400
    try:
        conn = get_db_conn()
        cur = conn.cursor()
        cur.execute("SELECT descripcion, valor, valor_desc_50, valor_desc_25 FROM comparendos_tarifas WHERE periodo = 2026 AND codigo = %s", (codigo,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return jsonify({"codigo": codigo, "descripcion": row[0], "valor": row[1], "valor_desc_50": row[2], "valor_desc_25": row[3]})
        return jsonify({"error": "No se encontro esa infraccion"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reportar", methods=["POST"])
def reportar():
    try:
        data      = request.get_json()
        tipo      = data.get("tipo", "").strip()
        comentario = data.get("comentario", "").strip()
        placa     = data.get("placa", "").strip().upper()
        municipio = data.get("municipio", "").strip().upper()
        pagina    = data.get("pagina", "").strip()
        if not tipo:
            return jsonify({"ok": False, "error": "Tipo requerido"}), 400
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO reportes_usuarios (tipo, comentario, placa, municipio, pagina)
            VALUES (%s, %s, %s, %s, %s)
        """, (tipo, comentario, placa, municipio, pagina))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/reportar/lista", methods=["GET"])
def reportar_lista():
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT id, tipo, comentario, placa, municipio, pagina, creado_en
            FROM reportes_usuarios
            ORDER BY creado_en DESC
            LIMIT 100
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"reportes": [{
            "id": r[0], "tipo": r[1], "comentario": r[2],
            "placa": r[3], "municipio": r[4], "pagina": r[5],
            "fecha": str(r[6])
        } for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/reportar/eliminar/<int:reporte_id>", methods=["DELETE"])
def reportar_eliminar(reporte_id):
    try:
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("DELETE FROM reportes_usuarios WHERE id = %s", (reporte_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500





@app.route("/ocr-tarjeta", methods=["POST"])
def ocr_tarjeta():
    try:
        data = request.get_json()
        if not data or "imagen" not in data:
            return jsonify({"error": "No se recibio imagen"}), 400

        def preparar_archivo(img_data):
            es_pdf = "data:application/pdf" in img_data
            media_type = "application/pdf" if es_pdf else "image/jpeg"
            if not es_pdf:
                if "data:image/png" in img_data:
                    media_type = "image/png"
                elif "data:image/webp" in img_data:
                    media_type = "image/webp"
            if "," in img_data:
                img_data = img_data.split(",")[1]
            if es_pdf:
                return {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": img_data}}
            return {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_data}}

        # Puede venir un solo archivo (imagen o PDF con ambas caras), o dos
        # archivos separados (ej: foto de la cara frontal + foto de la cara
        # trasera, subidas por separado). Ambos se envian juntos a Claude.
        archivos_content = [preparar_archivo(data["imagen"])]
        if data.get("imagen2"):
            archivos_content.append(preparar_archivo(data["imagen2"]))

        anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
        if not anthropic_key:
            return jsonify({"error": "API Key de Anthropic no configurada"}), 500

        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": anthropic_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-opus-4-5", "max_tokens": 600, "messages": [{"role": "user", "content": archivos_content + [
                {"type": "text", "text": "Eres un experto en leer tarjetas/licencias de tránsito y documentos de impuestos vehiculares de Colombia. Puedes recibir UNO o DOS archivos (imagenes y/o PDFs): si recibes dos, son la cara frontal y la cara trasera de la MISMA tarjeta, subidas por separado — analiza ambos como si fueran las dos caras de un mismo documento. Cada archivo puede incluir SOLO la cara frontal, SOLO la cara trasera, o AMBAS caras. Los archivos pueden estar rotados o de lado (esto es MUY comun en fotos, especialmente de la cara trasera) — gira mentalmente el texto para leerlo sin importar su orientacion. Analiza TODOS los caracteres con mucho cuidado especialmente los numeros. Extrae: 1. PLACA (exactamente 3 letras + 3 numeros, verifica cada caracter) 2. MARCA del vehiculo 3. LINEA del vehiculo 4. MODELO (anno 4 digitos) 5. CLASE (automovil, motocicleta, campero, camioneta, etc) 6. SERVICIO (particular, publico, oficial) 7. CAPACIDAD: el numero que aparece junto a la etiqueta 'CAPACIDAD Kg/PSJ' (o similar, puede venir abreviada distinto) -- es el numero de pasajeros o de carga en Kg segun el tipo de vehiculo 8. CILINDRADA (numero en cc) 9. TIPO_DOCUMENTO (uno de: C.C, NIT, P.P.T, T.I, R.C - aparece debajo de IDENTIFICACION al lado izquierdo del numero) 10. CEDULA (numero de identificacion, verifica TODOS los digitos uno por uno, no omitas ninguno) 11. NOMBRE_COMPLETO del propietario: la tarjeta suele tener dos campos separados, 'APELLIDOS' y 'NOMBRES' -- combina AMBOS en un solo texto (ej. si dice Apellidos: GARCIA PEREZ y Nombres: JUAN CARLOS, responde 'GARCIA PEREZ JUAN CARLOS'). Si solo encuentras un campo con el nombre completo ya junto, cópialo tal cual. Nunca dejes fuera los nombres de pila, solo los apellidos no es una respuesta completa. 12. MUNICIPIO: este dato SOLO aparece en la cara TRASERA. Hay dos formatos posibles, busca cualquiera de los dos: (a) un campo llamado 'MUNICIPIO DE MATRICULA' (o similar, ej. 'Municipio Matricula') — en este caso el valor de ese campo ES DIRECTAMENTE el nombre del municipio, cópialo tal cual. (b) un campo llamado 'ORGANISMO DE TRANSITO', que casi siempre viene ABREVIADO de forma variable, por ejemplo 'STRIA TTEYTTO ENVIGADO' o 'STRIA DE TTOYTTE MEDELLIN' — ambos significan 'Secretaria de Transito y Transporte de <MUNICIPIO>'. El patron general es: unas siglas abreviadas de 'Secretaria de Transito y Transporte' seguidas del NOMBRE DEL MUNICIPIO al final del texto. En este caso extrae SOLAMENTE el nombre del municipio (la ultima palabra o palabras), sin ninguna de las siglas ni abreviaturas que la preceden. Si la cara trasera no esta visible en ninguno de los archivos, deja este campo vacio, NO lo inventes ni lo asumas. 13. LIMITACION_PROPIEDAD: este dato tambien esta SOLO en la cara TRASERA. SOLO debes responder algo distinto de vacio si encuentras EXACTAMENTE uno de estos dos patrones — si no encuentras ninguno de los dos (por ejemplo la cara trasera no esta visible, o esta borrosa, o no aparece ninguno de estos campos), deja el valor VACIO, NUNCA asumas ni adivines: (a) un campo llamado 'LIMITACION A LA PROPIEDAD' — debajo de ese titulo puede aparecer una serie de ASTERISCOS (ej: '******'), lo que significa que el vehiculo NO tiene ningun gravamen (responde 'NINGUNA', nunca copies los asteriscos tal cual); o puede aparecer el nombre de una persona natural o juridica (ej: 'PRENDA - BANCO FINANDINA', 'PRENDA - BANCO DE OCCIDENTE'), lo que significa que SI tiene gravamen (copia el valor completo tal cual aparece). (b) un campo llamado 'GRAVAMENES A LA PROPIEDAD' con una respuesta directa 'SI' o 'NO' — si dice 'SI' responde 'SI' (tiene gravamen), si dice 'NO' responde 'NINGUNA' (no tiene gravamen). Recuerda: si no ves con claridad ninguno de estos dos campos, deja limitacion_propiedad vacio — es preferible dejarlo vacio a arriesgarte a decir que no tiene gravamen cuando en realidad no pudiste verificarlo. 14. ES_DECLARACION_ANTIOQUIA: true SOLO si alguno de los archivos es un documento titulado 'DECLARACION SUGERIDA DE IMPUESTOS SOBRE VEHICULOS AUTOMOTORES' emitido por la Gobernacion de Antioquia. Si es asi, todos los campos 1-13 de arriba se extraen IGUAL (los mismos datos existen en este documento, solo que organizados en casillas tipo D.1 PLACA, D.2 MARCA, C.1 NOMBRE, C.3 APELLIDOS, etc en vez del formato de tarjeta -- usa tu criterio para ubicarlos ahi). Ademas: (a) el SERVICIO de este vehiculo SIEMPRE es 'PARTICULAR' (este tipo de documento solo se emite para vehiculos particulares), pon SERVICIO='PARTICULAR' sin importar lo que digan otros campos. (b) DECLARACION_VIGENCIA: el año de la vigencia que se esta declarando/pagando, normalmente aparece en la zona superior izquierda del documento (ej: 'Vigencia 2026' o similar — copia solo el numero de 4 digitos del año). (c) DECLARACION_PAGADO: true si el documento tiene un SELLO DE BANCO O ENTIDAD FINANCIERA que indique que fue pagado (busca sellos, timbres, o textos como 'PAGADO', 'RECIBIDO', nombre de un banco estampado, codigos de transaccion bancaria, etc). false si no hay ningun sello o indicio de pago. (d) DECLARACION_AVALUO: el valor de la casilla '1. AVALUO COMERCIAL DEL VEHICULO' (solo el numero, sin simbolos de moneda ni puntos de miles). (e) DECLARACION_CAJA: la sigla junto a la casilla 'CAJA' (ej MT, AT, CVT). (f) DECLARACION_TRACCION: el valor junto a 'TRACCION' o 'COMBUSTION/TRACCION' (ej 4X2, 4X4). (g) DECLARACION_CELULAR: casilla C.4 (CELULAR). (h) DECLARACION_EMAIL: casilla C.6 (E-MAIL). (i) DECLARACION_DIRECCION: casilla C.7 (DIRECCION). (j) DECLARACION_IMPUESTO: el valor de la casilla '2. IMPUESTO SOBRE VEHICULOS AUTOMOTORES' (solo el numero). (k) DECLARACION_SANCIONES: el valor de la casilla '3. MAS SANCIONES' (solo el numero). (l) DECLARACION_TOTAL_PAGAR: el valor de la casilla '11. TOTAL A PAGAR' (solo el numero). Si alguno de estos campos adicionales (e-l) no es visible, deja ese campo vacio, NUNCA lo inventes. Si el archivo NO es este tipo de documento, deja ES_DECLARACION_ANTIOQUIA en false y los demas campos de declaracion vacios. Responde SOLO en JSON sin explicaciones: {\"placa\": \"\", \"marca\": \"\", \"linea\": \"\", \"modelo\": \"\", \"clase\": \"\", \"servicio\": \"\", \"capacidad\": \"\", \"cilindrada\": \"\", \"carroceria\": \"\", \"tipo_documento\": \"\", \"cedula\": \"\", \"apellidos\": \"\", \"municipio\": \"\", \"limitacion_propiedad\": \"\", \"es_declaracion_antioquia\": false, \"declaracion_vigencia\": \"\", \"declaracion_pagado\": false, \"declaracion_avaluo\": \"\", \"declaracion_caja\": \"\", \"declaracion_traccion\": \"\", \"declaracion_celular\": \"\", \"declaracion_email\": \"\", \"declaracion_direccion\": \"\", \"declaracion_impuesto\": \"\", \"declaracion_sanciones\": \"\", \"declaracion_total_pagar\": \"\"}"}
            ]}]},
            timeout=120
        )
        if response.status_code != 200:
            return jsonify({"error": f"Error Claude API: {response.status_code}"}), 500
        resp_data = response.json()
        texto = resp_data["content"][0]["text"].strip()
        import json as json_lib, re as re_module
        texto_clean = texto.replace("```json", "").replace("```", "").strip()
        json_match = re_module.search(r'\{[^{}]*\}', texto_clean, re_module.DOTALL)
        if not json_match:
            return jsonify({"error": "No se pudo parsear respuesta de Claude"}), 500
        resultado = json_lib.loads(json_match.group())
        placa                = resultado.get("placa", "").upper().replace(" ", "").replace("-", "")
        marca                = resultado.get("marca", "").upper().strip()
        linea                = resultado.get("linea", "").upper().strip()
        modelo               = resultado.get("modelo", "").strip()
        clase                = resultado.get("clase", "").upper().strip()
        servicio             = resultado.get("servicio", "").upper().strip()
        capacidad            = resultado.get("capacidad", "").strip()
        cilindrada           = resultado.get("cilindrada", "").strip()
        carroceria           = resultado.get("carroceria", "").upper().strip()
        tipo_documento       = resultado.get("tipo_documento", "").upper().strip()
        cedula               = resultado.get("cedula", "").strip()
        apellidos            = resultado.get("apellidos", "").upper().strip()
        municipio            = resultado.get("municipio", "").upper().strip()
        limitacion_propiedad = resultado.get("limitacion_propiedad", "").strip()

        # Declaración Sugerida de Impuestos sobre Vehículos Automotores (Gobernación de Antioquia)
        es_declaracion_antioquia = bool(resultado.get("es_declaracion_antioquia"))
        paz_salvo_detectado = False
        declaracion_extra = {}
        if es_declaracion_antioquia:
            # 1. Este tipo de documento solo se emite para vehiculos particulares
            servicio = "PARTICULAR"

            # 2. Si tiene sello de pago Y la vigencia es el año actual -> paz y salvo
            declaracion_pagado = bool(resultado.get("declaracion_pagado"))
            declaracion_vigencia_raw = str(resultado.get("declaracion_vigencia", "")).strip()
            declaracion_avaluo_raw = str(resultado.get("declaracion_avaluo", "")).strip()
            anio_actual = datetime.now().year  # siempre el año actual real, nunca fijo

            try:
                declaracion_vigencia = int(re.sub(r"[^\d]", "", declaracion_vigencia_raw)) if declaracion_vigencia_raw else 0
            except ValueError:
                declaracion_vigencia = 0
            try:
                declaracion_avaluo = int(re.sub(r"[^\d]", "", declaracion_avaluo_raw)) if declaracion_avaluo_raw else 0
            except ValueError:
                declaracion_avaluo = 0

            if declaracion_pagado and declaracion_vigencia == anio_actual and placa and declaracion_avaluo > 0:
                cache_antioquia_guardar_paz_salvo(placa, declaracion_avaluo, {})
                paz_salvo_detectado = True

            # 3. Campos adicionales que solo trae este tipo de documento -- se
            # devuelven al frontend por si se quieren usar para prellenar
            # otros formularios (ej. la Declaracion Manual).
            def _num_declaracion(campo):
                crudo = str(resultado.get(campo, "")).strip()
                try:
                    return int(re.sub(r"[^\d]", "", crudo)) if crudo else 0
                except ValueError:
                    return 0

            declaracion_extra = {
                "vigencia": declaracion_vigencia,
                "pagado": declaracion_pagado,
                "avaluo": declaracion_avaluo,
                "caja": str(resultado.get("declaracion_caja", "")).upper().strip(),
                "traccion": str(resultado.get("declaracion_traccion", "")).upper().strip(),
                "celular": str(resultado.get("declaracion_celular", "")).strip(),
                "email": str(resultado.get("declaracion_email", "")).strip(),
                "direccion": str(resultado.get("declaracion_direccion", "")).strip(),
                "impuesto": _num_declaracion("declaracion_impuesto"),
                "sanciones": _num_declaracion("declaracion_sanciones"),
                "total_pagar": _num_declaracion("declaracion_total_pagar"),
            }

        return jsonify({"placa": placa, "marca": marca, "linea": linea, "modelo": modelo, "clase": clase, "servicio": servicio, "capacidad": capacidad, "cilindrada": cilindrada, "carroceria": carroceria, "tipo_documento": tipo_documento, "cedula": cedula, "apellidos": apellidos, "municipio": municipio, "limitacion_propiedad": limitacion_propiedad, "paz_salvo_antioquia_detectado": paz_salvo_detectado, "es_declaracion_antioquia": es_declaracion_antioquia, "declaracion": declaracion_extra, "desde_cache": False})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/ocr-runt-texto", methods=["POST"])
def ocr_runt_texto():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No se recibieron datos"}), 400
        texto_placa  = (data.get("texto_placa") or "").strip()
        texto_cedula = (data.get("texto_cedula") or "").strip()
        if not texto_placa and not texto_cedula:
            return jsonify({"error": "Debes pegar al menos un texto"}), 400

        anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
        if not anthropic_key:
            return jsonify({"error": "API Key de Anthropic no configurada"}), 500

        texto_combinado = ""
        if texto_placa:
            texto_combinado += "=== TEXTO COPIADO DEL RUNT — CONSULTA POR PLACA (datos del VEHICULO) ===\n" + texto_placa + "\n\n"
        if texto_cedula:
            texto_combinado += "=== TEXTO COPIADO DEL RUNT — CONSULTA POR CEDULA (datos del PROPIETARIO/CONDUCTOR) ===\n" + texto_cedula

        prompt = (
            "Eres un experto en interpretar texto copiado y pegado directamente del portal RUNT "
            "(Registro Unico Nacional de Transito de Colombia). Te voy a dar el texto plano que un "
            "usuario copio de la pagina de resultados del RUNT — puede ser de una consulta por PLACA "
            "(trae datos del vehiculo), por CEDULA (trae datos del propietario/conductor), o ambas. "
            "El texto puede venir desordenado, con saltos de linea irregulares, o con texto de menus/"
            "botones de la pagina mezclado — ignora ese ruido y concentrate en los datos reales. "
            "Extrae los siguientes datos si estan presentes en cualquiera de los dos textos: "
            "1. PLACA 2. MARCA 3. LINEA 4. MODELO (año) 5. CLASE 6. SERVICIO 7. CAPACIDAD "
            "8. CILINDRADA (cc) 9. TIPO_DOCUMENTO (C.C, NIT, C.E, T.I, R.C, P.P.T) "
            "10. CEDULA (numero de identificacion del propietario) 11. APELLIDOS (y nombres) del "
            "propietario 12. MUNICIPIO: hay dos formatos posibles, busca cualquiera de los dos: "
            "(a) un campo llamado 'MUNICIPIO DE MATRICULA' (o similar, ej. 'Municipio Matricula') — en "
            "este caso el valor de ese campo ES DIRECTAMENTE el nombre del municipio, cópialo tal cual. "
            "(b) un campo llamado 'Organismo de Transito' u 'Organismo de Transito Matricula', casi "
            "siempre ABREVIADO de forma variable, por ejemplo "
            "'STRIA TTEYTTO ENVIGADO' o 'STRIA DE TTOYTTE MEDELLIN' — ambos significan 'Secretaria de "
            "Transito y Transporte de <MUNICIPIO>'. El patron general es: unas siglas abreviadas de "
            "'Secretaria de Transito y Transporte' seguidas del NOMBRE DEL MUNICIPIO al final del texto. "
            "En este caso extrae SOLAMENTE el nombre del municipio (la ultima palabra o palabras), sin "
            "ninguna de las siglas ni abreviaturas que la preceden (nunca dejes 'STRIA', 'TTEYTTO', "
            "'SRIA', 'SECRETARIA' ni similares como parte del valor). Si no aparece ese dato en el "
            "texto, deja el campo vacio, no lo inventes. 13. LIMITACION_PROPIEDAD (gravamenes, prenda "
            "a favor de alguna entidad, o 'NINGUNA' si no tiene). SOLO debes responder algo distinto de "
            "vacio si encuentras EXACTAMENTE uno de estos dos patrones — si no encuentras ninguno, deja "
            "el valor VACIO, nunca asumas ni adivines: (a) un campo 'Limitacion a la Propiedad' — si "
            "aparece con una serie de asteriscos como '******', eso significa 'NINGUNA' (nunca copies "
            "los asteriscos tal cual); si aparece el nombre de una persona o entidad, copia ese valor "
            "tal cual (tiene gravamen). (b) un campo 'Gravamenes a la Propiedad' con respuesta directa "
            "SI o NO — si dice SI responde 'SI' (tiene gravamen), si dice NO responde 'NINGUNA' (no "
            "tiene gravamen). Es preferible dejarlo vacio a arriesgarte a decir que no tiene gravamen "
            "cuando en realidad no pudiste verificarlo. Si un dato no aparece en el texto, "
            "deja ese campo vacio, NO lo inventes ni lo asumas. Responde SOLO en JSON sin explicaciones: "
            "{\"placa\": \"\", \"marca\": \"\", \"linea\": \"\", \"modelo\": \"\", \"clase\": \"\", "
            "\"servicio\": \"\", \"capacidad\": \"\", \"cilindrada\": \"\", \"carroceria\": \"\", "
            "\"tipo_documento\": \"\", \"cedula\": \"\", \"apellidos\": \"\", \"municipio\": \"\", "
            "\"limitacion_propiedad\": \"\"}\n\nTEXTO A ANALIZAR:\n" + texto_combinado
        )

        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": anthropic_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-opus-4-5", "max_tokens": 600, "messages": [{"role": "user", "content": prompt}]},
            timeout=90
        )
        if response.status_code != 200:
            return jsonify({"error": f"Error Claude API: {response.status_code}"}), 500
        resp_data = response.json()
        texto_resp = resp_data["content"][0]["text"].strip()
        import json as json_lib, re as re_module
        texto_clean = texto_resp.replace("```json", "").replace("```", "").strip()
        json_match = re_module.search(r'\{[^{}]*\}', texto_clean, re_module.DOTALL)
        if not json_match:
            return jsonify({"error": "No se pudo parsear respuesta"}), 500
        resultado = json_lib.loads(json_match.group())
        placa                = resultado.get("placa", "").upper().replace(" ", "").replace("-", "")
        marca                = resultado.get("marca", "").upper().strip()
        linea                = resultado.get("linea", "").upper().strip()
        modelo               = resultado.get("modelo", "").strip()
        clase                = resultado.get("clase", "").upper().strip()
        servicio             = resultado.get("servicio", "").upper().strip()
        capacidad            = resultado.get("capacidad", "").strip()
        cilindrada           = resultado.get("cilindrada", "").strip()
        carroceria           = resultado.get("carroceria", "").upper().strip()
        tipo_documento       = resultado.get("tipo_documento", "").upper().strip()
        cedula               = resultado.get("cedula", "").strip()
        apellidos            = resultado.get("apellidos", "").upper().strip()
        municipio            = resultado.get("municipio", "").upper().strip()
        limitacion_propiedad = resultado.get("limitacion_propiedad", "").strip()
        return jsonify({"placa": placa, "marca": marca, "linea": linea, "modelo": modelo, "clase": clase, "servicio": servicio, "capacidad": capacidad, "cilindrada": cilindrada, "carroceria": carroceria, "tipo_documento": tipo_documento, "cedula": cedula, "apellidos": apellidos, "municipio": municipio, "limitacion_propiedad": limitacion_propiedad})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/ocr-guardar-municipio", methods=["POST"])
def ocr_guardar_municipio():
    try:
        data      = request.get_json()
        placa     = data.get("placa", "").upper().strip()
        municipio = data.get("municipio", "").upper().strip()
        if not placa or not municipio:
            return jsonify({"ok": False}), 400
        conn = get_db_conn()
        cur  = conn.cursor()
        cur.execute("UPDATE cache_tarjetas SET municipio=%s, actualizado_en=NOW() WHERE placa=%s", (municipio, placa))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ============================================================
# SIBGA — Avalúos motos bajo cilindraje (≤125cc)
# Datos en tabla retefuente_bajocilindraje
# ============================================================

SIBGA_PERIODO = 2024

def _sibga_col_anio(modelo):
    try:
        anio = int(str(modelo).strip())
    except:
        return "anio_2001_ant"
    if anio <= 2001: return "anio_2001_ant"
    if anio > 2024:  return "anio_2024"
    return f"anio_{anio}"


@app.route("/sibga/marcas", methods=["GET"])
def sibga_marcas():
    """Marcas de motos bajo cilindraje desde BD."""
    try:
        conn = get_db_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT marca FROM retefuente_bajocilindraje
            WHERE cilindraje <= 125 AND cilindraje > 0
            ORDER BY marca
        """)
        rows = cur.fetchall(); cur.close(); conn.close()
        return jsonify({"marcas": [r[0] for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/sibga/lineas", methods=["GET"])
def sibga_lineas():
    """Líneas de una marca desde BD."""
    marca = request.args.get("marca", "").upper().strip()
    if not marca:
        return jsonify({"error": "marca requerida"}), 400
    try:
        conn = get_db_conn(); cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT linea_id, linea FROM retefuente_bajocilindraje
            WHERE marca=%s AND cilindraje <= 125 AND cilindraje > 0
            ORDER BY linea
        """, (marca,))
        rows = cur.fetchall(); cur.close(); conn.close()
        return jsonify({"lineas": [{"id": r[0], "nombre": r[1]} for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/sibga/avaluo", methods=["GET"])
def sibga_avaluo():
    """Avalúo de moto bajo cilindraje desde BD."""
    linea_id = request.args.get("linea_id", type=int)
    modelo   = request.args.get("modelo", type=int)
    if not linea_id or not modelo:
        return jsonify({"error": "linea_id y modelo requeridos"}), 400

    col_anio = _sibga_col_anio(modelo)

    try:
        conn = get_db_conn(); cur = conn.cursor()
        cur.execute(f"""
            SELECT {col_anio}, linea, cilindraje, marca
            FROM retefuente_bajocilindraje
            WHERE linea_id=%s
        """, (linea_id,))
        row = cur.fetchone(); cur.close(); conn.close()

        if not row or not row[0]:
            return jsonify({"error": "No se encontró avalúo para esa línea y modelo"}), 404

        return jsonify({
            "avaluo":     row[0],
            "linea":      row[1],
            "cilindraje": row[2],
            "marca":      row[3],
            "modelo":     modelo,
            "periodo":    SIBGA_PERIODO,
            "fuente":     "bd"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/sibga/opciones", methods=["GET"])
def sibga_opciones():
    """Devuelve opciones de avalúo para motos bajo cilindraje — igual que retefuente/opciones."""
    marca  = request.args.get("marca", "").strip().upper()
    linea  = request.args.get("linea", "").strip().upper()
    modelo = request.args.get("modelo", type=int, default=2020)
    if not marca:
        return jsonify({"error": "marca requerida"}), 400

    col_anio = _sibga_col_anio(modelo)

    try:
        conn = get_db_conn(); cur = conn.cursor()

        # Buscar por marca + palabras de la línea
        cil_sibga = int(re.sub(r'[^0-9]', '', request.args.get('cilindraje','0') or '0') or 0)
        where  = ["marca = %s", f"{col_anio} > 0", "cilindraje <= 125", "cilindraje > 0"]
        params = [marca]

        if linea:
            # Separar letras y números pegados: AK125 -> AK 125
            linea_sep = re.sub(r'([A-Za-z])(\d)', r'\1 \2', linea)
            linea_sep = re.sub(r'(\d)([A-Za-z])', r'\1 \2', linea_sep)
            palabras = [p for p in linea_sep.split() if len(p) > 1][:5]
            if palabras:
                or_conds = []
                for p in palabras:
                    or_conds.append("REPLACE(linea, ' ', '') ILIKE %s")
                    params.append(f'%{p}%')
                where.append("(" + " OR ".join(or_conds) + ")")

        sql = f"""
            SELECT linea_id, linea, cilindraje, {col_anio} as avaluo
            FROM retefuente_bajocilindraje
            WHERE {' AND '.join(where)}
            ORDER BY cilindraje ASC
            LIMIT 200
        """
        cur.execute(sql, params)
        rows = cur.fetchall()

        # Si no hay resultados con filtro de línea, buscar solo por marca
        if not rows and linea:
            cur.execute(f"""
                SELECT linea_id, linea, cilindraje, {col_anio} as avaluo
                FROM retefuente_bajocilindraje
                WHERE marca=%s AND {col_anio} > 0 AND cilindraje <= 125 AND cilindraje > 0
                ORDER BY cilindraje ASC
                LIMIT 40
            """, (marca,))
            rows = cur.fetchall()

        # Ordenar: cilindraje más cercano primero, luego mayor coincidencia con línea
        if linea:
            linea_sep2 = re.sub(r'([A-Za-z])(\d)', r'\1 \2', linea)
            linea_sep2 = re.sub(r'(\d)([A-Za-z])', r'\1 \2', linea_sep2)
            linea_words_s = [w.upper() for w in linea_sep2.split() if len(w) > 1][:5]
        else:
            linea_words_s = []
        def score_sibga(r):
            cil_r = r[2] or 0
            cil_dist = abs(cil_r - cil_sibga) if cil_sibga > 0 else cil_r
            linea_db_sin_esp = (r[1] or '').upper().replace(' ', '')
            lin_score = sum(1 for w in linea_words_s if w in linea_db_sin_esp)
            return (cil_dist, -lin_score)
        rows = sorted(rows, key=score_sibga)[:20]

        cur.close(); conn.close()

        opciones = [{
            "linea_id":   r[0],
            "linea":      r[1],
            "cilindraje": r[2],
            "avaluo":     r[3],
            "retefuente": round(r[3] / 100) if r[3] else 0,
        } for r in rows]

        return jsonify({"opciones": opciones})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
